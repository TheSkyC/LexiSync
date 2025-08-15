# Copyright (c) 2025, TheSkyC
# SPDX-License-Identifier: Apache-2.0

import os
import threading
from copy import deepcopy
from rapidfuzz import fuzz
from openpyxl import Workbook, load_workbook
import polib
import weakref
import traceback
import re
import logging
logger = logging.getLogger(__name__)

from PySide6.QtWidgets import (
    QMainWindow, QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLineEdit, QTextEdit, QCheckBox, QFileDialog,
    QMessageBox, QInputDialog, QStatusBar, QProgressBar,
    QMenu, QTableView, QHeaderView, QDockWidget, QLabel,
    QAbstractItemView, QFrame, QComboBox, QListWidgetItem
)
from PySide6.QtCore import (
    Qt, QModelIndex, Signal, QObject, QTimer, QByteArray, QEvent,
    QRunnable, QThreadPool, QItemSelectionModel, QSize, QDir
)
from PySide6.QtGui import QAction, QKeySequence, QFont, QPalette, QColor, QSyntaxHighlighter, QTextCharFormat

from models.translatable_string import TranslatableString
from models.translatable_strings_model import TranslatableStringsModel, TranslatableStringsProxyModel
from plugins.plugin_manager import PluginManager
from ui_components.comment_status_panel import CommentStatusPanel
from ui_components.context_panel import ContextPanel
from ui_components.custom_cell_delegate import CustomCellDelegate
from ui_components.custom_table_view import CustomTableView
from ui_components.details_panel import DetailsPanel
from ui_components.elided_label import ElidedLabel
from ui_components.file_explorer_panel import FileExplorerPanel
from ui_components.tm_panel import TMPanel
from ui_components.glossary_panel import GlossaryPanel


from dialogs.font_settings_dialog import FontSettingsDialog
from dialogs.keybinding_dialog import KeybindingDialog
from dialogs.language_pair_dialog import LanguagePairDialog
from dialogs.new_project_dialog import NewProjectDialog
from dialogs.pot_drop_dialog import POTDropDialog
from dialogs.extraction_pattern_dialog import ExtractionPatternManagerDialog
from dialogs.prompt_manager_dialog import PromptManagerDialog
from dialogs.diff_dialog import DiffDialog
from dialogs.statistics_dialog import StatisticsDialog
from dialogs.settings_dialog import SettingsDialog
from dialogs.search_dialog import AdvancedSearchDialog

from services import language_service
from services import export_service, po_file_service
from services.ai_translator import AITranslator
from services.code_file_service import extract_translatable_strings, save_translated_code
from services.project_service import create_project, load_project, save_project
from services.prompt_service import generate_prompt_from_structure
from services.validation_service import run_validation_on_all, placeholder_regex
from services.expansion_ratio_service import ExpansionRatioService
from services.tm_service import TMService
from services.glossary_service import GlossaryService
from services.glossary_worker import GlossaryAnalysisWorker

from utils import config_manager
from utils.constants import *
from utils.enums import WarningType
from utils.localization import _, lang_manager
from utils.text_utils import get_linguistic_length
from utils.path_utils import get_app_data_path

try:
    import requests
except ImportError:
    requests = None
    logger.warning("提示: requests 未找到, AI翻译功能不可用。pip install requests")

class GlossaryHighlighter(QSyntaxHighlighter):
    def __init__(self, parent, matches_list):
        super().__init__(parent)
        self.highlight_format = QTextCharFormat()
        self.highlight_format.setUnderlineColor(QColor("teal"))
        self.highlight_format.setUnderlineStyle(QTextCharFormat.WaveUnderline)
        self.rules = []
        if matches_list:
            keywords = [re.escape(match['source']) for match in matches_list]
            if keywords:
                pattern = r'\b(' + '|'.join(keywords) + r')\b'
                self.rules.append((re.compile(pattern, re.IGNORECASE), self.highlight_format))

    def highlightBlock(self, text):
        for pattern, format in self.rules:
            for match in pattern.finditer(text):
                start, end = match.span()
                self.setFormat(start, end - start, format)

class AITranslationWorker(QRunnable):
    def __init__(self, app_instance, ts_id, original_text, target_language, context_dict, plugin_placeholders, is_batch_item):
        super().__init__()
        self.app_ref = weakref.ref(app_instance)
        self.ts_id = ts_id
        self.original_text = original_text
        self.target_language = target_language
        self.context_dict = context_dict
        self.plugin_placeholders = plugin_placeholders
        self.is_batch_item = is_batch_item

    def run(self):
        app = self.app_ref()
        if not app:
            return
        app.running_workers.add(self)

        try:
            placeholder_spans = [m.span() for m in app.placeholder_regex.finditer(self.original_text)]
            def is_inside_placeholder(pos):
                for start, end in placeholder_spans:
                    if start <= pos < end:
                        return True
                return False
            original_words = set(re.findall(r'\b\w+\b', self.original_text.lower()))
            potential_glossary_terms = {}
            if original_words:
                source_lang = app.source_language
                target_lang_code = app.current_target_language if app.is_project_mode else app.target_language
                potential_glossary_terms = app.glossary_service.get_translations_batch(
                    words=list(original_words),
                    source_lang=source_lang,
                    target_lang=target_lang_code,
                    include_reverse=False
                )
            valid_glossary_terms = {}
            if potential_glossary_terms:
                for word, term_info in potential_glossary_terms.items():
                    is_valid_term = False
                    try:
                        for match in re.finditer(r'\b' + re.escape(word) + r'\b', self.original_text, re.IGNORECASE):
                            if not is_inside_placeholder(match.start()):
                                is_valid_term = True
                                break
                    except re.error:
                        continue

                    if is_valid_term:
                        valid_glossary_terms[word] = term_info
            glossary_prompt_part = ""
            if valid_glossary_terms:
                header = f"| {_('Source Term')} | {_('Should be Translated As')} |\n|---|---|\n"
                rows = []
                for word, term_info in valid_glossary_terms.items():
                    targets = " or ".join(f"'{t['target']}'" for t in term_info['translations'])
                    source_term_escaped = word.replace('|', '\\|')
                    rows.append(f"| {source_term_escaped} | {targets} |")
                glossary_prompt_part = f"{header}" + "\n".join(rows)

            placeholders = {
                '[Target Language]': self.target_language,
                '[Untranslated Context]': self.context_dict.get("original_context", ""),
                '[Translated Context]': self.context_dict.get("translation_context", ""),
                '[Glossary]': glossary_prompt_part
            }
            if self.plugin_placeholders:
                placeholders.update(self.plugin_placeholders)
            prompt_structure = app.config.get("ai_prompt_structure", DEFAULT_PROMPT_STRUCTURE)
            final_prompt = generate_prompt_from_structure(prompt_structure, placeholders)
            logger.debug("=" * 20 + " AI PROMPT" + "=" * 20)
            logger.debug(f"Original Text to Translate:\n---\n{self.original_text}\n---")
            logger.debug(f"\nFinal System Prompt Sent to AI:\n---\n{final_prompt}\n---")
            logger.debug("=" * 57)
            translated_text = app.ai_translator.translate(self.original_text, final_prompt)
            logger.debug(f"Raw response from AI for ts_id '{self.ts_id}': '{translated_text}'")
            app.thread_signals.handle_ai_result.emit(self.ts_id, translated_text, None, self.is_batch_item)
        except Exception as e:
            app = self.app_ref()
            if app:
                app.thread_signals.handle_ai_result.emit(self.ts_id, None, str(e), self.is_batch_item)
        finally:
            app = self.app_ref()
            if app:
                if self.is_batch_item and app.ai_batch_semaphore is not None:
                    app.ai_batch_semaphore.release()
                    app.thread_signals.decrement_active_threads.emit()
                app.running_workers.discard(self)

class ThreadSafeSignals(QObject):
    handle_ai_result = Signal(str, str, str, bool)
    decrement_active_threads = Signal()

class LexiSyncApp(QMainWindow):
    language_changed = Signal()
    def __init__(self, config):
        super().__init__()
        self.config = config
        if "window_geometry" in self.config and self.config["window_geometry"]:
            self.restoreGeometry(QByteArray.fromBase64(self.config["window_geometry"].encode('utf-8')))
        else:
            self.setGeometry(100, 100, 1600, 900)
        self.setWindowTitle(_("LexiSync - v{version}").format(version=APP_VERSION))
        self.thread_signals = ThreadSafeSignals()
        self.thread_signals.handle_ai_result.connect(self._handle_ai_translation_result)
        self.thread_signals.decrement_active_threads.connect(self._decrement_active_threads_and_dispatch_more)
        self.running_workers = set()
        self.current_project_path = None
        self.current_code_file_path = None
        self.current_po_file_path = None
        self.is_project_mode = False
        self.project_config = {}
        self.source_language = self.config.get("default_source_language", "en")
        self.target_languages = []
        self.current_target_language = self.config.get("default_target_language", "zh")
        self.original_raw_code_content = ""
        self.current_project_modified = False
        self.is_po_mode = False
        self.current_po_metadata = None
        self.source_comment = ""
        self.current_focused_ts_id = None
        self.default_window_state = None
        self.neighbor_select_timer = QTimer(self)
        self.neighbor_select_timer.setSingleShot(True)

        self.source_language = self.config.get("default_source_language", "en")
        self.target_language = self.config.get("default_target_language", "zh")

        self.stats_update_timer = QTimer(self)
        self.stats_update_timer.setSingleShot(True)
        self.stats_update_timer.timeout.connect(self._update_details_panel_stats)

        self.tm_update_timer = QTimer(self)
        self.tm_update_timer.setSingleShot(True)
        self.tm_update_timer.timeout.connect(self.perform_tm_update)
        self.last_tm_query = ""

        self.filter_update_timer = QTimer(self)
        self.filter_update_timer.setSingleShot(True)
        self.filter_update_timer.timeout.connect(self.refresh_sheet_preserve_selection)

        self.quick_search_timer = QTimer(self)
        self.quick_search_timer.setSingleShot(True)
        self.quick_search_timer.timeout.connect(self._perform_delayed_search_filter)
        self._last_quick_search_text = ""

        self.translatable_objects = []
        self.tm_service = TMService()
        self.setup_tm_service()

        self.glossary_service = GlossaryService()
        self.setup_glossary_service()
        self.glossary_analysis_cache = {}
        self.original_highlighter = None

        self.undo_history = []
        self.redo_history = []
        self.current_selected_ts_id = None

        self.last_search_term = ""
        self.last_replace_term = ""
        self.last_search_options = {
            "case_sensitive": False,
            "in_original": True,
            "in_translation": True,
            "in_comment": True
        }
        self.find_highlight_indices = set()
        self.current_find_highlight_index = None
        self.ai_translator = AITranslator(
            api_key=self.config.get("ai_api_key"),
            model_name=self.config.get("ai_model_name", "deepseek-chat"),
            api_url=self.config.get("ai_api_base_url", DEFAULT_API_URL)
        )
        self.ai_translation_batch_ids_queue = []
        self.is_ai_translating_batch = False
        self.ai_batch_total_items = 0
        self.ai_batch_dispatched_count = 0
        self.ai_batch_completed_count = 0
        self.ai_batch_successful_translations_for_undo = []
        self.ai_batch_semaphore = None
        self.ai_batch_next_item_index = 0
        self.ai_batch_active_threads = 0
        self.ai_thread_pool = QThreadPool.globalInstance()
        self.is_finalizing_batch_translation = False


        self.filter_actions = {}
        self.filter_checkboxes = {}
        self.source_language = self.config.get("default_source_language", "en")
        self.target_language = self.config.get("default_target_language", "zh")
        self.show_ignored_var = self.config.get("show_ignored", True)
        self.show_untranslated_var = self.config.get("show_untranslated", False)
        self.show_translated_var = self.config.get("show_translated", False)
        self.show_unreviewed_var = self.config.get("show_unreviewed", False)
        self.use_static_sorting_var = self.config.get("use_static_sorting", False)

        self.search_term_var = self.config.get("ui_state", {}).get("search_term", "")


        self.auto_save_tm_var = self.config.get("auto_save_tm", False)
        self.auto_backup_tm_on_save_var = self.config.get("auto_backup_tm_on_save", True)
        self.auto_compile_mo_var = self.config.get("auto_compile_mo_on_save", True)
        self.auto_save_interval_sec = self.config.get("auto_save_interval_sec", 0)
        self.auto_save_timer = QTimer(self)
        self.auto_save_timer.timeout.connect(self.auto_save_project)
        self.setup_auto_save_timer()

        self.placeholder_regex = placeholder_regex
        self._placeholder_validation_job = None

        self.last_sort_column = "seq_id"
        self.last_sort_reverse = False


        self.language_changed.connect(self.update_ui_texts)
        self.setAcceptDrops(True)
        self.setStyleSheet("QPushButton:focus { outline: none; }")
        self.plugin_manager = PluginManager(self)
        self.UI_initialization()
        self.file_explorer_panel.file_double_clicked.connect(self.open_file_from_explorer)
        last_path = self.config.get('last_file_explorer_path')
        if last_path and os.path.isdir(last_path):
            self.file_explorer_panel.set_root_path(last_path)

        ExpansionRatioService.initialize()
        QTimer.singleShot(100, self.prewarm_dependencies)
        if hasattr(self, 'plugin_manager'):
            QTimer.singleShot(0, lambda: self.plugin_manager.run_hook('on_app_ready'))

    def prewarm_dependencies(self):
        self.update_statusbar(_("Initializing services in the background..."), persistent=True)

        try:
            import langid
            langid.classify("This is a pre-warming text to load the language model.")
            from services import po_file_service, project_service
            from dialogs import language_pair_dialog

            self.update_statusbar(_("Ready"))

        except Exception as e:
            error_msg = f"Failed to pre-warm dependencies: {e}"
            logger.error(error_msg)
            self.update_statusbar(error_msg)

    def UI_initialization(self):
        self._setup_ui()
        self.proxy_model.set_static_sorting_enabled(self.use_static_sorting_var)
        self.update_ui_state_after_file_load()
        self.update_ai_related_ui_state()
        self.update_counts_display()
        self.update_recent_files_menu()
        self.restore_window_state()

    def _setup_ui(self):
        self._setup_menu()
        self._setup_toolbars()
        self._setup_main_layout()
        self._setup_statusbar()
        self._setup_dock_widgets()
        self._setup_keybindings()

    def _setup_toolbars(self):
        self.project_toolbar = self.addToolBar(_("Project"))
        self.project_toolbar.setObjectName("projectToolbar")
        self.target_lang_label = QLabel(_("Target Language:"))
        self.target_lang_combo = QComboBox()
        self.target_lang_combo.setMinimumWidth(150)
        self.project_toolbar.addWidget(self.target_lang_label)
        self.project_toolbar.addWidget(self.target_lang_combo)
        self.project_toolbar.setVisible(False)  # Initially hidden
        self.target_lang_combo.currentIndexChanged.connect(self._on_target_language_changed)

    def _setup_menu(self):
        self.file_menu = self.menuBar().addMenu(_("&File"))
        self.edit_menu = self.menuBar().addMenu(_("&Edit"))
        self.view_menu = self.menuBar().addMenu(_("&View"))
        self.tools_menu = self.menuBar().addMenu(_("&Tools"))
        self.settings_menu = self.menuBar().addMenu(_("&Settings"))
        self.plugin_menu = self.menuBar().addMenu(_("&Plugins"))
        self.help_menu = self.menuBar().addMenu(_("&Help"))

        # File Menu
        self.action_open_code_file = QAction(_("Open File..."), self)
        self.action_open_code_file.triggered.connect(self.open_code_file_dialog)
        self.file_menu.addAction(self.action_open_code_file)

        self.action_new_project = QAction(_("New Project..."), self)
        self.action_new_project.triggered.connect(self.new_project_dialog)
        self.file_menu.addAction(self.action_new_project)

        self.action_open_project = QAction(_("Open Project..."), self)
        self.action_open_project.triggered.connect(self.open_project_dialog)
        self.file_menu.addAction(self.action_open_project)
        self.file_menu.addSeparator()

        self.action_compare_new_version = QAction(_("Compare/Import New Version..."), self)
        self.action_compare_new_version.triggered.connect(self.compare_with_new_version)
        self.action_compare_new_version.setEnabled(False)
        self.file_menu.addAction(self.action_compare_new_version)
        self.file_menu.addSeparator()

        self.action_save_current_file = QAction(_("Save"), self)
        self.action_save_current_file.triggered.connect(self.save_current_file)
        self.action_save_current_file.setEnabled(False)
        self.file_menu.addAction(self.action_save_current_file)

        self.action_save_current_file_as = QAction(_("Save As..."), self)
        self.action_save_current_file_as.triggered.connect(self.save_current_file_as)
        self.action_save_current_file_as.setEnabled(False)
        self.file_menu.addAction(self.action_save_current_file_as)
        self.file_menu.addSeparator()

        self.action_save_code_file = QAction(_("Save Translation to New Code File"), self)
        self.action_save_code_file.triggered.connect(self.save_code_file)
        self.action_save_code_file.setEnabled(False)
        self.file_menu.addAction(self.action_save_code_file)
        self.file_menu.addSeparator()

        # IMPORT MENU
        self.import_menu = QMenu(_("Import"), self)
        self.file_menu.addMenu(self.import_menu)

        self.action_import_excel = QAction(_("Import Translations from Excel"), self)
        self.action_import_excel.triggered.connect(self.import_project_translations_from_excel)
        self.action_import_excel.setEnabled(False)
        self.import_menu.addAction(self.action_import_excel)

        self.action_import_po = QAction(_("Import Translations from PO File..."), self)
        self.action_import_po.triggered.connect(self.import_po_file_dialog)
        self.import_menu.addAction(self.action_import_po)

        # EXPORT MENU
        self.export_menu = QMenu(_("Export"), self)
        self.file_menu.addMenu(self.export_menu)

        self.action_export_excel = QAction(_("Export to Excel"), self)
        self.action_export_excel.triggered.connect(self.export_project_translations_to_excel)
        self.action_export_excel.setEnabled(False)
        self.export_menu.addAction(self.action_export_excel)

        self.action_export_json = QAction(_("Export to JSON"), self)
        self.action_export_json.triggered.connect(self.export_project_translations_to_json)
        self.action_export_json.setEnabled(False)
        self.export_menu.addAction(self.action_export_json)

        self.action_export_yaml = QAction(_("Export to YAML"), self)
        self.action_export_yaml.triggered.connect(self.export_project_translations_to_yaml)
        self.action_export_yaml.setEnabled(False)
        self.export_menu.addAction(self.action_export_yaml)

        self.action_export_po = QAction(_("Export to PO File..."), self)
        self.action_export_po.triggered.connect(self.export_to_po_file_dialog)
        self.action_export_po.setEnabled(False)
        self.export_menu.addAction(self.action_export_po)

        # PLUGINS IMPORTERS & EXPORTERS
        if hasattr(self, 'plugin_manager'):
            importers = self.plugin_manager.run_hook('register_importers')
            if importers:
                self.import_menu.addSeparator()
                for menu_text, callback in importers.items():
                    action = QAction(menu_text, self)
                    action.triggered.connect(callback)
                    self.import_menu.addAction(action)

            exporters = self.plugin_manager.run_hook('register_exporters')
            if exporters:
                self.export_menu.addSeparator()
                for menu_text, callback in exporters.items():
                    action = QAction(menu_text, self)
                    action.triggered.connect(callback)
                    action.setEnabled(bool(self.translatable_objects))
                    self.export_menu.addAction(action)
        self.file_menu.addSeparator()

        self.action_extract_pot = QAction(_("Extract POT Template from Code..."), self)
        self.action_extract_pot.triggered.connect(self.extract_to_pot_dialog)
        self.file_menu.addAction(self.action_extract_pot)
        self.file_menu.addSeparator()


        self.recent_files_menu = QMenu(_("Recent Files"), self)
        self.file_menu.addMenu(self.recent_files_menu)
        self.file_menu.addSeparator()

        # Edit Menu
        self.action_exit = QAction(_("Exit"), self)
        self.action_exit.triggered.connect(self.close)
        self.file_menu.addAction(self.action_exit)

        self.action_undo = QAction(_("Undo"), self)
        self.action_undo.triggered.connect(self.undo_action)
        self.action_undo.setEnabled(False)
        self.edit_menu.addAction(self.action_undo)

        self.action_redo = QAction(_("Redo"), self)
        self.action_redo.triggered.connect(self.redo_action)
        self.action_redo.setEnabled(False)
        self.edit_menu.addAction(self.action_redo)
        self.edit_menu.addSeparator()

        self.action_find_replace = QAction(_("Find/Replace..."), self)
        self.action_find_replace.triggered.connect(self.show_advanced_search_dialog)
        self.action_find_replace.setEnabled(False)
        self.edit_menu.addAction(self.action_find_replace)
        self.edit_menu.addSeparator()

        self.action_copy_original = QAction(_("Copy Original"), self)
        self.action_copy_original.triggered.connect(self.cm_copy_original)
        self.action_copy_original.setEnabled(False)
        self.edit_menu.addAction(self.action_copy_original)

        self.action_paste_translation = QAction(_("Paste to Translation"), self)
        self.action_paste_translation.triggered.connect(self.cm_paste_to_translation)
        self.action_paste_translation.setEnabled(False)
        self.edit_menu.addAction(self.action_paste_translation)

        # View Menu
        self.action_show_ignored = QAction(_("Show Ignored"), self, checkable=True)
        self.action_show_ignored.setChecked(self.show_ignored_var)
        self.action_show_ignored.triggered.connect(lambda checked: self.set_filter_var('show_ignored', checked))
        self.view_menu.addAction(self.action_show_ignored)
        self.filter_actions['show_ignored'] = self.action_show_ignored

        self.action_show_untranslated = QAction(_("Show Untranslated"), self, checkable=True)
        self.action_show_untranslated.setChecked(self.show_untranslated_var)
        self.action_show_untranslated.triggered.connect(lambda checked: self.set_filter_var('show_untranslated', checked))
        self.view_menu.addAction(self.action_show_untranslated)
        self.filter_actions['show_untranslated'] = self.action_show_untranslated

        self.action_show_translated = QAction(_("Show Translated"), self, checkable=True)
        self.action_show_translated.setChecked(self.show_translated_var)
        self.action_show_translated.triggered.connect(lambda checked: self.set_filter_var('show_translated', checked))
        self.view_menu.addAction(self.action_show_translated)
        self.filter_actions['show_translated'] = self.action_show_translated

        self.action_show_unreviewed = QAction(_("Show Unreviewed"), self, checkable=True)
        self.action_show_unreviewed.setChecked(self.show_unreviewed_var)
        self.action_show_unreviewed.triggered.connect(lambda checked: self.set_filter_var('show_unreviewed', checked))
        self.view_menu.addAction(self.action_show_unreviewed)
        self.filter_actions['show_unreviewed'] = self.action_show_unreviewed
        self.view_menu.addSeparator()

        self.action_restore_layout = QAction(_("Restore Default Layout"), self)
        self.action_restore_layout.triggered.connect(self.restore_default_layout)
        self.view_menu.addAction(self.action_restore_layout)

        # Tools Menu
        self.action_apply_tm_to_untranslated = QAction(_("Apply TM to Untranslated"), self)
        self.action_apply_tm_to_untranslated.triggered.connect(lambda: self.apply_tm_to_all_current_strings(only_if_empty=True, confirm=True))
        self.action_apply_tm_to_untranslated.setEnabled(False)
        self.tools_menu.addAction(self.action_apply_tm_to_untranslated)

        self.action_ai_translate_selected = QAction(_("AI Translate Selected"), self)
        self.action_ai_translate_selected.triggered.connect(self.cm_ai_translate_selected)
        self.action_ai_translate_selected.setEnabled(False)
        self.tools_menu.addAction(self.action_ai_translate_selected)

        self.action_ai_translate_all_untranslated = QAction(_("AI Translate All Untranslated"), self)
        self.action_ai_translate_all_untranslated.triggered.connect(self.ai_translate_all_untranslated)
        self.action_ai_translate_all_untranslated.setEnabled(False)
        self.tools_menu.addAction(self.action_ai_translate_all_untranslated)

        self.action_stop_ai_batch_translation = QAction(_("Stop AI Batch Translation"), self)
        self.action_stop_ai_batch_translation.triggered.connect(self.stop_batch_ai_translation)
        self.action_stop_ai_batch_translation.setEnabled(False)
        self.tools_menu.addAction(self.action_stop_ai_batch_translation)
        self.tools_menu.addSeparator()

        self.tools_menu.addSeparator()
        self.action_run_validation_on_all = QAction(_("Re-validate All Entries"), self)
        self.action_run_validation_on_all.triggered.connect(self._run_and_refresh_with_validation)
        self.action_run_validation_on_all.setEnabled(False)
        self.tools_menu.addAction(self.action_run_validation_on_all)


        self.action_reload_translatable_text = QAction(_("Reload Translatable Text"), self)
        self.action_reload_translatable_text.triggered.connect(self.reload_translatable_text)
        self.action_reload_translatable_text.setEnabled(False)
        self.tools_menu.addAction(self.action_reload_translatable_text)
        self.tools_menu.addSeparator()

        self.action_show_statistics = QAction(_("Project Statistics..."), self)
        self.action_show_statistics.triggered.connect(self.show_statistics_dialog)
        self.action_show_statistics.setEnabled(False)
        self.tools_menu.addAction(self.action_show_statistics)

        # Settings Menu
        self.action_show_settings = QAction(_("Settings..."), self)
        self.action_show_settings.triggered.connect(self.show_settings_dialog)
        self.settings_menu.addAction(self.action_show_settings)

        self.settings_menu.addSeparator()

        self.action_language_pair_settings = QAction(_("Language Pair Settings..."), self)
        self.action_language_pair_settings.triggered.connect(self.show_language_pair_dialog)
        self.settings_menu.addAction(self.action_language_pair_settings)

        # Plugins Menu
        if hasattr(self, 'plugin_manager'):
            self.plugin_menu.addAction(_("Plugin Marketplace..."), self.plugin_manager.show_marketplace_dialog)
            self.plugin_menu.addSeparator()
            self.plugin_manager.setup_plugin_ui()

        # Help Menu
        self.action_about = QAction(_("About"), self)
        self.action_about.triggered.connect(self.about)
        self.help_menu.addAction(self.action_about)

    def setup_auto_save_timer(self):
        if self.auto_save_interval_sec > 0:
            self.auto_save_timer.start(self.auto_save_interval_sec * 1000)
        else:
            self.auto_save_timer.stop()

    def setup_tm_service(self):
        global_tm_dir = os.path.join(get_app_data_path(), "tm")
        project_tm_dir = None
        use_global = True
        if self.is_project_mode and self.current_project_path:
            project_tm_dir = os.path.join(self.current_project_path, "tm")
            project_settings = self.project_config.get("settings", {})
            use_global = project_settings.get("use_global_tm", True)
        final_global_tm_dir = global_tm_dir if use_global else None
        self.tm_service.connect_databases(final_global_tm_dir, project_tm_dir)

    def setup_glossary_service(self):
        global_glossary_dir = os.path.join(get_app_data_path(), "glossary")
        project_glossary_dir = None
        if self.is_project_mode and self.current_project_path:
            project_glossary_dir = os.path.join(self.current_project_path, "glossary")

        self.glossary_service.connect_databases(global_glossary_dir, project_glossary_dir)

    def update_ui_texts(self):
        self.setWindowTitle(_("LexiSync - v{version}").format(version=APP_VERSION))
        self.update_title()

        self.file_menu.setTitle(_("&File"))
        self.edit_menu.setTitle(_("&Edit"))
        self.view_menu.setTitle(_("&View"))
        self.tools_menu.setTitle(_("&Tools"))
        self.settings_menu.setTitle(_("&Settings"))
        if hasattr(self, 'plugin_menu'):
            self.plugin_menu.setTitle(_("&Plugins"))
        self.help_menu.setTitle(_("&Help"))

        #File Menu
        self.action_open_code_file.setText(_("Open..."))
        self.action_open_project.setText(_("Open Project..."))
        self.action_compare_new_version.setText(_("Compare/Import New Version..."))
        self.action_save_current_file.setText(_("Save"))
        self.action_save_current_file_as.setText(_("Save As..."))
        self.action_save_code_file.setText(_("Save Translation to New Code File"))

        if hasattr(self, 'import_menu'):
            self.import_menu.setTitle(_("Import"))
        if hasattr(self, 'export_menu'):
            self.export_menu.setTitle(_("Export"))

        self.action_import_excel.setText(_("Import Translations from Excel"))
        self.action_export_excel.setText(_("Export to Excel"))
        self.action_export_json.setText(_("Export to JSON"))
        self.action_export_yaml.setText(_("Export to YAML"))
        self.action_extract_pot.setText(_("Extract POT Template from Code..."))
        self.action_import_po.setText(_("Import Translations from PO File..."))
        self.action_export_po.setText(_("Export to PO File..."))
        self.recent_files_menu.setTitle(_("Recent Files"))
        self.action_exit.setText(_("Exit"))

        #Edit Menu
        self.action_undo.setText(_("Undo"))
        self.action_redo.setText(_("Redo"))
        self.action_find_replace.setText(_("Find/Replace..."))
        self.action_copy_original.setText(_("Copy Original"))
        self.action_paste_translation.setText(_("Paste to Translation"))

        #View Menu

        self.action_show_ignored.setText(_("Show Ignored"))
        self.action_show_untranslated.setText(_("Show Untranslated"))
        self.action_show_translated.setText(_("Show Translated"))
        self.action_show_unreviewed.setText(_("Show Unreviewed"))
        self.action_toggle_file_explorer.setText(_("File Explorer Panel"))
        self.action_toggle_details_panel.setText(_("Edit && Details Panel"))
        self.action_toggle_comment_status_panel.setText(_("Comment && Status Panel"))
        self.action_toggle_context_panel.setText(_("Context Preview Panel"))
        self.action_toggle_tm_panel.setText(_("Translation Memory Panel"))
        self.action_restore_layout.setText(_("Restore Default Layout"))

        #Tools Menu
        self.action_apply_tm_to_untranslated.setText(_("Apply TM to Untranslated"))
        self.action_ai_translate_selected.setText(_("AI Translate Selected"))
        self.action_ai_translate_all_untranslated.setText(_("AI Translate All Untranslated"))
        self.action_stop_ai_batch_translation.setText(_("Stop AI Batch Translation"))
        self.action_run_validation_on_all.setText(_("Re-validate All Entries"))
        self.action_reload_translatable_text.setText(_("Reload Translatable Text"))
        self.action_show_statistics.setText(_("Project Statistics..."))

        #Settings Menu
        self.action_show_settings.setText(_("Settings..."))
        self.action_language_pair_settings.setText(_("Language Pair Settings..."))

        #Help Menu
        self.action_about.setText(_("About"))

        self.sheet_model.setHeaderData(0, Qt.Horizontal, "#")
        self.sheet_model.setHeaderData(1, Qt.Horizontal, "S")
        self.sheet_model.setHeaderData(2, Qt.Horizontal, _("Original"))
        self.sheet_model.setHeaderData(3, Qt.Horizontal, _("Translation"))
        self.sheet_model.setHeaderData(4, Qt.Horizontal, _("Comment"))
        self.sheet_model.setHeaderData(5, Qt.Horizontal, "✔")
        self.sheet_model.setHeaderData(6, Qt.Horizontal, _("Line"))
        self.sheet_model.headerDataChanged.emit(Qt.Horizontal, 0, 6)

        self.filter_label.setText(_("Filter:"))
        self.ignored_checkbox.setText(_("Ignored"))
        self.untranslated_checkbox.setText(_("Untranslated"))
        self.translated_checkbox.setText(_("Translated"))
        self.unreviewed_checkbox.setText(_("Unreviewed"))
        self.search_button.setText(_("Find"))
        self.search_entry.textChanged.disconnect(self.search_filter_changed)
        self.search_entry.setText("")
        self.search_entry.textChanged.connect(self.search_filter_changed)
        self.search_entry.setPlaceholderText(_("Quick search..."))
        self.on_search_focus_out()
        self.on_search_focus_out()

        if hasattr(self, 'file_explorer_dock'):
            self.file_explorer_dock.setWindowTitle(_("File Explorer"))
        if hasattr(self, 'details_dock'):
            self.details_dock.setWindowTitle(_("Edit && Details"))
        if hasattr(self, 'context_dock'):
            self.context_dock.setWindowTitle(_("Context Preview"))
        if hasattr(self, 'tm_dock'):
            self.tm_dock.setWindowTitle(_("Translation Memory Matches"))
        if hasattr(self, 'comment_status_dock'):
            self.comment_status_dock.setWindowTitle(_("Comment && Status"))

        if hasattr(self, 'file_explorer_dock'):
            self.file_explorer_panel.update_ui_texts()
        if hasattr(self, 'details_dock'):
            self.details_panel.update_ui_texts()
        if hasattr(self, 'details_dock'):
            self.comment_status_panel.update_ui_texts()
        if hasattr(self, 'tm_dock'):
            self.context_panel.update_ui_texts()
        if hasattr(self, 'comment_status_dock'):
            self.tm_panel.update_ui_texts()
        if hasattr(self, 'plugin_manager'):
            self.plugin_manager.setup_plugin_ui()

        self.update_statusbar(_("Ready"), persistent=True)
        self.update_counts_display()
        self.update_recent_files_menu()

    def set_filter_var(self, var_name, value):
        setattr(self, f"{var_name}_var", value)

        action = self.filter_actions.get(var_name)
        checkbox = self.filter_checkboxes.get(var_name)
        if action and action.isChecked() != value:
            action.setChecked(value)
        if checkbox and checkbox.isChecked() != value:
            checkbox.setChecked(value)
        self.filter_update_timer.start(150)
        self.save_config()

    def set_config_var(self, var_name, value):
        setattr(self, f"{var_name}_var", value)
        self.config[var_name] = value
        self.save_config()

    def execute_action(self, action_name: str, path: str = None) -> bool:
        if action_name == "open_code_file_dialog":
            return self.open_code_file_dialog()

        if action_name == "new_project":
            return self.new_project_dialog()

        if action_name == "open_project_dialog":
            return self.open_project_dialog()

        if action_name == "open_recent_file" or action_name == "open_specific_project":
            if not path:
                return False

            if not os.path.exists(path):
                QMessageBox.critical(self, _("File not found"),
                                     _("File or project '{path}' does not exist.").format(path=path))
                recent_files = self.config.get("recent_files", [])
                if path in recent_files:
                    recent_files.remove(path)
                    self.config["recent_files"] = recent_files
                    self.update_recent_files_menu()
                return False

            if os.path.isdir(path) and os.path.exists(os.path.join(path, "project.json")):
                self.open_project(path)
                return True
            elif os.path.isfile(path):
                ext = os.path.splitext(path)[1].lower()
                if ext in ['.ow', '.txt']:
                    self.open_code_file_path(path)
                    return True
                elif ext in ['.po', '.pot']:
                    self.import_po_file_dialog_with_path(path)
                    return True
            QMessageBox.warning(self, _("Unsupported Type"),
                                _("Cannot open '{path}'. It is not a valid file or project.").format(path=path))
            return False

        if action_name == "open_specific_file":
            if path and os.path.isfile(path):
                ext = os.path.splitext(path)[1].lower()
                if ext in ['.ow', '.txt']:
                    self.open_code_file_path(path)
                    return True
                elif ext in ['.po', '.pot']:
                    self.import_po_file_dialog_with_path(path)
                    return True
            elif path and os.path.isdir(path):
                if os.path.exists(os.path.join(path, "project.json")):
                    self.open_project(path)
                    return True
                else:
                    self.file_explorer_panel.set_root_path(path)
                    return False
            return False

        if action_name == "show_marketplace":
            self.plugin_manager.show_marketplace_dialog()
            return True

        if action_name == "show_settings":
            self.show_settings_dialog()
            return True
        return False

    def _setup_keybindings(self):
        bindings = self.config.get('keybindings', DEFAULT_KEYBINDINGS)
        self.action_open_code_file.setShortcut(QKeySequence(bindings.get('open_code_file', '')))
        self.action_new_project.setShortcut(QKeySequence(bindings.get('new_project', '')))
        self.action_open_project.setShortcut(QKeySequence(bindings.get('open_project', '')))
        self.action_save_current_file.setShortcut(QKeySequence(bindings.get('save_current_file', '')))
        self.action_save_code_file.setShortcut(QKeySequence(bindings.get('save_code_file', '')))
        self.action_undo.setShortcut(QKeySequence(bindings.get('undo', '')))
        self.action_redo.setShortcut(QKeySequence(bindings.get('redo', '')))
        self.action_find_replace.setShortcut(QKeySequence(bindings.get('find_replace', '')))
        self.action_copy_original.setShortcut(QKeySequence(bindings.get('copy_original', '')))
        self.action_paste_translation.setShortcut(QKeySequence(bindings.get('paste_translation', '')))
        self.action_ai_translate_selected.setShortcut(QKeySequence(bindings.get('ai_translate_selected', '')))
        self.ACTION_MAP_FOR_DIALOG = {
            'open_code_file': self.action_open_code_file,
            'new_project': self.action_new_project,
            'open_project': self.action_open_project,
            'save_current_file': self.action_save_current_file,
            'save_code_file': self.action_save_code_file,
            'undo': self.action_undo,
            'redo': self.action_redo,
            'find_replace': self.action_find_replace,
            'copy_original': self.action_copy_original,
            'paste_translation': self.action_paste_translation,
            'ai_translate_selected': self.action_ai_translate_selected,
        }

        global_actions = {
            'toggle_reviewed': self.cm_toggle_reviewed_status,
            'toggle_ignored': self.cm_toggle_ignored_status,
            'apply_and_next': self.apply_and_select_next_untranslated,
            'refresh_sort': self.refresh_sort,
        }

        for name, slot in global_actions.items():
            shortcut_str = bindings.get(name, '')
            if not shortcut_str:
                continue


            action = QAction(self)
            action.setShortcut(QKeySequence(shortcut_str))
            action.triggered.connect(slot)


            self.addAction(action)

            self.ACTION_MAP_FOR_DIALOG[name] = action

    def _toggle_static_sorting_mode(self, checked: bool):
        self.use_static_sorting_var = checked
        self.proxy_model.set_static_sorting_enabled(checked)
        self.set_config_var('use_static_sorting', checked)
        if checked:
            self.update_statusbar(_("Static sorting enabled. Press F5 to refresh."))
        else:
            self.update_statusbar(_("Dynamic sorting enabled."))
            self.proxy_model.invalidate()

    def _get_global_tm_path(self):
        app_data_dir = get_app_data_path()
        tm_dir = os.path.join(app_data_dir, "tm")
        os.makedirs(tm_dir, exist_ok=True)
        return tm_dir

    def refresh_sort(self):
        if self.use_static_sorting_var:
            self.proxy_model.invalidate()
            self.update_statusbar(_("View refreshed."))

    def detect_language_from_data(self, text_type: str) -> str | None:
        if not self.translatable_objects:
            QMessageBox.information(self, _("Info"), _("No text available to detect language from."))
            return None

        if text_type == 'source':
            strings_to_check = [ts.original_semantic for ts in self.translatable_objects]
            return language_service.detect_source_language(strings_to_check)
        elif text_type == 'target':
            strings_to_check = [ts.translation for ts in self.translatable_objects]
            return language_service.detect_source_language(strings_to_check)
        return None

    def show_language_pair_dialog(self):
        dialog = LanguagePairDialog(self, self.source_language, self.target_language, self)
        if dialog.exec():
            if (self.source_language != dialog.source_lang or
                    self.target_language != dialog.target_lang):
                self.source_language = dialog.source_lang
                self.target_language = dialog.target_lang
                self.config["default_source_language"] = self.source_language
                self.config["default_target_language"] = self.target_language
                self.save_config()
                if self.translatable_objects:
                    self.mark_project_modified()
                    self.update_statusbar(_("Language pair updated. Re-validating all entries..."), persistent=True)
                    QApplication.processEvents()
                    self._run_and_refresh_with_validation()
                    self.update_statusbar(
                        _("Validation complete. Language pair set to {src} -> {tgt}.").format(src=self.source_language,
                                                                                              tgt=self.target_language))
                else:
                    self.update_statusbar(_("Language pair set to {src} -> {tgt}.").format(src=self.source_language,
                                                                                           tgt=self.target_language))

    def show_keybinding_dialog(self):
        dialog = KeybindingDialog(self, _("Keybinding Settings"), self)
        if dialog.exec():
            self._setup_keybindings()

    def show_font_settings_dialog(self):
        dialog = FontSettingsDialog(self, _("Font Settings"), self)
        if dialog.exec():
            QMessageBox.information(self, _("Restart Required"),
                                    _("Font settings have been changed. Please restart the application for the changes to take effect."))

    def auto_save_project(self):
        if not self.current_project_modified:
            return
        if not (self.current_project_path or self.current_po_file_path or self.current_code_file_path):
            return
        focused_widget = QApplication.focusWidget()
        self.update_statusbar(_("Auto-saving..."), persistent=True)
        QApplication.processEvents()
        try:
            self.save_current_file()
            self.update_statusbar(_("Project auto-saved."), persistent=False)
        finally:
            if focused_widget and QApplication.focusWidget() != focused_widget:
                focused_widget.setFocus()

    def change_language(self, new_lang_code):
        if new_lang_code != self.config.get('language'):
            self.config['language'] = new_lang_code
            lang_manager.setup_translation(new_lang_code)
            self.save_config()
            if hasattr(self, 'plugin_manager'):
                self.plugin_manager.on_main_app_language_changed()
            self.language_changed.emit()

    def _setup_main_layout(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(5)

        # Filter Toolbar
        toolbar_frame = QWidget()
        toolbar_layout = QHBoxLayout(toolbar_frame)
        toolbar_layout.setContentsMargins(0, 0, 0, 0)

        self.filter_label = QLabel(_("Filter:"))
        toolbar_layout.addWidget(self.filter_label)

        self.ignored_checkbox = QCheckBox(_("Ignored"))
        self.ignored_checkbox.setChecked(self.show_ignored_var)
        self.ignored_checkbox.stateChanged.connect(lambda state: self.set_filter_var('show_ignored', bool(state)))
        toolbar_layout.addWidget(self.ignored_checkbox)
        self.filter_checkboxes['show_ignored'] = self.ignored_checkbox

        self.untranslated_checkbox = QCheckBox(_("Untranslated"))
        self.untranslated_checkbox.setChecked(self.show_untranslated_var)
        self.untranslated_checkbox.stateChanged.connect(lambda state: self.set_filter_var('show_untranslated', bool(state)))
        toolbar_layout.addWidget(self.untranslated_checkbox)
        self.filter_checkboxes['show_untranslated'] = self.untranslated_checkbox

        self.translated_checkbox = QCheckBox(_("Translated"))
        self.translated_checkbox.setChecked(self.show_translated_var)
        self.translated_checkbox.stateChanged.connect(lambda state: self.set_filter_var('show_translated', bool(state)))
        toolbar_layout.addWidget(self.translated_checkbox)
        self.filter_checkboxes['show_translated'] = self.translated_checkbox

        self.unreviewed_checkbox = QCheckBox(_("Unreviewed"))
        self.unreviewed_checkbox.setChecked(self.show_unreviewed_var)
        self.unreviewed_checkbox.stateChanged.connect(lambda state: self.set_filter_var('show_unreviewed', bool(state)))
        toolbar_layout.addWidget(self.unreviewed_checkbox)
        self.filter_checkboxes['show_unreviewed'] = self.unreviewed_checkbox

        # Plugins
        separator = QFrame()
        separator.setFrameShape(QFrame.VLine)
        separator.setFrameShadow(QFrame.Sunken)
        toolbar_layout.addWidget(separator)

        if hasattr(self, 'plugin_manager'):
            self.plugin_manager.run_hook('on_main_toolbar_setup', toolbar_layout=toolbar_layout)

        toolbar_layout.addStretch(1)

        self.search_entry = QLineEdit()
        self.search_entry.setPlaceholderText(_("Quick search..."))
        self.search_entry.setText(self.search_term_var)
        self.search_entry.textChanged.connect(self.search_filter_changed)
        self.search_entry.returnPressed.connect(self.find_string_from_toolbar)
        self.search_entry.setFixedWidth(200)
        self.search_entry.installEventFilter(self)
        toolbar_layout.addWidget(self.search_entry)

        self.search_button = QPushButton(_("Find"))
        self.search_button.clicked.connect(self.find_string_from_toolbar)
        toolbar_layout.addWidget(self.search_button)
        main_layout.addWidget(toolbar_frame)

        # Table View
        self.table_view = CustomTableView(self, self)
        self.table_view.setVerticalScrollMode(QAbstractItemView.ScrollPerItem)
        self.table_view.setIconSize(QSize(32, 32))
        self.table_view.setAttribute(Qt.WA_Hover, False)
        palette = self.table_view.palette()
        palette.setColor(QPalette.Base, QColor("white"))
        self.table_view.setPalette(palette)
        palette = self.table_view.palette()
        palette.setColor(QPalette.Highlight, QColor(51, 153, 255, 45))
        palette.setColor(QPalette.HighlightedText, palette.color(QPalette.Text))
        self.table_view.setPalette(palette)
        self.table_view.setStyleSheet("""
            QTableView {
                gridline-color: transparent;
                alternate-background-color: rgba(247, 247, 247, 211);
            }
            /* 被选择行 */
            QTableView::item:selected {
                background-color: rgba(51, 153, 255, 45); /* 半透明蓝色背景 */
                border: 1px solid (51, 153, 255, 145); /* 细边框 */
                border-right: none;
                border-left: none;
            }
            /* 焦点行 */
            QTableView::item:selected:focus, QTableView::item:focus {
                background-color: rgba(51, 153, 255, 255); /* 半透明蓝色 */
                border: 1px solid rgba(255, 0, 0, 200); /* 红色边框 */
            }

        """)
        self.table_view.setAlternatingRowColors(True)
        self.sheet_model = TranslatableStringsModel(self.translatable_objects, self)
        self.proxy_model = TranslatableStringsProxyModel(self)
        self.proxy_model.setSourceModel(self.sheet_model)
        self.table_view.setModel(self.proxy_model)

        self.table_view.setSelectionBehavior(QTableView.SelectRows)
        self.table_view.setSelectionMode(QTableView.ExtendedSelection)
        self.table_view.setSortingEnabled(True)
        self.table_view.sortByColumn(0, Qt.AscendingOrder)

        header = self.table_view.horizontalHeader()
        # 固定宽度的列
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # 原文
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # 译文
        # 可以自由拉伸的列
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)  # #
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)  # S
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)  # 注释
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Interactive)  # ✔
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Interactive)  # Line

        self.table_view.setColumnWidth(0, 40)  # #
        self.table_view.setColumnWidth(1, 35)  # S
        self.table_view.setColumnWidth(5, 30)  # ✔
        self.table_view.setColumnWidth(6, 50)  # Line
        self.table_view.setColumnWidth(4, 120)  # Comment
        self.table_view.verticalHeader().setVisible(False)

        self.table_view.horizontalHeader().sectionClicked.connect(self._sort_sheet_column)
        self.table_view.doubleClicked.connect(self.on_sheet_double_click)
        self.table_view.selectionModel().currentChanged.connect(self.on_sheet_select)
        self.table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self.show_sheet_context_menu)
        self.table_view.setItemDelegate(CustomCellDelegate(self.table_view, self))
        self.table_view.selectionModel().selectionChanged.connect(self._on_selection_model_changed)
        main_layout.addWidget(self.table_view)

    def eventFilter(self, obj, event):
        if obj is self.search_entry:
            if event.type() == QEvent.FocusIn:
                if self.search_entry.text() == _("Quick search..."):
                    self.search_entry.setText("")
                    self.search_entry.setStyleSheet("color: black;")
            elif event.type() == QEvent.FocusOut:
                if not self.search_entry.text():
                    self.search_entry.setText(_("Quick search..."))
                    self.search_entry.setStyleSheet("color: grey;")
        return super().eventFilter(obj, event)

    def _setup_dock_widgets(self):
        self.details_panel = DetailsPanel(self)
        self.context_panel = ContextPanel(self)
        self.tm_panel = TMPanel(self, app_instance=self)
        self.comment_status_panel = CommentStatusPanel(self)
        self.glossary_panel = GlossaryPanel(self)

        #FileExplorerPanel
        self.file_explorer_panel = FileExplorerPanel(self, self)
        self.file_explorer_dock = QDockWidget(_("File Explorer"), self)
        self.file_explorer_dock.setObjectName("fileExplorerDock")
        self.file_explorer_dock.setWidget(self.file_explorer_panel)
        self.addDockWidget(Qt.LeftDockWidgetArea, self.file_explorer_dock)

        # GlossaryPanel
        self.glossary_dock = QDockWidget(_("Glossary"), self)
        self.glossary_dock.setObjectName("glossaryDock")
        self.glossary_dock.setWidget(self.glossary_panel)
        self.addDockWidget(Qt.LeftDockWidgetArea, self.glossary_dock)

        self.splitDockWidget(self.file_explorer_dock, self.glossary_dock, Qt.Vertical)
        self.resizeDocks([self.file_explorer_dock, self.glossary_dock], [400, 200], Qt.Vertical)

        # DetailsPanel
        self.details_panel.apply_translation_signal.connect(self.apply_translation_from_button)
        self.details_panel.translation_text_changed_signal.connect(self.schedule_placeholder_validation)
        self.details_panel.translation_text_changed_signal.connect(self.schedule_details_panel_stats_update)
        self.details_panel.translation_focus_out_signal.connect(self.apply_translation_focus_out)
        self.details_panel.ai_translate_signal.connect(self.ai_translate_selected_from_button)

        # CommentStatusPanel
        self.comment_status_panel.apply_comment_signal.connect(self.apply_comment_from_button)
        self.comment_status_panel.comment_focus_out_signal.connect(self.apply_comment_focus_out)
        self.comment_status_panel.ignore_checkbox.stateChanged.connect(self.toggle_ignore_selected_checkbox)
        self.comment_status_panel.reviewed_checkbox.stateChanged.connect(self.toggle_reviewed_selected_checkbox)

        # TMPanel
        self.tm_panel.apply_tm_suggestion_signal.connect(self.apply_tm_suggestion_from_listbox)
        self.tm_panel.update_tm_signal.connect(self.update_tm_for_selected_string)
        self.tm_panel.clear_tm_signal.connect(self.clear_tm_for_selected_string)


        # DetailsPanel
        self.details_dock = QDockWidget(_("Edit && Details"), self)
        self.details_dock.setObjectName("detailsDock")
        self.details_dock.setWidget(self.details_panel)
        self.details_dock.setFeatures(
            QDockWidget.DockWidgetFloatable | QDockWidget.DockWidgetMovable | QDockWidget.DockWidgetClosable)
        self.details_dock.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, False)
        self.addDockWidget(Qt.BottomDockWidgetArea, self.details_dock)

        # ContextPanel
        self.context_dock = QDockWidget(_("Context Preview"), self)
        self.context_dock.setObjectName("contextDock")
        self.context_dock.setWidget(self.context_panel)
        self.context_dock.setFeatures(
            QDockWidget.DockWidgetFloatable | QDockWidget.DockWidgetMovable | QDockWidget.DockWidgetClosable)
        self.context_dock.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, False)
        self.addDockWidget(Qt.RightDockWidgetArea, self.context_dock)

        # TMPanel
        self.tm_dock = QDockWidget(_("Translation Memory Matches"), self)
        self.tm_dock.setObjectName("tmDock")
        self.tm_dock.setWidget(self.tm_panel)
        self.tm_dock.setFeatures(
            QDockWidget.DockWidgetFloatable | QDockWidget.DockWidgetMovable | QDockWidget.DockWidgetClosable)
        self.tm_dock.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, False)
        self.addDockWidget(Qt.RightDockWidgetArea, self.tm_dock)
        self.splitDockWidget(self.context_dock, self.tm_dock, Qt.Vertical)

        # CommentStatusPanel
        self.comment_status_dock = QDockWidget(_("Comment && Status"), self)
        self.comment_status_dock.setObjectName("commentDock")
        self.comment_status_dock.setWidget(self.comment_status_panel)
        self.comment_status_dock.setFeatures(
            QDockWidget.DockWidgetFloatable | QDockWidget.DockWidgetMovable | QDockWidget.DockWidgetClosable)
        self.comment_status_dock.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, False)
        self.addDockWidget(Qt.RightDockWidgetArea, self.comment_status_dock)
        self.splitDockWidget(self.tm_dock, self.comment_status_dock, Qt.Vertical)

        # File Explorer Panel Action
        self.action_toggle_file_explorer = self.file_explorer_dock.toggleViewAction()
        self.action_toggle_file_explorer.setText(_("File Explorer Panel"))

        # Details Panel Action
        self.action_toggle_details_panel = self.details_dock.toggleViewAction()
        self.action_toggle_details_panel.setText(_("Edit && Details Panel"))

        # Context Panel Action
        self.action_toggle_context_panel = self.context_dock.toggleViewAction()
        self.action_toggle_context_panel.setText(_("Context Preview Panel"))

        # TM Panel Action
        self.action_toggle_tm_panel = self.tm_dock.toggleViewAction()
        self.action_toggle_tm_panel.setText(_("Translation Memory Panel"))

        # Comment Status Panel Action
        self.action_toggle_comment_status_panel = self.comment_status_dock.toggleViewAction()
        self.action_toggle_comment_status_panel.setText(_("Comment && Status Panel"))

        # Glossary Panel Action
        self.action_toggle_glossary_panel = self.glossary_dock.toggleViewAction()
        self.action_toggle_glossary_panel.setText(_("Glossary Panel"))
        self.view_menu.addAction(self.action_toggle_glossary_panel)

        self.view_menu.addAction(self.action_toggle_file_explorer)
        self.view_menu.addAction(self.action_toggle_details_panel)
        self.view_menu.addAction(self.action_toggle_context_panel)
        self.view_menu.addAction(self.action_toggle_tm_panel)
        self.view_menu.addAction(self.action_toggle_comment_status_panel)

        if self.default_window_state is None:
            self.default_window_state = self.saveState()

    def restore_default_layout(self):
        if self.default_window_state:
            self.restoreState(self.default_window_state)
        docks_to_show = [
            self.file_explorer_dock, self.glossary_dock, self.details_dock,
            self.context_dock, self.tm_dock, self.comment_status_dock
        ]
        for dock in docks_to_show:
            dock.setVisible(True)
        main_width = self.size().width()
        main_height = self.size().height()
        left_width = max(200, int(main_width * 0.15))
        self.resizeDocks([self.file_explorer_dock, self.glossary_dock], [left_width, left_width], Qt.Horizontal)
        self.resizeDocks([self.file_explorer_dock, self.glossary_dock], [int(main_height * 0.6), int(main_height * 0.4)], Qt.Vertical)
        right_width = max(200, int(main_width * 0.15))
        right_docks = [self.context_dock, self.tm_dock, self.glossary_dock, self.comment_status_dock]
        self.resizeDocks(right_docks, [right_width] * len(right_docks), Qt.Horizontal)
        self.resizeDocks(right_docks, [int(main_height * 0.25)] * len(right_docks), Qt.Vertical)
        bottom_height = max(200, int(main_height * 0.30))
        self.resizeDocks([self.details_dock], [bottom_height], Qt.Vertical)

        self.update_statusbar(_("The layout has been restored to its default state."))

    def _setup_statusbar(self):
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)

        self.statusbar_label = ElidedLabel()
        self.statusbar_label.setText(_("Ready"))
        self.statusBar.addWidget(self.statusbar_label, 1)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setVisible(False)
        self.statusBar.addPermanentWidget(self.progress_bar)

        self.counts_label = QLabel()
        self.statusBar.addPermanentWidget(self.counts_label)
        self.update_counts_display()

        if hasattr(self, 'plugin_manager'):
            plugin_widgets = self.plugin_manager.run_hook('add_statusbar_widgets')
            if plugin_widgets:
                flat_widgets = [widget for sublist in plugin_widgets for widget in sublist]
                for widget in flat_widgets:
                    self.statusBar.addPermanentWidget(widget)

        extra_info = []
        if not requests: extra_info.append(_("Hint: requests not found, AI translation is disabled."))
        if extra_info: self.update_statusbar(
            self.statusbar_label.text() + " | " + _("Hint: ") + ", ".join(extra_info) + ".")

    def update_statusbar(self, text, persistent=False):
        self.statusbar_label.setText(text)
        QApplication.processEvents()
        if not persistent:
            QTimer.singleShot(5000, lambda: self.clear_statusbar_if_unchanged(text))

    def clear_statusbar_if_unchanged(self, original_text):
        if self.statusbar_label.text() == original_text:
            self.statusbar_label.setText(_("Ready"))

    def update_counts_display(self):
        if not hasattr(self, 'translatable_objects') or not self.proxy_model:
            self.counts_label.setText(_("Displayed: 0/0 | Translated: 0 | Untranslated: 0 | Ignored: 0"))
            return

        displayed_count = self.proxy_model.rowCount()
        total_count = len(self.translatable_objects)

        translated_visible = 0
        untranslated_visible = 0
        ignored_visible = 0

        for i in range(self.proxy_model.rowCount()):
            index = self.proxy_model.index(i, 0)
            ts_obj = self.proxy_model.data(index, Qt.UserRole)
            if ts_obj:
                if ts_obj.is_ignored:
                    ignored_visible += 1
                elif ts_obj.translation.strip():
                    translated_visible += 1
                else:
                    untranslated_visible += 1

        self.counts_label.setText(
            _("Displayed: {displayed_count}/{total_count} | Translated: {translated_visible} | Untranslated: {untranslated_visible} | Ignored: {ignored_visible}").format(
                displayed_count=displayed_count,
                total_count=total_count,
                translated_visible=translated_visible,
                untranslated_visible=untranslated_visible,
                ignored_visible=ignored_visible
            )
        )

    def update_title(self):
        base_title = f"LexiSync - v{APP_VERSION}"
        name_part = ""
        modified_indicator = "*" if self.current_project_modified else ""

        if self.is_project_mode:
            project_name = self.project_config.get('name', os.path.basename(self.current_project_path or ""))
            lang_name = self.target_lang_combo.currentText()
            name_part = f"{project_name} ({lang_name})"
        else:
            if self.current_po_file_path:
                name_part = os.path.basename(self.current_po_file_path)
            elif self.current_code_file_path:
                name_part = os.path.basename(self.current_code_file_path)

        if name_part:
            self.setWindowTitle(f"{base_title} - {name_part}{modified_indicator}")
        else:
            self.setWindowTitle(base_title)

    def update_ui_state_after_file_load(self, file_or_project_loaded=False):
        has_content = bool(self.translatable_objects) and file_or_project_loaded

        self.action_save_current_file.setEnabled(has_content)
        self.action_save_current_file_as.setEnabled(has_content)
        self.action_compare_new_version.setEnabled(has_content)

        if self.is_po_mode:
            self.action_save_code_file.setEnabled(False)
            self.action_export_po.setEnabled(has_content)
        else:
            can_save_to_code = bool(self.original_raw_code_content) and has_content
            self.action_save_code_file.setEnabled(can_save_to_code)
            self.action_export_po.setEnabled(False)

        self.action_import_excel.setEnabled(has_content)
        self.action_export_excel.setEnabled(has_content)
        self.action_export_json.setEnabled(has_content)
        self.action_export_yaml.setEnabled(has_content)

        self.action_find_replace.setEnabled(has_content)
        self.action_copy_original.setEnabled(has_content and self.current_selected_ts_id is not None)
        self.action_paste_translation.setEnabled(has_content and self.current_selected_ts_id is not None)

        self.action_undo.setEnabled(bool(self.undo_history))
        self.action_redo.setEnabled(bool(self.redo_history))

        self.action_apply_tm_to_untranslated.setEnabled(has_content)
        self.action_reload_translatable_text.setEnabled(
            bool(self.original_raw_code_content or self.current_code_file_path)
        )
        self.action_show_statistics.setEnabled(has_content)
        self.action_run_validation_on_all.setEnabled(has_content)

        self.update_ai_related_ui_state()
        self.update_title()

    def update_ai_related_ui_state(self):
        ai_available = requests is not None
        file_loaded_and_has_strings = bool(self.translatable_objects)
        item_selected = self.current_selected_ts_id is not None
        can_start_ai_ops = ai_available and file_loaded_and_has_strings and not self.is_ai_translating_batch

        self.action_ai_translate_selected.setEnabled(can_start_ai_ops and item_selected)
        self.action_ai_translate_all_untranslated.setEnabled(can_start_ai_ops)
        self.action_stop_ai_batch_translation.setEnabled(self.is_ai_translating_batch)

        self.details_panel.ai_translate_current_btn.setEnabled(can_start_ai_ops and item_selected)

        if self.is_ai_translating_batch:
            self.progress_bar.setVisible(True)
        else:
            self.progress_bar.setVisible(False)

    def mark_project_modified(self, modified=True):
        if self.current_project_modified != modified:
            self.current_project_modified = modified
            self.update_title()

    def add_to_undo_history(self, action_type, data):
        self.undo_history.append({'type': action_type, 'data': deepcopy(data)})
        if len(self.undo_history) > MAX_UNDO_HISTORY:
            self.undo_history.pop(0)
        self.redo_history.clear()
        self.action_undo.setEnabled(True)
        self.action_redo.setEnabled(False)
        self.mark_project_modified()

    def _find_ts_obj_by_id(self, obj_id):
        for ts_obj in self.translatable_objects:
            if ts_obj.id == obj_id:
                return ts_obj
        return None

    def undo_action(self):
        focused_widget = QApplication.focusWidget()
        if isinstance(focused_widget, QTextEdit) and focused_widget.document().isUndoAvailable():
            focused_widget.undo()
            return

        if not self.undo_history:
            self.update_statusbar(_("No more actions to undo"))
            return
        action_to_undo = self.undo_history[-1]
        was_handled_by_plugin = self.plugin_manager.run_hook(
            'on_before_undo_redo',
            action_type=action_to_undo['type'],
            is_undo=True,
            action_data=action_to_undo['data']
        )
        if was_handled_by_plugin:
            return
        action_log = self.undo_history.pop()
        action_type, action_data = action_log['type'], action_log['data']
        redo_payload_data = None
        changed_ids = set()

        if action_type == 'single_change':
            obj_id = action_data['string_id']
            field = action_data['field']
            val_to_restore = action_data['old_value']

            ts_obj = self._find_ts_obj_by_id(obj_id)
            if ts_obj:
                current_val_before_undo = getattr(ts_obj, field) if field != 'translation' else ts_obj.get_translation_for_storage_and_tm()
                if field == 'translation':
                    ts_obj.set_translation_internal(val_to_restore.replace("\\n", "\n"))
                else:
                    setattr(ts_obj, field, val_to_restore)
                redo_payload_data = {'string_id': obj_id, 'field': field, 'old_value': val_to_restore, 'new_value': current_val_before_undo}
                self.update_statusbar(_("Undo: {field} for ID {id} -> '{value}'").format(field=field, id=str(obj_id)[:8] + "...", value=str(val_to_restore)[:30]))
                changed_ids.add(obj_id)
            else:
                self.update_statusbar(_("Undo error: Object ID {obj_id} not found").format(obj_id=obj_id))
                self.action_redo.setEnabled(bool(self.redo_history))
                return

        elif action_type in ['bulk_change', 'bulk_excel_import', 'bulk_ai_translate', 'bulk_context_menu', 'bulk_replace_all']:
            temp_redo_changes = []
            for item_change in action_data['changes']:
                obj_id, field, val_to_restore = item_change['string_id'], item_change['field'], item_change['old_value']
                ts_obj = self._find_ts_obj_by_id(obj_id)
                if ts_obj:
                    current_val_before_undo = getattr(ts_obj, field) if field != 'translation' else ts_obj.get_translation_for_storage_and_tm()
                    if field == 'translation':
                        ts_obj.set_translation_internal(val_to_restore.replace("\\n", "\n"))
                    else:
                        setattr(ts_obj, field, val_to_restore)
                    temp_redo_changes.append({'string_id': obj_id, 'field': field, 'old_value': val_to_restore, 'new_value': current_val_before_undo})
                    changed_ids.add(obj_id)
            redo_payload_data = {'changes': temp_redo_changes}
            self.update_statusbar(_("Undo: Bulk change ({count} items)").format(count=len(temp_redo_changes)))

        if redo_payload_data:
            self.redo_history.append({'type': action_type, 'data': redo_payload_data})

        if self.use_static_sorting_var:
            self._update_view_for_ids(changed_ids)
        else:
            self._run_and_refresh_with_validation()

        if self.current_selected_ts_id in changed_ids:
            self.force_refresh_ui_for_current_selection()

        self.action_undo.setEnabled(bool(self.undo_history))
        self.action_redo.setEnabled(bool(self.redo_history))
        self.mark_project_modified()

    def redo_action(self):
        focused_widget = QApplication.focusWidget()
        if isinstance(focused_widget, QTextEdit) and focused_widget.document().isRedoAvailable():
            focused_widget.redo()
            return

        if not self.redo_history:
            self.update_statusbar(_("No more actions to redo"))
            return
        action_to_redo = self.redo_history[-1]
        was_handled_by_plugin = self.plugin_manager.run_hook(
            'on_before_undo_redo',
            action_type=action_to_redo['type'],
            is_undo=False,
            action_data=action_to_redo['data']
        )
        if was_handled_by_plugin:
            return
        action_log = self.redo_history.pop()
        action_type, action_data_to_apply = action_log['type'], action_log['data']
        undo_payload_data = None
        changed_ids = set()

        if action_type == 'single_change':
            obj_id = action_data_to_apply['string_id']
            field = action_data_to_apply['field']
            val_to_set = action_data_to_apply['new_value']

            ts_obj = self._find_ts_obj_by_id(obj_id)
            if ts_obj:
                current_val_before_redo = getattr(ts_obj, field) if field != 'translation' else ts_obj.get_translation_for_storage_and_tm()
                if field == 'translation':
                    ts_obj.set_translation_internal(val_to_set.replace("\\n", "\n"))
                else:
                    setattr(ts_obj, field, val_to_set)
                undo_payload_data = {'string_id': obj_id, 'field': field, 'old_value': current_val_before_redo, 'new_value': val_to_set}
                self.update_statusbar(_("Redo: {field} for ID {id} -> '{value}'").format(field=field, id=str(obj_id)[:8] + "...", value=str(val_to_set)[:30]))
                changed_ids.add(obj_id)
            else:
                self.update_statusbar(_("Redo error: Object ID {obj_id} not found").format(obj_id=obj_id))
                self.action_undo.setEnabled(bool(self.undo_history))
                return

        elif action_type in ['bulk_change', 'bulk_excel_import', 'bulk_ai_translate', 'bulk_context_menu', 'bulk_replace_all']:
            temp_undo_changes = []
            for item_change in action_data_to_apply['changes']:
                obj_id, field, val_to_set = item_change['string_id'], item_change['field'], item_change['new_value']
                ts_obj = self._find_ts_obj_by_id(obj_id)
                if ts_obj:
                    current_val_before_redo = getattr(ts_obj, field) if field != 'translation' else ts_obj.get_translation_for_storage_and_tm()
                    if field == 'translation':
                        ts_obj.set_translation_internal(val_to_set.replace("\\n", "\n"))
                    else:
                        setattr(ts_obj, field, val_to_set)
                    temp_undo_changes.append({'string_id': obj_id, 'field': field, 'old_value': current_val_before_redo, 'new_value': val_to_set})
                    changed_ids.add(obj_id)
            undo_payload_data = {'changes': temp_undo_changes}
            self.update_statusbar(_("Redo: Bulk change ({count} items)").format(count=len(temp_undo_changes)))

        if undo_payload_data:
            self.undo_history.append({'type': action_type, 'data': undo_payload_data})
            if len(self.undo_history) > MAX_UNDO_HISTORY:
                self.undo_history.pop(0)

        if self.use_static_sorting_var:
            self._update_view_for_ids(changed_ids)
        else:
            self._run_and_refresh_with_validation()

        if self.current_selected_ts_id in changed_ids:
            self.force_refresh_ui_for_current_selection()

        self.action_redo.setEnabled(bool(self.redo_history))
        self.action_undo.setEnabled(bool(self.undo_history))
        self.mark_project_modified()

    def closeEvent(self, event):
        if not self.prompt_save_if_modified():
            event.ignore()
            return

        if self.is_ai_translating_batch:
            reply = QMessageBox.question(self, _("AI Translation in Progress"),
                                         _("AI batch translation is still in progress. Are you sure you want to exit?\nUnfinished translations will be lost."),
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                event.ignore()
                return
            else:
                self.stop_batch_ai_translation(silent=True)
        self.tm_service.disconnect_databases()
        self.glossary_service.disconnect_databases()
        self.save_config()
        self.save_window_state()
        if hasattr(self, 'plugin_manager'):
            self.plugin_manager.run_hook('on_app_shutdown')
        event.accept()

    def save_config(self):
        self.config["show_ignored"] = self.ignored_checkbox.isChecked()
        self.config["show_untranslated"] = self.untranslated_checkbox.isChecked()
        self.config["show_translated"] = self.translated_checkbox.isChecked()
        self.config["show_unreviewed"] = self.unreviewed_checkbox.isChecked()
        self.config["use_static_sorting"] = self.use_static_sorting_var
        self.config["auto_save_tm"] = self.auto_save_tm_var
        self.config["auto_backup_tm_on_save"] = self.auto_backup_tm_on_save_var
        self.config["auto_compile_mo_on_save"] = self.auto_compile_mo_var
        self.config["auto_save_interval_sec"] = self.auto_save_interval_sec
        current_search_text = self.search_entry.text()
        if current_search_text == _("Quick search..."):
            self.config["ui_state"]["search_term"] = ""
        else:
            self.config["ui_state"]["search_term"] = current_search_text
        self.config["ui_state"]["selected_ts_id"] = self.current_selected_ts_id or ""

        config_manager.save_config(self)

    def save_window_state(self):
        self.config["window_state"] = self.saveState().toBase64().data().decode('utf-8')
        self.config["window_geometry"] = self.saveGeometry().toBase64().data().decode('utf-8')
        config_manager.save_config(self)

    def restore_window_state(self):
        if "window_state" in self.config and self.config["window_state"]:
            self.restoreState(QByteArray.fromBase64(self.config["window_state"].encode('utf-8')))
        else:
            QTimer.singleShot(0, self.restore_default_layout)


    def add_to_recent_files(self, filepath):
        if not filepath: return
        recent_files = self.config.get("recent_files", [])
        if filepath in recent_files:
            recent_files.remove(filepath)
        recent_files.insert(0, filepath)
        self.config["recent_files"] = recent_files[:10]
        self.update_recent_files_menu()
        self.save_config()

    def update_recent_files_menu(self):
        self.recent_files_menu.clear()
        recent_files = self.config.get("recent_files", [])
        if not recent_files:
            self.recent_files_menu.setEnabled(False)
            self.recent_files_menu.addAction(QAction(_("No History"), self, enabled=False))
            return
        self.recent_files_menu.setEnabled(True)

        for i, filepath in enumerate(recent_files):
            label = f"{i + 1}: {os.path.basename(filepath)}"
            action = QAction(label, self)
            action.triggered.connect(lambda checked, p=filepath: self.open_recent_file(p))
            self.recent_files_menu.addAction(action)
        self.recent_files_menu.addSeparator()
        clear_action = QAction(_("Clear History"), self)
        clear_action.triggered.connect(self.clear_recent_files)
        self.recent_files_menu.addAction(clear_action)

    def open_recent_file(self, filepath):
        if not os.path.exists(filepath):
            QMessageBox.critical(self, _("File not found"),
                                 _("File '{filepath}' does not exist.").format(filepath=filepath))
            recent_files = self.config.get("recent_files", [])
            if filepath in recent_files:
                recent_files.remove(filepath)
                self.config["recent_files"] = recent_files
                self.update_recent_files_menu()
            return False

        if not self.prompt_save_if_modified():
            return False
        return self.open_file_by_path(filepath)

    def open_file_by_path(self, filepath: str) -> bool:
        if not filepath or not os.path.exists(filepath):
            return False
        lower_path = filepath.lower()
        try:
            if os.path.isdir(filepath) and os.path.exists(os.path.join(filepath, "project.json")):
                self.open_project(filepath)
                return True
            if lower_path.endswith((".ow", ".txt")):
                self.open_code_file_path(filepath)
            elif lower_path.endswith((".po", ".pot")):
                self.import_po_file_dialog_with_path(filepath)
            else:
                if hasattr(self, 'plugin_manager'):
                    if self.plugin_manager.run_hook('on_file_dropped', filepath):
                        return True
                QMessageBox.warning(self, _("Unsupported File"),
                                    _("The file type of '{filename}' is not supported.").format(
                                        filename=os.path.basename(filepath)))
                return False
            return bool(self.translatable_objects)
        except Exception as e:
            QMessageBox.critical(self, _("Error"), _("Failed to open file '{filename}':\n{error}").format(
                filename=os.path.basename(filepath), error=str(e)))
            return False

    def clear_recent_files(self):
        reply = QMessageBox.question(self, _("Confirmation"),
                                     _("Are you sure you want to clear all recent file history?"),
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.config["recent_files"] = []
            self.update_recent_files_menu()
            self.save_config()

    def about(self):
        QMessageBox.about(self, _("About"),
                          _("LexiSync\n\n"
                            "Version: {version}\n"
                            "Author: TheSkyC\n"
                            "China Server ID: 小鸟游六花#56683 / Asia Server: 小鳥游六花#31665").format(
                              version=APP_VERSION))

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        dropped_files = [url.toLocalFile() for url in event.mimeData().urls() if url.isLocalFile()]
        if not dropped_files:
            event.acceptProposedAction()
            return
        was_handled_by_plugins = False
        if hasattr(self, 'plugin_manager'):
            was_handled_by_plugins = self.plugin_manager.run_hook('on_files_dropped', dropped_files)
        if was_handled_by_plugins:
            event.acceptProposedAction()
            return
        filepath = dropped_files[0]
        if len(dropped_files) > 1:
            self.update_statusbar(_("Multiple files dropped. Opening the first one: {filename}").format(
                filename=os.path.basename(filepath)))

        if os.path.isfile(filepath):
            if filepath.lower().endswith(".pot"):
                self.handle_pot_file_drop(filepath)
            elif filepath.lower().endswith((".ow", ".txt", ".po")):
                if self.prompt_save_if_modified():
                    self.open_file_by_path(filepath)
            else:
                was_handled = False
                if hasattr(self, 'plugin_manager'):
                    was_handled = self.plugin_manager.run_hook('on_file_dropped', filepath)

                if not was_handled:
                    self.update_statusbar(_("Drag and drop failed: Invalid file type '{filename}'").format(
                        filename=os.path.basename(filepath)))
        elif os.path.isdir(filepath):
            if os.path.exists(os.path.join(filepath, "project.json")):
                if self.prompt_save_if_modified():
                    self.open_project(filepath)
            else:
                self.file_explorer_panel.set_root_path(filepath)
                self.update_statusbar(
                    _("Set file explorer root to: {directory}").format(directory=os.path.basename(filepath)))
        else:
            self.update_statusbar(_("Drag and drop failed: '{filename}' is not a file.").format(
                filename=os.path.basename(filepath)))
        event.acceptProposedAction()
        event.acceptProposedAction()

    def new_project_dialog(self):
        if not self.prompt_save_if_modified():
            return False

        dialog = NewProjectDialog(self, self)
        if dialog.exec():
            data = dialog.get_data()
            project_name = data['name']
            project_location = data['location']

            for sf in data['source_files']:
                sf['patterns'] = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)

            try:
                project_path = os.path.join(project_location, project_name)
                new_project_path = create_project(
                    project_path,
                    project_name,
                    data['source_lang'],
                    data['target_langs'],
                    data['source_files'],
                    data['use_global_tm']
                )
                self.open_project(new_project_path)
                return True
            except Exception as e:
                QMessageBox.critical(self, _("Project Creation Failed"), str(e))
        return False

    def open_project(self, project_path):
        if self.is_ai_translating_batch:
            QMessageBox.warning(self, _("Operation Restricted"), _("AI batch translation is in progress."))
            return
        try:
            self._reset_app_state()
            loaded_data = load_project(project_path)

            self.current_project_path = project_path
            self.is_project_mode = True
            self.project_config = loaded_data["project_config"]
            self.translatable_objects = loaded_data["translatable_objects"]
            self.original_raw_code_content = loaded_data["original_raw_code_content"]

            self.source_language = self.project_config.get("source_language", "en")
            self.target_languages = self.project_config.get("target_languages", [])
            self.current_target_language = self.project_config.get("current_target_language")

            self.setup_tm_service()
            self.setup_glossary_service()

            self.plugin_manager.run_hook('on_project_loaded', self.translatable_objects)

            self.add_to_recent_files(project_path)
            self.config["last_dir"] = os.path.dirname(project_path)
            self.save_config()

            self._update_language_switcher()
            self._run_and_refresh_with_validation()
            self.mark_project_modified(False)

            self.update_statusbar(_("Project '{name}' loaded.").format(name=self.project_config.get('name')),
                                  persistent=True)
            self.update_ui_state_after_file_load(file_or_project_loaded=True)
            self.project_toolbar.setVisible(True)

        except Exception as e:
            QMessageBox.critical(self, _("Open Project Error"),
                                 _("Could not load project from '{path}': {error}").format(
                                     path=os.path.basename(project_path), error=e))
            self._reset_app_state()
            self.update_statusbar(_("Project loading failed."), persistent=True)
        self.update_counts_display()

    def _update_language_switcher(self):
        self.target_lang_combo.blockSignals(True)
        self.target_lang_combo.clear()
        for lang_code in self.target_languages:
            lang_name = next((name for name, code in SUPPORTED_LANGUAGES.items() if code == lang_code), lang_code)
            self.target_lang_combo.addItem(lang_name, lang_code)

        current_index = self.target_lang_combo.findData(self.current_target_language)
        if current_index != -1:
            self.target_lang_combo.setCurrentIndex(current_index)
        self.target_lang_combo.blockSignals(False)

    def _on_target_language_changed(self, index):
        new_lang = self.target_lang_combo.itemData(index)
        if not new_lang or new_lang == self.current_target_language:
            return

        if not self.prompt_save_if_modified():
            self.target_lang_combo.blockSignals(True)
            current_index = self.target_lang_combo.findData(self.current_target_language)
            self.target_lang_combo.setCurrentIndex(current_index)
            self.target_lang_combo.blockSignals(False)
            return
        if self.current_project_modified:
            save_project(self.current_project_path, self)
        self.current_target_language = new_lang
        self.open_project(self.current_project_path)

    def open_code_file_dialog(self):
        if not self.prompt_save_if_modified(): return False
        dialog_title = _("Open File for Quick Edit")
        file_filters = (
                _("All Supported Files (*.ow *.txt *.po *.pot);;") +
                _("Overwatch Workshop Files (*.ow *.txt);;") +
                _("PO Translation Files (*.po *.pot);;") +
                _("All Files (*.*)")
        )
        filepath, selected_filter = QFileDialog.getOpenFileName(
            self,
            dialog_title,
            self.config.get("last_dir", os.getcwd()),
            file_filters
        )
        if filepath:
            file_ext = os.path.splitext(filepath)[1].lower()
            if file_ext in ['.ow', '.txt']:
                self.open_code_file_path(filepath)
            elif file_ext in ['.po', '.pot']:
                self.import_po_file_dialog_with_path(filepath)
            else:
                self.open_code_file_path(filepath)
            return True
        return False

    def open_project_dialog(self):
        if not self.prompt_save_if_modified():
            return False
        project_path = QFileDialog.getExistingDirectory(
            self,
            _("Open Project Folder"),
            self.config.get("last_dir", os.getcwd())
        )
        if project_path and os.path.exists(os.path.join(project_path, "project.json")):
            self.open_project(project_path)
            return True
        elif project_path:
            QMessageBox.warning(self, _("Invalid Project"), _("The selected folder is not a valid LexiSync project."))
            return False

        return False

    def open_code_file_path(self, filepath):
        if self.is_ai_translating_batch:
            QMessageBox.warning(self, _("Operation Restricted"),
                                _("AI batch translation is in progress. Please wait for it to complete or stop it before opening a new file."))
            return
        self._reset_app_state()
        try:
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                original_content = f.read()
            processed_content = self.plugin_manager.run_hook(
                'process_raw_content_before_extraction',
                original_content,
                filepath=filepath
            )
            if not isinstance(processed_content, str):
                logger.warning(
                    f"Warning: Plugin hook 'process_raw_content_before_extraction' returned a non-string value (type: {type(processed_content)}). Reverting to original content.")
                content = original_content
            else:
                content = processed_content
            content = content.replace('\r\n', '\n')
            self.original_raw_code_content = content

            self._update_file_explorer(filepath)
            self.current_code_file_path = filepath
            self.current_project_path = None
            self.current_po_file_path = None
            self.current_po_metadata = None
            self.add_to_recent_files(filepath)
            self.config["last_dir"] = os.path.dirname(filepath)
            self.save_config()
            self.update_statusbar(_("Extracting strings..."), persistent=True)
            QApplication.processEvents()
            extraction_patterns = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)
            extraction_patterns = self.plugin_manager.run_hook(
                'process_extraction_patterns',
                extraction_patterns,
                filepath=filepath,
                raw_content=self.original_raw_code_content
            )
            extraction_patterns = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)
            self.translatable_objects = extract_translatable_strings(self.original_raw_code_content,
                                                                     extraction_patterns)
            self.plugin_manager.run_hook('on_project_loaded', self.translatable_objects)
            detected_lang = language_service.detect_source_language([ts.original_semantic for ts in self.translatable_objects])
            self.source_language = detected_lang
            self.target_language = self.config.get("default_target_language", "zh")
            self.apply_tm_to_all_current_strings(silent=True, only_if_empty=True)
            self.undo_history.clear()
            self.redo_history.clear()
            self.current_selected_ts_id = None
            self.mark_project_modified(False)
            self.is_project_mode = False
            self.is_po_mode = False
            self._run_and_refresh_with_validation()
            self.update_statusbar(
                _("Loaded {count} translatable strings from {filename}").format(count=len(self.translatable_objects),
                                                                                filename=os.path.basename(filepath)),
                persistent=True)
            self.update_ui_state_after_file_load(file_or_project_loaded=True)

        except Exception as e:
            QMessageBox.critical(self, _("Error"), _("Could not open or parse code file '{filename}': {error}").format(
                filename=os.path.basename(filepath), error=e))
            self._reset_app_state()
            self.update_statusbar(_("Code file loading failed"), persistent=True)
        self.update_counts_display()

    def open_file_from_explorer(self, file_path):
        if not os.path.isfile(file_path):
            self.update_statusbar(
                _("Error: '{filename}' is not a valid file.").format(filename=os.path.basename(file_path)),
                persistent=True)
            return

        lower_path = file_path.lower()
        filename = os.path.basename(file_path)

        if not self.prompt_save_if_modified():
            self.update_statusbar(_("Open operation cancelled by user."), persistent=False)
            return

        self.update_statusbar(_("Opening '{filename}'...").format(filename=filename), persistent=True)
        QApplication.processEvents()

        try:
            self.open_file_by_path(file_path)
        except Exception as e:
            error_message = _("Failed to open '{filename}': {error}").format(filename=filename, error=str(e))
            self.update_statusbar(error_message, persistent=True)
            QMessageBox.critical(self, _("File Open Error"), error_message)

    def _reset_app_state(self):
        self.current_code_file_path = None
        self.current_project_path = None
        self.is_project_mode = False
        self.setup_tm_service()
        self.setup_glossary_service()
        self.project_config = {}
        self.current_po_file_path = None
        self.current_po_metadata = None
        self.original_raw_code_content = ""
        self.translatable_objects = []
        self.undo_history.clear()
        self.redo_history.clear()
        self.current_selected_ts_id = None
        self.mark_project_modified(False)
        self.refresh_sheet()
        self.clear_details_pane()
        self.update_ui_state_after_file_load(file_or_project_loaded=False)
        self.update_title()

    def handle_pot_file_drop(self, pot_filepath):
        if not self.translatable_objects:
            if self.prompt_save_if_modified():
                self.import_po_file_dialog_with_path(pot_filepath)
            return

        dialog = POTDropDialog(self)
        if dialog.exec():
            if dialog.result == "update":
                self.run_comparison_with_file(pot_filepath)
            elif dialog.result == "import":
                if self.prompt_save_if_modified():
                    self.import_po_file_dialog_with_path(pot_filepath)

    def run_comparison_with_file(self, filepath):
        self._run_comparison_logic(filepath)

    def prompt_save_if_modified(self):
        if self.current_project_modified:
            reply = QMessageBox.question(self, _("Unsaved Changes"),
                                         _("The current project has unsaved changes. Do you want to save?"),
                                         QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
            if reply == QMessageBox.Yes:
                return self.save_current_file()
            elif reply == QMessageBox.No:
                return True
            else:
                return False
        return True

    def refresh_sheet_preserve_selection(self, item_to_reselect_after=None):
        self.refresh_sheet(preserve_selection=True, item_to_reselect_after=item_to_reselect_after)

    def refresh_sheet(self, preserve_selection=True, item_to_reselect_after=None):
        old_selected_id = self.current_selected_ts_id
        if item_to_reselect_after:
            old_selected_id = item_to_reselect_after
        self.proxy_model.set_filters(
            show_ignored=self.show_ignored_var,
            show_untranslated=self.show_untranslated_var,
            show_translated=self.show_translated_var,
            show_unreviewed=self.show_unreviewed_var,
            search_term=self.search_entry.text() if self.search_entry.text() != _("Quick search...") else "",
            is_po_mode=self.is_po_mode
        )
        self.update_counts_display()
        if preserve_selection and old_selected_id:
            self.select_sheet_row_by_id(old_selected_id, see=True)
        elif not self.current_selected_ts_id and self.proxy_model.rowCount() > 0:
            first_index = self.proxy_model.index(0, 0)
            self.table_view.selectionModel().setCurrentIndex(first_index,
                                                             QItemSelectionModel.ClearAndSelect | QItemSelectionModel.Rows)

        if not self.current_selected_ts_id:
            self.clear_details_pane()

    def search_filter_changed(self, text):
        self._last_quick_search_text = text
        self.quick_search_timer.start(80)

    def _perform_delayed_search_filter(self):
        search_term_to_use = self._last_quick_search_text
        self.proxy_model.set_filters(
            show_ignored=self.ignored_checkbox.isChecked(),
            show_untranslated=self.untranslated_checkbox.isChecked(),
            show_translated=self.translated_checkbox.isChecked(),
            show_unreviewed=self.unreviewed_checkbox.isChecked(),
            search_term=search_term_to_use if search_term_to_use != _("Quick search...") else "",
            is_po_mode=self.is_po_mode
        )

        self.update_counts_display()

    def find_string_from_toolbar(self):
        if self.quick_search_timer.isActive():
            self.quick_search_timer.stop()
        if self.search_entry.text().lower() != self.proxy_model.search_term:
            self._perform_delayed_search_filter()
        if self.proxy_model.rowCount() > 0:
            first_index = self.proxy_model.index(0, 0)
            self.table_view.selectionModel().setCurrentIndex(first_index,
                                                             QItemSelectionModel.ClearAndSelect | QItemSelectionModel.Rows)
            self.table_view.scrollTo(first_index, QAbstractItemView.PositionAtTop)
            self.update_statusbar(_("Filtered by '{search_term}'.").format(search_term=self.search_entry.text()))
        elif self.search_entry.text() and self.search_entry.text() != _("Quick search..."):
            self.update_statusbar(_("No matches found for '{search_term}' under current filters.").format(
                search_term=self.search_entry.text()))
        else:
            self.update_statusbar(_("Search cleared."))

        if not self.search_entry.text() and self.search_entry.placeholderText() != _("Quick search..."):
            self.search_entry.setPlaceholderText(_("Quick search..."))
            self.search_entry.setStyleSheet("color: grey;")

    def _on_validation_complete(self, ts_id, slow_warnings, slow_minor_warnings):
        ts_obj = self._find_ts_obj_by_id(ts_id)
        if not ts_obj:
            return
        ts_obj.warnings.extend(slow_warnings)
        ts_obj.minor_warnings.extend(slow_minor_warnings)
        ts_obj.update_style_cache()
        source_index = self.sheet_model.index_from_id(ts_id)
        if source_index.isValid():
            first_col = source_index.siblingAtColumn(0)
            last_col = source_index.siblingAtColumn(self.sheet_model.columnCount() - 1)
            self.sheet_model.dataChanged.emit(first_col, last_col)

    def on_search_focus_in(self, event):
        if self.search_entry.text() == _("Quick search..."):
            self.search_entry.setText("")
            self.search_entry.setStyleSheet("color: black;")
        QLineEdit.focusInEvent(self.search_entry, event)

    def on_search_focus_out(self, event=None):
        if not self.search_entry.text():
            self.search_entry.setText(_("Quick search..."))
            self.search_entry.setStyleSheet("color: grey;")
        if event:
            QLineEdit.focusOutEvent(self.search_entry, event)

    def _sort_sheet_column(self, logical_index):
        current_order = self.table_view.horizontalHeader().sortIndicatorOrder()
        self.table_view.sortByColumn(logical_index, current_order)
        self.update_counts_display()

    def show_sheet_context_menu(self, pos):
        index = self.table_view.indexAt(pos)
        if not index.isValid():
            return
        if not self.table_view.selectionModel().isRowSelected(index.row(), index.parent()):
            self.table_view.selectionModel().clearSelection()
            self.table_view.selectionModel().select(index, QItemSelectionModel.Select | QItemSelectionModel.Rows)
            self.on_sheet_select(index, index)

        context_menu = QMenu(self)
        context_menu.addAction(QAction(_("Copy Original"), self, triggered=self.cm_copy_original))
        context_menu.addAction(QAction(_("Copy Translation"), self, triggered=self.cm_copy_translation))
        context_menu.addSeparator()
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if selected_objs:
            first_obj = selected_objs[0]
            action_ignore = QAction(_("Mark as Ignored"), self, triggered=lambda: self.cm_set_ignored_status(True))
            action_unignore = QAction(_("Unmark as Ignored"), self, triggered=lambda: self.cm_set_ignored_status(False))
            if first_obj.is_ignored:
                context_menu.addAction(action_unignore)
            else:
                context_menu.addAction(action_ignore)
            context_menu.addSeparator()
            action_reviewed = QAction(_("Mark as Reviewed"), self, triggered=lambda: self.cm_set_reviewed_status(True))
            action_unreviewed = QAction(_("Mark as Unreviewed"), self,
                                        triggered=lambda: self.cm_set_reviewed_status(False))
            if first_obj.is_reviewed:
                context_menu.addAction(action_unreviewed)
            else:
                context_menu.addAction(action_reviewed)
            context_menu.addSeparator()
            action_ignore_warning = QAction(_("Ignore Warnings for Selected"), self,
                                            triggered=lambda: self.cm_set_warning_ignored_status(True))
            action_unignore_warning = QAction(_("Un-ignore Warnings for Selected"), self,
                                              triggered=lambda: self.cm_set_warning_ignored_status(False))
            if first_obj.is_warning_ignored:
                context_menu.addAction(action_unignore_warning)
            else:
                context_menu.addAction(action_ignore_warning)
            context_menu.addSeparator()

        context_menu.addAction(QAction(_("Edit Comment..."), self, triggered=self.cm_edit_comment))
        context_menu.addSeparator()
        context_menu.addAction(
            QAction(_("Apply Memory to Selected Items"), self, triggered=self.cm_apply_tm_to_selected))
        context_menu.addAction(
            QAction(_("Clear Selected Translations"), self, triggered=self.cm_clear_selected_translations))
        context_menu.addSeparator()
        context_menu.addAction(
            QAction(_("Use AI to Translate Selected Items"), self, triggered=self.cm_ai_translate_selected))

        if hasattr(self, 'plugin_manager'):
            plugin_menu_items = self.plugin_manager.run_hook('on_table_context_menu', selected_ts_objects=selected_objs)
            if plugin_menu_items:
                context_menu.addSeparator()
                self.plugin_manager._create_menu_from_structure(context_menu, plugin_menu_items)

        context_menu.exec(self.table_view.viewport().mapToGlobal(pos))

    def on_sheet_double_click(self, index):
        if not index.isValid(): return

        col = index.column()
        if col == 2:  # Original
            self.details_panel.original_text_display.setFocus()
        elif col == 3:  # Translation
            self.details_panel.translation_edit_text.setFocus()
        elif col == 4:  # Comment
            self.comment_status_panel.comment_edit_text.setFocus()

    def _on_selection_model_changed(self, selected, deselected):
        if hasattr(self, 'plugin_manager'):
            selected_objects = self._get_selected_ts_objects_from_sheet()
            self.plugin_manager.run_hook('on_selection_changed', selected_ts_objects=selected_objects)

    def on_sheet_select(self, current_index, previous_index):
        if self.neighbor_select_timer.isActive():
            self.neighbor_select_timer.stop()
        if self.table_view.is_dragging:
            return
        old_focused_id = self.current_focused_ts_id

        if not current_index.isValid():
            self.current_selected_ts_id = None
            self.current_focused_ts_id = None
            self.clear_details_pane()
            self.update_ui_state_for_selection(None)
            return
        else:
            ts_obj = self.proxy_model.data(current_index, Qt.UserRole)
            if not ts_obj:
                self.current_selected_ts_id = None
                self.current_focused_ts_id = None
                self.clear_details_pane()
                self.update_ui_state_for_selection(None)
                return

            newly_focused_id = ts_obj.id
            self.current_focused_ts_id = newly_focused_id
            if self.current_selected_ts_id != newly_focused_id:
                self.current_selected_ts_id = newly_focused_id
                self.force_refresh_ui_for_current_selection()
                self.update_ui_state_for_selection(self.current_selected_ts_id)
                self.trigger_glossary_analysis(ts_obj)
                self._update_details_panel_stats()
        self.tm_panel.clear_selected_tm_btn.setEnabled(True)
        old_source_index = self.sheet_model.index_from_id(old_focused_id)
        new_source_index = self.sheet_model.index_from_id(self.current_focused_ts_id)

        if old_source_index.isValid():
            self.sheet_model.dataChanged.emit(old_source_index,
                                              old_source_index.siblingAtColumn(self.sheet_model.columnCount() - 1))
        if new_source_index.isValid():
            self.sheet_model.dataChanged.emit(new_source_index,
                                              new_source_index.siblingAtColumn(self.sheet_model.columnCount() - 1))
        if self.is_finalizing_batch_translation:
            return
        status_message = _("Selected: \"{text}...\" (Line: {line_num})").format(
            text=ts_obj.original_semantic[:30].replace(chr(10), '↵'),
            line_num=ts_obj.line_num_in_file
        )
        warning_message = ""
        if ts_obj.warnings and not ts_obj.is_warning_ignored:
            messages = [msg for wt, msg in ts_obj.warnings]
            warning_message = "⚠️ " + " | ".join(messages)
        elif ts_obj.minor_warnings and not ts_obj.is_warning_ignored:
            messages = [msg for wt, msg in ts_obj.minor_warnings]
            warning_message = "💡 " + " | ".join(messages)
        self.update_statusbar(warning_message if warning_message else status_message, persistent=True)

    def schedule_details_panel_stats_update(self):
        self.stats_update_timer.start(100)

    def _update_file_explorer(self, file_path):
        if not file_path: return
        normalized_file_path = os.path.normpath(file_path)
        current_root = os.path.normpath(self.file_explorer_panel.source_model.rootPath())
        drive_file, _ = os.path.splitdrive(normalized_file_path)
        drive_root, _ = os.path.splitdrive(current_root)
        is_outside_current_root = False
        if drive_file.lower() != drive_root.lower():
            is_outside_current_root = True
        else:
            is_outside_current_root = os.path.commonpath([normalized_file_path, current_root]) != current_root
        if is_outside_current_root or current_root == os.path.normpath(QDir.rootPath()):
            root_path = None
            path_parts = normalized_file_path.split(os.sep)
            try:
                locales_index = path_parts.index('locales')
                root_path = os.sep.join(path_parts[:locales_index])
            except ValueError:
                root_path = os.path.dirname(normalized_file_path)

            if root_path and os.path.isdir(root_path):
                self.file_explorer_panel.set_root_path(root_path)
        QTimer.singleShot(100, lambda: self.file_explorer_panel.select_file(file_path))

    def _update_details_panel_stats(self):
        if not self.current_selected_ts_id:
            self.details_panel.update_stats_labels(None, None)
            return

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj:
            self.details_panel.update_stats_labels(None, None)
            return

        current_translation_text = self.details_panel.translation_edit_text.toPlainText()

        # 计算净化后的字符数
        orig_len = get_linguistic_length(ts_obj.original_semantic)
        trans_len = get_linguistic_length(current_translation_text)
        char_counts = (orig_len, trans_len)

        # 计算膨胀率
        actual_ratio = trans_len / orig_len if orig_len > 0 else None
        service = ExpansionRatioService.get_instance()
        expected_ratio = service.get_expected_ratio(
            self.source_language,
            self.target_language,
            ts_obj.original_semantic,
        )
        ratios = (actual_ratio, expected_ratio)

        self.details_panel.update_stats_labels(char_counts, ratios)

    def schedule_placeholder_validation(self):
        if self._placeholder_validation_job:
            self._placeholder_validation_job.stop()
        self._placeholder_validation_job = QTimer()
        self._placeholder_validation_job.setSingleShot(True)
        self._placeholder_validation_job.timeout.connect(self._update_all_highlights)
        self._placeholder_validation_job.start(100)

    def _update_all_highlights(self):
        if not self.current_selected_ts_id:
            return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj:
            return

        original_text_widget = self.details_panel.original_text_display
        translation_text_widget = self.details_panel.translation_edit_text

        original_placeholders = set(self.placeholder_regex.findall(ts_obj.original_semantic))
        translated_placeholders = set(self.placeholder_regex.findall(translation_text_widget.toPlainText()))

        self.details_panel.apply_placeholder_highlights(original_text_widget, translation_text_widget,
                                                        original_placeholders, translated_placeholders)

    def update_ui_state_for_selection(self, selected_id):
        state = True if selected_id else False
        self.action_copy_original.setEnabled(state)
        self.action_paste_translation.setEnabled(state)
        # DetailsPanel
        self.details_panel.apply_btn.setEnabled(state)
        can_ai_translate = state and self.config.get("ai_api_key") and requests is not None
        self.details_panel.ai_translate_current_btn.setEnabled(bool(can_ai_translate))
        # CommentPanel
        self.comment_status_panel.apply_comment_btn.setEnabled(state)
        self.comment_status_panel.ignore_checkbox.setEnabled(state)
        self.comment_status_panel.reviewed_checkbox.setEnabled(state)
        # TMPanel
        self.tm_panel.update_selected_tm_btn.setEnabled(state)

        if not selected_id:
            self.tm_panel.clear_selected_tm_btn.setEnabled(False)
        self.update_ai_related_ui_state()

    def trigger_glossary_analysis(self, ts_obj):
        if ts_obj.id in self.glossary_analysis_cache:
            self._handle_glossary_analysis_result(ts_obj.id, self.glossary_analysis_cache[ts_obj.id])
            return

        self.glossary_panel.clear_matches()
        worker = GlossaryAnalysisWorker(self, ts_obj.id, ts_obj.original_semantic)
        worker.signals.finished.connect(self._handle_glossary_analysis_result)
        self.ai_thread_pool.start(worker)

    def _handle_glossary_analysis_result(self, ts_id, matches):
        self.glossary_analysis_cache[ts_id] = matches
        if self.current_selected_ts_id == ts_id:
            self.glossary_panel.update_matches(matches)
            if self.original_highlighter:
                self.original_highlighter.setParent(None)

            self.original_highlighter = GlossaryHighlighter(self.details_panel.original_text_display.document(),
                                                            matches)

    def clear_details_pane(self):
        # 清空 DetailsPanel
        self.details_panel.original_text_display.setPlainText("")
        self.details_panel.translation_edit_text.setPlainText("")
        self.details_panel.update_stats_labels(None, None)
        self.details_panel.apply_btn.setEnabled(False)
        self.details_panel.ai_translate_current_btn.setEnabled(False)

        # 清空 CommentStatusPanel
        self.comment_status_panel.comment_edit_text.setPlainText("")
        self.comment_status_panel.apply_comment_btn.setEnabled(False)
        self.comment_status_panel.ignore_checkbox.setChecked(False)
        self.comment_status_panel.ignore_checkbox.setText(_("Ignore this string"))
        self.comment_status_panel.reviewed_checkbox.setChecked(False)
        self.comment_status_panel.ignore_checkbox.setEnabled(False)
        self.comment_status_panel.reviewed_checkbox.setEnabled(False)

        # 清空其他面板
        self.context_panel.set_context([])
        self.tm_panel.update_tm_suggestions(exact_match=None, fuzzy_matches=[])
        self.tm_panel.update_selected_tm_btn.setEnabled(False)
        self.tm_panel.clear_selected_tm_btn.setEnabled(False)

    def _run_and_refresh_with_validation(self):
        if not self.translatable_objects:
            self.sheet_model.set_translatable_objects([])
            self.proxy_model.invalidate()
            return
        self.update_statusbar(_("Validating all entries..."), persistent=True)
        QApplication.processEvents()
        run_validation_on_all(self.translatable_objects, self.config, self)

        for ts_obj in self.translatable_objects:
            ts_obj.update_style_cache()
        self.sheet_model.set_translatable_objects(self.translatable_objects)
        if self.use_static_sorting_var:
            self.proxy_model.invalidate()
        else:
            self.refresh_sheet_preserve_selection()
        self.force_refresh_ui_for_current_selection()
        self.update_statusbar(_("Validation complete."), persistent=False)

    def cm_set_warning_ignored_status(self, ignore_flag):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs:
            self.update_statusbar(_("No items selected to modify warning status."))
            return

        changes_for_undo = []
        for ts_obj in selected_objs:
            if ts_obj.is_warning_ignored != ignore_flag:
                old_value = ts_obj.is_warning_ignored

                changes_for_undo.append({
                    'string_id': ts_obj.id,
                    'field': 'is_warning_ignored',
                    'old_value': old_value,
                    'new_value': ignore_flag
                })

                ts_obj.is_warning_ignored = ignore_flag

        if changes_for_undo:
            self.add_to_undo_history('bulk_context_menu', {'changes': changes_for_undo})
            self.mark_project_modified()
            count = len(changes_for_undo)
            if ignore_flag:
                status_message = _("Ignored warnings for {count} item(s).").format(count=count)
            else:
                status_message = _("Un-ignored warnings for {count} item(s).").format(count=count)
            self.update_statusbar(status_message)
        else:
            self.update_statusbar(_("Selected item(s) already have the desired warning status."))

        self._run_and_refresh_with_validation()

    def _apply_translation_to_model(self, ts_obj, new_translation_from_ui, source="manual", run_validation=True):
        processed_translation = self.plugin_manager.run_hook(
            'process_string_for_save',
            new_translation_from_ui,
            ts_object=ts_obj,
            column='translation',
            source = source
        )
        if processed_translation == ts_obj.translation:
            return processed_translation, False
        old_translation_for_undo = ts_obj.get_translation_for_storage_and_tm()
        ids_to_update = {ts_obj.id}
        all_changes_for_undo_list = []
        ts_obj.set_translation_internal(processed_translation)
        self.plugin_manager.run_hook(
            'on_string_saved',
            ts_object=ts_obj,
            column='translation',
            new_value=processed_translation,
            old_value=old_translation_for_undo.replace("\\n", "\n")
        )
        new_translation_for_tm_storage = ts_obj.get_translation_for_storage_and_tm()
        primary_change_data = {
            'string_id': ts_obj.id, 'field': 'translation',
            'old_value': old_translation_for_undo, 'new_value': new_translation_for_tm_storage
        }
        all_changes_for_undo_list.append(primary_change_data)
        if new_translation_from_ui.strip():
            db_path_to_use = self.tm_service.project_db_path if self.is_project_mode else self.tm_service.global_db_path
            source_lang = self.source_language
            target_lang = self.current_target_language if self.is_project_mode else self.target_language

            if db_path_to_use:
                self.tm_service.update_tm_entry(
                    db_path_to_use, ts_obj.original_semantic,
                    ts_obj.get_translation_for_storage_and_tm(),
                    source_lang, target_lang
                )
        for other_ts_obj in self.translatable_objects:
            if other_ts_obj.id != ts_obj.id and \
                    other_ts_obj.original_semantic == ts_obj.original_semantic and \
                    other_ts_obj.translation != new_translation_from_ui:
                old_other_translation_for_undo = other_ts_obj.get_translation_for_storage_and_tm()
                other_ts_obj.set_translation_internal(new_translation_from_ui)
                self.plugin_manager.run_hook(
                    'on_string_saved',
                    ts_object=other_ts_obj,
                    column='translation',
                    new_value=new_translation_from_ui,
                    old_value=old_other_translation_for_undo.replace("\\n", "\n")
                )
                all_changes_for_undo_list.append({
                    'string_id': other_ts_obj.id, 'field': 'translation',
                    'old_value': old_other_translation_for_undo, 'new_value': new_translation_for_tm_storage
                })
                ids_to_update.add(other_ts_obj.id)
        undo_action_type = 'bulk_change' if len(all_changes_for_undo_list) > 1 else 'single_change'
        undo_data_payload = {'changes': all_changes_for_undo_list} if len(
            all_changes_for_undo_list) > 1 else primary_change_data

        if source not in ["ai_batch_item"]:
            self.add_to_undo_history(undo_action_type, undo_data_payload)
        self._update_view_for_ids(ids_to_update)

        self.update_statusbar(_("Translation applied: \"{original_semantic}...\"").format(
            original_semantic=ts_obj.original_semantic[:20].replace(chr(10), '↵')))
        self.mark_project_modified()
        return processed_translation, True

    def apply_translation_from_button(self):
        if not self.current_selected_ts_id: return
        if self.current_selected_ts_id == "##NEW_ENTRY##":
            new_original = self.details_panel.original_text_display.toPlainText().strip()
            if not new_original:
                QMessageBox.critical(self, _("Error"), _("Original text cannot be empty for a new entry."))
                return

            if any(ts.original_semantic == new_original for ts in self.translatable_objects):
                QMessageBox.critical(self, _("Error"), _("This original text already exists."))
                return

            new_ts = TranslatableString(
                original_raw=new_original, original_semantic=new_original,
                line_num=0, char_pos_start_in_file=0, char_pos_end_in_file=0, full_code_lines=[]
            )
            new_ts.translation = self.details_panel.translation_edit_text.toPlainText().strip()
            new_ts.comment = self.comment_status_panel.comment_edit_text.toPlainText().strip()

            self.translatable_objects.append(new_ts)
            self.mark_project_modified()
            self._run_and_refresh_with_validation()
            self.select_sheet_row_by_id(new_ts.id, see=True)
            return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return

        new_translation_ui = self.details_panel.translation_edit_text.toPlainText()
        final_text, changed = self._apply_translation_to_model(ts_obj, new_translation_ui, source="manual_button")
        if changed:
            self.details_panel.translation_edit_text.blockSignals(True)
            self.details_panel.translation_edit_text.setPlainText(final_text)
            self.details_panel.translation_edit_text.blockSignals(False)

    def apply_translation_focus_out(self):
        if not self.current_selected_ts_id: return

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return

        new_translation_ui = self.details_panel.translation_edit_text.toPlainText()
        if new_translation_ui != ts_obj.get_translation_for_ui():
            final_text, changed = self._apply_translation_to_model(ts_obj, new_translation_ui, source="manual_focus_out")
            if changed:
                self.details_panel.translation_edit_text.blockSignals(True)
                self.details_panel.translation_edit_text.setPlainText(final_text)
                self.details_panel.translation_edit_text.blockSignals(False)

    def _save_comment_from_ui(self):
        if not self.current_selected_ts_id: return False
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return False
        full_comment_text = self.comment_status_panel.comment_edit_text.toPlainText()

        old_po_comment_for_hook = ts_obj.po_comment
        old_user_comment_for_hook = ts_obj.comment

        old_po_lines = ts_obj.po_comment.splitlines()
        old_user_lines = ts_obj.comment.splitlines()
        old_full_text = "\n".join(old_po_lines + old_user_lines)
        new_po_lines = []
        new_user_lines = []
        lines = full_comment_text.splitlines()
        for line in lines:
            if line.strip().startswith('#'):
                new_po_lines.append(line)
            else:
                new_user_lines.append(line)
        has_fuzzy_in_new_comment = any('fuzzy' in line for line in new_po_lines if line.strip().startswith('#,'))
        new_is_fuzzy = has_fuzzy_in_new_comment
        if full_comment_text == old_full_text and new_is_fuzzy == ts_obj.is_fuzzy:
            return False
        new_po_comment = "\n".join(new_po_lines)
        new_user_comment = "\n".join(new_user_lines)
        old_is_fuzzy = ts_obj.is_fuzzy
        if full_comment_text == old_full_text and new_is_fuzzy == ts_obj.is_fuzzy:
            if hasattr(self.comment_status_panel, 'highlighter'):
                self.comment_status_panel.highlighter.rehighlight()
            return False
        self.add_to_undo_history('bulk_change', {
            'changes': [
                {'string_id': ts_obj.id, 'field': 'po_comment', 'old_value': ts_obj.po_comment,
                 'new_value': new_po_comment},
                {'string_id': ts_obj.id, 'field': 'comment', 'old_value': ts_obj.comment,
                 'new_value': new_user_comment},
                {'string_id': ts_obj.id, 'field': 'is_fuzzy', 'old_value': old_is_fuzzy, 'new_value': new_is_fuzzy}
            ]
        })
        ts_obj.po_comment = new_po_comment
        ts_obj.comment = new_user_comment
        ts_obj.is_fuzzy = new_is_fuzzy
        if old_user_comment_for_hook != new_user_comment:
            self.plugin_manager.run_hook(
                'on_string_saved',
                ts_object=ts_obj,
                column='comment',
                new_value=new_user_comment,
                old_value=old_user_comment_for_hook
            )
        if old_po_comment_for_hook != new_po_comment:
            self.plugin_manager.run_hook(
                'on_string_saved',
                ts_object=ts_obj,
                column='po_comment',
                new_value=new_po_comment,
                old_value=old_po_comment_for_hook
            )
        ts_obj.update_style_cache()
        source_index = self.sheet_model.index_from_id(ts_obj.id)
        if source_index.isValid():
            first_col_index = source_index.siblingAtColumn(0)
            last_col_index = source_index.siblingAtColumn(self.sheet_model.columnCount() - 1)
            self.sheet_model.dataChanged.emit(first_col_index, last_col_index)
        self.mark_project_modified()
        self.update_statusbar(_("Comment updated."))
        if hasattr(self.comment_status_panel, 'highlighter'):
            self.comment_status_panel.highlighter.rehighlight()
        return True


    def apply_comment_from_button(self):
        self._save_comment_from_ui()

    def apply_comment_focus_out(self):
        self._save_comment_from_ui()

    def _apply_comment_to_model(self, ts_obj, new_comment):
        if new_comment == ts_obj.comment: return False

        old_comment = ts_obj.comment
        ts_obj.comment = new_comment

        self.add_to_undo_history('single_change', {
            'string_id': ts_obj.id, 'field': 'comment',
            'old_value': old_comment, 'new_value': new_comment
        })
        self._run_and_refresh_with_validation()
        self.update_statusbar(_("Comment updated for ID {id}...").format(id=str(ts_obj.id)[:8]))
        self.mark_project_modified()
        return True

    def _update_view_for_ids(self, changed_ids: set):
        if not changed_ids:
            return

        from services.validation_service import validate_string
        for ts_id in changed_ids:
            ts_obj = self._find_ts_obj_by_id(ts_id)
            if ts_obj:
                validate_string(ts_obj, self.config, self)
                ts_obj.update_style_cache()
                source_index = self.sheet_model.index_from_id(ts_obj.id)
                if source_index.isValid():
                    first_col_index = source_index.siblingAtColumn(0)
                    last_col_index = source_index.siblingAtColumn(self.sheet_model.columnCount() - 1)
                    self.sheet_model.dataChanged.emit(first_col_index, last_col_index)

        if self.current_selected_ts_id in changed_ids:
            self.force_refresh_ui_for_current_selection()

        self.update_counts_display()

    def force_full_refresh(self, id_to_reselect=None):
        self.sheet_model.set_translatable_objects(self.translatable_objects)
        self.proxy_model.set_filters(
            show_ignored=self.show_ignored_var,
            show_untranslated=self.show_untranslated_var,
            show_translated=self.show_translated_var,
            show_unreviewed=self.show_unreviewed_var,
            search_term=self.search_entry.text() if self.search_entry.text() != _("Quick search...") else "",
            is_po_mode=self.is_po_mode
        )
        if id_to_reselect:
            self.select_sheet_row_by_id(id_to_reselect, see=True)
        self.force_refresh_ui_for_current_selection()
        self.update_counts_display()

    def _select_neighbor_or_first(self, removed_row_index):
        if removed_row_index < self.proxy_model.rowCount():
            neighbor_index = self.proxy_model.index(removed_row_index, 0)
        elif self.proxy_model.rowCount() > 0:
            neighbor_index = self.proxy_model.index(self.proxy_model.rowCount() - 1, 0)
        else:
            self.table_view.clearSelection()
            self.on_sheet_select(QModelIndex(), QModelIndex())
            return
        selected_obj = self.proxy_model.data(neighbor_index, Qt.UserRole)
        self.table_view.setCurrentIndex(neighbor_index)
        self.on_sheet_select(neighbor_index, QModelIndex())

    def _deferred_select_neighbor(self, neighbor_row_in_proxy):
        if self.neighbor_select_timer.isActive():
            self.neighbor_select_timer.stop()
        if neighbor_row_in_proxy < self.proxy_model.rowCount():
            new_index = self.proxy_model.index(neighbor_row_in_proxy, 0)
            self.table_view.setCurrentIndex(new_index)
        elif self.proxy_model.rowCount() > 0:
            new_index = self.proxy_model.index(self.proxy_model.rowCount() - 1, 0)
            self.table_view.setCurrentIndex(new_index)

    def toggle_ignore_selected_checkbox(self, state):
        new_ignore_state = bool(state)
        if not self.current_selected_ts_id: return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj or new_ignore_state == ts_obj.is_ignored: return
        neighbor_id_to_select = None
        will_disappear = not self.show_ignored_var and new_ignore_state
        if will_disappear:
            source_index = self.sheet_model.index_from_id(ts_obj.id)
            proxy_index = self.proxy_model.mapFromSource(source_index)
            if proxy_index.isValid():
                current_row = proxy_index.row()
                if current_row + 1 < self.proxy_model.rowCount():
                    neighbor_proxy_index = self.proxy_model.index(current_row + 1, 0)
                    neighbor_obj = self.proxy_model.data(neighbor_proxy_index, Qt.UserRole)
                    if neighbor_obj: neighbor_id_to_select = neighbor_obj.id
                elif current_row - 1 >= 0:
                    neighbor_proxy_index = self.proxy_model.index(current_row - 1, 0)
                    neighbor_obj = self.proxy_model.data(neighbor_proxy_index, Qt.UserRole)
                    if neighbor_obj: neighbor_id_to_select = neighbor_obj.id
        primary_change = {'string_id': ts_obj.id, 'field': 'is_ignored', 'old_value': ts_obj.is_ignored,
                          'new_value': new_ignore_state}
        self.add_to_undo_history('single_change', primary_change)
        ts_obj.is_ignored = new_ignore_state
        if not new_ignore_state: ts_obj.was_auto_ignored = False
        ts_obj.update_style_cache()
        self.mark_project_modified()
        self.update_statusbar(_("Ignore status for ID {id} -> {status}").format(id=str(ts_obj.id)[:8] + "...", status=_(
            'Yes') if new_ignore_state else _('No')))
        self._update_view_for_ids({ts_obj.id})
        def deferred_refresh():
            if will_disappear:
                self.force_full_refresh(id_to_reselect=neighbor_id_to_select)
        QTimer.singleShot(0, deferred_refresh)

    def toggle_reviewed_selected_checkbox(self, state):
        new_reviewed_state = bool(state)
        if not self.current_selected_ts_id: return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj or new_reviewed_state == ts_obj.is_reviewed: return

        neighbor_id_to_select = None
        will_disappear = self.unreviewed_checkbox.isChecked() and new_reviewed_state
        if will_disappear:
            source_index = self.sheet_model.index_from_id(ts_obj.id)
            proxy_index = self.proxy_model.mapFromSource(source_index)
            if proxy_index.isValid():
                current_row = proxy_index.row()
                if current_row + 1 < self.proxy_model.rowCount():
                    neighbor_proxy_index = self.proxy_model.index(current_row + 1, 0)
                    neighbor_obj = self.proxy_model.data(neighbor_proxy_index, Qt.UserRole)
                    if neighbor_obj: neighbor_id_to_select = neighbor_obj.id
                elif current_row - 1 >= 0:
                    neighbor_proxy_index = self.proxy_model.index(current_row - 1, 0)
                    neighbor_obj = self.proxy_model.data(neighbor_proxy_index, Qt.UserRole)
                    if neighbor_obj: neighbor_id_to_select = neighbor_obj.id
        primary_change = {'string_id': ts_obj.id, 'field': 'is_reviewed', 'old_value': ts_obj.is_reviewed,
                          'new_value': new_reviewed_state}
        self.add_to_undo_history('single_change', primary_change)
        ts_obj.is_reviewed = new_reviewed_state
        ts_obj.update_style_cache()
        self.mark_project_modified()
        self.update_statusbar(
            _("Review status for ID {id} -> {status}").format(id=str(ts_obj.id)[:8] + "...",
                                                              status=_('Yes') if new_reviewed_state else _('No')))
        self._update_view_for_ids({ts_obj.id})
        def deferred_refresh():
            if will_disappear:
                self.force_full_refresh(id_to_reselect=neighbor_id_to_select)
        QTimer.singleShot(0, deferred_refresh)

    def save_code_file_content(self, filepath_to_save):
        if not self.original_raw_code_content:
            QMessageBox.critical(self, _("Error"),
                                 _("There is no original code file content to save.\nPlease ensure the code file associated with the project is loaded."))
            return False
        try:
            save_translated_code(filepath_to_save, self.original_raw_code_content, self.translatable_objects, self)
            self.update_statusbar(
                _("Code file saved to: {filename}").format(filename=os.path.basename(filepath_to_save)),
                persistent=True)
            file_ext = os.path.splitext(filepath_to_save)[1].lstrip('.')
            self.plugin_manager.run_hook('on_after_project_save', filepath=filepath_to_save, file_format=file_ext or 'txt')
            return True
        except Exception as e_save:
            QMessageBox.critical(self, _("Save Error"), _("Could not save code file: {error}").format(error=e_save))
            return False

    def save_code_file(self):
        if not self.current_code_file_path:
            QMessageBox.critical(self, _("Error"), _("No original code file path."))
            return

        if not self.original_raw_code_content:
            QMessageBox.critical(self, _("Error"), _("No original code content to save."))
            return

        base, ext = os.path.splitext(self.current_code_file_path)
        new_filepath = f"{base}_translated{ext}"

        if os.path.exists(new_filepath):
            reply = QMessageBox.question(self, _("Confirm Overwrite"),
                                         _("File '{filename}' already exists. Overwrite? A backup file (.bak) will be created.").format(
                                             filename=os.path.basename(new_filepath)),
                                         QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.No:
                return

        self.save_code_file_content(new_filepath)

    def save_current_file(self):
        if self.is_project_mode:
            if not self.current_project_path:
                QMessageBox.critical(self, _("Error"), _("Project path is missing. Cannot save."))
                return False
            try:
                if save_project(self.current_project_path, self):
                    self.mark_project_modified(False)
                    self.update_statusbar(_("Project saved."), persistent=True)
                    return True
            except Exception as e:
                QMessageBox.critical(self, _("Save Error"), str(e))
            return False
        else:
            if self.is_po_mode:
                if self.current_po_file_path:
                    return self.save_po_file(self.current_po_file_path)
                else:
                    return self.save_po_as_dialog()
            else:
                return self.save_project_as_dialog()

    def save_current_file_as(self):
        if self.is_po_mode:
            return self.save_po_as_dialog()
        else:
            return self.save_project_as_dialog()

    def save_project_dialog(self):
        return self.save_current_file()

    def save_project_as_dialog(self):
        if self.is_project_mode:
            QMessageBox.information(self, _("Save As"),
                                    _("To save a copy of the project, please create a new project and import the source files."))
            return False

        if not self.translatable_objects:
            QMessageBox.information(self, _("Info"), _("There is no content to save as a project."))
            return False

        dialog = NewProjectDialog(self, self)

        source_path = self.current_code_file_path or self.current_po_file_path
        if source_path:
            project_name = os.path.splitext(os.path.basename(source_path))[0]
            dialog.project_name_edit.setText(project_name)
            dialog.location_edit.setText(os.path.dirname(source_path))

            file_type = 'po' if self.is_po_mode else 'code'
            dialog.source_files = [{'path': source_path, 'type': file_type}]
            dialog.source_files_list.addItem(QListWidgetItem(source_path))

        if dialog.exec():
            data = dialog.get_data()
            project_name = data['name']
            project_location = data['location']

            for sf in data['source_files']:
                sf['patterns'] = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)

            try:
                project_path = os.path.join(project_location, project_name)
                new_project_path = create_project(
                    project_path, project_name, data['source_lang'],
                    data['target_langs'], data['source_files'], data['use_global_tm']
                )
                self.open_project(new_project_path)
                return True
            except Exception as e:
                QMessageBox.critical(self, _("Project Creation Failed"), str(e))
        return False

    def save_po_file(self, filepath, compile_mo=True):
        try:
            original_file_name = os.path.basename(self.current_code_file_path or "source_code")
            po_file_service.save_to_po(filepath, self.translatable_objects, self.current_po_metadata,
                                       original_file_name, self)

            self.current_po_file_path = filepath
            self.mark_project_modified(False)
            self.update_statusbar(_("PO file saved to: {filename}").format(filename=os.path.basename(filepath)),
                                  persistent=True)
            self.update_title()
            self.plugin_manager.run_hook('on_after_project_save', filepath=filepath, file_format='po')
            if self.auto_compile_mo_var and compile_mo:
                try:
                    mo_filepath = os.path.splitext(filepath)[0] + ".mo"
                    po_file_to_compile = polib.pofile(filepath, encoding='utf-8')
                    po_file_to_compile.save_as_mofile(mo_filepath)
                    self.update_statusbar(_("PO file saved and MO file compiled: {filename}").format(filename=os.path.basename(mo_filepath)))
                except Exception as e_mo:
                    QMessageBox.critical(self, _("MO Compilation Failed"), _("Could not compile MO file: {error}").format(error=e_mo))
            return True
        except Exception as e:
            QMessageBox.critical(self, _("Save PO Error"), _("Failed to save PO file: {error}").format(error=e))
            return False

    def save_po_as_dialog(self):
        filepath, selected_filter = QFileDialog.getSaveFileName(
            self,
            _("Save PO File As"),
            self.config.get("last_dir", os.getcwd()),
            _("PO files (*.po);;All Files (*.*)")
        )
        if filepath:
            return self.save_po_file(filepath)
        return False

    def export_project_translations_to_excel(self):
        if not self.translatable_objects:
            QMessageBox.information(self, _("Info"), _("No data to export."))
            return

        default_filename = "project_translations.xlsx"
        if self.current_project_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_project_path))
            default_filename = f"{base}_translations.xlsx"
        elif self.current_code_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_code_file_path))
            default_filename = f"{base}_translations.xlsx"
        elif self.current_po_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_po_file_path))
            default_filename = f"{base}_translations.xlsx"

        filepath, selected_filter = QFileDialog.getSaveFileName(
            self,
            _("Export Project Translations to Excel"),
            default_filename,
            "Excel files (*.xlsx)"
        )
        if not filepath: return

        wb = Workbook()
        ws = wb.active
        ws.title = "Translations"
        headers = ["UUID", "Type", _("Original (Semantic)"), _("Translation"), _("Comment"), _("Reviewed"),
                   _("Ignored"), _("Source Line"),
                   _("Original (Raw)")]
        ws.append(headers)
        items_to_export = []
        for i in range(self.proxy_model.rowCount()):
            index = self.proxy_model.index(i, 0)
            ts_obj = self.proxy_model.data(index, Qt.UserRole)
            if ts_obj:
                items_to_export.append(ts_obj)

        if not items_to_export and self.translatable_objects:
            items_to_export = self.translatable_objects

        for ts_obj in items_to_export:
            ws.append([
                ts_obj.id,
                ts_obj.string_type,
                ts_obj.original_semantic,
                ts_obj.get_translation_for_storage_and_tm(),
                ts_obj.comment,
                _("Yes") if ts_obj.is_reviewed else _("No"),
                _("Yes") if ts_obj.is_ignored else _("No"),
                ts_obj.line_num_in_file,
                ts_obj.original_raw
            ])
        try:
            wb.save(filepath)
            self.update_statusbar(
                _("Project translations exported to: {filename}").format(filename=os.path.basename(filepath)))
        except Exception as e:
            QMessageBox.critical(self, _("Export Error"),
                                 _("Could not export project translations to Excel: {error}").format(error=e))

    def import_project_translations_from_excel(self):
        if not self.translatable_objects:
            QMessageBox.information(self, _("Info"),
                                    _("Please load a code file or project first to match imported translations."))
            return

        filepath, selected_filter = QFileDialog.getOpenFileName(
            self,
            _("Import Translations from Excel"),
            self.config.get("last_dir", os.getcwd()),
            _("Excel files (*.xlsx)")
        )
        if not filepath: return

        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active

            header_row_values = [cell.value for cell in ws[1]]
            if not header_row_values or not all(isinstance(h, str) for h in header_row_values if h is not None):
                QMessageBox.critical(self, _("Import Error"), _("Excel header format is incorrect or empty."))
                return

            try:
                uuid_col_idx = header_row_values.index("UUID")
                trans_col_idx = header_row_values.index(_("Translation"))
                comment_col_idx = header_row_values.index(_("Comment")) if _("Comment") in header_row_values else -1
                reviewed_col_idx = header_row_values.index(_("Reviewed")) if _("Reviewed") in header_row_values else -1
                ignored_col_idx = header_row_values.index(_("Ignored")) if _("Ignored") in header_row_values else -1
                orig_col_idx = header_row_values.index(
                    _("Original (Semantic)")) if _("Original (Semantic)") in header_row_values else -1
            except ValueError:
                QMessageBox.critical(self, _("Import Error"),
                                     _("Excel header must contain 'UUID' and '{translation_col}' columns.\nOptional columns: '{comment_col}', '{reviewed_col}', '{ignored_col}', '{original_semantic_col}'.").format(
                                         translation_col=_("Translation"), comment_col=_("Comment"),
                                         reviewed_col=_("Reviewed"),
                                         ignored_col=_("Ignored"), original_semantic_col=_("Original (Semantic)")))
                return

            imported_count = 0
            changes_for_undo = []

            for r_idx, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                try:
                    obj_id_from_excel = row_cells[uuid_col_idx]
                    if obj_id_from_excel is None: continue

                    ts_obj = self._find_ts_obj_by_id(str(obj_id_from_excel))
                    if not ts_obj:
                        continue

                    if orig_col_idx != -1 and row_cells[orig_col_idx] is not None:
                        if ts_obj.original_semantic != str(row_cells[orig_col_idx]):
                            logger.warning(
                                _("Warning: Excel row {row_num}, UUID {uuid} - Original text does not match the one in Excel. Data will still be imported.").format(
                                    row_num=r_idx + 2, uuid=obj_id_from_excel))

                    translation_from_excel_raw = str(row_cells[trans_col_idx]) if row_cells[
                                                                                      trans_col_idx] is not None else ""
                    translation_for_model = translation_from_excel_raw.replace("\\n", "\n")
                    if ts_obj.translation != translation_for_model:
                        changes_for_undo.append({'string_id': ts_obj.id, 'field': 'translation',
                                                 'old_value': ts_obj.get_translation_for_storage_and_tm(),
                                                 'new_value': translation_from_excel_raw})
                        ts_obj.set_translation_internal(translation_for_model)
                        if translation_for_model.strip():
                            db_path_to_use = self.tm_service.project_db_path if self.is_project_mode else self.tm_service.global_db_path
                            source_lang = self.source_language
                            target_lang = self.current_target_language if self.is_project_mode else self.target_language
                            if db_path_to_use:
                                self.tm_service.update_tm_entry(
                                    db_path_to_use, ts_obj.original_semantic,
                                    translation_from_excel_raw,
                                    source_lang, target_lang
                                )
                        imported_count += 1
                    if comment_col_idx != -1 and row_cells[comment_col_idx] is not None:
                        comment_from_excel = str(row_cells[comment_col_idx])
                        if ts_obj.comment != comment_from_excel:
                            changes_for_undo.append(
                                {'string_id': ts_obj.id, 'field': 'comment', 'old_value': ts_obj.comment,
                                 'new_value': comment_from_excel})
                            ts_obj.comment = comment_from_excel
                            if not imported_count: imported_count = 1

                    if reviewed_col_idx != -1 and row_cells[reviewed_col_idx] is not None:
                        reviewed_str = str(row_cells[reviewed_col_idx]).lower()
                        is_reviewed_excel = reviewed_str in [_("Yes").lower(), "true", "yes", "1"]
                        if ts_obj.is_reviewed != is_reviewed_excel:
                            changes_for_undo.append(
                                {'string_id': ts_obj.id, 'field': 'is_reviewed', 'old_value': ts_obj.is_reviewed,
                                 'new_value': is_reviewed_excel})
                            ts_obj.is_reviewed = is_reviewed_excel
                            if not imported_count: imported_count = 1

                    if ignored_col_idx != -1 and row_cells[ignored_col_idx] is not None:
                        ignored_str = str(row_cells[ignored_col_idx]).lower()
                        is_ignored_excel = ignored_str in [_("Yes").lower(), "true", "yes", "1"]
                        if ts_obj.is_ignored != is_ignored_excel:
                            changes_for_undo.append(
                                {'string_id': ts_obj.id, 'field': 'is_ignored', 'old_value': ts_obj.is_ignored,
                                 'new_value': is_ignored_excel})
                            ts_obj.is_ignored = is_ignored_excel
                            if not is_ignored_excel: ts_obj.was_auto_ignored = False
                            if not imported_count: imported_count = 1

                except Exception as cell_err:
                    logger.warning(
                        _("Error processing Excel row {row_num}: {error}. Skipping this row.").format(row_num=r_idx + 2,
                                                                                                      error=cell_err))

            if changes_for_undo:
                self.add_to_undo_history('bulk_excel_import', {'changes': changes_for_undo})
                self.mark_project_modified()

            self._run_and_refresh_with_validation()
            if self.current_selected_ts_id: self.force_refresh_ui_for_current_selection()

            self.update_statusbar(_("Imported/updated {field_count} fields for {item_count} items from Excel.").format(
                field_count=len(changes_for_undo), item_count=imported_count))

        except ValueError as ve:
            QMessageBox.critical(self, _("Import Error"),
                                 _("Error processing Excel file (possibly column names issue): {error}").format(
                                     error=ve))
        except Exception as e:
            QMessageBox.critical(self, _("Import Error"),
                                 _("Could not import project translations from Excel: {error}").format(error=e))

    def export_project_translations_to_json(self):
        if not self.translatable_objects:
            QMessageBox.information(self, _("Info"), _("No data to export."))
            return

        default_filename = "project_translations.json"
        if self.current_project_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_project_path))
            default_filename = f"{base}_translations.json"
        elif self.current_code_file_path:
            base, __extension = os.path.splitext(os.path.basename(self.current_code_file_path))
            default_filename = f"{base}_translations.json"
        elif self.current_po_file_path:
            base, __extension = os.path.splitext(os.path.basename(self.current_po_file_path))
            default_filename = f"{base}_translations.json"

        filepath, selected_filter = QFileDialog.getSaveFileName(
            self,
            _("Export Project Translations to JSON"),
            default_filename,
            _("JSON files (*.json);;All Files (*.*)")
        )
        if not filepath: return

        try:
            displayed_ids_order = []
            for i in range(self.proxy_model.rowCount()):
                index = self.proxy_model.index(i, 0)
                ts_obj = self.proxy_model.data(index, Qt.UserRole)
                if ts_obj:
                    displayed_ids_order.append(ts_obj.id)

            export_service.export_to_json(filepath, self.translatable_objects, displayed_ids_order,
                                          app_instance=self)
            self.update_statusbar(
                _("Project translations exported to: {filename}").format(filename=os.path.basename(filepath)))
        except Exception as e:
            QMessageBox.critical(self, _("Export Error"),
                                 _("Could not export project translations to JSON: {error}").format(error=e))

    def export_project_translations_to_yaml(self):
        if not self.translatable_objects:
            QMessageBox.information(self, _("Info"), _("No data to export."))
            return

        default_filename = "project_translations.yaml"
        if self.current_project_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_project_path))
            default_filename = f"{base}_translations.yaml"
        elif self.current_code_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_code_file_path))
            default_filename = f"{base}_translations.yaml"
        elif self.current_po_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_po_file_path))
            default_filename = f"{base}_translations.yaml"

        filepath, selected_filter = QFileDialog.getSaveFileName(
            self,
            _("Export Project Translations to YAML"),
            default_filename,
            _("YAML files (*.yaml *.yml);;All Files (*.*)")
        )
        if not filepath: return

        try:
            displayed_ids_order = []
            for i in range(self.proxy_model.rowCount()):
                index = self.proxy_model.index(i, 0)
                ts_obj = self.proxy_model.data(index, Qt.UserRole)
                if ts_obj:
                    displayed_ids_order.append(ts_obj.id)

            export_service.export_to_yaml(filepath, self.translatable_objects, displayed_ids_order,
                                          app_instance=self)
            self.update_statusbar(
                _("Project translations exported to: {filename}").format(filename=os.path.basename(filepath)))
        except Exception as e:
            QMessageBox.critical(self, _("Export Error"),
                                 _("Could not export project translations to YAML: {error}").format(error=e))

    def extract_to_pot_dialog(self):
        code_filepath, selected_filter = QFileDialog.getOpenFileName(
            self,
            _("Select Code File to Extract POT From"),
            self.config.get("last_dir", os.getcwd()),
            _("Overwatch Workshop Files (*.ow *.txt);;All Files (*.*)")
        )
        if not code_filepath: return

        pot_save_filepath, selected_filter = QFileDialog.getSaveFileName(
            self,
            _("Save POT Template File"),
            os.path.splitext(os.path.basename(code_filepath))[0] + ".pot",
            _("PO Template files (*.pot);;All files (*.*)")
        )
        if not pot_save_filepath: return

        try:
            with open(code_filepath, 'r', encoding='utf-8', errors='replace') as f:
                code_content = f.read()

            extraction_patterns = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)
            extraction_patterns = self.plugin_manager.run_hook(
                'process_extraction_patterns',
                extraction_patterns,
                filepath=code_filepath,
                raw_content=code_content
            )

            extraction_patterns = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)
            project_name = os.path.basename(code_filepath)

            pot_object = po_file_service.extract_to_pot(code_content, extraction_patterns, project_name, APP_VERSION,
                                                        os.path.basename(code_filepath))
            pot_object.save(pot_save_filepath)
            self.update_statusbar(
                _("POT template saved to: {filename}").format(filename=os.path.basename(pot_save_filepath)))
            self.config["last_dir"] = os.path.dirname(code_filepath)
            self.save_config()
        except Exception as e:
            QMessageBox.critical(self, _("POT Extraction Error"),
                                 _("Error extracting POT file: {error}").format(error=e))

    def import_po_file_dialog_with_path(self, po_filepath):
        self._reset_app_state()
        try:
            self.translatable_objects, self.current_po_metadata, po_lang_full = po_file_service.load_from_po(po_filepath)
            self.original_raw_code_content = ""
            self.current_code_file_path = None
            self.source_language = language_service.detect_source_language(
                [ts.original_semantic for ts in self.translatable_objects if ts.original_semantic.strip()]
            )
            target_lang_set = False
            if po_lang_full:
                full_code_match = next((code for name, code in SUPPORTED_LANGUAGES.items() if code.lower() == po_lang_full.lower().replace('-', '_')), None)
                if full_code_match:
                    self.target_language = full_code_match
                    target_lang_set = True
                else:
                    base_lang_code = po_lang_full.split('_')[0].split('-')[0].lower()
                    if base_lang_code in SUPPORTED_LANGUAGES.values():
                        self.target_language = base_lang_code
                        target_lang_set = True

            if not target_lang_set:
                translated_strings = [ts.translation for ts in self.translatable_objects if ts.translation.strip()]
                if translated_strings:
                    detected_target_lang = language_service.detect_source_language(translated_strings)
                    if detected_target_lang and detected_target_lang in SUPPORTED_LANGUAGES.values():
                        self.target_language = detected_target_lang
                        target_lang_set = True

            if not target_lang_set:
                self.target_language = self.config.get("default_target_language", "zh")
            self.current_project_path = None

            self.add_to_recent_files(po_filepath)
            self.config["last_dir"] = os.path.dirname(po_filepath)
            self.save_config()

            self.undo_history.clear()
            self.redo_history.clear()
            self.mark_project_modified(False)
            self.is_project_mode = False
            self.is_po_mode = True
            self._update_file_explorer(po_filepath)
            self.current_po_file_path = po_filepath
            self._run_and_refresh_with_validation()
            self.update_statusbar(
                _("Loaded {count} entries from PO file {filename}.").format(count=len(self.translatable_objects),
                                                                            filename=os.path.basename(po_filepath)),
                persistent=True)
            self.update_ui_state_after_file_load(file_or_project_loaded=True)

        except Exception as e:
            logger.error("--- AN EXCEPTION OCCURRED DURING PO IMPORT ---")
            traceback.print_exc()
            logger.error("---------------------------------------------")
            QMessageBox.critical(self, _("PO Import Error"), _("Error importing PO file '{filename}': {error}").format(
                filename=os.path.basename(po_filepath), error=e))
            self._reset_app_state()
            self.update_statusbar(_("PO file loading failed"), persistent=True)
        self.update_counts_display()

    def import_po_file_dialog(self):
        if not self.prompt_save_if_modified(): return

        po_filepath, selected_filter = QFileDialog.getOpenFileName(
            self,
            _("Select PO File to Import"),
            self.config.get("last_dir", os.getcwd()),
            _("PO files (*.po);;POT files (*.pot);;All files (*.*)")
        )
        if not po_filepath: return
        self.import_po_file_dialog_with_path(po_filepath)

    def export_to_po_file_dialog(self):
        if not self.translatable_objects:
            QMessageBox.information(self, _("Info"), _("No data to export."))
            return

        default_filename = "translations.po"
        if self.current_project_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_project_path))
            default_filename = f"{base}.po"
        elif self.current_code_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_code_file_path))
            default_filename = f"{base}.po"
        elif self.current_po_file_path:
            default_filename = os.path.basename(self.current_po_file_path)

        filepath, selected_filter = QFileDialog.getSaveFileName(
            self,
            _("Export to PO File..."),
            default_filename,
            _("PO files (*.po);;All Files (*.*)")
        )
        if not filepath: return

        try:
            original_file_name_for_po = os.path.basename(
                self.current_code_file_path) if self.current_code_file_path else "source_code"
            po_file_service.save_to_po(filepath, self.translatable_objects, self.current_po_metadata,
                                       original_file_name_for_po)
            self.update_statusbar(
                _("Translations exported to PO file: {filename}").format(filename=os.path.basename(filepath)))
            self.mark_project_modified(False)
        except Exception as e:
            QMessageBox.critical(self, _("Export Error"), _("Failed to export to PO file: {error}").format(error=e))

    def show_extraction_pattern_dialog(self):
        dialog = ExtractionPatternManagerDialog(self, _("Extraction Rule Manager"), self)
        if dialog.exec():
            if dialog.result and self.original_raw_code_content:
                reply = QMessageBox.question(self, _("Extraction Rules Updated"),
                                             _("Extraction rules updated. Do you want to reload the translatable text of the current code using the new rules immediately?"),
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                if reply == QMessageBox.Yes:
                    self.reload_translatable_text()

    def reload_translatable_text(self):
        if not self.original_raw_code_content and not self.current_code_file_path:
            QMessageBox.information(self, _("Info"), _("No code content loaded to reload."))
            return

        current_content_to_reextract = self.original_raw_code_content
        source_name = _("current in-memory code")
        if self.current_code_file_path and os.path.exists(self.current_code_file_path):
            try:
                with open(self.current_code_file_path, 'r', encoding='utf-8', errors='replace') as f:
                    current_content_to_reextract = f.read()
                source_name = _("file '{filename}'").format(filename=os.path.basename(self.current_code_file_path))
            except Exception as e:
                QMessageBox.warning(self, _("File Read Error"),
                                    _("Could not re-read {filepath} from disk.\nUsing in-memory version.\nError: {error}").format(
                                        filepath=self.current_code_file_path, error=e))

        if not current_content_to_reextract:
            QMessageBox.critical(self, _("Error"), _("Could not get code content for re-extraction."))
            return

        old_translations_map = {ts.original_semantic: {
            'translation': ts.translation,
            'comment': ts.comment,
            'is_reviewed': ts.is_reviewed,
            'is_ignored': ts.is_ignored,
            'was_auto_ignored': ts.was_auto_ignored
        } for ts in self.translatable_objects}

        if self.current_project_modified or old_translations_map:
            reply = QMessageBox.question(self, _("Confirm Reload"),
                                         _("This will re-extract strings from {source} using the new rules.\n"
                                           "Existing translations will be preserved where the original text matches, but unmatched translations and statuses may be lost.\n"
                                           "This action will clear the undo history. Continue?").format(
                                             source=source_name),
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return

        try:
            self.update_statusbar(_("Re-extracting strings with new rules..."), persistent=True)
            QApplication.processEvents()
            extraction_patterns = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)
            extraction_patterns = self.plugin_manager.run_hook(
                'process_extraction_patterns',
                extraction_patterns,
                filepath=self.current_code_file_path or "",
                raw_content=current_content_to_reextract
            )
            self.original_raw_code_content = current_content_to_reextract
            self.translatable_objects = extract_translatable_strings(self.original_raw_code_content,
                                                                     extraction_patterns)
            restored_count = 0
            for ts_obj in self.translatable_objects:
                if ts_obj.original_semantic in old_translations_map:
                    saved_state = old_translations_map[ts_obj.original_semantic]
                    ts_obj.set_translation_internal(saved_state['translation'])
                    ts_obj.comment = saved_state['comment']
                    ts_obj.is_reviewed = saved_state['is_reviewed']
                    if not saved_state['was_auto_ignored'] and saved_state['is_ignored']:
                        ts_obj.is_ignored = True
                        ts_obj.was_auto_ignored = False
                    restored_count += 1

            if restored_count > 0:
                self.update_statusbar(
                    _("Attempted to restore {count} old translations/statuses.").format(count=restored_count),
                    persistent=False)

            self.apply_tm_to_all_current_strings(silent=True, only_if_empty=True)

            self.undo_history.clear()
            self.redo_history.clear()
            self.current_selected_ts_id = None
            self.mark_project_modified(True)

            self.refresh_sheet()
            self.update_statusbar(
                _("Reloaded {count} strings from {source} using new rules.").format(
                    count=len(self.translatable_objects), source=source_name),
                persistent=True)
            self.update_ui_state_after_file_load(file_or_project_loaded=True)

        except Exception as e:
            QMessageBox.critical(self, _("Reload Error"),
                                 _("Error reloading translatable text: {error}").format(error=e))
            self.update_statusbar(_("Reload failed."), persistent=True)
        self.update_counts_display()

    def show_statistics_dialog(self):
        if not self.translatable_objects:
            QMessageBox.information(self, _("Statistics"), _("No project data loaded to generate statistics."))
            return
        dialog = StatisticsDialog(self, self.translatable_objects)
        dialog.locate_item_signal.connect(self.select_sheet_row_by_id_and_scroll)
        dialog.show()

    def select_sheet_row_by_id_and_scroll(self, ts_id):
        self.select_sheet_row_by_id(ts_id, see=True)
        self.activateWindow()

    def update_tm_for_selected_string(self):
        if not self.current_selected_ts_id:
            QMessageBox.information(self, _("Info"), _("Please select an item first."))
            return

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return

        current_translation_ui = self.details_panel.translation_edit_text.toPlainText()

        db_path_to_use = self.tm_service.project_db_path if self.is_project_mode else self.tm_service.global_db_path
        source_lang = self.source_language
        target_lang = self.current_target_language if self.is_project_mode else self.target_language

        self.tm_service.update_tm_entry(
            db_path_to_use,
            ts_obj.original_semantic,
            current_translation_ui.replace("\n", "\\n"),
            source_lang,
            target_lang
        )

        self.update_statusbar(
            _("TM updated for original: '{text}...'").format(text=ts_obj.original_semantic[:30].replace(chr(10), '↵')))

        self.perform_tm_update()
        self.mark_project_modified()

    def clear_tm_for_selected_string(self):
        if not self.current_selected_ts_id:
            QMessageBox.information(self, _("Info"), _("Please select an item first."))
            return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return
        source_lang = self.source_language
        target_lang = self.current_target_language if self.is_project_mode else self.target_language
        db_path_to_use = self.tm_service.project_db_path if self.is_project_mode else self.tm_service.global_db_path

        if not db_path_to_use:
            QMessageBox.warning(self, _("Warning"), _("No active TM database found to clear the entry from."))
            return
        existing_translation = self.tm_service.get_translation(ts_obj.original_semantic, source_lang, target_lang)
        if existing_translation is not None:
            reply = QMessageBox.question(self, _("Confirm Clear"),
                                         _("Are you sure you want to remove the TM entry for:\n'{text}...'?").format(
                                             text=ts_obj.original_semantic[:100].replace(chr(10), '↵')),
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                success = self.tm_service.delete_tm_entry(
                    db_path=db_path_to_use,
                    source_text=ts_obj.original_semantic,
                    source_lang=source_lang,
                    target_lang=target_lang
                )

                if success:
                    self.update_statusbar(_("TM entry cleared for selected item."))
                    self.perform_tm_update()
                else:
                    QMessageBox.critical(self, _("Error"), _("Failed to delete the TM entry from the database."))
        else:
            QMessageBox.information(self, _("Info"), _("The selected item has no entry in the current TM."))

    def apply_tm_suggestion_from_listbox(self, translation_text_ui):
        self.details_panel.translation_edit_text.setPlainText(translation_text_ui)

        if self.current_selected_ts_id:
            ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
            if ts_obj:
                self._apply_translation_to_model(ts_obj, translation_text_ui, source="tm_suggestion")

        self.update_statusbar(_("TM suggestion applied."))

    def apply_tm_to_all_current_strings(self, silent=False, only_if_empty=False, confirm=False):
        if not self.translatable_objects:
            if not silent: QMessageBox.information(self, _("Info"), _("No strings to apply TM to."))
            return 0

        if confirm and not only_if_empty:
            reply = QMessageBox.question(self, _("Confirm Operation"),
                                         _("This will apply TM to all matching strings, overwriting existing translations. Continue?"),
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return 0

        applied_count = 0
        bulk_changes_for_undo = []
        ids_to_update = set()

        source_lang = self.source_language
        target_lang = self.current_target_language if self.is_project_mode else self.target_language

        for ts_obj in self.translatable_objects:
            if ts_obj.is_ignored: continue
            if only_if_empty and ts_obj.translation.strip() != "": continue

            translation_from_tm = self.tm_service.get_translation(
                ts_obj.original_semantic, source_lang, target_lang
            )

            if translation_from_tm:
                tm_translation_ui = translation_from_tm.replace("\\n", "\n")
                if ts_obj.translation != tm_translation_ui:
                    old_val = ts_obj.get_translation_for_storage_and_tm()
                    ts_obj.set_translation_internal(tm_translation_ui)
                    bulk_changes_for_undo.append({
                        'string_id': ts_obj.id, 'field': 'translation',
                        'old_value': old_val, 'new_value': translation_from_tm
                    })
                    ids_to_update.add(ts_obj.id)
                    applied_count += 1

        if applied_count > 0:
            if bulk_changes_for_undo:
                self.add_to_undo_history('bulk_change', {'changes': bulk_changes_for_undo})
                self.mark_project_modified()
            self._update_view_for_ids(ids_to_update)

            if self.current_selected_ts_id:
                self.force_refresh_ui_for_current_selection()
            if not silent:
                QMessageBox.information(self, _("TM"), _("Applied TM to {count} strings.").format(count=applied_count))
            self.update_statusbar(_("Applied TM to {count} strings.").format(count=applied_count))
        elif not silent:
            QMessageBox.information(self, _("TM"), _("No applicable translations found in TM (or no changes needed)."))

        return applied_count

    def show_advanced_search_dialog(self):
        if not self.translatable_objects:
            QMessageBox.information(self, _("Info"), _("Please load a file or project first."))
            return
        dialog = AdvancedSearchDialog(self, _("Find and Replace"), self)
        dialog.exec()

    def cm_copy_original(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return
        text_to_copy = "\n".join([ts.original_semantic for ts in selected_objs])
        QApplication.clipboard().setText(text_to_copy)
        self.update_statusbar(_("Copied {count} original strings to clipboard.").format(count=len(selected_objs)))

    def cm_copy_translation(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return
        text_to_copy = "\n".join([ts.get_translation_for_ui() for ts in selected_objs])
        QApplication.clipboard().setText(text_to_copy)
        self.update_statusbar(_("Copied {count} translations to clipboard.").format(count=len(selected_objs)))

    def cm_paste_to_translation(self):
        if not self.current_selected_ts_id:
            self.update_statusbar(_("Please select an item."))
            return

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return

        clipboard_content = QApplication.clipboard().text()

        if isinstance(clipboard_content, str):
            self.details_panel.translation_edit_text.setPlainText(clipboard_content)
            cleaned_content = clipboard_content
            self._apply_translation_to_model(ts_obj, cleaned_content, source="manual_paste")
            self.update_statusbar(_("Clipboard content pasted to translation."))
        else:
            self.update_statusbar(_("Paste failed: Clipboard content is not text."))

    def show_prompt_manager(self):
        dialog = PromptManagerDialog(self, _("AI Prompt Manager"), self)
        dialog.exec()

    def show_settings_dialog(self):
        dialog = SettingsDialog(self)
        dialog.exec()

    def _check_ai_prerequisites(self, show_error=True):
        if not requests:
            if show_error:
                QMessageBox.critical(self, _("AI Feature Unavailable"),
                                     _("Python 'requests' library not found. Please install it (pip install requests) to use AI translation features."))
            return False
        if not self.config.get("ai_api_key"):
            if show_error:
                QMessageBox.critical(self, _("API Key Missing"),
                                     _("API Key is not set. Please configure it in 'Tools > AI Settings'."))
            return False
        return True

    def _decrement_active_threads_and_dispatch_more(self):
        if not self.is_ai_translating_batch:
            if self.ai_batch_active_threads > 0:
                self.ai_batch_active_threads -= 1
            if self.ai_batch_active_threads == 0:
                self._finalize_batch_ai_translation()
            return

        if self.ai_batch_active_threads > 0:
            self.ai_batch_active_threads -= 1
        if self.ai_batch_next_item_index < self.ai_batch_total_items:
            interval = self.config.get("ai_api_interval", 200)
            QTimer.singleShot(interval, self._dispatch_next_ai_batch_item)
        elif self.ai_batch_active_threads == 0 and self.ai_batch_completed_count >= self.ai_batch_total_items:
            self._finalize_batch_ai_translation()

    def apply_and_select_next_untranslated(self):
        if not self.current_selected_ts_id:
            return

        self.apply_translation_from_button()
        current_proxy_index = self.proxy_model.mapFromSource(
            self.sheet_model.index_from_id(self.current_selected_ts_id))
        if not current_proxy_index.isValid():
            return

        next_untranslated_id = None
        start_row = current_proxy_index.row() + 1
        for i in range(start_row, self.proxy_model.rowCount()):
            index = self.proxy_model.index(i, 0)
            ts_obj = self.proxy_model.data(index, Qt.UserRole)
            if ts_obj and not ts_obj.translation.strip() and not ts_obj.is_ignored:
                next_untranslated_id = ts_obj.id
                break
        if not next_untranslated_id:
            for i in range(0, start_row - 1):
                index = self.proxy_model.index(i, 0)
                ts_obj = self.proxy_model.data(index, Qt.UserRole)
                if ts_obj and not ts_obj.translation.strip() and not ts_obj.is_ignored:
                    next_untranslated_id = ts_obj.id
                    break

        if next_untranslated_id:
            self.select_sheet_row_by_id(next_untranslated_id, see=True)
            self.details_panel.translation_edit_text.setFocus()
        else:
            self.update_statusbar(_("No more untranslated items."))

    def _generate_ai_context_strings(self, current_ts_id_to_exclude):
        contexts = {
            "translation_context": "",
            "original_context": ""
        }

        try:
            current_item_index = \
                [i for i, ts in enumerate(self.translatable_objects) if ts.id == current_ts_id_to_exclude][0]
        except IndexError:
            return contexts

        # --- 处理 [Translated Context] ---
        if self.config.get("ai_use_translation_context", True):
            max_neighbors = self.config.get("ai_context_neighbors", 3)
            context_pairs = []
            count = 0
            for i in range(current_item_index - 1, -1, -1):
                if 0 < max_neighbors <= count: break
                ts = self.translatable_objects[i]
                if ts.translation.strip() and not ts.is_ignored:
                    context_pairs.insert(0, (ts.original_semantic, ts.get_translation_for_ui()))
                    count += 1
            count = 0
            for i in range(current_item_index + 1, len(self.translatable_objects)):
                if 0 < max_neighbors <= count: break
                ts = self.translatable_objects[i]
                if ts.translation.strip() and not ts.is_ignored:
                    context_pairs.append((ts.original_semantic, ts.get_translation_for_ui()))
                    count += 1
            if context_pairs:
                header = f"| {_('Original')} | {_('Translation')} |\n|---|---|\n"
                rows = [
                    f"| {orig.replace('|', '\\|').replace(chr(10), ' ')} | {trans.replace('|', '\\|').replace(chr(10), ' ')} |"
                    for orig, trans in context_pairs]
                contexts["translation_context"] = header + "\n".join(rows)

        # --- [Untranslated Context] ---
        if self.config.get("ai_use_original_context", True):
            max_neighbors = self.config.get("ai_original_context_neighbors", 3)
            context_items = []
            count = 0
            for i in range(current_item_index - 1, -1, -1):
                if 0 < max_neighbors <= count: break
                ts = self.translatable_objects[i]
                if not ts.is_ignored:
                    context_items.insert(0, ts.original_semantic)
                    count += 1
            count = 0
            for i in range(current_item_index + 1, len(self.translatable_objects)):
                if 0 < max_neighbors <= count: break
                ts = self.translatable_objects[i]
                if not ts.is_ignored:
                    context_items.append(ts.original_semantic)
                    count += 1
            if context_items:
                formatted_items = [f"- \"{item.replace(chr(10), ' ').strip()}\"" for item in context_items]
                contexts["original_context"] = "\n".join(formatted_items)

        return contexts
    def _initiate_single_ai_translation(self, ts_id_to_translate, called_from_cm=False):
        if not ts_id_to_translate:
            return False

        ts_obj = self._find_ts_obj_by_id(ts_id_to_translate)
        if not ts_obj: return False
        if hasattr(self, 'plugin_manager'):
            objects_to_process = self.plugin_manager.run_hook(
                'process_ai_translate_list',
                [ts_obj]
            )
            if not objects_to_process or objects_to_process[0].id != ts_obj.id:
                self.update_statusbar(_("AI translation for '{text}...' was skipped by a plugin.").format(text=ts_obj.original_semantic[:20]))
                return False
        if not called_from_cm and self.current_selected_ts_id == ts_id_to_translate:
            current_editor_text = self.details_panel.translation_edit_text.toPlainText()
            if current_editor_text != ts_obj.get_translation_for_ui():
                self._apply_translation_to_model(ts_obj, current_editor_text, source="pre_single_ai_save")

        if ts_obj.is_ignored:
            if not called_from_cm:
                QMessageBox.information(self, _("Ignored"),
                                        _("The selected string is marked as ignored and will not be AI translated."))
            return False

        if ts_obj.translation.strip():
            if called_from_cm and len(self._get_selected_ts_objects_from_sheet()) > 1:
                return False

            reply = QMessageBox.question(self, _("Overwrite Confirmation"),
                                         _("String \"{text}...\" already has a translation. Overwrite with AI translation?").format(
                                             text=ts_obj.original_semantic[:50]),
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return False

        if not called_from_cm:
            self.update_statusbar(
                _("AI is translating: \"{text}...\"").format(text=ts_obj.original_semantic[:30].replace(chr(10), '↵')))

        context_dict = self._generate_ai_context_strings(ts_obj.id)
        plugin_placeholders = self.plugin_manager.run_hook('get_ai_translation_context') or {}
        target_language_name = next((name for name, code in SUPPORTED_LANGUAGES.items() if code == self.target_language), self.target_language)
        worker = AITranslationWorker(self, ts_obj.id, ts_obj.original_semantic, target_language_name,
                                     context_dict, plugin_placeholders, False)
        self.ai_thread_pool.start(worker)
        return True

    def _dispatch_next_ai_batch_item(self):
        if not self.is_ai_translating_batch: return
        if self.ai_batch_next_item_index >= self.ai_batch_total_items: return

        if self.ai_batch_semaphore.acquire(blocking=False):
            if not self.is_ai_translating_batch:
                self.ai_batch_semaphore.release()
                return

            if self.ai_batch_next_item_index >= self.ai_batch_total_items:
                self.ai_batch_semaphore.release()
                return

            current_item_idx = self.ai_batch_next_item_index
            self.ai_batch_next_item_index += 1
            self.ai_batch_active_threads += 1

            ts_id = self.ai_translation_batch_ids_queue[current_item_idx]
            ts_obj = self._find_ts_obj_by_id(ts_id)

            self.update_statusbar(
                _("AI Batch: Processing {current}/{total} (Concurrency: {threads})...").format(
                    current=current_item_idx + 1, total=self.ai_batch_total_items,
                    threads=self.ai_batch_active_threads),
                persistent=True)

            if ts_obj and not ts_obj.is_ignored:
                context_dict = self._generate_ai_context_strings(ts_obj.id)
                plugin_context = self.plugin_manager.run_hook('get_ai_translation_context') or {}
                plugin_placeholders = self.plugin_manager.run_hook('get_ai_translation_context') or {}
                target_language_name = next(
                    (name for name, code in SUPPORTED_LANGUAGES.items() if code == self.target_language),
                    self.target_language)
                worker = AITranslationWorker(self, ts_obj.id, ts_obj.original_semantic, target_language_name,
                                             context_dict, plugin_placeholders , True)
                self.ai_thread_pool.start(worker)
            else:
                self.ai_batch_semaphore.release()
                self.ai_batch_active_threads -= 1
                self.ai_batch_completed_count += 1
                if self.is_ai_translating_batch:
                    if self.ai_batch_next_item_index < self.ai_batch_total_items:
                        QTimer.singleShot(0, self._dispatch_next_ai_batch_item)
                    elif self.ai_batch_active_threads == 0 and self.ai_batch_completed_count >= self.ai_batch_total_items:
                        self._finalize_batch_ai_translation()

    def cm_set_ignored_status(self, ignore_flag):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return

        will_any_disappear = ignore_flag and not self.show_ignored_var

        bulk_changes = []
        changed_ids = set()
        for ts_obj in selected_objs:
            if ts_obj.is_ignored != ignore_flag:
                old_val = ts_obj.is_ignored
                ts_obj.is_ignored = ignore_flag
                if not ignore_flag: ts_obj.was_auto_ignored = False
                ts_obj.update_style_cache()
                bulk_changes.append(
                    {'string_id': ts_obj.id, 'field': 'is_ignored', 'old_value': old_val, 'new_value': ignore_flag})
                changed_ids.add(ts_obj.id)

        if bulk_changes:
            self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
            self.mark_project_modified()
            self.update_statusbar(_("{count} items' ignore status updated.").format(count=len(bulk_changes)))

            if will_any_disappear:
                self.force_full_refresh(id_to_reselect=self.current_selected_ts_id)
            else:
                self._update_view_for_ids(changed_ids)

    def cm_set_reviewed_status(self, reviewed_flag):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return

        will_any_disappear = reviewed_flag and self.unreviewed_checkbox.isChecked()

        bulk_changes = []
        changed_ids = set()
        for ts_obj in selected_objs:
            if ts_obj.is_reviewed != reviewed_flag:
                old_val = ts_obj.is_reviewed
                ts_obj.is_reviewed = reviewed_flag
                ts_obj.update_style_cache()
                bulk_changes.append(
                    {'string_id': ts_obj.id, 'field': 'is_reviewed', 'old_value': old_val, 'new_value': reviewed_flag})
                changed_ids.add(ts_obj.id)

        if bulk_changes:
            self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
            self.mark_project_modified()
            self.update_statusbar(_("{count} items' review status updated.").format(count=len(bulk_changes)))

            if will_any_disappear:
                self.force_full_refresh(id_to_reselect=self.current_selected_ts_id)
            else:
                self._update_view_for_ids(changed_ids)

    def cm_toggle_ignored_status(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return

        will_any_disappear = (not self.show_ignored_var) and any(not ts_obj.is_ignored for ts_obj in selected_objs)

        bulk_changes = []
        changed_ids = set()
        for ts_obj in selected_objs:
            old_val = ts_obj.is_ignored
            new_val = not old_val
            ts_obj.is_ignored = new_val
            if not new_val: ts_obj.was_auto_ignored = False
            ts_obj.update_style_cache()
            bulk_changes.append(
                {'string_id': ts_obj.id, 'field': 'is_ignored', 'old_value': old_val, 'new_value': new_val})
            changed_ids.add(ts_obj.id)

        if bulk_changes:
            self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
            self.mark_project_modified()
            self.update_statusbar(_("{count} items' ignore status updated.").format(count=len(bulk_changes)))

            if will_any_disappear:
                self.force_full_refresh(id_to_reselect=self.current_selected_ts_id)
            else:
                self._update_view_for_ids(changed_ids)

    def cm_toggle_reviewed_status(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return

        will_any_disappear = self.unreviewed_checkbox.isChecked() and any(
            not ts_obj.is_reviewed for ts_obj in selected_objs)

        bulk_changes = []
        changed_ids = set()
        for ts_obj in selected_objs:
            old_val = ts_obj.is_reviewed
            new_val = not old_val
            ts_obj.is_reviewed = new_val
            ts_obj.update_style_cache()
            bulk_changes.append(
                {'string_id': ts_obj.id, 'field': 'is_reviewed', 'old_value': old_val, 'new_value': new_val})
            changed_ids.add(ts_obj.id)

        if bulk_changes:
            self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
            self.mark_project_modified()
            self.update_statusbar(_("{count} items' review status updated.").format(count=len(bulk_changes)))

            if will_any_disappear:
                self.force_full_refresh(id_to_reselect=self.current_selected_ts_id)
            else:
                self._update_view_for_ids(changed_ids)

    def __and_dispatch_more(self):
        if not self.is_aidecrement_active_threads_translating_batch:
            if self.ai_batch_active_threads > 0: self.ai_batch_active_threads -= 1
            if self.ai_batch_active_threads == 0:
                self._finalize_batch_ai_translation()
            return

        if self.ai_batch_active_threads > 0:
            self.ai_batch_active_threads -= 1

        if self.ai_batch_next_item_index < self.ai_batch_total_items:
            interval = self.config.get("ai_api_interval", 200)
            QTimer.singleShot(interval, self._dispatch_next_ai_batch_item)
        elif self.ai_batch_active_threads == 0 and self.ai_batch_completed_count >= self.ai_batch_total_items:
            self._finalize_batch_ai_translation()

    def ai_translate_selected_from_menu(self):
        self.cm_ai_translate_selected()

    def ai_translate_selected_from_button(self):
        if not self.current_selected_ts_id:
            self.update_statusbar(_("Please select an item to AI translate."))
            return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if ts_obj:
            self._start_ai_batch_translation([ts_obj])

    def ai_translate_all_untranslated(self):
        untranslated_objs = [
            ts for ts in self.translatable_objects
            if not ts.is_ignored and not ts.translation.strip()
        ]
        if not untranslated_objs:
            QMessageBox.information(self, _("No Translation Needed"),
                                    _("No untranslated and non-ignored strings found."))
            return

        self._start_ai_batch_translation(untranslated_objs)

    def _start_ai_batch_translation(self, items_to_translate):
        if not self._check_ai_prerequisites(): return
        if self.is_ai_translating_batch:
            QMessageBox.warning(self, _("Operation Restricted"), _("AI batch translation is already in progress."))
            return
        if hasattr(self, 'plugin_manager'):
            items_to_translate = self.plugin_manager.run_hook(
                'process_ai_translate_list',
                items_to_translate
            )
        untranslated_items = []
        already_translated_items = []

        for ts in items_to_translate:
            if not ts.is_ignored:
                if ts.translation.strip():
                    already_translated_items.append(ts)
                else:
                    untranslated_items.append(ts)
        items_to_process_after_confirmation = list(untranslated_items)

        if already_translated_items:
            if len(items_to_translate) == 1:
                ts_obj = items_to_translate[0]
                reply = QMessageBox.question(self, _("Overwrite Confirmation"),
                                             _("String \"{text}...\" already has a translation. Overwrite with AI translation?").format(
                                                 text=ts_obj.original_semantic[:50]),
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.Yes:
                    items_to_process_after_confirmation.extend(already_translated_items)
                else:
                    return
            else:
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle(_("Overwrite Confirmation"))
                msg_box.setText(
                    _("{count} item(s) already have translations.").format(count=len(already_translated_items)))
                msg_box.setInformativeText(_("Do you want to overwrite them, skip them, or cancel the operation?"))
                overwrite_btn = msg_box.addButton(_("Overwrite All"), QMessageBox.YesRole)
                skip_btn = msg_box.addButton(_("Skip Translated"), QMessageBox.NoRole)
                cancel_btn = msg_box.addButton(_("Cancel"), QMessageBox.RejectRole)
                msg_box.exec()

                if msg_box.clickedButton() == overwrite_btn:
                    items_to_process_after_confirmation.extend(already_translated_items)
                elif msg_box.clickedButton() == cancel_btn:
                    return
        unique_originals_to_translate = {}
        for ts in items_to_process_after_confirmation:
            if ts.original_semantic not in unique_originals_to_translate:
                unique_originals_to_translate[ts.original_semantic] = ts

        self.ai_translation_batch_ids_queue = [ts.id for ts in unique_originals_to_translate.values()]

        if not self.ai_translation_batch_ids_queue:
            QMessageBox.information(self, _("AI Translation"), _("No items to process."))
            return

        self.ai_batch_total_items = len(self.ai_translation_batch_ids_queue)
        api_interval_ms = self.config.get('ai_api_interval', 200)
        max_concurrency = self.config.get('ai_max_concurrent_requests', 1)
        avg_api_time_estimate_s = 3.0

        if max_concurrency == 1:
            estimated_time_s = self.ai_batch_total_items * (avg_api_time_estimate_s + api_interval_ms / 1000.0)
        else:
            estimated_time_s = (self.ai_batch_total_items / max_concurrency) * avg_api_time_estimate_s + \
                               (self.ai_batch_total_items / max_concurrency) * (
                                       api_interval_ms / 1000.0)
        if self.ai_batch_total_items > 50:
            reply = QMessageBox.question(self, _("Confirm Batch Translation"),
                                         _("You are about to AI translate {count} unique strings.\n"
                                           "This will be applied to all identical original texts.\n"
                                           "Estimated time: ~{time_s:.1f} seconds.\n"
                                           "Continue?").format(count=self.ai_batch_total_items,
                                                               time_s=estimated_time_s),
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return
        self.is_ai_translating_batch = True
        self.ai_batch_completed_count = 0
        self.ai_batch_successful_translations_for_undo = []
        self.ai_batch_next_item_index = 0
        self.ai_batch_active_threads = 0
        self.ai_batch_semaphore = threading.Semaphore(max_concurrency)

        self.progress_bar.setValue(0)
        self.update_ai_related_ui_state()
        self.update_statusbar(
            _("AI batch translation started for {count} unique strings...").format(count=self.ai_batch_total_items),
            persistent=True)

        for __ in range(max_concurrency):
            if self.ai_batch_next_item_index < self.ai_batch_total_items:
                self._dispatch_next_ai_batch_item()
            else:
                break

    def _finalize_batch_ai_translation(self):
        if not self.is_ai_translating_batch and self.ai_batch_active_threads > 0:
            return
        self.is_finalizing_batch_translation = True
        try:
            changed_ids = {change['string_id'] for change in self.ai_batch_successful_translations_for_undo}
            if self.ai_batch_successful_translations_for_undo:
                self.add_to_undo_history('bulk_ai_translate',
                                         {'changes': self.ai_batch_successful_translations_for_undo})
                self.mark_project_modified()
                self.check_batch_placeholder_mismatches()
            success_count = len(self.ai_batch_successful_translations_for_undo)
            processed_items = self.ai_batch_completed_count
            self.update_statusbar(
                _("AI batch translation complete. Successfully translated {success_count}/{processed_count} items (total {total_items} planned).").format(
                    success_count=success_count, processed_count=processed_items,
                    total_items=self.ai_batch_total_items),
                persistent=True)
            self.is_ai_translating_batch = False
            self.ai_translation_batch_ids_queue = []
            self.ai_batch_successful_translations_for_undo = []
            self.ai_batch_semaphore = None
            self.ai_batch_active_threads = 0
            self.ai_batch_next_item_index = 0
            self.ai_batch_completed_count = 0
            self.update_ai_related_ui_state()
            self._update_view_for_ids(changed_ids)

        finally:
            QTimer.singleShot(0, lambda: setattr(self, 'is_finalizing_batch_translation', False))

    def _handle_ai_translation_result(self, ts_id, translated_text, error_message, is_batch_item):
        trigger_ts_obj = self._find_ts_obj_by_id(ts_id)
        if not trigger_ts_obj:
            if is_batch_item: self.ai_batch_completed_count += 1
            return

        if error_message:
            error_msg_display = _("AI translation failed for \"{text}...\": {error}").format(
                text=trigger_ts_obj.original_semantic[:20].replace('\n', '↵'), error=error_message)
            self.update_statusbar(error_msg_display)
            if self.ai_batch_total_items == 1:
                QMessageBox.critical(self, _("AI Translation Error"),
                                     _("AI translation failed for \"{text}...\":\n{error}").format(
                                         text=trigger_ts_obj.original_semantic[:50], error=error_message))
        elif translated_text is not None and translated_text.strip():
            if hasattr(self, 'plugin_manager'):
                processed_text = self.plugin_manager.run_hook(
                    'process_ai_translated_text',
                    translated_text,
                    ts_object=trigger_ts_obj
                )
            else:
                processed_text = translated_text

            cleaned_translation = processed_text.strip()
            original_text_to_match = trigger_ts_obj.original_semantic
            changed_ids = set()

            single_translation_undo_changes = []

            for ts_obj in self.translatable_objects:
                if ts_obj.original_semantic == original_text_to_match and \
                        (not ts_obj.translation.strip() or ts_obj.id == trigger_ts_obj.id):

                    old_undo_val = ts_obj.get_translation_for_storage_and_tm()
                    new_undo_val = cleaned_translation.replace('\n', '\\n')

                    change_data = {
                        'string_id': ts_obj.id,
                        'field': 'translation',
                        'old_value': old_undo_val,
                        'new_value': new_undo_val
                    }

                    if is_batch_item:
                        self.ai_batch_successful_translations_for_undo.append(change_data)
                    else:
                        single_translation_undo_changes.append(change_data)

                    ts_obj.set_translation_internal(cleaned_translation)
                    changed_ids.add(ts_obj.id)
            if cleaned_translation:
                pass
            if not is_batch_item and single_translation_undo_changes:
                if len(single_translation_undo_changes) > 1:
                    self.add_to_undo_history('bulk_ai_translate', {'changes': single_translation_undo_changes})
                else:
                    self.add_to_undo_history('single_change', single_translation_undo_changes[0])
            if not is_batch_item:
                self._update_view_for_ids(changed_ids)
            if self.current_selected_ts_id == trigger_ts_obj.id:
                self.force_refresh_ui_for_current_selection()

            if self.ai_batch_total_items == 1:
                self.update_statusbar(_("AI translation successful: \"{text}...\"").format(
                    text=trigger_ts_obj.original_semantic[:20].replace('\n', '↵')))

        if is_batch_item:
            self.ai_batch_completed_count += 1
            if self.ai_batch_total_items > 0:
                progress_percent = (self.ai_batch_completed_count / self.ai_batch_total_items) * 100
                self.progress_bar.setValue(int(progress_percent))

            self.update_statusbar(
                _("AI Batch: {current}/{total} completed ({progress_percent:.0f}%).").format(
                    current=self.ai_batch_completed_count, total=self.ai_batch_total_items,
                    progress_percent=progress_percent),
                persistent=True)

            if self.ai_batch_completed_count >= self.ai_batch_total_items and self.ai_batch_active_threads == 0:
                self._finalize_batch_ai_translation()


    def check_batch_placeholder_mismatches(self):
        mismatched_items = []
        for change in self.ai_batch_successful_translations_for_undo:
            ts_obj = self._find_ts_obj_by_id(change['string_id'])
            if not ts_obj: continue

            original_placeholders = set(self.placeholder_regex.findall(ts_obj.original_semantic))
            translated_placeholders = set(self.placeholder_regex.findall(ts_obj.translation))

            if original_placeholders != translated_placeholders:
                mismatched_items.append(ts_obj)

        if mismatched_items:
            msg = _(
                "After AI batch translation, {count} items were found with placeholder mismatches.\nDo you want to add the comment \"Placeholder Mismatch\" to these items in bulk?").format(
                count=len(mismatched_items))
            reply = QMessageBox.question(self, _("Placeholder Mismatch"), msg, QMessageBox.Yes | QMessageBox.No,
                                         QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                bulk_comment_changes = []
                for ts_obj in mismatched_items:
                    old_comment = ts_obj.comment
                    new_comment = (old_comment + " " + _("Placeholder Mismatch")).strip()
                    if ts_obj.comment != new_comment:
                        ts_obj.comment = new_comment
                        bulk_comment_changes.append({
                            'string_id': ts_obj.id, 'field': 'comment',
                            'old_value': old_comment, 'new_value': new_comment
                        })
                if bulk_comment_changes:
                    self.add_to_undo_history('bulk_context_menu', {'changes': bulk_comment_changes})
                    self._run_and_refresh_with_validation()
                    self.update_statusbar(_("Added comments to {count} placeholder mismatched items.").format(
                        count=len(bulk_comment_changes)))

    def stop_batch_ai_translation(self, silent=False):
        if not self.is_ai_translating_batch:
            if not silent:
                QMessageBox.information(self, _("Info"), _("No AI batch translation task is in progress."))
            return

        was_translating = self.is_ai_translating_batch
        self.is_ai_translating_batch = False

        if not silent:
            QMessageBox.information(self, _("AI Batch Translation"),
                                    _("AI batch translation stop requested.\nDispatched tasks will continue to complete, please wait."))

        self.update_statusbar(_("AI batch translation stop requested. Finishing dispatched tasks..."), persistent=True)

        if was_translating and self.ai_batch_active_threads == 0:
            self._finalize_batch_ai_translation()
        else:
            self.update_ai_related_ui_state()




    def _get_selected_ts_objects_from_sheet(self):
        selected_objs = []
        selected_rows = self.table_view.selectionModel().selectedRows()
        if not selected_rows:
            if self.current_selected_ts_id:
                ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
                if ts_obj:
                    return [ts_obj]
            return []

        added_ids = set()
        for index in selected_rows:
            ts_obj = self.proxy_model.data(index, Qt.UserRole)
            if ts_obj and ts_obj.id not in added_ids:
                selected_objs.append(ts_obj)
                added_ids.add(ts_obj.id)
        return selected_objs

    def select_sheet_row_by_id(self, ts_id, see=False):
        source_index = self.sheet_model.index_from_id(ts_id)
        if source_index.isValid():
            proxy_index = self.proxy_model.mapFromSource(source_index)
            if proxy_index.isValid():
                self.table_view.selectionModel().clearSelection()
                self.table_view.selectionModel().select(proxy_index, QItemSelectionModel.Select | QItemSelectionModel.Rows)
                self.table_view.setCurrentIndex(proxy_index)

                if see:
                    self.table_view.scrollTo(proxy_index, QAbstractItemView.ScrollHint.EnsureVisible)
            else:
                self.table_view.clearSelection()
        else:
            self.table_view.clearSelection()

    def force_refresh_ui_for_current_selection(self):
        if not self.current_selected_ts_id:
            self.clear_details_pane()
            return

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj:
            self.clear_details_pane()
            return
        self.details_panel.original_text_display.blockSignals(True)
        self.details_panel.translation_edit_text.blockSignals(True)
        try:
            self.details_panel.original_text_display.setPlainText(ts_obj.original_semantic)
            self.details_panel.translation_edit_text.setPlainText(ts_obj.get_translation_for_ui())
        finally:
            self.details_panel.original_text_display.blockSignals(False)
            self.details_panel.translation_edit_text.blockSignals(False)

        po_comments = ts_obj.po_comment.splitlines()
        user_comments = ts_obj.comment.splitlines()
        all_comment_lines = po_comments + user_comments
        final_text = "\n".join(all_comment_lines)
        self.comment_status_panel.comment_edit_text.setPlainText(final_text)

        self.comment_status_panel.ignore_checkbox.setChecked(ts_obj.is_ignored)
        ignore_label = _("Ignore this string")
        if ts_obj.is_ignored and ts_obj.was_auto_ignored:
            ignore_label += _(" (Auto)")
        self.comment_status_panel.ignore_checkbox.setText(ignore_label)

        self.comment_status_panel.reviewed_checkbox.setChecked(ts_obj.is_reviewed)
        self.comment_status_panel.highlighter.rehighlight()

        self.context_panel.set_context(ts_obj)
        self.schedule_tm_update(ts_obj.original_semantic)
        self._update_all_highlights()

    def schedule_tm_update(self, original_text):
        self.last_tm_query = original_text
        self.tm_update_timer.start(250)

    def perform_tm_update(self):
        if self.last_tm_query:
            source_lang = self.source_language
            target_lang = self.current_target_language if self.is_project_mode else self.target_language

            exact_match = self.tm_service.get_translation(self.last_tm_query, source_lang, target_lang)

            fuzzy_matches = self.tm_service.get_fuzzy_matches(
                source_text=self.last_tm_query,
                source_lang=source_lang,
                target_lang=target_lang,
                limit=5,
                threshold=0.70
            )

            if exact_match and fuzzy_matches:
                fuzzy_matches = [
                    match for match in fuzzy_matches
                    if match['source_text'] != self.last_tm_query
                ]

            self.tm_panel.update_tm_suggestions(exact_match, fuzzy_matches)

    def cm_edit_comment(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return

        initial_comment = selected_objs[0].comment if len(selected_objs) == 1 else ""
        prompt_text = _("Enter comment for {count} selected items:").format(count=len(selected_objs)) if len(
            selected_objs) > 1 else _("Original:\n{original_semantic}...\n\nEnter comment:").format(
            original_semantic=selected_objs[0].original_semantic[:100])

        new_comment, ok = QInputDialog.getMultiLineText(self, _("Edit Comment..."), prompt_text, initial_comment)

        if ok and new_comment is not None:
            bulk_changes = []
            for ts_obj in selected_objs:
                if ts_obj.comment != new_comment:
                    old_comment = ts_obj.comment
                    ts_obj.comment = new_comment
                    bulk_changes.append({
                        'string_id': ts_obj.id, 'field': 'comment',
                        'old_value': old_comment, 'new_value': new_comment
                    })

            if bulk_changes:
                self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
                self._run_and_refresh_with_validation()
                if self.current_selected_ts_id in [c['string_id'] for c in bulk_changes]:
                    self.comment_status_panel.comment_edit_text.setPlainText(new_comment)
                self.update_statusbar(_("Updated comments for {count} items.").format(count=len(bulk_changes)))
                self.mark_project_modified()

    def cm_apply_tm_to_selected(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return

        applied_count = 0
        bulk_changes = []
        source_lang = self.source_language
        target_lang = self.current_target_language if self.is_project_mode else self.target_language

        for ts_obj in selected_objs:
            if ts_obj.is_ignored: continue

            translation_from_tm = self.tm_service.get_translation(
                ts_obj.original_semantic, source_lang, target_lang
            )

            if translation_from_tm:
                tm_translation_ui = translation_from_tm.replace("\\n", "\n")
                if ts_obj.translation != tm_translation_ui:
                    old_val = ts_obj.get_translation_for_storage_and_tm()
                    ts_obj.set_translation_internal(tm_translation_ui)
                    bulk_changes.append({'string_id': ts_obj.id, 'field': 'translation', 'old_value': old_val,
                                         'new_value': translation_from_tm})
                    applied_count += 1

        if bulk_changes:
            self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
            self._run_and_refresh_with_validation()
            self.force_refresh_ui_for_current_selection()
            self.update_statusbar(_("Applied TM to {count} selected items.").format(count=applied_count))
            self.mark_project_modified()
        elif selected_objs:
            QMessageBox.information(self, _("Info"),
                                    _("No matching TM entries or no changes needed for selected items."))

    def cm_clear_selected_translations(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs:
            return

        reply = QMessageBox.question(self, _("Confirm Clear"),
                                     _("Are you sure you want to clear the translations for the {count} selected items?").format(
                                         count=len(selected_objs)), QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No:
            return

        bulk_changes = []
        cleared_ids = set()

        for ts_obj in selected_objs:
            if ts_obj.translation != "":
                old_val = ts_obj.get_translation_for_storage_and_tm()
                ts_obj.set_translation_internal("")
                bulk_changes.append({
                    'string_id': ts_obj.id,
                    'field': 'translation',
                    'old_value': old_val,
                    'new_value': ""
                })
                cleared_ids.add(ts_obj.id)

        if not bulk_changes:
            self.update_statusbar(_("Translations for selected items are already empty, no clearing needed."))
            return
        self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
        self.mark_project_modified()
        if self.current_selected_ts_id in cleared_ids:
            self.details_panel.translation_edit_text.setPlainText("")
            self.details_panel.translation_edit_text.document().setModified(False)
        self._run_and_refresh_with_validation()
        self.update_statusbar(_("Cleared {count} translations.").format(count=len(bulk_changes)))

    def cm_ai_translate_selected(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs:
            self.update_statusbar(_("No items selected for AI translation."))
            return
        if len(selected_objs) == 1:
            ts_obj = selected_objs[0]
            self._initiate_single_ai_translation(ts_obj.id, called_from_cm=True)
        else:
            self._start_ai_batch_translation(selected_objs)

    def _run_comparison_logic(self, new_filepath):
        try:
            is_po_mode = self.is_po_mode
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.update_statusbar(_("Parsing new file..."), persistent=True)
            QApplication.processEvents()

            new_strings = []
            new_code_content = None

            if is_po_mode:
                pot_file = polib.pofile(new_filepath, encoding='utf-8')
                old_strings_map = {s.original_semantic: s for s in self.translatable_objects}
                used_old_strings_for_fuzzy_match = set()

                for entry in pot_file:
                    if entry.obsolete:
                        continue
                    new_obj = po_file_service._po_entry_to_translatable_string(entry)
                    if new_obj.original_semantic in old_strings_map:
                        # 精确匹配
                        old_obj = old_strings_map[new_obj.original_semantic]
                        new_obj.translation = old_obj.translation
                        new_obj.comment = old_obj.comment
                        new_obj.po_comment = old_obj.po_comment
                        new_obj.is_fuzzy = old_obj.is_fuzzy
                        new_obj.is_reviewed = old_obj.is_reviewed
                        new_obj.is_ignored = old_obj.is_ignored
                        used_old_strings_for_fuzzy_match.add(old_obj)
                    else:
                        # 模糊匹配
                        best_match_score = 0
                        best_match_old_s = None
                        for old_s in self.translatable_objects:
                            if old_s not in used_old_strings_for_fuzzy_match:
                                score = fuzz.ratio(new_obj.original_semantic,old_s.original_semantic) / 100
                                if score > best_match_score:
                                    best_match_score = score
                                    best_match_old_s = old_s

                        if best_match_score >= 0.85 and best_match_old_s:
                            new_obj.translation = best_match_old_s.translation
                            new_obj.comment = best_match_old_s.comment
                            new_obj.po_comment = best_match_old_s.po_comment
                            new_obj.is_fuzzy = True
                            new_obj.is_reviewed = False
                            used_old_strings_for_fuzzy_match.add(best_match_old_s)

                    new_strings.append(new_obj)
                new_map = {s.original_semantic: s for s in new_strings}
                diff_results = {'added': [], 'removed': [], 'modified': [], 'unchanged': []}

                for new_obj in new_strings:
                    if new_obj.original_semantic in old_strings_map:
                        old_obj = old_strings_map[new_obj.original_semantic]
                        if new_obj.is_fuzzy:
                            diff_results['modified'].append({'old_obj': old_obj, 'new_obj': new_obj,
                                                             'similarity': fuzz.ratio(new_obj.original_semantic,old_obj.original_semantic) / 100})
                        else:
                            diff_results['unchanged'].append({'old_obj': old_obj, 'new_obj': new_obj})
                    else:
                        diff_results['added'].append({'new_obj': new_obj})

                for old_obj in self.translatable_objects:
                    if old_obj.original_semantic not in new_map:
                        diff_results['removed'].append({'old_obj': old_obj})
                self.progress_bar.setValue(70)
                summary = (_("Comparison complete. Found ") + _("{added} new items, ").format(
                    added=len(diff_results['added'])) + _("{removed} removed items, ").format(
                    removed=len(diff_results['removed'])) + _("and {modified} modified/inherited items.").format(
                    modified=len(diff_results['modified'])))
                diff_results['summary'] = summary

                dialog = DiffDialog(self, _("Version Comparison Results"), diff_results)
                self.progress_bar.setVisible(False)

                if dialog.exec():
                    self.update_statusbar(_("Applying updates..."), persistent=True)
                    self.translatable_objects = new_strings
                    self.apply_tm_to_all_current_strings(silent=True, only_if_empty=True)
                    self._run_and_refresh_with_validation()
                    self.mark_project_modified()
                    self.update_statusbar(_("Project updated to new version."), persistent=True)
                else:
                    self.update_statusbar(_("Version update cancelled."))
                return
            else:
                with open(new_filepath, 'r', encoding='utf-8', errors='replace') as f:
                    new_code_content = f.read()
                extraction_patterns = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)
                new_strings = extract_translatable_strings(new_code_content, extraction_patterns)
            self.progress_bar.setValue(30)
            self.update_statusbar(_("Comparing versions..."), persistent=True)
            QApplication.processEvents()
            old_strings = self.translatable_objects
            old_map = {s.original_semantic: s for s in old_strings}
            new_map = {s.original_semantic: s for s in new_strings}
            diff_results = {'added': [], 'removed': [], 'modified': [], 'unchanged': []}
            used_old_strings_for_fuzzy_match = set()
            for new_obj in new_strings:
                if new_obj.original_semantic in old_map:
                    old_obj = old_map[new_obj.original_semantic]

                    new_obj.translation = old_obj.translation
                    new_obj.comment = old_obj.comment
                    new_obj.po_comment = old_obj.po_comment
                    new_obj.is_fuzzy = old_obj.is_fuzzy
                    new_obj.is_ignored = old_obj.is_ignored
                    new_obj.is_reviewed = old_obj.is_reviewed
                    new_obj.occurrences = old_obj.occurrences
                    new_obj.context_lines = old_obj.context_lines
                    new_obj.current_line_in_context_idx = old_obj.current_line_in_context_idx

                    diff_results['unchanged'].append({'old_obj': old_obj, 'new_obj': new_obj})
                else:
                    best_match_score = 0
                    best_match_old_s = None
                    for old_s in old_strings:
                        if old_s.original_semantic not in new_map and old_s not in used_old_strings_for_fuzzy_match:
                            score = fuzz.ratio(new_obj.original_semantic, old_s.original_semantic) / 100
                            if score > best_match_score:
                                best_match_score = score
                                best_match_old_s = old_s
                    if best_match_score >= 0.85 and best_match_old_s:
                        new_obj.translation = best_match_old_s.translation
                        new_obj.comment = best_match_old_s.comment
                        new_obj.po_comment = best_match_old_s.po_comment
                        new_obj.is_fuzzy = True
                        new_obj.is_reviewed = False
                        new_obj.occurrences = best_match_old_s.occurrences
                        new_obj.context_lines = best_match_old_s.context_lines
                        new_obj.current_line_in_context_idx = best_match_old_s.current_line_in_context_idx

                        if new_obj.translation.strip():
                            if not hasattr(new_obj, 'minor_warnings') or not isinstance(new_obj.minor_warnings, list):
                                new_obj.minor_warnings = []
                            new_obj.minor_warnings.append(
                                (WarningType.FUZZY_TRANSLATION, _("Fuzzy match, please review.")))

                        diff_results['modified'].append(
                            {'old_obj': best_match_old_s, 'new_obj': new_obj, 'similarity': best_match_score})
                        used_old_strings_for_fuzzy_match.add(best_match_old_s)
                    else:
                        diff_results['added'].append({'new_obj': new_obj})
            for old_obj in old_strings:
                if old_obj.original_semantic not in new_map:
                    if old_obj not in used_old_strings_for_fuzzy_match:
                        diff_results['removed'].append({'old_obj': old_obj})


            for old_obj in old_strings:
                if old_obj.original_semantic not in new_map:
                    was_modified = any(res['old_obj'] is old_obj for res in diff_results['modified'])
                    if not was_modified:
                        diff_results['removed'].append({'old_obj': old_obj})

            self.progress_bar.setValue(70)
            summary = (_("Comparison complete. Found ") + _("{added} new items, ").format(
                added=len(diff_results['added'])) + _("{removed} removed items, ").format(
                removed=len(diff_results['removed'])) + _("and {modified} modified/inherited items.").format(
                modified=len(diff_results['modified'])))
            diff_results['summary'] = summary

            dialog = DiffDialog(self, _("Version Comparison Results"), diff_results)
            self.progress_bar.setVisible(False)

            if dialog.exec():
                self.update_statusbar(_("Applying updates..."), persistent=True)
                self.translatable_objects = new_strings
                if not is_po_mode and new_code_content is not None:
                    self.original_raw_code_content = new_code_content
                    self.current_code_file_path = new_filepath
                elif is_po_mode:
                    # self.current_po_file_path = new_filepath
                    pass
                self.apply_tm_to_all_current_strings(silent=True, only_if_empty=True)
                self._run_and_refresh_with_validation()
                self.mark_project_modified()
                self.update_statusbar(_("Project updated to new version."), persistent=True)
            else:
                self.update_statusbar(_("Version update cancelled."))

        except Exception as e:
            self.progress_bar.setVisible(False)
            QMessageBox.critical(self, _("Comparison Failed"), _("An error occurred: {error}").format(error=e))
            self.update_statusbar(_("Version comparison failed."))

    def compare_with_new_version(self):
        if not self.translatable_objects:
            QMessageBox.critical(self, _("Error"), _("Please open a project or file first."))
            return

        is_po_mode = self.is_po_mode
        if is_po_mode:
            title = _("Select new POT template for comparison")
            filetypes = _("PO Template Files (*.pot);;All Files (*.*)")
            initial_dir = os.path.dirname(self.current_po_file_path) if self.current_po_file_path else self.config.get(
                "last_dir", os.getcwd())
        else:
            title = _("Select new version code file for comparison")
            filetypes = _("Overwatch Workshop Files (*.ow *.txt);;All Files (*.*)")
            initial_dir = os.path.dirname(
                self.current_code_file_path) if self.current_code_file_path else self.config.get("last_dir",
                                                                                                 os.getcwd())

        filepath, selected_filter = QFileDialog.getOpenFileName(self, title, initial_dir, filetypes)
        if filepath:
            self._run_comparison_logic(filepath)