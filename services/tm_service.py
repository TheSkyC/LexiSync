# Copyright (c) 2025, TheSkyC
# SPDX-License-Identifier: Apache-2.0

import sqlite3
import re
import os
import json
import hashlib
import logging
import threading
from rapidfuzz import fuzz
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from contextlib import contextmanager
from openpyxl import load_workbook
from utils.localization import _

logger = logging.getLogger(__name__)

MANIFEST_FILE = "manifest.json"
DB_FILE = "tm.db"


class TMService:
    def __init__(self):
        self.project_db_path: Optional[str] = None
        self.global_db_path: Optional[str] = None
        self._lock = threading.RLock()

    def connect_databases(self, global_tm_path: Optional[str], project_tm_path: Optional[str] = None):
        with self._lock:
            self.disconnect_databases()

            if global_tm_path:
                os.makedirs(global_tm_path, exist_ok=True)
                self.global_db_path = os.path.join(global_tm_path, DB_FILE)
                with self._get_db_connection(self.global_db_path) as conn:
                    self._create_schema(conn)

            if project_tm_path:
                os.makedirs(project_tm_path, exist_ok=True)
                self.project_db_path = os.path.join(project_tm_path, DB_FILE)
                with self._get_db_connection(self.project_db_path) as conn:
                    self._create_schema(conn)

    def disconnect_databases(self):
        with self._lock:
            self.project_db_path = None
            self.global_db_path = None

    @contextmanager
    def _get_db_connection(self, db_path: str):
        conn = None
        try:
            conn = sqlite3.connect(
                db_path,
                timeout=30.0,
                check_same_thread=False,
                isolation_level=None
            )
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA synchronous=NORMAL")
            conn.row_factory = sqlite3.Row
            yield conn
        except Exception as e:
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
            logger.error(f"TM Database connection error for {db_path}: {e}")
            raise
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass

    def _create_schema(self, conn: sqlite3.Connection):
        try:
            cursor = conn.cursor()
            cursor.execute("BEGIN IMMEDIATE TRANSACTION")
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS translation_units (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_lang TEXT NOT NULL,
                    target_lang TEXT NOT NULL,
                    source_text TEXT NOT NULL,
                    target_text TEXT NOT NULL,
                    source_manifest_key TEXT NOT NULL,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(source_lang, target_lang, source_text)
                );
            """)
            cursor.execute(
                "CREATE INDEX IF NOT EXISTS idx_tm_source ON translation_units (source_lang, target_lang, source_text);")
            cursor.execute("COMMIT")
        except Exception as e:
            cursor.execute("ROLLBACK")
            logger.error(f"TM Schema creation failed: {e}")
            raise

    def find_conflicts(self, db_path: str, source_texts: List[str], source_lang: str, target_lang: str) -> Dict[
        str, dict]:
        """
        Batch check for existing TM entries.
        Returns: { 'source_text': {'id': id, 'original_text': source_text, 'existing_targets': ['target_text']} }
        Note: TM usually enforces unique (source, source_lang, target_lang), so existing_targets will have at most 1 item.
        """
        if not db_path or not os.path.exists(db_path):
            return {}

        conflicts = {}
        unique_sources = list(set(source_texts))

        try:
            with self._get_db_connection(db_path) as conn:
                cursor = conn.cursor()
                chunk_size = 900
                for i in range(0, len(unique_sources), chunk_size):
                    chunk = unique_sources[i:i + chunk_size]
                    placeholders = ','.join('?' for _ in chunk)

                    query = f"""
                        SELECT id, source_text, target_text
                        FROM translation_units
                        WHERE source_text IN ({placeholders})
                          AND source_lang = ?
                          AND target_lang = ?
                    """
                    cursor.execute(query, chunk + [source_lang, target_lang])

                    rows = cursor.fetchall()
                    for row in rows:
                        term_id = row['id']
                        src_text = row['source_text']
                        tgt_text = row['target_text']

                        # Use exact source text as key
                        if src_text not in conflicts:
                            conflicts[src_text] = {
                                'id': term_id,
                                'original_text': src_text,
                                'existing_targets': []
                            }

                        conflicts[src_text]['existing_targets'].append(tgt_text)

        except Exception as e:
            logger.error(f"Failed to find TM conflicts: {e}")

        return conflicts

    def delete_tm_entry(self, db_path: str, source_text: str, source_lang: str, target_lang: str) -> bool:
        if not source_text.strip() or not db_path:
            return False

        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")
                    cursor.execute("""
                        DELETE FROM translation_units
                        WHERE source_text = ? AND source_lang = ? AND target_lang = ?
                    """, (source_text, source_lang, target_lang))
                    rowcount = cursor.rowcount
                    cursor.execute("COMMIT")

                    if rowcount > 0:
                        logger.info(f"Deleted TM entry for '{source_text}'")
                        return True
                    else:
                        return False
            except Exception as e:
                logger.error(f"Failed to delete TM entry: {e}", exc_info=True)
                return False

    def get_translation(self, source_text: str, source_lang: str, target_lang: str, db_to_check: str = 'all') -> \
    Optional[str]:
        with self._lock:
            if db_to_check in ('all', 'project') and self.project_db_path:
                try:
                    with self._get_db_connection(self.project_db_path) as conn:
                        result = self._query_translation_in_db(conn, source_text, source_lang, target_lang)
                        if result is not None:
                            return result
                except Exception:
                    pass

            if db_to_check in ('all', 'global') and self.global_db_path:
                try:
                    with self._get_db_connection(self.global_db_path) as conn:
                        return self._query_translation_in_db(conn, source_text, source_lang, target_lang)
                except Exception:
                    pass
            return None

    def get_entry_count_by_source(self, dir_path: str, source_key: str) -> int:
        if not dir_path or not os.path.exists(dir_path): return 0
        db_path = os.path.join(dir_path, DB_FILE)
        if not os.path.exists(db_path): return 0

        try:
            with self._get_db_connection(db_path) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT COUNT(*) FROM translation_units WHERE source_manifest_key = ?",
                    (source_key,)
                )
                result = cursor.fetchone()
                return result[0] if result else 0
        except Exception as e:
            logger.error(f"Failed to get entry count: {e}")
            return 0

    def _query_translation_in_db(self, conn: sqlite3.Connection, source_text: str, source_lang: str,
                                 target_lang: str) -> Optional[str]:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT target_text FROM translation_units WHERE source_text = ? AND source_lang = ? AND target_lang = ?",
            (source_text, source_lang, target_lang)
        )
        row = cursor.fetchone()
        return row['target_text'] if row else None

    def update_tm_entry(self, db_path: str, source_text: str, target_text: str, source_lang: str, target_lang: str,
                        source_key: str = "manual"):
        if not source_text.strip() or not db_path: return

        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")
                    cursor.execute("""
                        INSERT OR REPLACE INTO translation_units 
                        (source_lang, target_lang, source_text, target_text, source_manifest_key)
                        VALUES (?, ?, ?, ?, ?)
                    """, (source_lang, target_lang, source_text, target_text, source_key))
                    cursor.execute("COMMIT")
            except Exception as e:
                logger.error(f"Failed to update TM entry: {e}")

    def import_from_file(self, source_filepath: str, tm_dir_path: str, source_lang: str, target_lang: str,
                         progress_callback=None) -> Tuple[bool, str]:
        with self._lock:
            db_path = os.path.join(tm_dir_path, DB_FILE)
            manifest_path = os.path.join(tm_dir_path, MANIFEST_FILE)

            try:
                os.makedirs(tm_dir_path, exist_ok=True)
                manifest = self._read_manifest(manifest_path)
                file_stats = self._get_file_stats(source_filepath)
                filename = os.path.basename(source_filepath)

                if filename in manifest.get("imported_sources", {}):
                    existing_stats = manifest["imported_sources"][filename]
                    if all(existing_stats.get(k) == file_stats.get(k) for k in
                           ["filesize", "last_modified", "checksum"]):
                        return True, _("This TM file has already been imported and has not changed.")

                if progress_callback: progress_callback(_("Parsing TM file..."))
                tus_to_import = self._parse_tm_file(source_filepath, source_lang, target_lang)

                if progress_callback: progress_callback(_("Connecting to database..."))
                with self._get_db_connection(db_path) as conn:
                    self._create_schema(conn)
                    self._merge_tus_into_db(conn, tus_to_import, filename, progress_callback)

                file_stats['import_date'] = datetime.now().isoformat() + "Z"
                file_stats['tu_count'] = len(tus_to_import)
                file_stats['source_lang'] = source_lang
                file_stats['target_lang'] = target_lang
                manifest.setdefault("imported_sources", {})[filename] = file_stats
                self._write_manifest(manifest_path, manifest)

                return True, _("Successfully imported {count} TM entries.").format(count=len(tus_to_import))
            except Exception as e:
                logger.error(f"Failed to import TM file '{source_filepath}': {e}", exc_info=True)
                return False, str(e)

    def get_fuzzy_matches(self, source_text: str, source_lang: str, target_lang: str, limit: int = 5,
                          threshold: float = 0.7) -> List[Dict]:
        all_matches = []
        with self._lock:
            if self.project_db_path:
                try:
                    with self._get_db_connection(self.project_db_path) as conn:
                        matches = self._query_fuzzy_in_db(conn, source_text, source_lang, target_lang, limit)
                        all_matches.extend(matches)
                except Exception:
                    pass

            if self.global_db_path:
                try:
                    with self._get_db_connection(self.global_db_path) as conn:
                        matches = self._query_fuzzy_in_db(conn, source_text, source_lang, target_lang, limit)
                        all_matches.extend(matches)
                except Exception:
                    pass

        if not all_matches:
            return []

        # Deduplicate and score
        unique_matches = {(m['source_text'], m['target_text']): m for m in all_matches}.values()
        scored_matches = []
        for match in unique_matches:
            score = fuzz.ratio(source_text, match['source_text']) / 100.0
            if score >= threshold:
                scored_matches.append({
                    "score": score,
                    "source_text": match['source_text'],
                    "target_text": match['target_text']
                })
        scored_matches.sort(key=lambda x: x['score'], reverse=True)

        return scored_matches[:limit]

    def _query_fuzzy_in_db(self, conn: sqlite3.Connection, source_text: str, source_lang: str, target_lang: str,
                           limit: int) -> List[Dict]:
        cursor = conn.cursor()
        keywords = [word for word in re.findall(r'\b\w+\b', source_text) if len(word) > 3]
        if not keywords:
            keywords = [source_text[:10]]

        keywords = keywords[:10]

        like_clauses = " OR ".join(["source_text LIKE ?"] * len(keywords))
        like_params = [f"%{kw}%" for kw in keywords]

        query = f"""
            SELECT source_text, target_text
            FROM translation_units
            WHERE source_lang = ? AND target_lang = ? AND ({like_clauses})
            LIMIT ?
        """
        params = [source_lang, target_lang] + like_params + [limit * 10]
        cursor.execute(query, params)
        return [dict(row) for row in cursor.fetchall()]

    def _parse_tm_file(self, filepath: str, source_lang: str, target_lang: str) -> List[Dict]:
        __, ext = os.path.splitext(filepath)
        ext = ext.lower()

        if ext == '.xlsx':
            return self._parse_xlsx(filepath, source_lang, target_lang)
        else:
            raise ValueError(_.format(ext=ext))

    def _parse_xlsx(self, filepath: str, source_lang: str, target_lang: str) -> List[Dict]:
        tus = []
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2 and row[0] is not None and row[1] is not None:
                source_text = str(row[0])
                target_text = str(row[1])
                tus.append({
                    "source_lang": source_lang,
                    "target_lang": target_lang,
                    "source_text": source_text,
                    "target_text": target_text
                })
        return tus

    def _merge_tus_into_db(self, conn: sqlite3.Connection, tus: List[Dict], source_key: str, progress_callback=None):
        cursor = conn.cursor()
        data_to_insert = [
            (
                tu['source_lang'], tu['target_lang'], tu['source_text'],
                tu['target_text'], source_key
            )
            for tu in tus if tu.get('source_text') and tu.get('target_text')
        ]

        if not data_to_insert:
            return

        try:
            cursor.execute("BEGIN IMMEDIATE TRANSACTION;")
            if progress_callback: progress_callback(
                _("Inserting {count} TM entries...").format(count=len(data_to_insert)))

            cursor.executemany("""
                INSERT OR REPLACE INTO translation_units 
                (source_lang, target_lang, source_text, target_text, source_manifest_key)
                VALUES (?, ?, ?, ?, ?)
            """, data_to_insert)

            cursor.execute("COMMIT")
        except Exception as e:
            cursor.execute("ROLLBACK")
            raise IOError(f"Database operation failed: {e}")

    def remove_source(self, source_key: str, tm_dir_path: str) -> Tuple[bool, str]:
        with self._lock:
            db_path = os.path.join(tm_dir_path, DB_FILE)
            manifest_path = os.path.join(tm_dir_path, MANIFEST_FILE)

            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION;")
                    cursor.execute("DELETE FROM translation_units WHERE source_manifest_key = ?", (source_key,))
                    rows_deleted = cursor.rowcount
                    cursor.execute("COMMIT")

                manifest = self._read_manifest(manifest_path)
                if "imported_sources" in manifest and source_key in manifest["imported_sources"]:
                    del manifest["imported_sources"][source_key]
                    self._write_manifest(manifest_path, manifest)

                return True, _("Successfully removed {count} TM entries from source '{source}'.").format(
                    count=rows_deleted, source=source_key)
            except Exception as e:
                logger.error(f"Failed to remove TM source '{source_key}': {e}", exc_info=True)
                return False, str(e)

    def query_entries(self, db_path: str, page: int = 1, page_size: int = 50,
                      source_key: str = None, src_lang: str = None, tgt_lang: str = None,
                      search_term: str = None) -> List[Dict]:
        if not db_path or not os.path.exists(db_path): return []

        offset = (page - 1) * page_size
        query = "SELECT id, source_text, target_text, source_lang, target_lang, source_manifest_key, created_at FROM translation_units WHERE 1=1"
        params = []

        if source_key and source_key != "All":
            query += " AND source_manifest_key = ?"
            params.append(source_key)
        if src_lang and src_lang != "All":
            query += " AND source_lang = ?"
            params.append(src_lang)
        if tgt_lang and tgt_lang != "All":
            query += " AND target_lang = ?"
            params.append(tgt_lang)
        if search_term:
            query += " AND (source_text LIKE ? OR target_text LIKE ?)"
            wildcard = f"%{search_term}%"
            params.extend([wildcard, wildcard])

        query += " ORDER BY id DESC LIMIT ? OFFSET ?"
        params.extend([page_size, offset])

        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute(query, params)
                    return [dict(row) for row in cursor.fetchall()]
            except Exception as e:
                logger.error(f"TM query failed: {e}")
                return []

    def count_entries(self, db_path: str, source_key: str = None, src_lang: str = None,
                      tgt_lang: str = None, search_term: str = None) -> int:
        if not db_path or not os.path.exists(db_path): return 0

        query = "SELECT COUNT(*) FROM translation_units WHERE 1=1"
        params = []
        if source_key and source_key != "All":
            query += " AND source_manifest_key = ?"
            params.append(source_key)
        if src_lang and src_lang != "All":
            query += " AND source_lang = ?"
            params.append(src_lang)
        if tgt_lang and tgt_lang != "All":
            query += " AND target_lang = ?"
            params.append(tgt_lang)
        if search_term:
            query += " AND (source_text LIKE ? OR target_text LIKE ?)"
            wildcard = f"%{search_term}%"
            params.extend([wildcard, wildcard])

        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute(query, params)
                    return cursor.fetchone()[0]
            except Exception as e:
                logger.error(f"TM count failed: {e}")
                return 0

    def update_entry_target(self, db_path: str, entry_id: int, new_target: str) -> bool:
        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")
                    cursor.execute("UPDATE translation_units SET target_text = ? WHERE id = ?", (new_target, entry_id))
                    cursor.execute("COMMIT")
                    return True
            except Exception as e:
                logger.error(f"TM update failed: {e}")
                return False

    def delete_entry_by_id(self, db_path: str, entry_id: int) -> bool:
        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")
                    cursor.execute("DELETE FROM translation_units WHERE id = ?", (entry_id,))
                    cursor.execute("COMMIT")
                    return True
            except Exception as e:
                logger.error(f"TM delete failed: {e}")
                return False

    def get_distinct_languages(self, db_path: str) -> Tuple[List[str], List[str]]:
        if not db_path or not os.path.exists(db_path): return [], []
        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT DISTINCT source_lang FROM translation_units")
                    srcs = [r[0] for r in cursor.fetchall()]
                    cursor.execute("SELECT DISTINCT target_lang FROM translation_units")
                    tgts = [r[0] for r in cursor.fetchall()]
                    return sorted(srcs), sorted(tgts)
            except Exception:
                return [], []

    def batch_update_tm(self, db_path: str, entries: List[Dict], source_lang: str, target_lang: str,
                        source_key: str, display_name: str, strategy: str = 'overwrite') -> Tuple[bool, str]:
        """
        Batch update TM.
        entries: List of {'source': str, 'target': str, 'action': 'new'|'overwrite'|'skip'}
        strategy: Global strategy fallback ('overwrite' or 'skip')
        """
        if not db_path or not os.path.exists(db_path):
            return False, _("Target database not found.")

        if not entries:
            return False, _("No entries to save.")

        # Prepare data
        data_to_insert = []

        for e in entries:
            action = e.get('action', strategy)
            if action == 'skip':
                continue

            data_to_insert.append((
                source_lang, target_lang, e['source'], e['target'], source_key
            ))

        if not data_to_insert:
            return True, _("No entries to update (all skipped).")

        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")

                    sql = """
                        INSERT OR REPLACE INTO translation_units 
                        (source_lang, target_lang, source_text, target_text, source_manifest_key)
                        VALUES (?, ?, ?, ?, ?)
                    """

                    cursor.executemany(sql, data_to_insert)

                    cursor.execute(
                        "SELECT COUNT(*) FROM translation_units WHERE source_manifest_key = ?",
                        (source_key,)
                    )
                    actual_count = cursor.fetchone()[0]

                    # Update manifest
                    manifest_path = os.path.join(os.path.dirname(db_path), MANIFEST_FILE)
                    manifest = self._read_manifest(manifest_path)

                    manifest.setdefault("imported_sources", {})[source_key] = {
                        "filepath": display_name,
                        "filesize": 0,
                        "last_modified": datetime.utcnow().isoformat() + "Z",
                        "checksum": "batch_save",
                        "import_date": datetime.utcnow().isoformat() + "Z",
                        "tu_count": actual_count,
                        "source_lang": source_lang,
                        "target_lang": target_lang
                    }

                    self._write_manifest(manifest_path, manifest)

                    cursor.execute("COMMIT")

                return True, _("Successfully processed {count} entries.").format(count=len(data_to_insert))
            except Exception as e:
                logger.error(f"Batch TM update failed: {e}", exc_info=True)
                return False, str(e)

    def _read_manifest(self, manifest_path: str) -> Dict:
        try:
            with open(manifest_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            return {"version": 1, "imported_sources": {}}

    def _write_manifest(self, manifest_path: str, manifest_data: Dict):
        with open(manifest_path, 'w', encoding='utf-8') as f:
            json.dump(manifest_data, f, indent=4, ensure_ascii=False)

    def _get_file_stats(self, filepath: str) -> Dict:
        stat = os.stat(filepath)
        normalized_filepath = filepath.replace('\\', '/')
        return {
            "filepath": normalized_filepath,
            "filesize": stat.st_size,
            "last_modified": datetime.fromtimestamp(stat.st_mtime).isoformat() + "Z",
            "checksum": self._calculate_checksum(filepath)
        }

    def _calculate_checksum(self, filepath: str, hash_algo="sha256") -> str:
        h = hashlib.new(hash_algo)
        with open(filepath, 'rb') as f:
            while chunk := f.read(8192):
                h.update(chunk)
        return h.hexdigest()