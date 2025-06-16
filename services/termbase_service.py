import os
from openpyxl import Workbook, load_workbook
from utils import constants

class TermbaseService:
    def __init__(self, tb_file=constants.TB_FILE_EXCEL):
        self.tb_file = tb_file
        self.termbase = {}
        self.load_termbase()

    def load_termbase(self):
        self.termbase = {}
        if not os.path.exists(self.tb_file):
            self._create_default_tb_file()
            return

        try:
            wb = load_workbook(self.tb_file, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2 and row[0] and row[1]:
                    self.termbase[str(row[0])] = str(row[1])
        except Exception as e:
            print(f"Error loading termbase: {e}")

    def get_term(self, term):
        return self.termbase.get(term)

    def get_all_terms(self):
        return self.termbase.keys()

    def get_mappings_for_text(self, text):
        mappings = []
        for term, translation in self.termbase.items():
            if term in text:
                mappings.append(f"'{term}' -> '{translation}'")
        return ", ".join(mappings)

    def _create_default_tb_file(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Termbase"
            ws.append(["Term", "Translation", "Notes"])
            ws.append(["Genji", "源氏", "Example hero name"])
            ws.append(["Payload", "运载目标", "Example objective name"])
            wb.save(self.tb_file)
        except Exception as e:
            print(f"Could not create default termbase file: {e}")