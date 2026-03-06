# Copyright (c) 2025, TheSkyC
# SPDX-License-Identifier: Apache-2.0

import json
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils.localization import _


def export_to_json(filepath, translatable_objects, displayed_ids_order=None, app_instance=None):
    items_to_export_data = []
    export_obj_list = []
    if displayed_ids_order and app_instance:
        obj_map = {obj.id: obj for obj in app_instance.translatable_objects}
        export_obj_list = [obj_map[ts_id] for ts_id in displayed_ids_order if ts_id in obj_map]
    elif app_instance: # Fallback to all if no specific order/filter is provided
        export_obj_list = app_instance.translatable_objects
    else:
        export_obj_list = translatable_objects

    for ts_obj in export_obj_list:
        items_to_export_data.append({
            "id": ts_obj.id,
            "string_type": ts_obj.string_type,
            "original_semantic": ts_obj.original_semantic,
            "translation": ts_obj.get_translation_for_storage_and_tm(),
            "comment": ts_obj.comment,
            "is_reviewed": ts_obj.is_reviewed,
            "is_ignored": ts_obj.is_ignored,
            "line_num_in_file": ts_obj.line_num_in_file,
            "original_raw": ts_obj.original_raw,
            "is_plural": ts_obj.is_plural,
            "original_plural": ts_obj.original_plural,
            "plural_translations": ts_obj.plural_translations
        })

    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(items_to_export_data, f, indent=4, ensure_ascii=False)


def export_to_yaml(filepath, translatable_objects, displayed_ids_order=None, app_instance=None):
    items_to_export_data = []

    export_obj_list = []
    if displayed_ids_order and app_instance:
        obj_map = {obj.id: obj for obj in app_instance.translatable_objects}
        export_obj_list = [obj_map[ts_id] for ts_id in displayed_ids_order if ts_id in obj_map]
    elif app_instance: # Fallback to all if no specific order/filter is provided
        export_obj_list = app_instance.translatable_objects
    else:
        export_obj_list = translatable_objects

    for ts_obj in export_obj_list:
        items_to_export_data.append({
            "id": ts_obj.id,
            "string_type": ts_obj.string_type,
            "original_semantic": ts_obj.original_semantic,
            "translation": ts_obj.get_translation_for_storage_and_tm(),
            "comment": ts_obj.comment,
            "is_reviewed": ts_obj.is_reviewed,
            "is_ignored": ts_obj.is_ignored,
            "line_num_in_file": ts_obj.line_num_in_file,
            "original_raw": ts_obj.original_raw,
            "is_plural": ts_obj.is_plural,
            "original_plural": ts_obj.original_plural,
            "plural_translations": ts_obj.plural_translations
        })

    with open(filepath, 'w', encoding='utf-8') as f:
        yaml.dump(items_to_export_data, f, allow_unicode=True, sort_keys=False)


def export_qa_report(filepath, issues, project_name="LexiSync Project"):
    """
    生成带格式的 QA 报告
    issues: List of dicts {severity, file, line, source, translation, type, message, ignored}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "QA Issues"

    # 定义样式
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    error_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # 浅红
    warning_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")  # 浅黄
    info_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # 浅蓝

    headers = [
        _("Severity"), _("File"), _("Line/ID"), _("Source"),
        _("Translation"), _("Issue Type"), _("Description"), _("Status")
    ]
    ws.append(headers)

    # 应用表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for issue in issues:
        row_data = [
            issue['severity'].upper(),
            issue['file'],
            issue['line'],
            issue['source'],
            issue['translation'],
            issue['type'],
            issue['message'],
            _("Ignored") if issue['ignored'] else _("Active")
        ]
        ws.append(row_data)

        # 应用行样式
        last_row = ws.max_row
        fill = None
        if issue['severity'] == 'error':
            fill = error_fill
        elif issue['severity'] == 'warning':
            fill = warning_fill
        elif issue['severity'] == 'info':
            fill = info_fill

        if fill:
            for cell in ws[last_row]:
                cell.fill = fill

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 最高50
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)