# Copyright (c) 2025, TheSkyC
# SPDX-License-Identifier: Apache-2.0

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog, scrolledtext
import re
import os
import shutil
import json
import datetime
import time
import threading
from copy import deepcopy
from difflib import SequenceMatcher
from openpyxl import Workbook, load_workbook
import tksheet
import polib

from models.translatable_string import TranslatableString
from dialogs.ai_settings_dialog import AISettingsDialog
from dialogs.search_dialog import AdvancedSearchDialog
from dialogs.font_settings_dialog import FontSettingsDialog
from services import export_service, po_file_service
from services.ai_translator import AITranslator
from services.code_file_service import extract_translatable_strings, save_translated_code
from services.project_service import load_project, save_project
from services.prompt_service import generate_prompt_from_structure
from services.validation_service import run_validation_on_all, placeholder_regex as enhanced_placeholder_regex

from utils.constants import *
from utils import config_manager
from utils.localization import lang_manager, _
from utils.constants import DEFAULT_EXTRACTION_PATTERNS, EXTRACTION_PATTERN_PRESET_EXTENSION

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except ImportError:
    TkinterDnD = None
    print("提示: tkinterdnd2 未找到, 文件拖放功能不可用。pip install tkinterdnd2-universal")

try:
    import requests
except ImportError:
    requests = None
    print("提示: requests 未找到, AI翻译功能不可用。pip install requests")

initial_config = config_manager.load_config()
language_code = initial_config.get('language')
lang_manager.setup_translation(language_code)
if not language_code:
    language_code = lang_manager.get_best_match_language()
    initial_config['language'] = language_code
lang_manager.setup_translation(language_code)

class OverwatchLocalizerApp:
    def __init__(self, root):
        self.root = root
        self.config = initial_config
        if TkinterDnD and isinstance(root, TkinterDnD.Tk):
            pass
        elif TkinterDnD:
            self.root = TkinterDnD.DnDWrapper(self.root)

        self.root.title(_("Overwatch Localizer - v{version}").format(version=APP_VERSION))
        self.root.geometry("1600x900")

        self.ACTION_MAP = {
            'open_code_file': {'method': self.open_code_file_dialog, 'desc': _('Open Code File')},
            'open_project': {'method': self.open_project_dialog, 'desc': _('Open Project')},
            'save_current_file': {'method': self.save_current_file, 'desc': _('Save')},
            'save_code_file': {'method': self.save_code_file, 'desc': _('Save Translation to New Code File')},
            'undo': {'method': self.undo_action, 'desc': _('Undo')},
            'redo': {'method': self.redo_action, 'desc': _('Redo')},
            'find_replace': {'method': self.show_advanced_search_dialog, 'desc': _('Find/Replace')},
            'copy_original': {'method': self.copy_selected_original_text_menu, 'desc': _('Copy Original')},
            'paste_translation': {'method': self.paste_clipboard_to_selected_translation_menu,
                                  'desc': _('Paste to Translation')},
            'ai_translate_selected': {'method': self.ai_translate_selected_from_menu,
                                      'desc': _('AI Translate Selected')},
            'toggle_reviewed': {'method': self.cm_toggle_reviewed_status, 'desc': _('Toggle Reviewed Status')},
            'toggle_ignored': {'method': self.cm_toggle_ignored_status, 'desc': _('Toggle Ignored Status')},
            'apply_and_next': {'method': self.apply_and_select_next_untranslated,
                               'desc': _('Apply and Go to Next Untranslated')},
        }
        self.current_code_file_path = None
        self.current_project_file_path = None
        self.current_po_file_path = None
        self.original_raw_code_content = ""
        self.current_project_modified = False
        self.is_po_mode = False
        self.project_custom_instructions = ""
        self.current_po_metadata = None

        self.translatable_objects = []
        self.displayed_string_ids = []

        self.translation_memory = {}
        self.current_tm_file = None

        self.undo_history = []
        self.redo_history = []
        self.current_selected_ts_id = None

        self.config = config_manager.load_config()
        self.ai_translator = AITranslator(
            api_key=self.config.get("ai_api_key"),
            model_name=self.config.get("ai_model_name", "deepseek-chat"),
            api_url=self.config.get("ai_api_base_url", DEFAULT_API_URL)
        )
        self.ai_translation_batch_ids_queue = []
        self.is_ai_translating_batch = False
        self.ai_batch_total_items = 0
        self.ai_batch_dispatched_count = 0
        self.ai_batch_completed_count = 0
        self.ai_batch_successful_translations_for_undo = []
        self.ai_batch_semaphore = None
        self.ai_batch_next_item_index = 0
        self.ai_batch_active_threads = 0

        self.deduplicate_strings_var = tk.BooleanVar(value=self.config.get("deduplicate", False))
        self.show_ignored_var = tk.BooleanVar(value=self.config.get("show_ignored", True))
        self.show_untranslated_var = tk.BooleanVar(value=self.config.get("show_untranslated", False))
        self.show_translated_var = tk.BooleanVar(value=self.config.get("show_translated", False))
        self.show_unreviewed_var = tk.BooleanVar(value=self.config.get("show_unreviewed", False))
        self.search_var = tk.StringVar()

        self.auto_save_tm_var = tk.BooleanVar(value=self.config.get("auto_save_tm", False))
        self.auto_backup_tm_on_save_var = tk.BooleanVar(value=self.config.get("auto_backup_tm_on_save", True))

        font_settings = self.config["font_settings"]
        if font_settings["override_default_fonts"]:
            lang_code = lang_manager.get_current_language()
            script_type = 'latin'
            if lang_code.startswith('zh') or lang_code.startswith('ja') or lang_code.startswith('ko'):
                script_type = 'cjk'
            elif lang_code.startswith('ru'):
                script_type = 'cyrillic'

            main_cfg = font_settings["scripts"].get(script_type, font_settings["scripts"]["latin"])
            code_cfg = font_settings["code_context"]

            self.app_font = (main_cfg["family"], main_cfg["size"], main_cfg["style"])
            self.search_font = (main_cfg["family"], main_cfg["size"] - 1, main_cfg["style"])
            self.context_font = (code_cfg["family"], code_cfg["size"], code_cfg["style"])
        else:
            try:
                primary_font_family = "Source Han Sans"
                tk.font.Font(family=primary_font_family, size=10).actual()
            except tk.TclError:
                primary_font_family = "TkDefaultFont"
            self.app_font = (primary_font_family, 10, "normal")
            self.app_font_header = (primary_font_family, 10, "bold")
            self.search_font = (primary_font_family, 9, "normal")
            self.context_font = ("Consolas", 9, "normal")

        self.app_font_bold = (self.app_font[0], self.app_font[1], "bold")

        self._ignored_tag_font = None
        self.icons = self._load_icons()
        self.placeholder_regex = enhanced_placeholder_regex
        self._placeholder_validation_job = None

        self.last_sort_column = "seq_id"
        self.last_sort_reverse = False

        self._apply_theme()
        self._setup_menu()
        self._setup_main_layout()
        self._setup_statusbar()
        self._setup_drag_drop()
        self._setup_sheet_context_menu()

        self._load_default_tm_excel()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.update_ui_state_after_file_load()
        self.update_ai_related_ui_state()
        self.update_counts_display()
        self.update_title()
        self.update_recent_files_menu()

    def _load_icons(self):
        return {}

    def _apply_theme(self):
        style = ttk.Style(self.root)
        try:
            available_themes = style.theme_names()
            if 'clam' in available_themes:
                style.theme_use('clam')
            elif 'vista' in available_themes:
                style.theme_use('vista')
            elif 'aqua' in available_themes:
                style.theme_use('aqua')
            elif 'alt' in available_themes:
                style.theme_use('alt')

            style.configure("Treeview.Heading", font=(self.app_font, 10, 'bold'))
            style.configure("TNotebook.Tab", padding=[10, 5], font=(self.app_font, 10))
            style.configure("Status.TFrame", relief=tk.SUNKEN, borderwidth=1)
            style.configure("Filter.TFrame")
            style.configure("Toolbar.TButton", padding=5)
        except tk.TclError:
            print("TTK 主题不可用或应用失败。")

    def _setup_drag_drop(self):
        if TkinterDnD:
            self.root.drop_target_register(DND_FILES)
            self.root.dnd_bind('<<Drop>>', self.handle_drop)

    def handle_drop(self, event):
        if event.data:
            try:
                files = self.root.tk.splitlist(event.data)
                if files:
                    filepath = files[0]
                    if os.path.isfile(filepath):
                        if filepath.lower().endswith(".pot"):
                            self.handle_pot_file_drop(filepath)
                            return
                        elif filepath.lower().endswith((".ow", ".txt")):
                            if self.prompt_save_if_modified():
                                self.open_code_file_path(filepath)
                        elif filepath.lower().endswith(PROJECT_FILE_EXTENSION):
                            if self.prompt_save_if_modified():
                                self.open_project_file(filepath)
                        elif filepath.lower().endswith((".po", ".pot")):
                            if self.prompt_save_if_modified():
                                self.import_po_file_dialog_with_path(filepath)
                        else:
                            self.update_statusbar(_("Drag and drop failed: Invalid file type '{filename}'").format(
                                filename=os.path.basename(filepath)))
                    else:
                        self.update_statusbar(_("Drag and drop failed: '{filename}' is not a file.").format(
                            filename=os.path.basename(filepath)))
            except Exception as e:
                messagebox.showerror(_("Drag and Drop Error"),
                                     _("Error processing dropped file: {error}").format(error=e), parent=self.root)
                self.update_statusbar(_("Drag and drop processing error"))

    def save_config(self):
        config_manager.save_config(self)

    def add_to_recent_files(self, filepath):
        if not filepath: return
        recent_files = self.config.get("recent_files", [])
        if filepath in recent_files:
            recent_files.remove(filepath)
        recent_files.insert(0, filepath)
        self.config["recent_files"] = recent_files[:10]
        self.update_recent_files_menu()
        self.save_config()

    def update_recent_files_menu(self):
        self.recent_files_menu.delete(0, tk.END)
        recent_files = self.config.get("recent_files", [])
        if not recent_files:
            self.recent_files_menu.add_command(label=_("No History"), state=tk.DISABLED)
            return

        for i, filepath in enumerate(recent_files):
            label = f"{i + 1}: {filepath}"
            self.recent_files_menu.add_command(label=label, command=lambda p=filepath: self.open_recent_file(p))
        self.recent_files_menu.add_separator()
        self.recent_files_menu.add_command(label=_("Clear History"), command=self.clear_recent_files)

    def open_recent_file(self, filepath):
        if not os.path.exists(filepath):
            messagebox.showerror(_("File not found"), _("File '{filepath}' does not exist.").format(filepath=filepath),
                                 parent=self.root)
            recent_files = self.config.get("recent_files", [])
            if filepath in recent_files:
                recent_files.remove(filepath)
                self.config["recent_files"] = recent_files
                self.update_recent_files_menu()
            return

        if not self.prompt_save_if_modified():
            return

        if filepath.lower().endswith(PROJECT_FILE_EXTENSION):
            self.open_project_file(filepath)
        elif filepath.lower().endswith((".ow", ".txt")):
            self.open_code_file_path(filepath)

    def clear_recent_files(self):
        if messagebox.askyesno(_("Confirmation"), _("Are you sure you want to clear all recent file history?"),
                               parent=self.root):
            self.config["recent_files"] = []
            self.update_recent_files_menu()
            self.save_config()

    def about(self):
        messagebox.showinfo(_("About Overwatch Localizer"),
                            _("Overwatch Custom Code Translation Tool\n\n"
                              "Version: {version}\n"
                              "Author: TheSkyC\n"
                              "China Server ID: 小鸟游六花#56683 / Asia Server: 小鳥游六花#31665").format(
                                version=APP_VERSION), parent=self.root)

    def on_closing(self):
        if not self.prompt_save_if_modified():
            return

        if self.is_ai_translating_batch:
            if messagebox.askyesno(_("AI Translation in Progress"),
                                   _("AI batch translation is still in progress. Are you sure you want to exit?\nUnfinished translations will be lost."),
                                   parent=self.root):
                self.stop_batch_ai_translation(silent=True)
            else:
                return

        if self.current_tm_file and self.translation_memory:
            self.save_tm_to_excel(self.current_tm_file, silent=True, backup=self.auto_backup_tm_on_save_var.get())
        elif self.translation_memory:
            default_tm_path = self._get_default_tm_excel_path()
            if default_tm_path:
                self.save_tm_to_excel(default_tm_path, silent=True, backup=self.auto_backup_tm_on_save_var.get())

        self.save_config()
        self.root.destroy()

    def _setup_menu(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=_("File"), menu=file_menu)
        file_menu.add_command(label=_("Open Code File..."), command=self.ACTION_MAP['open_code_file']['method'])
        file_menu.add_command(label=_("Open Project..."), command=self.ACTION_MAP['open_project']['method'])
        file_menu.add_separator()
        file_menu.add_command(label=_("Compare/Import New Version..."), command=self.compare_with_new_version,
                              state=tk.DISABLED)
        file_menu.add_separator()
        file_menu.add_command(label=_("Save"), command=self.save_current_file, state=tk.DISABLED)
        file_menu.add_command(label=_("Save As..."), command=self.save_current_file_as, state=tk.DISABLED)
        file_menu.add_separator()
        file_menu.add_command(label=_("Save Translation to New Code File"),
                              command=self.ACTION_MAP['save_code_file']['method'],
                              state=tk.DISABLED)

        file_menu.add_separator()
        file_menu.add_command(label=_("Import Translations from Excel"),
                              command=self.import_project_translations_from_excel,
                              state=tk.DISABLED)
        file_menu.add_command(label=_("Export to Excel"), command=self.export_project_translations_to_excel,
                              state=tk.DISABLED)
        file_menu.add_command(label=_("Export to JSON"), command=self.export_project_translations_to_json,
                              state=tk.DISABLED)
        file_menu.add_command(label=_("Export to YAML"), command=self.export_project_translations_to_yaml,
                              state=tk.DISABLED)
        file_menu.add_separator()
        file_menu.add_command(label=_("Extract POT Template from Code..."), command=self.extract_to_pot_dialog)
        file_menu.add_command(label=_("Import Translations from PO File..."), command=self.import_po_file_dialog)
        file_menu.add_command(label=_("Export to PO File..."), command=self.export_to_po_file_dialog, state=tk.DISABLED)
        file_menu.add_separator()
        file_menu.add_command(label=_("Import TM (Excel)"), command=self.import_tm_excel_dialog)
        file_menu.add_command(label=_("Export Current TM (Excel)"), command=self.export_tm_excel_dialog)
        file_menu.add_separator()
        self.recent_files_menu = tk.Menu(file_menu, tearoff=0)
        file_menu.add_cascade(label=_("Recent Files"), menu=self.recent_files_menu)
        file_menu.add_separator()
        file_menu.add_command(label=_("Exit"), command=self.on_closing)
        self.file_menu = file_menu

        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=_("Edit"), menu=edit_menu)
        edit_menu.add_command(label=_("Undo"), command=self.ACTION_MAP['undo']['method'], state=tk.DISABLED)
        edit_menu.add_command(label=_("Redo"), command=self.ACTION_MAP['redo']['method'], state=tk.DISABLED)
        edit_menu.add_separator()
        edit_menu.add_command(label=_("Find/Replace..."), command=self.ACTION_MAP['find_replace']['method'],
                              state=tk.DISABLED)
        edit_menu.add_separator()
        edit_menu.add_command(label=_("Copy Original"), command=self.ACTION_MAP['copy_original']['method'],
                              state=tk.DISABLED)
        edit_menu.add_command(label=_("Paste to Translation"), command=self.ACTION_MAP['paste_translation']['method'],
                              state=tk.DISABLED)
        self.edit_menu = edit_menu

        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=_("View"), menu=view_menu)
        view_menu.add_checkbutton(label=_("Deduplicate Strings"), variable=self.deduplicate_strings_var,
                                  command=self.refresh_sheet_preserve_selection)
        view_menu.add_checkbutton(label=_("Show Ignored"), variable=self.show_ignored_var,
                                  command=self.refresh_sheet_preserve_selection)
        view_menu.add_checkbutton(label=_("Show Untranslated"), variable=self.show_untranslated_var,
                                  command=self.refresh_sheet_preserve_selection)
        view_menu.add_checkbutton(label=_("Show Translated"), variable=self.show_translated_var,
                                  command=self.refresh_sheet_preserve_selection)
        view_menu.add_checkbutton(label=_("Show Unreviewed"), variable=self.show_unreviewed_var,
                                  command=self.refresh_sheet_preserve_selection)

        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=_("Tools"), menu=tools_menu)
        tools_menu.add_command(label=_("Apply TM to Untranslated"),
                               command=lambda: self.apply_tm_to_all_current_strings(only_if_empty=True, confirm=True),
                               state=tk.DISABLED)
        tools_menu.add_command(label=_("Clear TM (in-memory)"),
                               command=self.clear_entire_translation_memory)
        tools_menu.add_separator()
        tools_menu.add_command(label=_("AI Translate Selected"),
                               command=self.ACTION_MAP['ai_translate_selected']['method'],
                               state=tk.DISABLED)
        tools_menu.add_command(label=_("AI Translate All Untranslated"), command=self.ai_translate_all_untranslated,
                               state=tk.DISABLED)
        tools_menu.add_command(label=_("Stop AI Batch Translation"), command=lambda: self.stop_batch_ai_translation(),
                               state=tk.DISABLED)
        tools_menu.add_separator()
        tools_menu.add_command(label=_("Project-specific Instructions..."),
                               command=self.show_project_custom_instructions_dialog,
                               state=tk.DISABLED)
        tools_menu.add_command(label=_("AI Settings..."), command=self.show_ai_settings_dialog)
        tools_menu.add_separator()
        tools_menu.add_command(label=_("Extraction Rule Manager..."), command=self.show_extraction_pattern_dialog)
        tools_menu.add_command(label=_("Reload Translatable Text"), command=self.reload_translatable_text,
                               state=tk.DISABLED)
        self.tools_menu = tools_menu

        settings_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=_("Settings"), menu=settings_menu)
        settings_menu.add_checkbutton(label=_("Auto-backup TM on Save"), variable=self.auto_backup_tm_on_save_var,
                                      command=self.save_config)
        self.language_var = tk.StringVar(value=self.config.get('language', 'en_US'))
        language_menu = tk.Menu(settings_menu, tearoff=0)
        settings_menu.add_cascade(label=_("Language"), menu=language_menu)

        available_langs = lang_manager.get_available_languages()
        for lang_code in available_langs:
            lang_name = {
                'en_US': 'English',  # Support
                'zh_CN': '简体中文',  # Support
                'ja_JP': '日本語',  # Support
                'ko_KR': '한국어',  # Support
                'fr_FR': 'le français',  # Support
                'de_DE': 'Deutsch',  # Support
                'ru_RU': 'русский язык',  # Support
                'es_ES': 'español (España)',  # Support
                'es_MX': 'español (Latinoamérica)',
                'pt_BR': 'português (Brasil)',
                'pt_PT': 'português (Portugal)',
                'it_IT': 'italiano',  # Support
                'pl_PL': 'polski',
                'tr_TR': 'Türkçe',
                'ar_SA': 'العربية',
                'zh_TW': '繁體中文'
            }.get(lang_code, lang_code)
            language_menu.add_radiobutton(
                label=lang_name,
                variable=self.language_var,
                value=lang_code,
                command=self.change_language
            )
        settings_menu.add_command(label=_("Keybinding Settings..."), command=self.show_keybinding_dialog)
        settings_menu.add_command(label=_("Font Settings..."), command=self.show_font_settings_dialog)
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=_("Help"), menu=help_menu)
        help_menu.add_command(label=_("About"), command=self.about)

        self._setup_keybindings()
        self.update_menu_accelerators()

    def _setup_keybindings(self):
        for action in self.ACTION_MAP.keys():
            for key_seq in self.config.get('keybindings', {}).values():
                if key_seq: self.root.unbind_all(key_seq)
            for key_seq in DEFAULT_KEYBINDINGS.values():
                if key_seq: self.root.unbind_all(key_seq)

        for action, key_sequence in self.config.get('keybindings', {}).items():
            if key_sequence and action in self.ACTION_MAP:
                command = self.ACTION_MAP[action]['method']
                self.root.bind_all(key_sequence, lambda e, cmd=command: cmd(e) or "break")

    def update_menu_accelerators(self):
        bindings = self.config.get('keybindings', {})
        self.file_menu.entryconfig(_("Open Code File..."), accelerator=bindings.get('open_code_file', ''))
        self.file_menu.entryconfig(_("Open Project..."), accelerator=bindings.get('open_project', ''))
        self.file_menu.entryconfig(_("Save"), accelerator=bindings.get('save_current_file', ''))
        self.file_menu.entryconfig(_("Save Translation to New Code File"),
                                   accelerator=bindings.get('save_code_file', ''))

        self.edit_menu.entryconfig(_("Undo"), accelerator=bindings.get('undo', ''))
        self.edit_menu.entryconfig(_("Redo"), accelerator=bindings.get('redo', ''))
        self.edit_menu.entryconfig(_("Find/Replace..."), accelerator=bindings.get('find_replace', ''))
        self.edit_menu.entryconfig(_("Copy Original"), accelerator=bindings.get('copy_original', ''))
        self.edit_menu.entryconfig(_("Paste to Translation"), accelerator=bindings.get('paste_translation', ''))

        self.tools_menu.entryconfig(_("AI Translate Selected"),
                                    accelerator=bindings.get('ai_translate_selected', ''))

    def show_keybinding_dialog(self):
        from dialogs.keybinding_dialog import KeybindingDialog
        KeybindingDialog(self.root, _("Keybinding Settings"), self)

    def show_font_settings_dialog(self):
        FontSettingsDialog(self.root, _("Font Settings"), self)

    def change_language(self):
        new_lang = self.language_var.get()
        if new_lang != self.config.get('language'):
            self.config['language'] = new_lang
            self.save_config()
            messagebox.showinfo(
                _("Restart Required"),
                _("Language settings have been changed. Please restart the application for the changes to take effect."),
                parent=self.root
            )

    def _setup_main_layout(self):
        main_frame = ttk.Frame(self.root, padding="5")
        main_frame.pack(expand=True, fill=tk.BOTH)

        self.paned_window = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        self.paned_window.pack(expand=True, fill=tk.BOTH, pady=(5, 0))

        self.left_pane = ttk.Frame(self.paned_window)
        self.paned_window.add(self.left_pane, weight=7)

        self._setup_filter_toolbar(self.left_pane)
        self._setup_sheet_panel(self.left_pane)

        self.right_pane = ttk.Frame(self.paned_window)
        self.paned_window.add(self.right_pane, weight=3)
        self._setup_details_pane(self.right_pane)

    def _setup_filter_toolbar(self, parent):
        toolbar = ttk.Frame(parent, style="Filter.TFrame", padding=5)
        toolbar.pack(side=tk.TOP, fill=tk.X, pady=(0, 5))

        ttk.Label(toolbar, text=_("Filter:")).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Checkbutton(toolbar, text=_("Deduplicate"), variable=self.deduplicate_strings_var,
                        command=self.refresh_sheet_preserve_selection).pack(side=tk.LEFT, padx=3)
        ttk.Checkbutton(toolbar, text=_("Ignored"), variable=self.show_ignored_var,
                        command=self.refresh_sheet_preserve_selection).pack(side=tk.LEFT, padx=3)
        ttk.Checkbutton(toolbar, text=_("Untranslated"), variable=self.show_untranslated_var,
                        command=self.refresh_sheet_preserve_selection).pack(side=tk.LEFT, padx=3)
        ttk.Checkbutton(toolbar, text=_("Translated"), variable=self.show_translated_var,
                        command=self.refresh_sheet_preserve_selection).pack(side=tk.LEFT, padx=3)
        ttk.Checkbutton(toolbar, text=_("Unreviewed"), variable=self.show_unreviewed_var,
                        command=self.refresh_sheet_preserve_selection).pack(side=tk.LEFT, padx=3)

        search_frame = ttk.Frame(toolbar)
        search_frame.pack(side=tk.RIGHT, padx=5)

        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=25, font=(self.app_font, 9))
        self.search_entry.pack(side=tk.LEFT, padx=(0, 5))
        self.search_entry.bind("<Return>", lambda e: self.find_string_from_toolbar())

        self.search_entry.bind("<FocusIn>", self.on_search_focus_in)
        self.search_entry.bind("<FocusOut>", self.on_search_focus_out)

        self.on_search_focus_out(None)

        search_button = ttk.Button(search_frame, text=_("Find"), command=self.find_string_from_toolbar,
                                   style="Toolbar.TButton")
        search_button.pack(side=tk.LEFT)

    def _setup_sheet_panel(self, parent):
        sheet_frame = ttk.Frame(parent)
        sheet_frame.pack(expand=True, fill=tk.BOTH, padx=0, pady=0)

        self.sheet = tksheet.Sheet(
            sheet_frame,
            headers=["#", "S", _("Original"), _("Translation"), _("Comment"), "✔", _("col_line")],
            font=self.app_font,
            header_font=self.app_font_bold,
            show_x_scrollbar=True,
            show_y_scrollbar=True,
            show_top_left=False,
            show_row_index=False,
            show_header=True,
            show_headings=True,
            right_click_select_row=False,
            data=[]
        )
        self.sheet.pack(expand=True, fill=tk.BOTH)

        self.sheet.enable_bindings(
            "single_select",
            "row_select",
            "arrowkeys",
            "column_header_select",
            "column_width_resize",
            "drag_select",
        )

        self.sheet.extra_bindings([
            ("column_header_select", self._sort_sheet_column)
        ])

        self.sheet.bind("<ButtonRelease-3>", self.show_sheet_context_menu)
        self.sheet.bind("<Double-1>", self.on_sheet_double_click)
        self.sheet.bind("<ButtonRelease-1>", self.on_sheet_select)
        self.sheet.bind("<KeyRelease-Up>", self.on_sheet_select)
        self.sheet.bind("<KeyRelease-Down>", self.on_sheet_select)

        self.sheet.column_width(column=0, width=40)
        self.sheet.column_width(column=1, width=30)
        self.sheet.column_width(column=2, width=300)
        self.sheet.column_width(column=3, width=300)
        self.sheet.column_width(column=4, width=90)
        self.sheet.column_width(column=5, width=30)
        self.sheet.column_width(column=6, width=50)

        self.sheet.align_columns(columns=[0], align="e")
        self.sheet.align_columns(columns=[1, 5, 6], align="center")

        self.sheet.readonly_columns(list(range(7)))

    def _setup_sheet_context_menu(self):
        self.sheet_context_menu = tk.Menu(self.sheet, tearoff=0)
        self.sheet_context_menu.add_command(label=_("Copy Original"), command=self.cm_copy_original)
        self.sheet_context_menu.add_command(label=_("Copy Translation"), command=self.cm_copy_translation)
        self.sheet_context_menu.add_separator()
        self.sheet_context_menu.add_command(label=_("Mark as Ignored"),
                                            command=lambda: self.cm_set_ignored_status(True))
        self.sheet_context_menu.add_command(label=_("Unmark as Ignored"),
                                            command=lambda: self.cm_set_ignored_status(False))
        self.sheet_context_menu.add_separator()
        self.sheet_context_menu.add_command(label=_("Mark as Reviewed"),
                                            command=lambda: self.cm_set_reviewed_status(True))
        self.sheet_context_menu.add_command(label=_("Mark as Unreviewed"),
                                            command=lambda: self.cm_set_reviewed_status(False))
        self.sheet_context_menu.add_separator()
        self.sheet_context_menu.add_command(label=_("Edit Comment..."), command=self.cm_edit_comment)
        self.sheet_context_menu.add_separator()
        self.sheet_context_menu.add_command(label=_("Apply Memory to Selected Items"),
                                            command=self.cm_apply_tm_to_selected)
        self.sheet_context_menu.add_command(label=_("Clear Selected Translations"),
                                            command=self.cm_clear_selected_translations)
        self.sheet_context_menu.add_separator()
        self.sheet_context_menu.add_command(label=_("Use AI to Translate Selected Items"),
                                            command=self.cm_ai_translate_selected)
        self.sheet_context_menu.add_separator()
        self.sheet_context_menu.add_command(label=_("Ignore Warnings for Selected"), command=lambda: self.cm_set_warning_ignored_status(True))
        self.sheet_context_menu.add_command(label=_("Un-ignore Warnings for Selected"), command=lambda: self.cm_set_warning_ignored_status(False))

    def show_sheet_context_menu(self, event):
        clicked_row = self.sheet.identify_row(event=event, allow_end=False)
        if clicked_row is None:
            return
        selection_boxes = self.sheet.get_all_selection_boxes_with_types()
        selected_rows = set()
        for box, _ in selection_boxes:
            for r in range(box[0], box[2]):
                selected_rows.add(r)

        if clicked_row not in selected_rows:
            self.sheet.deselect("all")
            self.sheet.select_row(row=clicked_row)
            self.on_sheet_select()

        self.sheet_context_menu.post(event.x_root, event.y_root)

    def on_sheet_double_click(self, event):
        clicked_col = self.sheet.identify_column(event=event)

        if clicked_col == 2:
            self.original_text_display.focus_set()
        elif clicked_col == 3:
            self.translation_edit_text.focus_set()
        elif clicked_col == 4:
            self.comment_edit_text.focus_set()

    def _sort_sheet_column(self, event=None, column_index=None):
        if event:
            column_index = event.column

        if column_index is None:
            return

        col_map = {0: "seq_id", 1: "status", 2: "original", 3: "translation", 4: "comment", 5: "reviewed", 6: "line"}
        col_key = col_map.get(column_index)
        if not col_key:
            return

        if self.last_sort_column == col_key:
            self.last_sort_reverse = not self.last_sort_reverse
        else:
            self.last_sort_reverse = False
        self.last_sort_column = col_key

        def get_sort_key(ts_obj):
            if col_key == "seq_id":
                return self.displayed_string_ids.index(ts_obj.id) if ts_obj.id in self.displayed_string_ids else float(
                    'inf')
            elif col_key == "line":
                return ts_obj.line_num_in_file
            elif col_key == "reviewed":
                return ts_obj.is_reviewed
            elif col_key == "status":
                if ts_obj.is_ignored: return 3
                if ts_obj.translation.strip(): return 2
                return 1
            elif col_key == "original":
                return ts_obj.original_semantic.lower()
            elif col_key == "translation":
                return ts_obj.get_translation_for_ui().lower()
            elif col_key == "comment":
                return ts_obj.comment.lower()
            return 0

        self.translatable_objects.sort(key=get_sort_key, reverse=self.last_sort_reverse)
        self.refresh_sheet(preserve_selection=True)

    def _setup_details_pane(self, parent_frame):
        details_outer_frame = ttk.LabelFrame(parent_frame, text=_("Edit & Details"), padding="5")
        details_outer_frame.pack(expand=True, fill=tk.BOTH, padx=5, pady=0)

        details_paned_window = ttk.PanedWindow(details_outer_frame, orient=tk.VERTICAL)
        details_paned_window.pack(fill=tk.BOTH, expand=True)

        top_section_frame = ttk.Frame(details_paned_window, padding=5)
        details_paned_window.add(top_section_frame, weight=4)
        top_section_frame.columnconfigure(0, weight=1)

        ttk.Label(top_section_frame, text=_("Original (Ctrl+Shift+C to copy):")).pack(anchor=tk.W, padx=5, pady=(0, 2))
        orig_frame = ttk.Frame(top_section_frame)
        orig_frame.pack(fill=tk.X, expand=False, padx=5, pady=(0, 5))
        orig_frame.grid_rowconfigure(0, weight=1)
        orig_frame.grid_columnconfigure(0, weight=1)

        self.original_text_display = tk.Text(orig_frame, height=3, wrap=tk.WORD, state=tk.DISABLED, relief=tk.SOLID,
                                             borderwidth=1, font=self.app_font)
        self.original_text_display.grid(row=0, column=0, sticky="nsew")

        orig_scrollbar = ttk.Scrollbar(orig_frame, orient="vertical", command=self.original_text_display.yview)
        orig_scrollbar.grid(row=0, column=1, sticky="ns")
        self.original_text_display.config(yscrollcommand=orig_scrollbar.set)

        ttk.Label(top_section_frame, text=_("Translation (Ctrl+Shift+V to paste):")).pack(anchor=tk.W, padx=5,
                                                                                          pady=(5, 2))
        trans_frame = ttk.Frame(top_section_frame)
        trans_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        trans_frame.grid_rowconfigure(0, weight=1)
        trans_frame.grid_columnconfigure(0, weight=1)

        self.translation_edit_text = tk.Text(trans_frame, height=5, wrap=tk.WORD, relief=tk.SOLID, borderwidth=1,
                                             undo=True, font=self.app_font)
        self.translation_edit_text.grid(row=0, column=0, sticky="nsew")
        self.translation_edit_text.bind("<FocusOut>", self.apply_translation_focus_out)
        self.translation_edit_text.bind("<KeyRelease>", self.schedule_placeholder_validation)

        trans_scrollbar = ttk.Scrollbar(trans_frame, orient="vertical", command=self.translation_edit_text.yview)
        trans_scrollbar.grid(row=0, column=1, sticky="ns")
        self.translation_edit_text.config(yscrollcommand=trans_scrollbar.set)

        trans_actions_frame = ttk.Frame(top_section_frame)
        trans_actions_frame.pack(fill=tk.X, padx=5, pady=(0, 5))
        self.apply_btn = ttk.Button(trans_actions_frame, text=_("Apply Translation"),
                                    command=self.apply_translation_from_button,
                                    state=tk.DISABLED, style="Toolbar.TButton")
        self.apply_btn.pack(side=tk.LEFT, padx=(0, 10))
        self.ai_translate_current_btn = ttk.Button(trans_actions_frame, text=_("AI Translate Selected"),
                                                   command=self.ai_translate_selected_from_button, state=tk.DISABLED,
                                                   style="Toolbar.TButton")
        self.ai_translate_current_btn.pack(side=tk.RIGHT, padx=5)

        ttk.Label(top_section_frame, text=_("Comment:")).pack(anchor=tk.W, padx=5, pady=(5, 2))
        comment_frame = ttk.Frame(top_section_frame)
        comment_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        comment_frame.grid_rowconfigure(0, weight=1)
        comment_frame.grid_columnconfigure(0, weight=1)

        self.comment_edit_text = tk.Text(comment_frame, height=3, wrap=tk.WORD, relief=tk.SOLID, borderwidth=1,
                                         undo=True, font=self.app_font)
        self.comment_edit_text.grid(row=0, column=0, sticky="nsew")
        self.comment_edit_text.bind("<FocusOut>", self.apply_comment_focus_out)

        comment_scrollbar = ttk.Scrollbar(comment_frame, orient="vertical", command=self.comment_edit_text.yview)
        comment_scrollbar.grid(row=0, column=1, sticky="ns")
        self.comment_edit_text.config(yscrollcommand=comment_scrollbar.set)

        comment_actions_frame = ttk.Frame(top_section_frame)
        comment_actions_frame.pack(fill=tk.X, padx=5, pady=(0, 5))
        self.apply_comment_btn = ttk.Button(comment_actions_frame, text=_("Apply Comment"),
                                            command=self.apply_comment_from_button, state=tk.DISABLED,
                                            style="Toolbar.TButton")
        self.apply_comment_btn.pack(side=tk.LEFT)

        status_frame = ttk.Frame(top_section_frame)
        status_frame.pack(fill=tk.X, padx=5, pady=5)
        self.ignore_var = tk.BooleanVar()
        self.toggle_ignore_btn = ttk.Checkbutton(status_frame, text=_("Ignore this string"), variable=self.ignore_var,
                                                 command=self.toggle_ignore_selected_checkbox, state=tk.DISABLED)
        self.toggle_ignore_btn.pack(side=tk.LEFT, padx=5)
        self.reviewed_var = tk.BooleanVar()
        self.toggle_reviewed_btn = ttk.Checkbutton(status_frame, text=_("Reviewed"), variable=self.reviewed_var,
                                                   command=self.toggle_reviewed_selected_checkbox, state=tk.DISABLED)
        self.toggle_reviewed_btn.pack(side=tk.LEFT, padx=15)

        context_section_frame = ttk.Frame(details_paned_window, padding=5)
        details_paned_window.add(context_section_frame, weight=2)
        context_section_frame.columnconfigure(0, weight=1)
        context_section_frame.rowconfigure(1, weight=1)
        ttk.Label(context_section_frame, text=_("Context Preview:")).pack(anchor=tk.W, padx=5, pady=(0, 2))

        ctx_frame = ttk.Frame(context_section_frame)
        ctx_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        ctx_frame.grid_rowconfigure(0, weight=1)
        ctx_frame.grid_columnconfigure(0, weight=1)

        self.context_text_display = tk.Text(ctx_frame, height=6, wrap=tk.WORD, state=tk.DISABLED, relief=tk.SOLID,
                                            borderwidth=1, font=("Consolas", 9))
        self.context_text_display.grid(row=0, column=0, sticky="nsew")
        self.context_text_display.tag_config("highlight", background="yellow", foreground="black")

        ctx_scrollbar = ttk.Scrollbar(ctx_frame, orient="vertical", command=self.context_text_display.yview)
        ctx_scrollbar.grid(row=0, column=1, sticky="ns")
        self.context_text_display.config(yscrollcommand=ctx_scrollbar.set)

        tm_section_frame = ttk.Frame(details_paned_window, padding=5)
        details_paned_window.add(tm_section_frame, weight=1)
        tm_section_frame.columnconfigure(0, weight=1)
        tm_section_frame.rowconfigure(1, weight=1)
        ttk.Label(tm_section_frame, text=_("Translation Memory Matches:")).pack(anchor=tk.W, pady=(0, 2), padx=5)
        self.tm_suggestions_listbox = tk.Listbox(tm_section_frame, height=4, relief=tk.SOLID, borderwidth=1,
                                                 font=self.app_font)
        self.tm_suggestions_listbox.pack(fill=tk.BOTH, expand=True, pady=(0, 5), padx=5)
        self.tm_suggestions_listbox.bind("<Double-1>", self.apply_tm_suggestion_from_listbox)
        tm_actions_frame = ttk.Frame(tm_section_frame)
        tm_actions_frame.pack(fill=tk.X, pady=(0, 0), padx=5)
        self.update_selected_tm_btn = ttk.Button(tm_actions_frame, text=_("Update TM for Selected"),
                                                 command=self.update_tm_for_selected_string, state=tk.DISABLED,
                                                 style="Toolbar.TButton")
        self.update_selected_tm_btn.pack(side=tk.LEFT, padx=(0, 5))
        self.clear_selected_tm_btn = ttk.Button(tm_actions_frame, text=_("Clear TM for Selected"),
                                                command=self.clear_tm_for_selected_string, state=tk.DISABLED,
                                                style="Toolbar.TButton")
        self.clear_selected_tm_btn.pack(side=tk.LEFT, padx=5)

        self.original_text_display.tag_configure('placeholder', foreground='orange red')
        self.original_text_display.tag_configure('placeholder_missing', background='#FFDDDD', foreground='red')
        self.translation_edit_text.tag_configure('placeholder', foreground='orange red')
        self.translation_edit_text.tag_configure('placeholder_extra', background='#FFDDDD', foreground='red')



    def _setup_statusbar(self):
        self.statusbar_frame = ttk.Frame(self.root, style="Status.TFrame")
        self.statusbar_frame.pack(side=tk.BOTTOM, fill=tk.X)

        self.statusbar_text = tk.StringVar()
        self.statusbar_text.set(_("Ready"))
        statusbar_label = ttk.Label(self.statusbar_frame, textvariable=self.statusbar_text, anchor=tk.W, padding=(5, 2))
        statusbar_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.counts_text = tk.StringVar()
        self.counts_label_widget = ttk.Label(self.statusbar_frame, textvariable=self.counts_text, anchor=tk.E,
                                             padding=(5, 2))
        self.counts_label_widget.pack(side=tk.RIGHT, padx=10)

        self.progress_bar = ttk.Progressbar(self.statusbar_frame, orient=tk.HORIZONTAL, length=150, mode='determinate')

        self.update_counts_display()

        extra_info = []
        if not TkinterDnD: extra_info.append(_("Hint: tkinterdnd2 not found, drag & drop is disabled."))
        if not requests: extra_info.append(_("Hint: requests not found, AI translation is disabled."))
        if extra_info: self.update_statusbar(
            self.statusbar_text.get() + " | " + _("Hint: ") + ", ".join(extra_info) + ".")

    def update_statusbar(self, text, persistent=False):
        self.statusbar_text.set(text)
        self.root.update_idletasks()
        if not persistent:
            self.root.after(5000, lambda: self.clear_statusbar_if_unchanged(text))

    def clear_statusbar_if_unchanged(self, original_text):
        if self.statusbar_text.get() == original_text:
            self.statusbar_text.set(_("Ready"))

    def update_counts_display(self):
        if not hasattr(self, 'translatable_objects'):
            self.counts_text.set(_("Displayed: 0/0 | Translated: 0 | Untranslated: 0 | Ignored: 0"))
            return

        displayed_count = len(self.displayed_string_ids)
        total_count = len(self.translatable_objects)

        translated_visible = 0
        untranslated_visible = 0
        ignored_visible = 0

        for ts_id in self.displayed_string_ids:
            ts_obj = self._find_ts_obj_by_id(ts_id)
            if ts_obj:
                if ts_obj.is_ignored:
                    ignored_visible += 1
                elif ts_obj.translation.strip():
                    translated_visible += 1
                else:
                    untranslated_visible += 1

        self.counts_text.set(
            _("Displayed: {displayed_count}/{total_count} | Translated: {translated_visible} | Untranslated: {untranslated_visible} | Ignored: {ignored_visible}").format(
                displayed_count=displayed_count,
                total_count=total_count,
                translated_visible=translated_visible,
                untranslated_visible=untranslated_visible,
                ignored_visible=ignored_visible
            )
        )

    def update_title(self):
        base_title = f"Overwatch Localizer - v{APP_VERSION}"
        file_name_part = ""
        if self.current_project_file_path:
            file_name_part = os.path.basename(self.current_project_file_path)
        elif self.current_code_file_path:
            file_name_part = os.path.basename(self.current_code_file_path)

        modified_indicator = "*" if self.current_project_modified else ""

        if file_name_part:
            self.root.title(_(f"{base_title} - {file_name_part}{modified_indicator}"))
        else:
            self.root.title(base_title)

    def update_ui_state_after_file_load(self, file_or_project_loaded=False):
        has_content = bool(self.translatable_objects) and file_or_project_loaded
        state = tk.NORMAL if has_content else tk.DISABLED

        self.file_menu.entryconfig(_("Save"), state=state)
        self.file_menu.entryconfig(_("Save As..."), state=state)

        self.file_menu.entryconfig(_("Compare/Import New Version..."), state=state)

        if self.is_po_mode:
            self.file_menu.entryconfig(_("Save Translation to New Code File"), state=tk.DISABLED)
        else:
            can_save_to_code = self.original_raw_code_content and has_content
            self.file_menu.entryconfig(_("Save Translation to New Code File"),
                                       state=tk.NORMAL if can_save_to_code else tk.DISABLED)

        self.file_menu.entryconfig(_("Import Translations from Excel"), state=state)
        self.file_menu.entryconfig(_("Export to Excel"), state=state)
        self.file_menu.entryconfig(_("Export to JSON"), state=state)
        self.file_menu.entryconfig(_("Export to YAML"), state=state)
        self.file_menu.entryconfig(_("Export to PO File..."), state=state)

        self.edit_menu.entryconfig(_("Find/Replace..."), state=state)
        self.edit_menu.entryconfig(_("Copy Original"), state=tk.DISABLED if not has_content else (
            tk.NORMAL if self.current_selected_ts_id else tk.DISABLED))
        self.edit_menu.entryconfig(_("Paste to Translation"), state=tk.DISABLED if not has_content else (
            tk.NORMAL if self.current_selected_ts_id else tk.DISABLED))

        self.edit_menu.entryconfig(_("Undo"), state=tk.NORMAL if self.undo_history else tk.DISABLED)
        self.edit_menu.entryconfig(_("Redo"), state=tk.NORMAL if self.redo_history else tk.DISABLED)

        self.tools_menu.entryconfig(_("Apply TM to Untranslated"), state=state)
        self.tools_menu.entryconfig(_("Project-specific Instructions..."),
                                    state=tk.NORMAL if self.current_project_file_path else tk.DISABLED)
        self.tools_menu.entryconfig(_("Reload Translatable Text"),
                                    state=tk.NORMAL if self.original_raw_code_content or self.current_code_file_path else tk.DISABLED)

        self.update_ai_related_ui_state()
        self.update_title()

    def update_ai_related_ui_state(self):
        ai_available = requests is not None
        file_loaded_and_has_strings = bool(self.translatable_objects)
        item_selected = self.current_selected_ts_id is not None
        can_start_ai_ops = ai_available and file_loaded_and_has_strings and not self.is_ai_translating_batch
        try:
            self.tools_menu.entryconfig(_("AI Translate Selected"),
                                        state=tk.NORMAL if can_start_ai_ops and item_selected else tk.DISABLED)
            self.tools_menu.entryconfig(_("AI Translate All Untranslated"),
                                        state=tk.NORMAL if can_start_ai_ops else tk.DISABLED)
            self.tools_menu.entryconfig(_("Stop AI Batch Translation"),
                                        state=tk.NORMAL if self.is_ai_translating_batch else tk.DISABLED)
            self.tools_menu.entryconfig(_("AI Settings..."),
                                        state=tk.NORMAL if ai_available else tk.DISABLED)
        except tk.TclError as e:
            print(f"Error updating AI menu states: {e}")

        if hasattr(self, 'ai_translate_current_btn'):
            self.ai_translate_current_btn.config(state=tk.NORMAL if can_start_ai_ops and item_selected else tk.DISABLED)

        if hasattr(self, 'progress_bar') and hasattr(self, 'counts_label_widget'):
            if self.is_ai_translating_batch:
                if not self.progress_bar.winfo_ismapped():
                    self.progress_bar.pack(side=tk.RIGHT, padx=5, pady=2, before=self.counts_label_widget)
                self.progress_bar.config(mode='determinate')
            else:
                if self.progress_bar.winfo_ismapped():
                    self.progress_bar.pack_forget()

    def mark_project_modified(self, modified=True):
        if self.current_project_modified != modified:
            self.current_project_modified = modified
            self.update_title()

    def add_to_undo_history(self, action_type, data):
        self.undo_history.append({'type': action_type, 'data': deepcopy(data)})
        if len(self.undo_history) > MAX_UNDO_HISTORY:
            self.undo_history.pop(0)
        self.redo_history.clear()
        try:
            self.edit_menu.entryconfig(_("Undo"), state=tk.NORMAL)
            self.edit_menu.entryconfig(_("Redo"), state=tk.DISABLED)
        except tk.TclError:
            pass
        self.mark_project_modified()

    def _find_ts_obj_by_id(self, obj_id):
        for ts_obj in self.translatable_objects:
            if ts_obj.id == obj_id:
                return ts_obj
        return None

    def undo_action(self, event=None):
        focused = self.root.focus_get()
        if event and isinstance(focused, (tk.Text, scrolledtext.ScrolledText, ttk.Entry)):
            is_main_editor = False
            if hasattr(self.translation_edit_text, 'text') and focused == self.translation_edit_text.text:
                is_main_editor = True
            elif hasattr(self, 'comment_edit_text') and hasattr(self.comment_edit_text,
                                                                'text') and focused == self.comment_edit_text.text:
                is_main_editor = True

            if is_main_editor:
                try:
                    focused.edit_undo()
                    return
                except tk.TclError:
                    pass

        if not self.undo_history:
            self.update_statusbar(_("No more actions to undo"))
            return

        action_log = self.undo_history.pop()
        action_type, action_data = action_log['type'], action_log['data']
        redo_payload_data = None
        changed_ids = set()

        if action_type == 'single_change':
            obj_id = action_data['string_id']
            field = action_data['field']
            val_to_restore = action_data['old_value']

            ts_obj = self._find_ts_obj_by_id(obj_id)
            if ts_obj:
                current_val_before_undo = getattr(ts_obj,
                                                  field) if field != 'translation' else ts_obj.get_translation_for_storage_and_tm()

                if field == 'translation':
                    ts_obj.set_translation_internal(val_to_restore.replace("\\n", "\n"))
                else:
                    setattr(ts_obj, field, val_to_restore)

                redo_payload_data = {'string_id': obj_id, 'field': field,
                                     'old_value': val_to_restore,
                                     'new_value': current_val_before_undo}
                self.update_statusbar(
                    _("Undo: {field} for ID {id} -> '{value}'").format(field=field, id=str(obj_id)[:8] + "...",
                                                                       value=str(val_to_restore)[:30]))
                changed_ids.add(obj_id)
            else:
                self.update_statusbar(_("Undo error: Object ID {obj_id} not found").format(obj_id=obj_id))
                self.edit_menu.entryconfig(_("Redo"), state=tk.NORMAL if self.redo_history else tk.DISABLED)
                return

        elif action_type in ['bulk_change', 'bulk_excel_import', 'bulk_ai_translate', 'bulk_context_menu',
                             'bulk_replace_all']:
            temp_redo_changes = []
            for item_change in action_data['changes']:
                obj_id, field, val_to_restore = item_change['string_id'], item_change['field'], item_change['old_value']
                ts_obj = self._find_ts_obj_by_id(obj_id)
                if ts_obj:
                    current_val_before_undo = getattr(ts_obj,
                                                      field) if field != 'translation' else ts_obj.get_translation_for_storage_and_tm()
                    if field == 'translation':
                        ts_obj.set_translation_internal(val_to_restore.replace("\\n", "\n"))
                    else:
                        setattr(ts_obj, field, val_to_restore)
                    temp_redo_changes.append({'string_id': obj_id, 'field': field,
                                              'old_value': val_to_restore,
                                              'new_value': current_val_before_undo})
                    changed_ids.add(obj_id)
            redo_payload_data = {'changes': temp_redo_changes}
            self.update_statusbar(_("Undo: Bulk change ({count} items)").format(count=len(temp_redo_changes)))

        if redo_payload_data:
            self.redo_history.append({'type': action_type, 'data': redo_payload_data})

        self.refresh_sheet(preserve_selection=True)

        if self.current_selected_ts_id in changed_ids:
            self.force_refresh_ui_for_current_selection()

        if not self.undo_history: self.edit_menu.entryconfig(_("Undo"), state=tk.DISABLED)
        self.edit_menu.entryconfig(_("Redo"), state=tk.NORMAL if self.redo_history else tk.DISABLED)
        self.mark_project_modified()

    def redo_action(self, event=None):
        focused = self.root.focus_get()
        if event and isinstance(focused, (tk.Text, scrolledtext.ScrolledText, ttk.Entry)):
            is_main_editor = False
            if hasattr(self.translation_edit_text, 'text') and focused == self.translation_edit_text.text:
                is_main_editor = True
            elif hasattr(self, 'comment_edit_text') and hasattr(self.comment_edit_text,
                                                                'text') and focused == self.comment_edit_text.text:
                is_main_editor = True

            if is_main_editor:
                try:
                    focused.edit_redo()
                    return
                except tk.TclError:
                    pass

        if not self.redo_history:
            self.update_statusbar(_("No more actions to redo"))
            return

        action_log = self.redo_history.pop()
        action_type, action_data_to_apply = action_log['type'], action_log['data']
        undo_payload_data = None
        changed_ids = set()

        if action_type == 'single_change':
            obj_id = action_data_to_apply['string_id']
            field = action_data_to_apply['field']

            val_to_set = action_data_to_apply['new_value']

            ts_obj = self._find_ts_obj_by_id(obj_id)
            if ts_obj:
                current_val_before_redo = getattr(ts_obj,
                                                  field) if field != 'translation' else ts_obj.get_translation_for_storage_and_tm()

                if field == 'translation':
                    ts_obj.set_translation_internal(val_to_set.replace("\\n", "\n"))
                else:
                    setattr(ts_obj, field, val_to_set)

                undo_payload_data = {'string_id': obj_id, 'field': field,
                                     'old_value': current_val_before_redo,
                                     'new_value': val_to_set}
                self.update_statusbar(
                    _("Redo: {field} for ID {id} -> '{value}'").format(field=field, id=str(obj_id)[:8] + "...",
                                                                       value=str(val_to_set)[:30]))
                changed_ids.add(obj_id)
            else:
                self.update_statusbar(_("Redo error: Object ID {obj_id} not found").format(obj_id=obj_id))
                self.edit_menu.entryconfig(_("Undo"), state=tk.NORMAL if self.undo_history else tk.DISABLED)
                return

        elif action_type in ['bulk_change', 'bulk_excel_import', 'bulk_ai_translate', 'bulk_context_menu',
                             'bulk_replace_all']:
            temp_undo_changes = []
            for item_change in action_data_to_apply['changes']:
                obj_id, field, val_to_set = item_change['string_id'], item_change['field'], item_change['new_value']
                ts_obj = self._find_ts_obj_by_id(obj_id)
                if ts_obj:
                    current_val_before_redo = getattr(ts_obj,
                                                      field) if field != 'translation' else ts_obj.get_translation_for_storage_and_tm()
                    if field == 'translation':
                        ts_obj.set_translation_internal(val_to_set.replace("\\n", "\n"))
                    else:
                        setattr(ts_obj, field, val_to_set)
                    temp_undo_changes.append({'string_id': obj_id, 'field': field,
                                              'old_value': current_val_before_redo,
                                              'new_value': val_to_set})
                    changed_ids.add(obj_id)
            undo_payload_data = {'changes': temp_undo_changes}
            self.update_statusbar(_("Redo: Bulk change ({count} items)").format(count=len(temp_undo_changes)))

        if undo_payload_data:
            self.undo_history.append({'type': action_type, 'data': undo_payload_data})
            if len(self.undo_history) > MAX_UNDO_HISTORY:
                self.undo_history.pop(0)

        self.refresh_sheet(preserve_selection=True)
        if self.current_selected_ts_id in changed_ids:
            self.force_refresh_ui_for_current_selection()

        if not self.redo_history: self.edit_menu.entryconfig(_("Redo"), state=tk.DISABLED)
        self.edit_menu.entryconfig(_("Undo"), state=tk.NORMAL if self.undo_history else tk.DISABLED)
        self.mark_project_modified()

    def open_code_file_dialog(self, event=None):
        if not self.prompt_save_if_modified(): return

        filepath = filedialog.askopenfilename(
            title=_("Open Code File"),
            filetypes=(("Overwatch Workshop Files", "*.ow;*.txt"), ("All Files", "*.*")),
            initialdir=self.config.get("last_dir", os.getcwd()),
            parent=self.root
        )
        if filepath:
            self.open_code_file_path(filepath)

    def open_code_file_path(self, filepath):
        if self.is_ai_translating_batch:
            messagebox.showwarning(_("Operation Restricted"),
                                   _("AI batch translation is in progress. Please wait for it to complete or stop it before opening a new file."),
                                   parent=self.root)
            return

        try:
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                self.original_raw_code_content = f.read()
            self.current_code_file_path = filepath
            self.current_project_file_path = None
            self.current_po_metadata = None
            self.project_custom_instructions = ""
            self.add_to_recent_files(filepath)
            self.config["last_dir"] = os.path.dirname(filepath)
            self.save_config()

            self.update_statusbar(_("Extracting strings..."), persistent=True)
            self.root.update_idletasks()
            extraction_patterns = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)
            self.translatable_objects = extract_translatable_strings(self.original_raw_code_content,
                                                                     extraction_patterns)
            self.apply_tm_to_all_current_strings(silent=True, only_if_empty=True)
            self._run_and_refresh_with_validation()


            self.undo_history.clear()
            self.redo_history.clear()
            self.current_selected_ts_id = None
            self.mark_project_modified(False)
            self.is_po_mode = False
            self.refresh_sheet()
            self.update_statusbar(
                _("Loaded {count} translatable strings from {filename}").format(count=len(self.translatable_objects),
                                                                                filename=os.path.basename(filepath)),
                persistent=True)
            self.update_ui_state_after_file_load(file_or_project_loaded=True)

        except Exception as e:
            messagebox.showerror(_("Error"), _("Could not open or parse code file '{filename}': {error}").format(
                filename=os.path.basename(filepath), error=e),
                                 parent=self.root)
            self._reset_app_state()
            self.update_statusbar(_("Code file loading failed"), persistent=True)
        self.update_counts_display()

    def open_project_dialog(self, event=None):
        if not self.prompt_save_if_modified(): return

        filepath = filedialog.askopenfilename(
            title=_("Open Project File"),
            filetypes=(("Overwatch Project Files", f"*{PROJECT_FILE_EXTENSION}"), ("All Files", "*.*")),
            initialdir=self.config.get("last_dir", os.getcwd()),
            parent=self.root
        )
        if filepath:
            self.open_project_file(filepath)

    def open_project_file(self, project_filepath):
        if self.is_ai_translating_batch:
            messagebox.showwarning(_("Operation Restricted"), _("AI batch translation is in progress."),
                                   parent=self.root)
            return

        try:
            loaded_data = load_project(project_filepath)
            project_data = loaded_data["project_data"]

            self.current_code_file_path = loaded_data["original_code_file_path"]
            self.original_raw_code_content = loaded_data["original_raw_code_content"]
            self.translatable_objects = loaded_data["translatable_objects"]
            self.current_po_metadata = project_data.get("po_metadata")
            self.project_custom_instructions = project_data.get("project_custom_instructions", "")

            tm_path_from_project = project_data.get("current_tm_file_path")
            if tm_path_from_project and os.path.exists(tm_path_from_project):
                self.load_tm_from_excel(tm_path_from_project, silent=True)
            elif tm_path_from_project:
                messagebox.showwarning(_("Project Warning"),
                                       _("Project's associated TM file '{tm_path}' not found.").format(
                                           tm_path=tm_path_from_project),
                                       parent=self.root)

            filter_settings = project_data.get("filter_settings", {})
            self.deduplicate_strings_var.set(filter_settings.get("deduplicate", False))
            self.show_ignored_var.set(filter_settings.get("show_ignored", True))
            self.show_untranslated_var.set(filter_settings.get("show_untranslated", False))
            self.show_translated_var.set(filter_settings.get("show_translated", False))
            self.show_unreviewed_var.set(filter_settings.get("show_unreviewed", False))

            self.current_project_file_path = project_filepath
            self.add_to_recent_files(project_filepath)
            self.config["last_dir"] = os.path.dirname(project_filepath)
            self.save_config()
            self._run_and_refresh_with_validation()
            self.undo_history.clear()
            self.redo_history.clear()
            self.current_selected_ts_id = None
            self.mark_project_modified(False)

            ui_state = project_data.get("ui_state", {})
            self.search_var.set(ui_state.get("search_term", ""))
            self.is_po_mode = False
            self.refresh_sheet()

            selected_id_from_proj = ui_state.get("selected_ts_id")
            if selected_id_from_proj:
                self.select_sheet_row_by_id(selected_id_from_proj, see=True)
                self.on_sheet_select(None)

            self.update_statusbar(_("Project '{filename}' loaded.").format(filename=os.path.basename(project_filepath)),
                                  persistent=True)
            self.update_ui_state_after_file_load(file_or_project_loaded=True)

        except Exception as e:
            messagebox.showerror(_("Open Project Error"), _("Could not load project file '{filename}': {error}").format(
                filename=os.path.basename(project_filepath), error=e),
                                 parent=self.root)
            self._reset_app_state()
            self.update_statusbar(_("Project file loading failed."), persistent=True)
        self.update_counts_display()

    def _reset_app_state(self):
        self.current_code_file_path = None
        self.current_project_file_path = None
        self.current_po_metadata = None
        self.original_raw_code_content = ""
        self.project_custom_instructions = ""
        self.translatable_objects = []
        self.redo_history.clear()
        self.current_selected_ts_id = None
        self.mark_project_modified(False)
        self.refresh_sheet()
        self.clear_details_pane()
        self.update_ui_state_after_file_load(file_or_project_loaded=False)
        self.update_title()

    def handle_pot_file_drop(self, pot_filepath):
        if not self.translatable_objects:
            if self.prompt_save_if_modified():
                self.import_po_file_dialog_with_path(pot_filepath)
            return

        from dialogs.pot_drop_dialog import POTDropDialog
        dialog = POTDropDialog(self.root, title=_("POT File Detected"))

        if dialog.result == "update":
            self.run_comparison_with_file(pot_filepath)
        elif dialog.result == "import":
            if self.prompt_save_if_modified():
                self.import_po_file_dialog_with_path(pot_filepath)

    def run_comparison_with_file(self, filepath):
        print(f"Running comparison with {filepath}")

    def prompt_save_if_modified(self):
        if self.current_project_modified:
            response = messagebox.askyesnocancel(_("Unsaved Changes"),
                                                 _("The current project has unsaved changes. Do you want to save?"),
                                                 parent=self.root)
            if response is True:
                return self.save_project_dialog()
            elif response is False:
                return True
            else:
                return False
        return True

    def refresh_sheet_preserve_selection(self, item_to_reselect_after=None):
        self.refresh_sheet(preserve_selection=True, item_to_reselect_after=item_to_reselect_after)

    def refresh_sheet(self, preserve_selection=True, item_to_reselect_after=None):
        old_selected_id = self.current_selected_ts_id
        if item_to_reselect_after:
            old_selected_id = item_to_reselect_after

        if self.translatable_objects:
            def sort_key(ts_obj):
                if ts_obj.warnings and not ts_obj.is_warning_ignored:
                    primary_key = 0
                elif ts_obj.is_ignored:
                    primary_key = 2
                else:
                    primary_key = 1
                secondary_key = ts_obj.line_num_in_file
                return (primary_key, secondary_key)

            self.translatable_objects.sort(key=sort_key)

        filtered_objects = []
        processed_originals_for_dedup = set()
        search_term = self.search_var.get().lower()
        is_searching = search_term and search_term != _("Quick search...").lower()

        for ts_obj in self.translatable_objects:
            if self.deduplicate_strings_var.get() and ts_obj.original_semantic in processed_originals_for_dedup:
                continue
            if not self.show_ignored_var.get() and ts_obj.is_ignored:
                continue

            has_translation = bool(ts_obj.translation.strip())
            if self.show_untranslated_var.get() and has_translation and not ts_obj.is_ignored:
                continue
            if self.show_translated_var.get() and not has_translation and not ts_obj.is_ignored:
                continue
            if self.show_unreviewed_var.get() and ts_obj.is_reviewed:
                continue

            if is_searching:
                if not (search_term in ts_obj.original_semantic.lower() or
                        search_term in ts_obj.get_translation_for_ui().lower() or
                        search_term in ts_obj.comment.lower()):
                    continue

            if self.deduplicate_strings_var.get():
                processed_originals_for_dedup.add(ts_obj.original_semantic)

            filtered_objects.append(ts_obj)

        data_for_sheet = []
        self.displayed_string_ids = []
        seq_id_counter = 1
        for ts_obj in filtered_objects:
            status_char = ""
            if ts_obj.warnings and not ts_obj.is_warning_ignored:
                status_char = "⚠️"
            elif ts_obj.is_ignored:
                status_char = "I"
                if ts_obj.was_auto_ignored:
                    status_char = "A"
            elif ts_obj.translation.strip():
                status_char = "T"
            else:
                status_char = "U"

            row_data = [
                seq_id_counter, status_char,
                ts_obj.original_semantic.replace("\n", "↵"),
                ts_obj.get_translation_for_ui().replace("\n", "↵"),
                ts_obj.comment.replace("\n", "↵")[:50],
                "✔" if ts_obj.is_reviewed else "",
                ts_obj.line_num_in_file
            ]
            data_for_sheet.append(row_data)
            self.displayed_string_ids.append(ts_obj.id)
            seq_id_counter += 1

        if self.is_po_mode:  # Check is_po_mode instead of metadata
            self.new_entry_id = "##NEW_ENTRY##"
            new_row_data = ["NEW", "", "", "", "", "", ""]
            data_for_sheet.append(new_row_data)
            self.displayed_string_ids.append(self.new_entry_id)

        self.sheet.set_sheet_data(data=data_for_sheet, redraw=False)

        self.sheet.column_width(column=0, width=40)
        self.sheet.column_width(column=1, width=30)
        self.sheet.column_width(column=2, width=300)
        self.sheet.column_width(column=3, width=300)
        self.sheet.column_width(column=4, width=150)
        self.sheet.column_width(column=5, width=30)
        self.sheet.column_width(column=6, width=70)
        self.sheet.align_columns(columns=[0], align="e")
        self.sheet.align_columns(columns=[1, 5, 6], align="center")

        self._apply_row_highlighting()
        self.sheet.redraw()

        if preserve_selection and old_selected_id:
            self.select_sheet_row_by_id(old_selected_id, see=True)

        self.update_counts_display()
        if not self.current_selected_ts_id:
            self.on_sheet_select(None)

    def _apply_row_highlighting(self):
        for row_idx, ts_id in enumerate(self.displayed_string_ids):
            ts_obj = self._find_ts_obj_by_id(ts_id)
            if not ts_obj: continue
            fg = "black"
            if ts_obj.warnings and not ts_obj.is_warning_ignored:
                fg = "orange red"
            elif ts_obj.is_ignored:
                fg = "#707070"
                if ts_obj.was_auto_ignored:
                    fg = "#a0a0a0"
            elif ts_obj.translation.strip():
                fg = "darkblue"
                if ts_obj.is_reviewed:
                    fg = "darkgreen"
            else:
                fg = "darkred"

            self.sheet.highlight_rows(rows=[row_idx], fg=fg, redraw=False)

    def find_string_from_toolbar(self):
        search_term = self.search_var.get()
        if search_term == _("Quick search..."):
            self.search_var.set("")

        self.refresh_sheet_preserve_selection()

        if self.displayed_string_ids:
            first_match_id = self.displayed_string_ids[0]
            self.select_sheet_row_by_id(first_match_id, see=True)
            self.on_sheet_select(None)
            self.update_statusbar(_("Filtered by '{search_term}'.").format(search_term=self.search_var.get()))
        elif self.search_var.get() and self.search_var.get() != _("Quick search..."):
            self.update_statusbar(_("No matches found for '{search_term}' under current filters.").format(
                search_term=self.search_var.get()))
        else:
            self.update_statusbar(_("Search cleared."))

        if not self.search_var.get() and hasattr(self.search_entry, 'insert'):
            self.search_entry.insert(0, _("Quick search..."))
            self.search_entry.config(foreground="grey")

    def on_search_focus_in(self, event):
        if self.search_var.get() == _("Quick search..."):
            self.search_var.set("")
            self.search_entry.config(foreground="black")

    def on_search_focus_out(self, event):
        if not self.search_var.get():
            self.search_var.set(_("Quick search..."))
            self.search_entry.config(foreground="grey")

    def refresh_ui_for_current_selection(self):
        current_id = self.current_selected_ts_id
        if current_id:
            self.current_selected_ts_id = None
            try:
                row_idx = self.displayed_string_ids.index(current_id)
                class MockSelection:
                    def __init__(self, row, col):
                        self.row = row
                        self.column = col

                self.sheet.set_currently_selected(row_idx, 2)
                self.on_sheet_select()

            except (ValueError, IndexError):
                self.clear_details_pane()

    def force_refresh_ui_for_current_selection(self):
        current_id = self.current_selected_ts_id
        if not current_id:
            self.clear_details_pane()
            return

        ts_obj = self._find_ts_obj_by_id(current_id)
        if not ts_obj:
            self.clear_details_pane()
            return

        self.original_text_display.config(state=tk.NORMAL)
        self.original_text_display.delete("1.0", tk.END)
        self.original_text_display.insert("1.0", ts_obj.original_semantic)
        self.original_text_display.config(state=tk.DISABLED)

        self.translation_edit_text.delete("1.0", tk.END)
        self.translation_edit_text.insert("1.0", ts_obj.get_translation_for_ui())
        self.translation_edit_text.edit_reset()

        self._update_placeholder_highlights()

        self.comment_edit_text.delete("1.0", tk.END)
        self.comment_edit_text.insert("1.0", ts_obj.comment)
        self.comment_edit_text.edit_reset()

        self.ignore_var.set(ts_obj.is_ignored)
        ignore_label = _("Ignore this string")
        if ts_obj.is_ignored and ts_obj.was_auto_ignored:
            ignore_label += _(" (Auto)")
        self.toggle_ignore_btn.config(text=ignore_label)

        self.reviewed_var.set(ts_obj.is_reviewed)
        self.update_tm_suggestions_for_text(ts_obj.original_semantic)

    def on_sheet_select(self, event=None):
        current_selection = self.sheet.get_currently_selected()
        if not current_selection or current_selection.row is None:
            return

        focused_row_index = current_selection.row
        if focused_row_index >= len(self.displayed_string_ids):
            return

        newly_selected_ts_id = self.displayed_string_ids[focused_row_index]

        if self.current_selected_ts_id == newly_selected_ts_id:
            return

        if self.current_selected_ts_id and self.current_selected_ts_id != newly_selected_ts_id:
            ts_obj_before_change = self._find_ts_obj_by_id(self.current_selected_ts_id)
            if ts_obj_before_change:
                current_editor_text = self.translation_edit_text.get("1.0", tk.END).rstrip('\n')
                if current_editor_text != ts_obj_before_change.get_translation_for_ui():
                    self._apply_translation_to_model(ts_obj_before_change, current_editor_text, source="manual_focus_out")

        self.current_selected_ts_id = newly_selected_ts_id

        if self.current_selected_ts_id == getattr(self, 'new_entry_id', None):
            self.clear_details_pane()
            self.original_text_display.config(state=tk.NORMAL)
            self.original_text_display.focus_set()
            self.update_ui_state_for_selection(self.new_entry_id)
            return
        else:
            self.original_text_display.config(state=tk.DISABLED)

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj:
            self.clear_details_pane()
            return

        self.original_text_display.config(state=tk.NORMAL)
        self.original_text_display.delete("1.0", tk.END)
        self.original_text_display.insert("1.0", ts_obj.original_semantic)
        self.original_text_display.config(state=tk.DISABLED)

        self.translation_edit_text.delete("1.0", tk.END)
        self.translation_edit_text.insert("1.0", ts_obj.get_translation_for_ui())
        self.translation_edit_text.edit_reset()

        self._update_placeholder_highlights()

        self.comment_edit_text.delete("1.0", tk.END)
        self.comment_edit_text.insert("1.0", ts_obj.comment)
        self.comment_edit_text.edit_reset()

        self.context_text_display.config(state=tk.NORMAL)
        self.context_text_display.delete("1.0", tk.END)
        self.context_text_display.tag_remove("highlight", "1.0", tk.END)
        if ts_obj.context_lines:
            for i, line_text in enumerate(ts_obj.context_lines):
                self.context_text_display.insert(tk.END, line_text + "\n")
                if i == ts_obj.current_line_in_context_idx:
                    self.context_text_display.tag_add("highlight", f"{i + 1}.0", f"{i + 1}.end")
            if ts_obj.current_line_in_context_idx >= 0:
                self.context_text_display.see(f"{ts_obj.current_line_in_context_idx + 1}.0")
        self.context_text_display.config(state=tk.DISABLED)

        self.ignore_var.set(ts_obj.is_ignored)
        ignore_label = _("Ignore this string")
        if ts_obj.is_ignored and ts_obj.was_auto_ignored:
            ignore_label += _(" (Auto)")
        self.toggle_ignore_btn.config(text=ignore_label)

        self.reviewed_var.set(ts_obj.is_reviewed)
        self.update_tm_suggestions_for_text(ts_obj.original_semantic)

        if hasattr(self, 'clear_selected_tm_btn'):
            self.clear_selected_tm_btn.config(
                state=tk.NORMAL if ts_obj.original_semantic in self.translation_memory else tk.DISABLED)

        self.update_statusbar(
            _("Selected: \"{text}...\" (Line: {line_num})").format(
                text=ts_obj.original_semantic[:30].replace(chr(10), '↵'),
                line_num=ts_obj.line_num_in_file
            ),
            persistent=True
        )
        if ts_obj.warnings and not ts_obj.is_warning_ignored:
            warning_text = "⚠️ " + " | ".join(ts_obj.warnings)
            self.update_statusbar(warning_text, persistent=True)
        else:
            self.update_statusbar(
                _("Selected: \"{text}...\" (Line: {line_num})").format(
                    text=ts_obj.original_semantic[:30].replace(chr(10), '↵'),
                    line_num=ts_obj.line_num_in_file
                ),
                persistent=True
            )

        self.update_ui_state_for_selection(self.current_selected_ts_id)

    def schedule_placeholder_validation(self, event=None):
        if self._placeholder_validation_job:
            self.root.after_cancel(self._placeholder_validation_job)
        self._placeholder_validation_job = self.root.after(150, self._update_placeholder_highlights)

    def _update_placeholder_highlights(self):
        if not self.current_selected_ts_id:
            return

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj:
            return

        original_text = ts_obj.original_semantic
        translated_text = self.translation_edit_text.get("1.0", tk.END)

        self.original_text_display.tag_configure('whitespace', background='#DDEEFF')
        self.translation_edit_text.tag_configure('whitespace', background='#DDEEFF')
        self.original_text_display.tag_configure('newline', foreground='#007ACC', font=(self.app_font, 10, 'italic'))
        self.translation_edit_text.tag_configure('newline', foreground='#007ACC', font=(self.app_font, 10, 'italic'))

        self._highlight_specific_chars(self.original_text_display, original_text)
        self._highlight_specific_chars(self.translation_edit_text, translated_text)

        original_placeholders = set(self.placeholder_regex.findall(original_text))
        translated_placeholders = set(self.placeholder_regex.findall(translated_text))
        missing_in_translation = original_placeholders - translated_placeholders
        extra_in_translation = translated_placeholders - original_placeholders

        self.original_text_display.config(state=tk.NORMAL)
        try:
            self.original_text_display.tag_remove('placeholder', '1.0', tk.END)
            self.original_text_display.tag_remove('placeholder_missing', '1.0', tk.END)
            for match in self.placeholder_regex.finditer(original_text):
                start, end = match.span()
                tag = 'placeholder_missing' if match.group(1) in missing_in_translation else 'placeholder'
                start_coord = f"1.0+{start}c"
                end_coord = f"1.0+{end}c"
                self.original_text_display.tag_add(tag, start_coord, end_coord)
        finally:
            self.original_text_display.config(state=tk.DISABLED)

        self.translation_edit_text.tag_remove('placeholder', '1.0', tk.END)
        self.translation_edit_text.tag_remove('placeholder_extra', '1.0', tk.END)
        for match in self.placeholder_regex.finditer(translated_text):
            start, end = match.span()
            tag = 'placeholder_extra' if match.group(1) in extra_in_translation else 'placeholder'
            start_coord = f"1.0+{start}c"
            end_coord = f"1.0+{end}c"
            self.translation_edit_text.tag_add(tag, start_coord, end_coord)

        self.root.update_idletasks()

    def _highlight_specific_chars(self, text_widget, text_content):
        text_widget.tag_remove('whitespace', '1.0', tk.END)
        text_widget.tag_remove('newline', '1.0', tk.END)

        if text_content.startswith(' '):
            end_offset = len(text_content) - len(text_content.lstrip(' '))
            text_widget.tag_add('whitespace', '1.0', f'1.{end_offset}')

        if text_content.endswith(' '):
            start_offset = len(text_content.rstrip(' '))
            if start_offset < len(text_content):
                text_widget.tag_add('whitespace', f'1.{start_offset}', 'end-1c')

        start = "1.0"
        while True:
            pos = text_widget.search(r'\\n', start, stopindex=tk.END, regexp=True)
            if not pos:
                break
            end = f"{pos}+2c"
            text_widget.tag_add('newline', pos, end)
            start = end

    def update_ui_state_for_selection(self, selected_id):
        state = tk.NORMAL if selected_id else tk.DISABLED

        try:
            self.edit_menu.entryconfig(_("Copy Original"), state=state)
            self.edit_menu.entryconfig(_("Paste to Translation"), state=state)
        except tk.TclError:
            pass

        if hasattr(self, 'apply_btn'): self.apply_btn.config(state=state)
        if hasattr(self, 'toggle_ignore_btn'): self.toggle_ignore_btn.config(state=state)
        if hasattr(self, 'toggle_reviewed_btn'): self.toggle_reviewed_btn.config(state=state)
        if hasattr(self, 'apply_comment_btn'): self.apply_comment_btn.config(state=state)
        if hasattr(self, 'update_selected_tm_btn'): self.update_selected_tm_btn.config(state=state)

        if not selected_id and hasattr(self, 'clear_selected_tm_btn'):
            self.clear_selected_tm_btn.config(state=tk.DISABLED)
        self.update_ai_related_ui_state()

    def clear_details_pane(self):
        self.translation_edit_text.delete(1.0, tk.END)
        self.comment_edit_text.delete(1.0, tk.END)
        self.reviewed_var.set(False)

        self.original_text_display.config(state=tk.NORMAL)
        self.original_text_display.delete("1.0", tk.END)
        self.original_text_display.config(state=tk.DISABLED)

        self.translation_edit_text.delete("1.0", tk.END)
        self.comment_edit_text.delete("1.0", tk.END)

        self.context_text_display.config(state=tk.NORMAL)
        self.context_text_display.delete("1.0", tk.END)
        self.context_text_display.config(state=tk.DISABLED)

        self.ignore_var.set(False)
        self.reviewed_var.set(False)

        self.apply_btn["state"] = tk.DISABLED
        self.apply_comment_btn["state"] = tk.DISABLED
        self.toggle_ignore_btn["state"] = tk.DISABLED
        self.toggle_reviewed_btn["state"] = tk.DISABLED
        self.tm_suggestions_listbox.delete(0, tk.END)

    def _run_and_refresh_with_validation(self):
        if not self.translatable_objects:
            return
        run_validation_on_all(self.translatable_objects)
        self.refresh_sheet_preserve_selection()
        self.refresh_ui_for_current_selection()

    def cm_set_warning_ignored_status(self, ignore_flag):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs:
            self.update_statusbar(_("No items selected to modify warning status."))
            return

        changes_for_undo = []
        for ts_obj in selected_objs:
            if ts_obj.is_warning_ignored != ignore_flag:
                old_value = ts_obj.is_warning_ignored

                changes_for_undo.append({
                    'string_id': ts_obj.id,
                    'field': 'is_warning_ignored',
                    'old_value': old_value,
                    'new_value': ignore_flag
                })

                ts_obj.is_warning_ignored = ignore_flag

        if changes_for_undo:
            self.add_to_undo_history('bulk_context_menu', {'changes': changes_for_undo})
            self.mark_project_modified()
            count = len(changes_for_undo)
            if ignore_flag:
                status_message = _("Ignored warnings for {count} item(s).").format(count=count)
            else:
                status_message = _("Un-ignored warnings for {count} item(s).").format(count=count)
            self.update_statusbar(status_message)
        else:
            self.update_statusbar(_("Selected item(s) already have the desired warning status."))

        self._run_and_refresh_with_validation()

    def _apply_translation_to_model(self, ts_obj, new_translation_from_ui, source="manual"):
        if new_translation_from_ui == ts_obj.translation:
            return False

        old_translation_for_undo = ts_obj.get_translation_for_storage_and_tm()
        ts_obj.set_translation_internal(new_translation_from_ui)
        new_translation_for_tm_storage = ts_obj.get_translation_for_storage_and_tm()

        primary_change_data = {
            'string_id': ts_obj.id,
            'field': 'translation',
            'old_value': old_translation_for_undo,
            'new_value': new_translation_for_tm_storage
        }

        if ts_obj.original_semantic not in self.translation_memory:
            if new_translation_from_ui.strip():
                self.translation_memory[ts_obj.original_semantic] = new_translation_for_tm_storage
        else:
            pass

        undo_action_type = 'single_change'
        undo_data_payload = primary_change_data

        all_changes_for_undo_list = [primary_change_data]
        for other_ts_obj in self.translatable_objects:
            if other_ts_obj.id != ts_obj.id and \
                    other_ts_obj.original_semantic == ts_obj.original_semantic and \
                    other_ts_obj.translation != new_translation_from_ui:
                old_other_translation_for_undo = other_ts_obj.get_translation_for_storage_and_tm()
                other_ts_obj.set_translation_internal(new_translation_from_ui)
                all_changes_for_undo_list.append({
                    'string_id': other_ts_obj.id,
                    'field': 'translation',
                    'old_value': old_other_translation_for_undo,
                    'new_value': new_translation_for_tm_storage
                })
        if len(all_changes_for_undo_list) > 1:
            undo_action_type = 'bulk_change'
            undo_data_payload = {'changes': all_changes_for_undo_list}

        if source not in ["ai_batch_item"]:
            self.add_to_undo_history(undo_action_type, undo_data_payload)
        else:
            return primary_change_data

        self.refresh_sheet(preserve_selection=True)
        self.update_statusbar(_("Translation applied: \"{original_semantic}...\"").format(
            original_semantic=ts_obj.original_semantic[:20].replace(chr(10), '↵')))

        if self.current_selected_ts_id == ts_obj.id:
            tm_exists_for_selected = ts_obj.original_semantic in self.translation_memory
            self.clear_selected_tm_btn.config(state=tk.NORMAL if tm_exists_for_selected else tk.DISABLED)
            self.update_tm_suggestions_for_text(ts_obj.original_semantic)

        self._run_and_refresh_with_validation()
        self.mark_project_modified()
        return True
        return True

    def apply_translation_from_button(self):
        if not self.current_selected_ts_id: return
        if self.current_selected_ts_id == self.new_entry_id:
            new_original = self.original_text_display.get("1.0", tk.END).strip()
            if not new_original:
                messagebox.showerror(_("Error"), _("Original text cannot be empty for a new entry."), parent=self.root)
                return

            if any(ts.original_semantic == new_original for ts in self.translatable_objects):
                messagebox.showerror(_("Error"), _("This original text already exists."), parent=self.root)
                return

            new_ts = TranslatableString(
                original_raw=new_original, original_semantic=new_original,
                line_num=0, char_pos_start_in_file=0, char_pos_end_in_file=0, full_code_lines=[]
            )
            new_ts.translation = self.translation_edit_text.get("1.0", tk.END).strip()
            new_ts.comment = self.comment_edit_text.get("1.0", tk.END).strip()

            self.translatable_objects.append(new_ts)
            self.mark_project_modified()
            self._run_and_refresh_with_validation()
            self.select_sheet_row_by_id(new_ts.id, see=True)
            return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return

        new_translation_ui = self.translation_edit_text.get("1.0", tk.END).rstrip('\n')
        self._apply_translation_to_model(ts_obj, new_translation_ui, source="manual_button")

    def apply_translation_focus_out(self, event=None):
        if not self.current_selected_ts_id: return
        if event and event.widget != self.translation_edit_text:
            return

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return

        new_translation_ui = self.translation_edit_text.get("1.0", tk.END).rstrip('\n')
        if new_translation_ui != ts_obj.get_translation_for_ui():
            self._apply_translation_to_model(ts_obj, new_translation_ui, source="manual_focus_out")

    def apply_comment_from_button(self):
        if not self.current_selected_ts_id: return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return
        new_comment = self.comment_edit_text.get("1.0", tk.END).rstrip('\n')
        self._apply_comment_to_model(ts_obj, new_comment)

    def apply_comment_focus_out(self, event=None):
        if not self.current_selected_ts_id: return
        if event and event.widget != self.comment_edit_text: return

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return
        new_comment = self.comment_edit_text.get("1.0", tk.END).rstrip('\n')
        if new_comment != ts_obj.comment:
            self._apply_comment_to_model(ts_obj, new_comment)

    def _apply_comment_to_model(self, ts_obj, new_comment):
        if new_comment == ts_obj.comment: return False

        old_comment = ts_obj.comment
        ts_obj.comment = new_comment

        self.add_to_undo_history('single_change', {
            'string_id': ts_obj.id, 'field': 'comment',
            'old_value': old_comment, 'new_value': new_comment
        })
        self.refresh_sheet(preserve_selection=True)
        self.update_statusbar(_("Comment updated for ID {id}...").format(id=str(ts_obj.id)[:8]))
        self.mark_project_modified()
        return True

    def toggle_ignore_selected_checkbox(self):
        if not self.current_selected_ts_id: return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return

        new_ignore_state = self.ignore_var.get()
        if new_ignore_state == ts_obj.is_ignored: return

        primary_change = {
            'string_id': ts_obj.id, 'field': 'is_ignored',
            'old_value': ts_obj.is_ignored, 'new_value': new_ignore_state
        }

        ts_obj.is_ignored = new_ignore_state
        if not new_ignore_state:
            ts_obj.was_auto_ignored = False

        all_changes_for_undo = [primary_change]
        for other_ts_obj in self.translatable_objects:
            if other_ts_obj.id != ts_obj.id and \
                    other_ts_obj.original_semantic == ts_obj.original_semantic and \
                    other_ts_obj.is_ignored != new_ignore_state:

                old_other_ignore = other_ts_obj.is_ignored
                other_ts_obj.is_ignored = new_ignore_state
                if not new_ignore_state: other_ts_obj.was_auto_ignored = False
                all_changes_for_undo.append({
                    'string_id': other_ts_obj.id, 'field': 'is_ignored',
                    'old_value': old_other_ignore, 'new_value': new_ignore_state
                })

        undo_action_type = 'bulk_change' if len(all_changes_for_undo) > 1 else 'single_change'
        undo_data_payload = {'changes': all_changes_for_undo} if undo_action_type == 'bulk_change' else primary_change

        self.add_to_undo_history(undo_action_type, undo_data_payload)

        self.refresh_sheet_and_select_neighbor(ts_obj.id)

        self.update_statusbar(_("Ignore status for ID {id} -> {status}").format(id=str(ts_obj.id)[:8] + "...", status=_(
            'Yes') if new_ignore_state else _('No')))
        self.mark_project_modified()

    def toggle_reviewed_selected_checkbox(self):
        if not self.current_selected_ts_id: return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return

        new_reviewed_state = self.reviewed_var.get()
        if new_reviewed_state == ts_obj.is_reviewed: return
        old_reviewed_state = ts_obj.is_reviewed
        old_warning_ignored_state = ts_obj.is_warning_ignored
        ts_obj.is_reviewed = new_reviewed_state
        ts_obj.is_warning_ignored = new_reviewed_state
        changes_for_undo = [
            {
                'string_id': ts_obj.id, 'field': 'is_reviewed',
                'old_value': old_reviewed_state, 'new_value': new_reviewed_state
            },
            {
                'string_id': ts_obj.id, 'field': 'is_warning_ignored',
                'old_value': old_warning_ignored_state, 'new_value': ts_obj.is_warning_ignored
            }
        ]

        self.add_to_undo_history('bulk_context_menu', {'changes': changes_for_undo})
        self._run_and_refresh_with_validation()

        self.update_statusbar(_("Review status for ID {id} -> {status}").format(id=str(ts_obj.id)[:8] + "...", status=_(
            'Yes') if new_reviewed_state else _('No')))
        self.mark_project_modified()

    def refresh_sheet_and_select_neighbor(self, removed_item_id):
        all_iids_before = self.displayed_string_ids
        neighbor_to_select = None
        if removed_item_id in all_iids_before:
            try:
                idx = all_iids_before.index(removed_item_id)
                if idx + 1 < len(all_iids_before):
                    neighbor_to_select = all_iids_before[idx + 1]
                elif idx - 1 >= 0:
                    neighbor_to_select = all_iids_before[idx - 1]
            except ValueError:
                pass

        self.refresh_sheet(preserve_selection=True, item_to_reselect_after=neighbor_to_select)

    def save_code_file_content(self, filepath_to_save):
        if not self.original_raw_code_content:
            messagebox.showerror(_("Error"),
                                 _("There is no original code file content to save.\nPlease ensure the code file associated with the project is loaded."),
                                 parent=self.root)
            return False
        try:
            save_translated_code(filepath_to_save, self.original_raw_code_content, self.translatable_objects, self)
            self.update_statusbar(
                _("Code file saved to: {filename}").format(filename=os.path.basename(filepath_to_save)),
                persistent=True)
            return True
        except Exception as e_save:
            messagebox.showerror(_("Save Error"), _("Could not save code file: {error}").format(error=e_save),
                                 parent=self.root)
            return False

    def save_code_file(self, event=None):
        if not self.current_code_file_path:
            messagebox.showerror(_("Error"), _("No original code file path."), parent=self.root)
            return

        if not self.original_raw_code_content:
            messagebox.showerror(_("Error"), _("No original code content to save."), parent=self.root)
            return

        base, ext = os.path.splitext(self.current_code_file_path)
        new_filepath = f"{base}_translated{ext}"

        if os.path.exists(new_filepath):
            if not messagebox.askyesno(_("Confirm Overwrite"),
                                       _("File '{filename}' already exists. Overwrite? A backup file (.bak) will be created.").format(
                                           filename=os.path.basename(new_filepath)),
                                       parent=self.root):
                return

        self.save_code_file_content(new_filepath)

    def save_current_file(self, event=None):
        if self.is_po_mode:
            if self.current_po_file_path:
                return self.save_po_file(self.current_po_file_path)
            else:
                return self.save_po_as_dialog()
        else:
            if self.current_project_file_path:
                return self.save_project_file(self.current_project_file_path)
            else:
                return self.save_project_as_dialog()

    def save_current_file_as(self, event=None):
        if self.is_po_mode:
            return self.save_po_as_dialog()
        else:
            return self.save_project_as_dialog()

    def save_project_dialog(self, event=None):
        if self.is_po_mode:
            if self.current_po_file_path:
                return self.save_po_file(self.current_po_file_path)
            else:
                return self.save_po_as_dialog()
        else:
            if self.current_project_file_path:
                return self.save_project_file(self.current_project_file_path)
            else:
                return self.save_project_as_dialog()

    def save_project_as_dialog(self, event=None):
        if self.is_po_mode:
            return self.save_po_as_dialog()
        if not self.translatable_objects and not self.current_code_file_path:
            messagebox.showinfo(_("Info"),
                                _("There is no content to save as a project. Please open a code file first."),
                                parent=self.root)
            return False

        initial_dir = os.path.dirname(
            self.current_project_file_path or self.current_code_file_path or self.config.get("last_dir", os.getcwd()))

        default_proj_name = "my_project"
        if self.current_project_file_path:
            default_proj_name = os.path.splitext(os.path.basename(self.current_project_file_path))[0]
        elif self.current_code_file_path:
            default_proj_name = os.path.splitext(os.path.basename(self.current_code_file_path))[0]

        initial_file = default_proj_name + PROJECT_FILE_EXTENSION

        filepath = filedialog.asksaveasfilename(
            defaultextension=PROJECT_FILE_EXTENSION,
            filetypes=(("Overwatch Project Files", f"*{PROJECT_FILE_EXTENSION}"), ("All Files", "*.*")),
            initialdir=initial_dir,
            initialfile=initial_file,
            title=_("Save Project As"),
            parent=self.root
        )
        if filepath:
            return self.save_project_file(filepath)
        return False

    def save_po_file(self, filepath):
        try:
            original_file_name = os.path.basename(self.current_code_file_path or "source_code")
            po_file_service.save_to_po(filepath, self.translatable_objects, self.current_po_metadata,
                                       original_file_name)

            self.current_po_file_path = filepath
            self.mark_project_modified(False)
            self.update_statusbar(_("PO file saved to: {filename}").format(filename=os.path.basename(filepath)),
                                  persistent=True)
            self.update_title()
            return True
        except Exception as e:
            messagebox.showerror(_("Save PO Error"), _("Failed to save PO file: {error}").format(error=e),
                                 parent=self.root)
            return False

    def save_po_as_dialog(self, event=None):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".po",
            filetypes=(("PO files", "*.po"), ("All Files", "*.*")),
            initialdir=self.config.get("last_dir", os.getcwd()),
            title=_("Save PO File As"),
            parent=self.root
        )
        if filepath:
            return self.save_po_file(filepath)
        return False

    def save_project_file(self, project_filepath):
        project_data_dict = {
            "version": APP_VERSION,
            "original_code_file_path": self.current_code_file_path or "",
            "translatable_objects_data": [ts.to_dict() for ts in self.translatable_objects],
            "project_custom_instructions": self.project_custom_instructions,
            "current_tm_file_path": self.current_tm_file or "",
            "filter_settings": {
                "deduplicate": self.deduplicate_strings_var.get(),
                "show_ignored": self.show_ignored_var.get(),
                "show_untranslated": self.show_untranslated_var.get(),
                "show_translated": self.show_translated_var.get(),
                "show_unreviewed": self.show_unreviewed_var.get(),
            },
            "ui_state": {
                "search_term": self.search_var.get() if self.search_var.get() != _("Quick search...") else "",
                "selected_ts_id": self.current_selected_ts_id or ""
            },
        }
        if self.current_po_metadata:
            project_data_dict["po_metadata"] = self.current_po_metadata

        if save_project(project_filepath, self):
            self.current_project_file_path = project_filepath
            self.add_to_recent_files(project_filepath)
            self.mark_project_modified(False)
            self.update_statusbar(_("Project saved to: {filename}").format(filename=os.path.basename(project_filepath)),
                                  persistent=True)
            self.update_title()
            self.config["last_dir"] = os.path.dirname(project_filepath)
            self.save_config()
            self.tools_menu.entryconfig(_("Project-specific Instructions..."), state=tk.NORMAL)
            return True
        return False

    def export_project_translations_to_excel(self):
        if not self.translatable_objects:
            messagebox.showinfo(_("Info"), _("No data to export."), parent=self.root)
            return

        default_filename = "project_translations.xlsx"
        if self.current_project_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_project_file_path))
            default_filename = f"{base}_translations.xlsx"
        elif self.current_code_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_code_file_path))
            default_filename = f"{base}_translations.xlsx"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=(("Excel files", "*.xlsx"),),
            initialfile=default_filename,
            title=_("Export Project Translations to Excel"),
            parent=self.root
        )
        if not filepath: return

        wb = Workbook()
        ws = wb.active
        ws.title = "Translations"
        headers = ["UUID", "Type", _("Original (Semantic)"), _("Translation"), _("Comment"), _("Reviewed"),
                   _("Ignored"), _("Source Line"),
                   _("Original (Raw)")]
        ws.append(headers)

        items_to_export = [self._find_ts_obj_by_id(ts_id) for ts_id in self.displayed_string_ids if
                           self._find_ts_obj_by_id(ts_id)]
        if not items_to_export and self.translatable_objects:
            items_to_export = self.translatable_objects

        for ts_obj in items_to_export:
            ws.append([
                ts_obj.id,
                ts_obj.string_type,
                ts_obj.original_semantic,
                ts_obj.get_translation_for_storage_and_tm(),
                ts_obj.comment,
                _("Yes") if ts_obj.is_reviewed else _("No"),
                _("Yes") if ts_obj.is_ignored else _("No"),
                ts_obj.line_num_in_file,
                ts_obj.original_raw
            ])
        try:
            wb.save(filepath)
            self.update_statusbar(
                _("Project translations exported to: {filename}").format(filename=os.path.basename(filepath)))
        except Exception as e:
            messagebox.showerror(_("Export Error"),
                                 _("Could not export project translations to Excel: {error}").format(error=e),
                                 parent=self.root)

    def import_project_translations_from_excel(self):
        if not self.translatable_objects:
            messagebox.showinfo(_("Info"),
                                _("Please load a code file or project first to match imported translations."),
                                parent=self.root)
            return

        filepath = filedialog.askopenfilename(
            filetypes=(("Excel files", "*.xlsx"),),
            title=_("Import Translations from Excel"),
            parent=self.root
        )
        if not filepath: return

        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active

            header_row_values = [cell.value for cell in ws[1]]
            if not header_row_values or not all(isinstance(h, str) for h in header_row_values if h is not None):
                messagebox.showerror(_("Import Error"), _("Excel header format is incorrect or empty."),
                                     parent=self.root)
                return

            try:
                uuid_col_idx = header_row_values.index("UUID")
                trans_col_idx = header_row_values.index(_("Translation"))
                comment_col_idx = header_row_values.index(_("Comment")) if _("Comment") in header_row_values else -1
                reviewed_col_idx = header_row_values.index(_("Reviewed")) if _("Reviewed") in header_row_values else -1
                ignored_col_idx = header_row_values.index(_("Ignored")) if _("Ignored") in header_row_values else -1
                orig_col_idx = header_row_values.index(
                    _("Original (Semantic)")) if _("Original (Semantic)") in header_row_values else -1
            except ValueError:
                messagebox.showerror(_("Import Error"),
                                     _("Excel header must contain 'UUID' and '{translation_col}' columns.\nOptional columns: '{comment_col}', '{reviewed_col}', '{ignored_col}', '{original_semantic_col}'.").format(
                                         translation_col=_("Translation"), comment_col=_("Comment"),
                                         reviewed_col=_("Reviewed"),
                                         ignored_col=_("Ignored"), original_semantic_col=_("Original (Semantic)")),
                                     parent=self.root)
                return

            imported_count = 0
            changes_for_undo = []

            for r_idx, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                try:
                    obj_id_from_excel = row_cells[uuid_col_idx]
                    if obj_id_from_excel is None: continue

                    ts_obj = self._find_ts_obj_by_id(str(obj_id_from_excel))
                    if not ts_obj:
                        continue

                    if orig_col_idx != -1 and row_cells[orig_col_idx] is not None:
                        if ts_obj.original_semantic != str(row_cells[orig_col_idx]):
                            print(
                                _("Warning: Excel row {row_num}, UUID {uuid} - Original text does not match the one in Excel. Data will still be imported.").format(
                                    row_num=r_idx + 2, uuid=obj_id_from_excel))

                    translation_from_excel_raw = str(row_cells[trans_col_idx]) if row_cells[
                                                                                      trans_col_idx] is not None else ""
                    translation_for_model = translation_from_excel_raw.replace("\\n", "\n")
                    if ts_obj.translation != translation_for_model:
                        changes_for_undo.append({'string_id': ts_obj.id, 'field': 'translation',
                                                 'old_value': ts_obj.get_translation_for_storage_and_tm(),
                                                 'new_value': translation_from_excel_raw})
                        ts_obj.set_translation_internal(translation_for_model)
                        if translation_for_model.strip():
                            self.translation_memory[ts_obj.original_semantic] = translation_from_excel_raw
                        imported_count += 1

                    if comment_col_idx != -1 and row_cells[comment_col_idx] is not None:
                        comment_from_excel = str(row_cells[comment_col_idx])
                        if ts_obj.comment != comment_from_excel:
                            changes_for_undo.append(
                                {'string_id': ts_obj.id, 'field': 'comment', 'old_value': ts_obj.comment,
                                 'new_value': comment_from_excel})
                            ts_obj.comment = comment_from_excel
                            if not imported_count: imported_count = 1

                    if reviewed_col_idx != -1 and row_cells[reviewed_col_idx] is not None:
                        reviewed_str = str(row_cells[reviewed_col_idx]).lower()
                        is_reviewed_excel = reviewed_str in [_("Yes").lower(), "true", "yes", "1"]
                        if ts_obj.is_reviewed != is_reviewed_excel:
                            changes_for_undo.append(
                                {'string_id': ts_obj.id, 'field': 'is_reviewed', 'old_value': ts_obj.is_reviewed,
                                 'new_value': is_reviewed_excel})
                            ts_obj.is_reviewed = is_reviewed_excel
                            if not imported_count: imported_count = 1

                    if ignored_col_idx != -1 and row_cells[ignored_col_idx] is not None:
                        ignored_str = str(row_cells[ignored_col_idx]).lower()
                        is_ignored_excel = ignored_str in [_("Yes").lower(), "true", "yes", "1"]
                        if ts_obj.is_ignored != is_ignored_excel:
                            changes_for_undo.append(
                                {'string_id': ts_obj.id, 'field': 'is_ignored', 'old_value': ts_obj.is_ignored,
                                 'new_value': is_ignored_excel})
                            ts_obj.is_ignored = is_ignored_excel
                            if not is_ignored_excel: ts_obj.was_auto_ignored = False
                            if not imported_count: imported_count = 1

                except Exception as cell_err:
                    print(
                        _("Error processing Excel row {row_num}: {error}. Skipping this row.").format(row_num=r_idx + 2,
                                                                                                      error=cell_err))

            if changes_for_undo:
                self.add_to_undo_history('bulk_excel_import', {'changes': changes_for_undo})
                self.mark_project_modified()

            self.refresh_sheet(preserve_selection=True)
            if self.current_selected_ts_id: self.on_sheet_select(None)

            self.update_statusbar(_("Imported/updated {field_count} fields for {item_count} items from Excel.").format(
                field_count=len(changes_for_undo), item_count=imported_count))

        except ValueError as ve:
            messagebox.showerror(_("Import Error"),
                                 _("Error processing Excel file (possibly column names issue): {error}").format(
                                     error=ve), parent=self.root)
        except Exception as e:
            messagebox.showerror(_("Import Error"),
                                 _("Could not import project translations from Excel: {error}").format(error=e),
                                 parent=self.root)

    def export_project_translations_to_json(self):
        if not self.translatable_objects:
            messagebox.showinfo(_("Info"), _("No data to export."), parent=self.root)
            return

        default_filename = "project_translations.json"
        if self.current_project_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_project_file_path))
            default_filename = f"{base}_translations.json"
        elif self.current_code_file_path:
            base, __extension = os.path.splitext(os.path.basename(self.current_code_file_path))
            default_filename = f"{base}_translations.json"
        elif self.current_selected_ts_id:
            default_filename = "po_export.json"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=(("JSON files", "*.json"), ("All Files", "*.*")),
            initialfile=default_filename,
            title=_("Export Project Translations to JSON"),
            parent=self.root
        )
        if not filepath: return

        try:
            export_service.export_to_json(filepath, self.translatable_objects, self.displayed_string_ids,
                                          app_instance=self)
            self.update_statusbar(
                _("Project translations exported to: {filename}").format(filename=os.path.basename(filepath)))
        except Exception as e:
            messagebox.showerror(_("Export Error"),
                                 _("Could not export project translations to JSON: {error}").format(error=e),
                                 parent=self.root)

    def export_project_translations_to_yaml(self):
        if not self.translatable_objects:
            messagebox.showinfo(_("Info"), _("No data to export."), parent=self.root)
            return

        default_filename = "project_translations.yaml"
        if self.current_project_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_project_file_path))
            default_filename = f"{base}_translations.yaml"
        elif self.current_code_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_code_file_path))
            default_filename = f"{base}_translations.yaml"
        elif self.current_selected_ts_id:
            default_filename = "po_export.yaml"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".yaml",
            filetypes=(("YAML files", "*.yaml;*.yml"), ("All Files", "*.*")),
            initialfile=default_filename,
            title=_("Export Project Translations to YAML"),
            parent=self.root
        )
        if not filepath: return

        try:
            export_service.export_to_yaml(filepath, self.translatable_objects, self.displayed_string_ids,
                                          app_instance=self)
            self.update_statusbar(
                _("Project translations exported to: {filename}").format(filename=os.path.basename(filepath)))
        except Exception as e:
            messagebox.showerror(_("Export Error"),
                                 _("Could not export project translations to YAML: {error}").format(error=e),
                                 parent=self.root)

    def extract_to_pot_dialog(self):
        code_filepath = filedialog.askopenfilename(
            title=_("Select Code File to Extract POT From"),
            filetypes=(("Overwatch Workshop Files", "*.ow;*.txt"), ("All Files", "*.*")),
            initialdir=self.config.get("last_dir", os.getcwd()),
            parent=self.root
        )
        if not code_filepath: return

        pot_save_filepath = filedialog.asksaveasfilename(
            title=_("Save POT Template File"),
            defaultextension=".pot",
            filetypes=(("PO Template files", "*.pot"), ("All files", "*.*")),
            initialfile=os.path.splitext(os.path.basename(code_filepath))[0] + ".pot",
            parent=self.root
        )
        if not pot_save_filepath: return

        try:
            with open(code_filepath, 'r', encoding='utf-8', errors='replace') as f:
                code_content = f.read()

            extraction_patterns = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)
            project_name = os.path.basename(code_filepath)

            pot_object = po_file_service.extract_to_pot(code_content, extraction_patterns, project_name, APP_VERSION,
                                                        os.path.basename(code_filepath))
            pot_object.save(pot_save_filepath)
            self.update_statusbar(
                _("POT template saved to: {filename}").format(filename=os.path.basename(pot_save_filepath)))
            self.config["last_dir"] = os.path.dirname(code_filepath)
            self.save_config()
        except Exception as e:
            messagebox.showerror(_("POT Extraction Error"), _("Error extracting POT file: {error}").format(error=e),
                                 parent=self.root)

    def import_po_file_dialog_with_path(self, po_filepath):
        original_code_for_context = None
        original_code_filepath_for_context = None
        if messagebox.askyesno(_("Associate Code File?"),
                               _("Do you want to associate an original code file to get context and line number information?"),
                               parent=self.root):
            code_context_filepath = filedialog.askopenfilename(
                title=_("Select Associated Code File (for context)"),
                filetypes=(("Overwatch Workshop Files", "*.ow;*.txt"), ("All Files", "*.*")),
                initialdir=os.path.dirname(po_filepath),
                parent=self.root
            )
            if code_context_filepath:
                try:
                    with open(code_context_filepath, 'r', encoding='utf-8', errors='replace') as f:
                        original_code_for_context = f.read()
                    original_code_filepath_for_context = code_context_filepath
                except Exception as e:
                    messagebox.showwarning(_("Code File Load Failed"),
                                           _("Could not load associated code file: {error}").format(error=e),
                                           parent=self.root)
        try:
            self.translatable_objects, self.current_po_metadata = po_file_service.load_from_po(
                po_filepath, original_code_for_context, original_code_filepath_for_context
            )
            self._run_and_refresh_with_validation()
            self.original_raw_code_content = original_code_for_context if original_code_for_context else ""
            self.current_code_file_path = original_code_filepath_for_context
            self.current_project_file_path = None
            self.project_custom_instructions = ""

            self.add_to_recent_files(po_filepath)
            self.config["last_dir"] = os.path.dirname(po_filepath)
            self.save_config()

            self.undo_history.clear()
            self.redo_history.clear()
            self.current_selected_ts_id = None
            self.mark_project_modified(False)
            self.is_po_mode = True
            self.current_po_file_path = po_filepath
            self.refresh_sheet()
            self.update_statusbar(
                _("Loaded {count} entries from PO file {filename}.").format(count=len(self.translatable_objects),
                                                                            filename=os.path.basename(po_filepath)),
                persistent=True)
            self.update_ui_state_after_file_load(file_or_project_loaded=True)

        except Exception as e:
            messagebox.showerror(_("PO Import Error"), _("Error importing PO file '{filename}': {error}").format(
                filename=os.path.basename(po_filepath), error=e),
                                 parent=self.root)
            self._reset_app_state()
            self.update_statusbar(_("PO file loading failed"), persistent=True)
        self.update_counts_display()

    def import_po_file_dialog(self):
        if not self.prompt_save_if_modified(): return

        po_filepath = filedialog.askopenfilename(
            title=_("Select PO File to Import"),
            filetypes=(("PO files", "*.po"), ("POT files", "*.pot"), ("All files", "*.*")),
            initialdir=self.config.get("last_dir", os.getcwd()),
            parent=self.root
        )
        if not po_filepath: return
        self.import_po_file_dialog_with_path(po_filepath)

    def export_to_po_file_dialog(self):
        if not self.translatable_objects:
            messagebox.showinfo(_("Info"), _("No data to export."), parent=self.root)
            return

        default_filename = "translations.po"
        if self.current_project_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_project_file_path))
            default_filename = f"{base}.po"
        elif self.current_code_file_path:
            base, _extension = os.path.splitext(os.path.basename(self.current_code_file_path))
            default_filename = f"{base}.po"
        elif self.recent_files and (
                self.recent_files[0].lower().endswith(".po") or self.recent_files[0].lower().endswith(".pot")):
            default_filename = os.path.basename(self.recent_files[0])

        filepath = filedialog.asksaveasfilename(
            defaultextension=".po",
            filetypes=(("PO files", "*.po"), ("All Files", "*.*")),
            initialfile=default_filename,
            title=_("Export to PO File..."),
            parent=self.root
        )
        if not filepath: return

        try:
            original_file_name_for_po = os.path.basename(
                self.current_code_file_path) if self.current_code_file_path else "source_code"
            po_file_service.save_to_po(filepath, self.translatable_objects, self.current_po_metadata,
                                       original_file_name_for_po)
            self.update_statusbar(
                _("Translations exported to PO file: {filename}").format(filename=os.path.basename(filepath)))
            self.mark_project_modified(False)
        except Exception as e:
            messagebox.showerror(_("Export Error"), _("Failed to export to PO file: {error}").format(error=e),
                                 parent=self.root)

    def show_extraction_pattern_dialog(self):
        from dialogs.extraction_pattern_dialog import ExtractionPatternManagerDialog
        dialog = ExtractionPatternManagerDialog(self.root, _("Extraction Rule Manager"), self)
        if dialog.result:
            if self.original_raw_code_content:
                if messagebox.askyesno(_("Extraction Rules Updated"),
                                       _("Extraction rules updated. Do you want to reload the translatable text of the current code using the new rules immediately?"),
                                       parent=self.root):
                    self.reload_translatable_text()

    def reload_translatable_text(self, event=None):
        if not self.original_raw_code_content and not self.current_code_file_path:
            messagebox.showinfo(_("Info"), _("No code content loaded to reload."), parent=self.root)
            return

        current_content_to_reextract = self.original_raw_code_content
        source_name = _("current in-memory code")
        if self.current_code_file_path and os.path.exists(self.current_code_file_path):
            try:
                with open(self.current_code_file_path, 'r', encoding='utf-8', errors='replace') as f:
                    current_content_to_reextract = f.read()
                source_name = _("file '{filename}'").format(filename=os.path.basename(self.current_code_file_path))
            except Exception as e:
                messagebox.showwarning(_("File Read Error"),
                                       _("Could not re-read {filepath} from disk.\nUsing in-memory version.\nError: {error}").format(
                                           filepath=self.current_code_file_path, error=e),
                                       parent=self.root)

        if not current_content_to_reextract:
            messagebox.showerror(_("Error"), _("Could not get code content for re-extraction."), parent=self.root)
            return

        old_translations_map = {ts.original_semantic: {
            'translation': ts.translation,
            'comment': ts.comment,
            'is_reviewed': ts.is_reviewed,
            'is_ignored': ts.is_ignored,
            'was_auto_ignored': ts.was_auto_ignored
        } for ts in self.translatable_objects}

        if self.current_project_modified or old_translations_map:
            if not messagebox.askyesno(_("Confirm Reload"),
                                       _("This will re-extract strings from {source} using the new rules.\n"
                                         "Existing translations will be preserved where the original text matches, but unmatched translations and statuses may be lost.\n"
                                         "This action will clear the undo history. Continue?").format(
                                           source=source_name),
                                       parent=self.root):
                return

        try:
            self.update_statusbar(_("Re-extracting strings with new rules..."), persistent=True)
            self.root.update_idletasks()
            extraction_patterns = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)
            self.original_raw_code_content = current_content_to_reextract
            self.translatable_objects = extract_translatable_strings(self.original_raw_code_content,
                                                                     extraction_patterns)
            restored_count = 0
            for ts_obj in self.translatable_objects:
                if ts_obj.original_semantic in old_translations_map:
                    saved_state = old_translations_map[ts_obj.original_semantic]
                    ts_obj.set_translation_internal(saved_state['translation'])
                    ts_obj.comment = saved_state['comment']
                    ts_obj.is_reviewed = saved_state['is_reviewed']
                    if not saved_state['was_auto_ignored'] and saved_state['is_ignored']:
                        ts_obj.is_ignored = True
                        ts_obj.was_auto_ignored = False
                    restored_count += 1

            if restored_count > 0:
                self.update_statusbar(
                    _("Attempted to restore {count} old translations/statuses.").format(count=restored_count),
                    persistent=False)

            self.apply_tm_to_all_current_strings(silent=True, only_if_empty=True)

            self.undo_history.clear()
            self.redo_history.clear()
            self.current_selected_ts_id = None
            self.mark_project_modified(True)

            self.refresh_sheet()
            self.update_statusbar(
                _("Reloaded {count} strings from {source} using new rules.").format(
                    count=len(self.translatable_objects), source=source_name),
                persistent=True)
            self.update_ui_state_after_file_load(file_or_project_loaded=True)

        except Exception as e:
            messagebox.showerror(_("Reload Error"), _("Error reloading translatable text: {error}").format(error=e),
                                 parent=self.root)
            self.update_statusbar(_("Reload failed."), persistent=True)
        self.update_counts_display()

    def _get_default_tm_excel_path(self):
        return os.path.join(os.getcwd(), TM_FILE_EXCEL)

    def _load_default_tm_excel(self):
        default_tm_path = self._get_default_tm_excel_path()
        if os.path.exists(default_tm_path):
            self.load_tm_from_excel(default_tm_path, silent=True)

    def load_tm_from_excel(self, filepath, silent=False):
        try:
            workbook = load_workbook(filepath, read_only=True)
            sheet = workbook.active
            loaded_count = 0
            new_tm_data = {}

            header = [cell.value for cell in sheet[1]]
            original_col_idx, translation_col_idx = -1, -1

            if header and len(header) >= 2:
                for i, col_name in enumerate(header):
                    if isinstance(col_name, str):
                        col_name_lower = col_name.lower()
                        if "original" in col_name_lower or _("original") in col_name_lower:
                            original_col_idx = i
                        if "translation" in col_name_lower or _("translation") in col_name_lower:
                            translation_col_idx = i

            if original_col_idx == -1 or translation_col_idx == -1:
                if not silent:
                    messagebox.showwarning(_("TM Load Warning"),
                                           _("Could not determine original/translation columns from '{filename}' header. "
                                             "Will try to use the first two columns by default (A=Original, B=Translation).").format(
                                               filename=os.path.basename(filepath)), parent=self.root)
                original_col_idx, translation_col_idx = 0, 1

            start_row = 2 if header else 1
            for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True)):
                if len(row) > max(original_col_idx, translation_col_idx):
                    original_val = row[original_col_idx]
                    translation_val = row[translation_col_idx]
                    if original_val is not None and translation_val is not None:
                        original = str(original_val)
                        translation_with_literal_slash_n = str(translation_val)

                        if original.strip():
                            new_tm_data[original] = translation_with_literal_slash_n
                            loaded_count += 1

            self.translation_memory.update(new_tm_data)

            if not silent:
                messagebox.showinfo(_("TM"),
                                    _("Loaded/merged {count} Excel TM records from '{filename}'.").format(
                                        count=loaded_count, filename=os.path.basename(filepath)),
                                    parent=self.root)
            self.current_tm_file = filepath
            self.update_statusbar(_("TM loaded from '{filename}' (Excel).").format(filename=os.path.basename(filepath)))

        except Exception as e:
            if not silent:
                messagebox.showerror(_("Error"), _("Failed to load Excel TM: {error}").format(error=e),
                                     parent=self.root)
            self.update_statusbar(_("Failed to load Excel TM: {error}").format(error=e))

    def save_tm_to_excel(self, filepath_to_save, silent=False, backup=True):
        if not self.translation_memory:
            if not silent:
                messagebox.showinfo(_("TM"), _("TM is empty, nothing to export."), parent=self.root)
            return

        if backup and self.auto_backup_tm_on_save_var.get() and os.path.exists(filepath_to_save):
            try:
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_dir = os.path.join(os.path.dirname(filepath_to_save), "tm_backups")
                os.makedirs(backup_dir, exist_ok=True)

                base_name, ext = os.path.splitext(os.path.basename(filepath_to_save))
                backup_filename = f"{base_name}_{timestamp}{ext}"
                backup_path = os.path.join(backup_dir, backup_filename)

                shutil.copy2(filepath_to_save, backup_path)
            except Exception as e_backup:
                if not silent:
                    if self.root.winfo_exists():
                        messagebox.showwarning(_("Backup Failed"),
                                               _("Could not create backup for TM: {error}").format(error=e_backup),
                                               parent=self.root)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "TranslationMemory"
        sheet['A1'] = "Original"
        sheet['B1'] = "Translation"

        row_num = 2
        for original, translation_with_literal_slash_n in self.translation_memory.items():
            sheet[f'A{row_num}'] = original
            sheet[f'B{row_num}'] = translation_with_literal_slash_n
            row_num += 1

        try:
            workbook.save(filepath_to_save)
            if not silent:
                messagebox.showinfo(_("TM"),
                                    _("TM saved to '{filename}'.").format(filename=os.path.basename(filepath_to_save)),
                                    parent=self.root)
            self.current_tm_file = filepath_to_save
            self.update_statusbar(_("TM saved to '{filename}'.").format(filename=os.path.basename(filepath_to_save)))
        except Exception as e_save:
            if not silent:
                messagebox.showerror(_("Error"), _("Failed to save TM: {error}").format(error=e_save), parent=self.root)

    def import_tm_excel_dialog(self):
        filepath = filedialog.askopenfilename(
            title=_("Import TM (Excel)"),
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            defaultextension=".xlsx",
            parent=self.root
        )
        if not filepath: return

        self.load_tm_from_excel(filepath)

        if self.translatable_objects and \
                messagebox.askyesno(_("Apply TM"),
                                    _("TM imported. Do you want to apply it to untranslated strings in the current project immediately?"),
                                    parent=self.root):
            self.apply_tm_to_all_current_strings(only_if_empty=True)

    def export_tm_excel_dialog(self):
        if not self.translation_memory:
            messagebox.showinfo(_("TM"), _("TM is empty, nothing to export."), parent=self.root)
            return

        initial_tm_filename = os.path.basename(
            self.current_tm_file if self.current_tm_file else self._get_default_tm_excel_path())
        filepath = filedialog.asksaveasfilename(
            title=_("Export Current TM (Excel)"),
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            defaultextension=".xlsx",
            initialfile=initial_tm_filename,
            parent=self.root
        )
        if not filepath: return
        self.save_tm_to_excel(filepath, backup=False)

    def clear_entire_translation_memory(self):
        if not self.translation_memory:
            messagebox.showinfo(_("Clear TM"), _("TM is already empty."), parent=self.root)
            return

        if messagebox.askyesno(_("Confirm Clear"),
                               _("Are you sure you want to clear all entries from the in-memory TM?\n"
                                 "This cannot be undone."), parent=self.root):
            self.translation_memory.clear()
            self.update_statusbar(_("In-memory TM has been cleared."))

            if self.current_selected_ts_id:
                ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
                if ts_obj:
                    self.update_tm_suggestions_for_text(ts_obj.original_semantic)
                if hasattr(self, 'clear_selected_tm_btn'):
                    self.clear_selected_tm_btn.config(state=tk.DISABLED)

    def update_tm_for_selected_string(self):
        if not self.current_selected_ts_id:
            messagebox.showinfo(_("Info"), _("Please select an item first."), parent=self.root)
            return

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj:
            messagebox.showerror(_("Error"), _("Could not find data for the selected item."), parent=self.root)
            return
        current_translation_ui = self.translation_edit_text.get("1.0", tk.END).rstrip('\n')

        if not current_translation_ui.strip():
            if messagebox.askyesno(_("Confirm Update TM"),
                                   _("The current translation is empty. Do you want to update the TM entry for:\n'{text}...' with an empty translation?").format(
                                       text=ts_obj.original_semantic[:100].replace(chr(10), '↵')),
                                   parent=self.root, icon='warning'):
                translation_for_tm_storage = ""
            else:
                self.update_statusbar(_("TM update cancelled."))
                return
        else:
            translation_for_tm_storage = current_translation_ui.replace("\n", "\\n")

        self.translation_memory[ts_obj.original_semantic] = translation_for_tm_storage
        self.update_statusbar(
            _("TM updated for original: '{text}...'").format(text=ts_obj.original_semantic[:30].replace(chr(10), '↵')))
        self.update_tm_suggestions_for_text(ts_obj.original_semantic)
        if hasattr(self, 'clear_selected_tm_btn'): self.clear_selected_tm_btn.config(
            state=tk.NORMAL)
        if hasattr(self, 'update_selected_tm_btn'): self.update_selected_tm_btn.config(state=tk.NORMAL)

        if self.auto_save_tm_var.get() and self.current_tm_file:
            self.save_tm_to_excel(self.current_tm_file, silent=True)
        elif self.auto_save_tm_var.get():
            self.save_tm_to_excel(self._get_default_tm_excel_path(), silent=True)
        self.mark_project_modified()

    def clear_tm_for_selected_string(self):
        if not self.current_selected_ts_id:
            messagebox.showinfo(_("Info"), _("Please select an item first."), parent=self.root)
            return
        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return

        if ts_obj.original_semantic in self.translation_memory:
            if messagebox.askyesno(_("Confirm Clear"),
                                   _("Are you sure you want to remove the TM entry for:\n'{text}...'?").format(
                                       text=ts_obj.original_semantic[:100].replace(chr(10), '↵')),
                                   parent=self.root):
                del self.translation_memory[ts_obj.original_semantic]
                self.update_statusbar(_("TM entry cleared for selected item."))
                self.update_tm_suggestions_for_text(ts_obj.original_semantic)
                if hasattr(self, 'clear_selected_tm_btn'): self.clear_selected_tm_btn.config(
                    state=tk.DISABLED)
                self.mark_project_modified()
        else:
            messagebox.showinfo(_("Info"), _("The selected item has no entry in the TM."), parent=self.root)

    def update_tm_suggestions_for_text(self, original_semantic_text):
        self.tm_suggestions_listbox.delete(0, tk.END)
        if not original_semantic_text: return

        if original_semantic_text in self.translation_memory:
            suggestion_from_tm = self.translation_memory[original_semantic_text]
            suggestion_for_ui = suggestion_from_tm.replace("\\n", "\n")
            self.tm_suggestions_listbox.insert(tk.END, f"(100% Exact Match): {suggestion_for_ui}")
            self.tm_suggestions_listbox.itemconfig(tk.END, {'fg': 'darkgreen'})

        original_lower = original_semantic_text.lower()
        for tm_orig, tm_trans_with_slash_n in self.translation_memory.items():
            if tm_orig.lower() == original_lower and tm_orig != original_semantic_text:
                suggestion_for_ui = tm_trans_with_slash_n.replace("\\n", "\n")
                self.tm_suggestions_listbox.insert(tk.END, f"(Case Mismatch): {suggestion_for_ui}")
                self.tm_suggestions_listbox.itemconfig(tk.END, {'fg': 'orange red'})
                break

        fuzzy_matches = []
        for tm_orig, tm_trans_with_slash_n in self.translation_memory.items():
            if tm_orig == original_semantic_text or tm_orig.lower() == original_lower:
                continue

            ratio = SequenceMatcher(None, original_semantic_text, tm_orig).ratio()

            if ratio > 0.65:
                fuzzy_matches.append((ratio, tm_orig, tm_trans_with_slash_n))

        fuzzy_matches.sort(key=lambda x: x[0], reverse=True)

        for ratio, orig_match_text, trans_match_text in fuzzy_matches[:3]:
            suggestion_for_ui = trans_match_text.replace("\\n", "\n")
            display_orig_match = orig_match_text[:40].replace("\n", "↵") + ("..." if len(orig_match_text) > 40 else "")
            self.tm_suggestions_listbox.insert(tk.END,
                                               f"({ratio * 100:.0f}% ~ {display_orig_match}): {suggestion_for_ui}")
            self.tm_suggestions_listbox.itemconfig(tk.END, {'fg': 'purple'})

    def apply_tm_suggestion_from_listbox(self, event):
        selected_indices = self.tm_suggestions_listbox.curselection()
        if not selected_indices: return

        selected_suggestion_full_ui = self.tm_suggestions_listbox.get(selected_indices[0])
        try:
            translation_text_ui = selected_suggestion_full_ui.split("): ", 1)[1].strip()
        except IndexError:
            translation_text_ui = selected_suggestion_full_ui.strip()

        self.translation_edit_text.delete('1.0', tk.END)
        self.translation_edit_text.insert('1.0', translation_text_ui)

        if self.current_selected_ts_id:
            ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
            if ts_obj:
                self._apply_translation_to_model(ts_obj, translation_text_ui, source="tm_suggestion")

        self.update_statusbar(_("TM suggestion applied."))

    def apply_tm_to_all_current_strings(self, silent=False, only_if_empty=False, confirm=False):
        if not self.translatable_objects:
            if not silent: messagebox.showinfo(_("Info"), _("No strings to apply TM to."), parent=self.root)
            return 0
        if not self.translation_memory:
            if not silent: messagebox.showinfo(_("Info"), _("TM is empty."), parent=self.root)
            return 0

        if confirm and not only_if_empty:
            if not messagebox.askyesno(_("Confirm Operation"),
                                       _("This will apply TM to all matching strings, overwriting existing translations. Continue?"),
                                       parent=self.root):
                return 0

        applied_count = 0
        bulk_changes_for_undo = []
        for ts_obj in self.translatable_objects:
            if ts_obj.is_ignored: continue

            if only_if_empty and ts_obj.translation.strip() != "":
                continue

            if ts_obj.original_semantic in self.translation_memory:
                translation_from_tm_storage = self.translation_memory[ts_obj.original_semantic]
                translation_for_model_ui = translation_from_tm_storage.replace("\\n", "\n")

                if ts_obj.translation != translation_for_model_ui:
                    old_translation_for_undo = ts_obj.get_translation_for_storage_and_tm()
                    ts_obj.set_translation_internal(translation_for_model_ui)

                    bulk_changes_for_undo.append({
                        'string_id': ts_obj.id, 'field': 'translation',
                        'old_value': old_translation_for_undo,
                        'new_value': translation_from_tm_storage
                    })
                    applied_count += 1

        if applied_count > 0:
            if bulk_changes_for_undo:
                self.add_to_undo_history('bulk_change', {'changes': bulk_changes_for_undo})
                self.mark_project_modified()

            self.refresh_sheet(preserve_selection=True)
            if self.current_selected_ts_id: self.on_sheet_select(None)

            if not silent:
                messagebox.showinfo(_("TM"), _("Applied TM to {count} strings.").format(count=applied_count),
                                    parent=self.root)
            self.update_statusbar(_("Applied TM to {count} strings.").format(count=applied_count))
        elif not silent:
            messagebox.showinfo(_("TM"), _("No applicable translations found in TM (or no changes needed)."),
                                parent=self.root)

        return applied_count

    def show_advanced_search_dialog(self, event=None):
        if not self.translatable_objects:
            messagebox.showinfo(_("Info"), _("Please load a file or project first."), parent=self.root)
            return
        AdvancedSearchDialog(self.root, _("Find and Replace"), self)

    def copy_selected_original_text_menu(self, event=None):
        self.cm_copy_original()
        return "break"

    def paste_clipboard_to_selected_translation_menu(self, event=None):
        self.cm_paste_to_translation()
        return "break"

    def show_project_custom_instructions_dialog(self):
        if not self.current_project_file_path:
            messagebox.showerror(_("Error"), _("This feature is only available when a project file is open."),
                                 parent=self.root)
            return

        new_instructions = simpledialog.askstring(_("Project-specific Instructions"),
                                                  _("Enter specific translation instructions for this project (e.g., 'Translate \"Hero\" as \"Agent\"', 'Use a lively and cute style').\nThese instructions will be used during AI translation."),
                                                  initialvalue=self.project_custom_instructions,
                                                  parent=self.root)

        if new_instructions is not None and new_instructions != self.project_custom_instructions:
            self.project_custom_instructions = new_instructions
            self.mark_project_modified()
            self.update_statusbar(_("Project-specific translation settings updated."))

    def show_ai_settings_dialog(self):
        if not requests:
            messagebox.showerror(_("Feature Unavailable"),
                                 _("The 'requests' library is not installed, AI translation is unavailable.\nPlease run: pip install requests"),
                                 parent=self.root)
            return
        AISettingsDialog(self.root, _("AI Settings"), self.config, self.save_config, self.ai_translator, self)

    def _check_ai_prerequisites(self, show_error=True):
        if not requests:
            if show_error:
                messagebox.showerror(_("AI Feature Unavailable"),
                                     _("Python 'requests' library not found. Please install it (pip install requests) to use AI translation features."),
                                     parent=self.root)
            return False
        if not self.config.get("ai_api_key"):
            if show_error:
                messagebox.showerror(_("API Key Missing"),
                                     _("API Key is not set. Please configure it in 'Tools > AI Settings'."),
                                     parent=self.root)
            return False
        return True

    def apply_and_select_next_untranslated(self, event=None):
        if not self.current_selected_ts_id:
            return

        self.apply_translation_from_button()

        try:
            current_idx = self.displayed_string_ids.index(self.current_selected_ts_id)
        except (ValueError, IndexError):
            return

        next_untranslated_id = None

        for i in range(current_idx + 1, len(self.displayed_string_ids)):
            next_id = self.displayed_string_ids[i]
            ts_obj = self._find_ts_obj_by_id(next_id)
            if ts_obj and not ts_obj.translation.strip() and not ts_obj.is_ignored:
                next_untranslated_id = next_id
                break

        if not next_untranslated_id:
            for i in range(0, current_idx):
                next_id = self.displayed_string_ids[i]
                ts_obj = self._find_ts_obj_by_id(next_id)
                if ts_obj and not ts_obj.translation.strip() and not ts_obj.is_ignored:
                    next_untranslated_id = next_id
                    break

        if next_untranslated_id:
            self.select_sheet_row_by_id(next_untranslated_id, see=True)
            self.on_sheet_select(None)
            self.translation_edit_text.focus_set()
        else:
            self.update_statusbar(_("No more untranslated items."))

    def _generate_ai_context_strings(self, current_ts_id_to_exclude):
        contexts = {
            "translation_context": "",
            "original_context": ""
        }

        try:
            current_item_index = \
                [i for i, ts in enumerate(self.translatable_objects) if ts.id == current_ts_id_to_exclude][0]
        except IndexError:
            return contexts

        if self.config.get("ai_use_translation_context", False):
            trans_context_items = []
            max_neighbors = self.config.get("ai_context_neighbors", 0)
            preceding_context = []
            count = 0
            for i in range(current_item_index - 1, -1, -1):
                if max_neighbors > 0 and count >= max_neighbors: break
                ts = self.translatable_objects[i]
                if ts.translation.strip() and not ts.is_ignored:
                    orig_for_ctx = ts.original_semantic.replace("|", " ").replace("\n", " ")[:100]
                    trans_for_ctx = ts.get_translation_for_storage_and_tm().replace("|", " ").replace("\\n", " ")[:100]
                    preceding_context.append(f"{orig_for_ctx} -> {trans_for_ctx}")
                    count += 1
            succeeding_context = []
            count = 0
            for i in range(current_item_index + 1, len(self.translatable_objects)):
                if max_neighbors > 0 and count >= max_neighbors: break
                ts = self.translatable_objects[i]
                if ts.translation.strip() and not ts.is_ignored:
                    orig_for_ctx = ts.original_semantic.replace("|", " ").replace("\n", " ")[:100]
                    trans_for_ctx = ts.get_translation_for_storage_and_tm().replace("|", " ").replace("\\n", " ")[:100]
                    succeeding_context.append(f"{orig_for_ctx} -> {trans_for_ctx}")
                    count += 1
            trans_context_items = list(reversed(preceding_context)) + succeeding_context
            contexts["translation_context"] = " ||| ".join(trans_context_items)

        if self.config.get("ai_use_original_context", True):
            orig_context_items = []
            max_neighbors = self.config.get("ai_original_context_neighbors", 3)
            start_idx = max(0, current_item_index - max_neighbors)
            end_idx = min(len(self.translatable_objects), current_item_index + max_neighbors + 1)

            for i in range(start_idx, end_idx):
                if i == current_item_index: continue
                ts = self.translatable_objects[i]
                if not ts.is_ignored:
                    orig_context_items.append(ts.original_semantic.replace("|", " ").replace("\n", " "))
            contexts["original_context"] = " ||| ".join(orig_context_items)

        return contexts

    def _perform_ai_translation_threaded(self, ts_id, original_text, target_language, context_dict,
                                         custom_instructions, is_batch_item):
        try:
            placeholders = {
                '[Target Language]': target_language,
                '[Custom Translate]': custom_instructions,
                '[Untranslated Context]': context_dict.get("original_context", ""),
                '[Translated Context]': context_dict.get("translation_context", "")
            }
            prompt_structure = self.config.get("ai_prompt_structure", DEFAULT_PROMPT_STRUCTURE)
            final_prompt = generate_prompt_from_structure(prompt_structure, placeholders)

            translated_text = self.ai_translator.translate(original_text, final_prompt)

            self.root.after(0, self._handle_ai_translation_result, ts_id, translated_text, None, is_batch_item)
        except Exception as e:
            self.root.after(0, self._handle_ai_translation_result, ts_id, None, e, is_batch_item)
        finally:
            if is_batch_item and self.ai_batch_semaphore is not None:
                self.ai_batch_semaphore.release()
                self.root.after(0, self._decrement_active_threads_and_dispatch_more)

    def show_prompt_manager_dialog(self):
        from dialogs.prompt_manager_dialog import PromptManagerDialog
        PromptManagerDialog(self.root, _("AI Prompt Manager"), self)

    def _initiate_single_ai_translation(self, ts_id_to_translate, called_from_cm=False):

        if not ts_id_to_translate:
            return False

        ts_obj = self._find_ts_obj_by_id(ts_id_to_translate)
        if not ts_obj: return False
        if not called_from_cm and self.current_selected_ts_id == ts_id_to_translate:
            current_editor_text = self.translation_edit_text.get("1.0", tk.END).rstrip('\n')
            if current_editor_text != ts_obj.get_translation_for_ui():
                self._apply_translation_to_model(ts_obj, current_editor_text, source="pre_single_ai_save")

        if ts_obj.is_ignored:
            if not called_from_cm:
                messagebox.showinfo(_("Ignored"),
                                    _("The selected string is marked as ignored and will not be AI translated."),
                                    parent=self.root)
            return False

        if ts_obj.translation.strip():
            if called_from_cm and len(self._get_selected_ts_objects_from_sheet()) > 1:
                return False

            if not messagebox.askyesno(_("Overwrite Confirmation"),
                                       _("String \"{text}...\" already has a translation. Overwrite with AI translation?").format(
                                           text=ts_obj.original_semantic[:50]),
                                       parent=self.root):
                return False

        if not called_from_cm:
            self.update_statusbar(
                _("AI is translating: \"{text}...\"").format(text=ts_obj.original_semantic[:30].replace(chr(10), '↵')))

        context_dict = self._generate_ai_context_strings(ts_obj.id)
        target_language = self.config.get("ai_target_language", _("Target Language"))

        thread = threading.Thread(target=self._perform_ai_translation_threaded,
                                  args=(ts_obj.id, ts_obj.original_semantic, target_language,
                                        context_dict, self.project_custom_instructions, False),
                                  daemon=True)
        thread.start()
        return True

    def _dispatch_next_ai_batch_item(self):
        if not self.is_ai_translating_batch: return
        if self.ai_batch_next_item_index >= self.ai_batch_total_items: return

        if self.ai_batch_semaphore.acquire(blocking=False):
            if not self.is_ai_translating_batch:
                self.ai_batch_semaphore.release()
                return

            if self.ai_batch_next_item_index >= self.ai_batch_total_items:
                self.ai_batch_semaphore.release()
                return

            current_item_idx = self.ai_batch_next_item_index
            self.ai_batch_next_item_index += 1
            self.ai_batch_active_threads += 1

            ts_id = self.ai_translation_batch_ids_queue[current_item_idx]
            ts_obj = self._find_ts_obj_by_id(ts_id)

            self.update_statusbar(
                _("AI Batch: Processing {current}/{total} (Concurrency: {threads})...").format(
                    current=current_item_idx + 1, total=self.ai_batch_total_items,
                    threads=self.ai_batch_active_threads),
                persistent=True)

            if ts_obj and not ts_obj.is_ignored and not ts_obj.translation.strip():
                context_dict = self._generate_ai_context_strings(ts_obj.id)
                target_language = self.config.get("ai_target_language", _("Target Language"))

                thread = threading.Thread(target=self._perform_ai_translation_threaded,
                                          args=(ts_obj.id, ts_obj.original_semantic, target_language,
                                                context_dict, self.project_custom_instructions, True),
                                          daemon=True)
                thread.start()
            else:
                self.ai_batch_semaphore.release()
                self.ai_batch_active_threads -= 1
                self.ai_batch_completed_count += 1
                if self.is_ai_translating_batch:
                    if self.ai_batch_next_item_index < self.ai_batch_total_items:
                        self.root.after(0, self._dispatch_next_ai_batch_item)
                    elif self.ai_batch_active_threads == 0 and self.ai_batch_completed_count >= self.ai_batch_total_items:
                        self._finalize_batch_ai_translation()

    def cm_toggle_reviewed_status(self, event=None):
        if self.current_selected_ts_id:
            self.cm_set_reviewed_status(not self.reviewed_var.get())

    def cm_toggle_ignored_status(self, event=None):
        if self.current_selected_ts_id:
            self.cm_set_ignored_status(not self.ignore_var.get())

    def _decrement_active_threads_and_dispatch_more(self):
        if not self.is_ai_translating_batch:
            if self.ai_batch_active_threads > 0: self.ai_batch_active_threads -= 1
            if self.ai_batch_active_threads == 0:
                self._finalize_batch_ai_translation()
            return

        if self.ai_batch_active_threads > 0:
            self.ai_batch_active_threads -= 1

        if self.ai_batch_next_item_index < self.ai_batch_total_items:
            interval = self.config.get("ai_api_interval", 200)
            self.root.after(interval, self._dispatch_next_ai_batch_item)
        elif self.ai_batch_active_threads == 0 and self.ai_batch_completed_count >= self.ai_batch_total_items:
            self._finalize_batch_ai_translation()

    def _handle_ai_translation_result(self, ts_id, translated_text, error_object, is_batch_item):
        ts_obj = self._find_ts_obj_by_id(ts_id)

        if not ts_obj:
            if is_batch_item: self.ai_batch_completed_count += 1
            return

        if error_object:
            error_msg = _("AI translation failed for \"{text}...\": {error}").format(
                text=ts_obj.original_semantic[:20].replace(chr(10), '↵'), error=error_object)
            self.update_statusbar(error_msg)
            if not is_batch_item:
                messagebox.showerror(_("AI Translation Error"),
                                     _("AI translation failed for \"{text}...\":\n{error}").format(
                                         text=ts_obj.original_semantic[:50], error=error_object),
                                     parent=self.root)
        elif translated_text is not None and translated_text.strip():
            apply_source = "ai_batch_item" if is_batch_item else "ai_selected"
            cleaned_translation = translated_text.strip()

            undo_change_data = None
            if is_batch_item:
                old_undo_val = ts_obj.get_translation_for_storage_and_tm()
                ts_obj.set_translation_internal(cleaned_translation)
                if cleaned_translation:
                    self.translation_memory[ts_obj.original_semantic] = ts_obj.get_translation_for_storage_and_tm()
                self.ai_batch_successful_translations_for_undo.append({
                    'string_id': ts_obj.id,
                    'field': 'translation',
                    'old_value': old_undo_val,
                    'new_value': ts_obj.get_translation_for_storage_and_tm()
                })
            else:
                change_applied = self._apply_translation_to_model(ts_obj, cleaned_translation, source=apply_source)

            try:
                row_idx = self.displayed_string_ids.index(ts_obj.id)
                self.sheet.set_cell_data(row_idx, 1, "T")
                self.sheet.set_cell_data(row_idx, 3, cleaned_translation.replace("\n", "↵"))
                self.sheet.highlight_rows(rows=[row_idx], fg="darkblue", redraw=True)
            except (ValueError, IndexError):
                pass

            if self.current_selected_ts_id == ts_obj.id:
                self.translation_edit_text.delete("1.0", tk.END)
                self.translation_edit_text.insert("1.0", cleaned_translation)
                self.schedule_placeholder_validation()
                self.update_tm_suggestions_for_text(ts_obj.original_semantic)
                if hasattr(self, 'clear_selected_tm_btn'):
                    self.clear_selected_tm_btn.config(
                        state=tk.NORMAL if ts_obj.original_semantic in self.translation_memory else tk.DISABLED)

            if not is_batch_item:
                self.update_statusbar(_("AI translation successful: \"{text}...\"").format(
                    text=ts_obj.original_semantic[:20].replace(chr(10), '↵')))

        elif translated_text is not None and not translated_text.strip():
            self.update_statusbar(_("AI returned empty translation for \"{text}...\"").format(
                text=ts_obj.original_semantic[:20].replace(chr(10), '↵')))
            if not is_batch_item and self.current_selected_ts_id == ts_obj.id:
                self.translation_edit_text.delete("1.0", tk.END)
                self._apply_translation_to_model(ts_obj, "", source="ai_selected_empty")

        if is_batch_item:
            self.ai_batch_completed_count += 1
            if self.ai_batch_total_items > 0:
                progress_percent = (self.ai_batch_completed_count / self.ai_batch_total_items) * 100
                if hasattr(self, 'progress_bar'): self.progress_bar['value'] = progress_percent
            else:
                progress_percent = 0

            self.update_statusbar(
                _("AI Batch: {current}/{total} completed ({progress_percent:.0f}%).").format(
                    current=self.ai_batch_completed_count, total=self.ai_batch_total_items,
                    progress_percent=progress_percent),
                persistent=True)

            if self.ai_batch_completed_count >= self.ai_batch_total_items and self.ai_batch_active_threads == 0:
                self._finalize_batch_ai_translation()

        if not self.is_ai_translating_batch and not is_batch_item:
            self.update_ai_related_ui_state()

    def ai_translate_selected_from_menu(self, event=None):
        self.cm_ai_translate_selected()

    def ai_translate_selected_from_button(self):
        self._initiate_single_ai_translation(self.current_selected_ts_id)

    def ai_translate_all_untranslated(self):
        if not self._check_ai_prerequisites(): return
        if self.is_ai_translating_batch:
            messagebox.showwarning(_("AI Translation in Progress"), _("AI batch translation is already in progress."),
                                   parent=self.root)
            return

        self.ai_translation_batch_ids_queue = [
            ts.id for ts in self.translatable_objects
            if not ts.is_ignored and not ts.translation.strip()
        ]

        if not self.ai_translation_batch_ids_queue:
            messagebox.showinfo(_("No Translation Needed"), _("No untranslated and non-ignored strings found."),
                                parent=self.root)
            return

        self.ai_batch_total_items = len(self.ai_translation_batch_ids_queue)
        api_interval_ms = self.config.get('ai_api_interval', 200)
        max_concurrency = self.config.get('ai_max_concurrent_requests', 1)

        avg_api_time_estimate_s = 3.0
        if max_concurrency == 1:
            estimated_time_s = self.ai_batch_total_items * (avg_api_time_estimate_s + api_interval_ms / 1000.0)
            concurrency_text = _("sequential execution")
        else:
            estimated_time_s = (self.ai_batch_total_items / max_concurrency) * avg_api_time_estimate_s + \
                               (self.ai_batch_total_items / max_concurrency) * (
                                       api_interval_ms / 1000.0)
            concurrency_text = _("up to {max_concurrency} concurrent").format(max_concurrency=max_concurrency)

        if not messagebox.askyesno(_("Confirm Batch Translation"),
                                   _("This will start AI translation for {count} untranslated strings ({concurrency_info}).\n"
                                     "API call interval {api_interval_ms}ms (minimum interval between tasks when concurrent).\n"
                                     "Estimated time: ~{time_s:.1f} seconds.\n"
                                     "Continue?").format(count=self.ai_batch_total_items,
                                                         concurrency_info=concurrency_text,
                                                         api_interval_ms=api_interval_ms, time_s=estimated_time_s),
                                   parent=self.root):
            self.ai_translation_batch_ids_queue = []
            return

        self.is_ai_translating_batch = True
        self.ai_batch_completed_count = 0
        self.ai_batch_successful_translations_for_undo = []
        self.ai_batch_next_item_index = 0
        self.ai_batch_active_threads = 0
        self.ai_batch_semaphore = threading.Semaphore(max_concurrency)

        if hasattr(self, 'progress_bar'): self.progress_bar['value'] = 0
        self.update_ai_related_ui_state()
        self.update_statusbar(
            _("AI batch translation started ({concurrency_info})...").format(concurrency_info=concurrency_text),
            persistent=True)

        for _extension in range(max_concurrency):
            if self.ai_batch_next_item_index < self.ai_batch_total_items:
                self._dispatch_next_ai_batch_item()
            else:
                break

    def _finalize_batch_ai_translation(self):
        if not self.is_ai_translating_batch and self.ai_batch_active_threads > 0:
            return

        if self.ai_batch_successful_translations_for_undo:
            self.add_to_undo_history('bulk_ai_translate', {'changes': self.ai_batch_successful_translations_for_undo})
            self.mark_project_modified()
            self.check_batch_placeholder_mismatches()

        success_count = len(self.ai_batch_successful_translations_for_undo)
        processed_items = self.ai_batch_completed_count

        self.update_statusbar(
            _("AI batch translation complete. Successfully translated {success_count}/{processed_count} items (total {total_items} planned).").format(
                success_count=success_count, processed_count=processed_items, total_items=self.ai_batch_total_items),
            persistent=True)

        self.is_ai_translating_batch = False
        self.ai_translation_batch_ids_queue = []
        self.ai_batch_successful_translations_for_undo = []
        self.ai_batch_semaphore = None
        self.ai_batch_active_threads = 0
        self.ai_batch_next_item_index = 0
        self.ai_batch_total_items = 0
        self.ai_batch_completed_count = 0

        self.update_ai_related_ui_state()
        self.refresh_sheet(preserve_selection=True)

    def check_batch_placeholder_mismatches(self):
        mismatched_items = []
        for change in self.ai_batch_successful_translations_for_undo:
            ts_obj = self._find_ts_obj_by_id(change['string_id'])
            if not ts_obj: continue

            original_placeholders = set(self.placeholder_regex.findall(ts_obj.original_semantic))
            translated_placeholders = set(self.placeholder_regex.findall(ts_obj.translation))

            if original_placeholders != translated_placeholders:
                mismatched_items.append(ts_obj)

        if mismatched_items:
            msg = _(
                "After AI batch translation, {count} items were found with placeholder mismatches.\nDo you want to add the comment \"Placeholder Mismatch\" to these items in bulk?").format(
                count=len(mismatched_items))
            if messagebox.askyesno(_("Placeholder Mismatch"), msg, parent=self.root):
                bulk_comment_changes = []
                for ts_obj in mismatched_items:
                    old_comment = ts_obj.comment
                    new_comment = (old_comment + " " + _("Placeholder Mismatch")).strip()
                    if ts_obj.comment != new_comment:
                        ts_obj.comment = new_comment
                        bulk_comment_changes.append({
                            'string_id': ts_obj.id, 'field': 'comment',
                            'old_value': old_comment, 'new_value': new_comment
                        })
                if bulk_comment_changes:
                    self.add_to_undo_history('bulk_context_menu', {'changes': bulk_comment_changes})
                    self.refresh_sheet_preserve_selection()
                    self.update_statusbar(_("Added comments to {count} placeholder mismatched items.").format(
                        count=len(bulk_comment_changes)))

    def stop_batch_ai_translation(self, silent=False):
        if not self.is_ai_translating_batch:
            if not silent:
                messagebox.showinfo(_("Info"), _("No AI batch translation task is in progress."), parent=self.root)
            return

        was_translating = self.is_ai_translating_batch
        self.is_ai_translating_batch = False

        if not silent:
            messagebox.showinfo(_("AI Batch Translation"),
                                _("AI batch translation stop requested.\nDispatched tasks will continue to complete, please wait."),
                                parent=self.root)

        self.update_statusbar(_("AI batch translation stop requested. Finishing dispatched tasks..."), persistent=True)

        if was_translating and self.ai_batch_active_threads == 0:
            self._finalize_batch_ai_translation()
        else:
            self.update_ai_related_ui_state()

    def _get_selected_ts_objects_from_sheet(self):
        selected_objs = []
        selection_boxes = self.sheet.get_all_selection_boxes_with_types()
        if not selection_boxes:
            if self.current_selected_ts_id:
                ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
                if ts_obj:
                    return [ts_obj]
            return []

        added_ids = set()

        for box, box_type in selection_boxes:
            start_row, _, end_row, _ = box
            for row_idx in range(start_row, end_row):
                if row_idx < len(self.displayed_string_ids):
                    ts_id = self.displayed_string_ids[row_idx]
                    if ts_id not in added_ids:
                        ts_obj = self._find_ts_obj_by_id(ts_id)
                        if ts_obj:
                            selected_objs.append(ts_obj)
                            added_ids.add(ts_id)

        return selected_objs

    def select_sheet_row_by_id(self, ts_id, see=False):
        try:
            row_idx = self.displayed_string_ids.index(ts_id)
            self.sheet.select_row(row=row_idx)
            if see:
                self.sheet.see(row=row_idx, keep_xscroll=True)
        except (ValueError, IndexError):
            self.sheet.deselect("all")

    def cm_copy_original(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return
        text_to_copy = "\n".join([ts.original_semantic for ts in selected_objs])
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(text_to_copy)
            self.update_statusbar(_("Copied {count} original strings to clipboard.").format(count=len(selected_objs)))
        except tk.TclError:
            self.update_statusbar(_("Copy failed. Could not access clipboard."))

    def cm_copy_translation(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return
        text_to_copy = "\n".join([ts.get_translation_for_ui() for ts in selected_objs])
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(text_to_copy)
            self.update_statusbar(_("Copied {count} translations to clipboard.").format(count=len(selected_objs)))
        except tk.TclError:
            self.update_statusbar(_("Copy failed. Could not access clipboard."))

    def cm_paste_to_translation(self):
        if not self.current_selected_ts_id:
            self.update_statusbar(_("Please select an item."))
            return

        ts_obj = self._find_ts_obj_by_id(self.current_selected_ts_id)
        if not ts_obj: return

        try:
            clipboard_content = self.root.clipboard_get()
        except tk.TclError:
            self.update_statusbar(_("Paste failed from clipboard."))
            return

        if isinstance(clipboard_content, str):
            self.translation_edit_text.delete('1.0', tk.END)
            self.translation_edit_text.insert('1.0', clipboard_content)
            cleaned_content = clipboard_content.rstrip('\n')
            self._apply_translation_to_model(ts_obj, cleaned_content, source="manual_paste")
            self.update_statusbar(_("Clipboard content pasted to translation."))
        else:
            self.update_statusbar(_("Paste failed: Clipboard content is not text."))

    def cm_set_ignored_status(self, ignore_flag):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return

        bulk_changes = []
        for ts_obj in selected_objs:
            if ts_obj.is_ignored != ignore_flag:
                old_val = ts_obj.is_ignored
                ts_obj.is_ignored = ignore_flag
                if not ignore_flag: ts_obj.was_auto_ignored = False
                bulk_changes.append(
                    {'string_id': ts_obj.id, 'field': 'is_ignored', 'old_value': old_val, 'new_value': ignore_flag})

        if bulk_changes:
            self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
            self.refresh_sheet_and_select_neighbor(selected_objs[0].id)
            self.update_statusbar(_("{count} items' ignore status updated.").format(count=len(bulk_changes)))
            self.mark_project_modified()

    def cm_set_reviewed_status(self, reviewed_flag):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return

        bulk_changes = []
        for ts_obj in selected_objs:
            if ts_obj.is_reviewed != reviewed_flag:
                old_val = ts_obj.is_reviewed
                ts_obj.is_reviewed = reviewed_flag
                bulk_changes.append(
                    {'string_id': ts_obj.id, 'field': 'is_reviewed', 'old_value': old_val, 'new_value': reviewed_flag})

        if bulk_changes:
            self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
            self.refresh_sheet_and_select_neighbor(selected_objs[0].id)
            self.update_statusbar(_("{count} items' review status updated.").format(count=len(bulk_changes)))
            self.mark_project_modified()

    def cm_edit_comment(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return

        initial_comment = selected_objs[0].comment if len(selected_objs) == 1 else ""
        prompt_text = _("Enter comment for {count} selected items:").format(count=len(selected_objs)) if len(
            selected_objs) > 1 else _("Original:\n{original_semantic}...\n\nEnter comment:").format(
            original_semantic=selected_objs[0].original_semantic[:100])

        new_comment = simpledialog.askstring(_("Edit Comment..."), prompt_text,
                                             initialvalue=initial_comment, parent=self.root)

        if new_comment is not None:
            bulk_changes = []
            for ts_obj in selected_objs:
                if ts_obj.comment != new_comment:
                    old_comment = ts_obj.comment
                    ts_obj.comment = new_comment
                    bulk_changes.append({
                        'string_id': ts_obj.id, 'field': 'comment',
                        'old_value': old_comment, 'new_value': new_comment
                    })

            if bulk_changes:
                self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
                self.refresh_sheet_preserve_selection()
                if self.current_selected_ts_id in [c['string_id'] for c in bulk_changes]:
                    self.comment_edit_text.delete("1.0", tk.END)
                    self.comment_edit_text.insert("1.0", new_comment)
                self.update_statusbar(_("Updated comments for {count} items.").format(count=len(bulk_changes)))
                self.mark_project_modified()

    def cm_apply_tm_to_selected(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs: return
        if not self.translation_memory:
            messagebox.showinfo(_("Info"), _("TM is empty."), parent=self.root)
            return

        applied_count = 0
        bulk_changes = []
        for ts_obj in selected_objs:
            if ts_obj.is_ignored: continue
            if ts_obj.original_semantic in self.translation_memory:
                tm_translation_storage = self.translation_memory[ts_obj.original_semantic]
                tm_translation_ui = tm_translation_storage.replace("\\n", "\n")
                if ts_obj.translation != tm_translation_ui:
                    old_val = ts_obj.get_translation_for_storage_and_tm()
                    ts_obj.set_translation_internal(tm_translation_ui)
                    bulk_changes.append({'string_id': ts_obj.id, 'field': 'translation', 'old_value': old_val,
                                         'new_value': tm_translation_storage})
                    applied_count += 1

        if bulk_changes:
            self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
            self.refresh_sheet_preserve_selection()
            self.on_sheet_select(None)
            self.update_statusbar(_("Applied TM to {count} selected items.").format(count=applied_count))
            self.mark_project_modified()
        elif selected_objs:
            messagebox.showinfo(_("Info"), _("No matching TM entries or no changes needed for selected items."),
                                parent=self.root)

    def cm_clear_selected_translations(self):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs:
            return

        if not messagebox.askyesno(_("Confirm Clear"),
                                   _("Are you sure you want to clear the translations for the {count} selected items?").format(
                                           count=len(selected_objs)), parent=self.root):
            return

        bulk_changes = []
        cleared_ids = set()

        for ts_obj in selected_objs:
            if ts_obj.translation != "":
                old_val = ts_obj.get_translation_for_storage_and_tm()
                ts_obj.set_translation_internal("")
                bulk_changes.append({
                    'string_id': ts_obj.id,
                    'field': 'translation',
                    'old_value': old_val,
                    'new_value': ""
                })
                cleared_ids.add(ts_obj.id)

        if not bulk_changes:
            self.update_statusbar(_("Translations for selected items are already empty, no clearing needed."))
            return
        self.add_to_undo_history('bulk_context_menu', {'changes': bulk_changes})
        self.mark_project_modified()
        if self.current_selected_ts_id in cleared_ids:
            self.translation_edit_text.delete('1.0', tk.END)
            self.translation_edit_text.edit_reset()
        self.refresh_sheet_preserve_selection()
        self.update_statusbar(_("Cleared {count} translations.").format(count=len(bulk_changes)))

    def cm_ai_translate_selected(self, event=None):
        selected_objs = self._get_selected_ts_objects_from_sheet()
        if not selected_objs:
            self.update_statusbar(_("No items selected for AI translation."))
            return

        if not self._check_ai_prerequisites(): return

        if self.is_ai_translating_batch:
            messagebox.showwarning(_("AI Translation in Progress"),
                                   _("AI batch translation is in progress. Please wait for it to complete or stop it."),
                                   parent=self.root)
            return

        items_actually_translated_count = 0

        for i, ts_obj in enumerate(selected_objs):
            if i > 0 and self.current_selected_ts_id == selected_objs[i - 1].id:
                prev_ts_obj = selected_objs[i - 1]
                current_editor_text = self.translation_edit_text.get("1.0", tk.END).rstrip('\n')
                if current_editor_text != prev_ts_obj.get_translation_for_ui():
                    self._apply_translation_to_model(prev_ts_obj, current_editor_text,
                                                     source="multi_ai_intermediate_save")

            if self.current_selected_ts_id != ts_obj.id:
                self.select_sheet_row_by_id(ts_obj.id, see=True)
                self.on_sheet_select(None)

            initiated = self._initiate_single_ai_translation(ts_obj.id, called_from_cm=True)
            if initiated:
                items_actually_translated_count += 1

            if initiated and i < len(selected_objs) - 1:
                interval_ms = self.config.get("ai_api_interval", 200)
                time.sleep(max(0.2, interval_ms / 1000.0 * 1.5))
                self.root.update_idletasks()

        if items_actually_translated_count > 0:
            self.update_statusbar(
                _("Started AI translation for {count} selected items.").format(count=items_actually_translated_count))
        elif selected_objs:
            self.update_statusbar(_("No eligible selected items for AI translation."))

    def compare_with_new_version(self, event=None):
        if not self.translatable_objects:
            messagebox.showerror(_("Error"), _("Please open a project or file first."), parent=self.root)
            return

        is_po_mode = self.is_po_mode

        if is_po_mode:
            title = _("Select new POT template for comparison")
            filetypes = (("PO Template Files", "*.pot"), ("All Files", "*.*"))
            initial_dir = os.path.dirname(self.current_po_file_path) if self.current_po_file_path else self.config.get(
                "last_dir", os.getcwd())
        else:
            title = _("Select new version code file for comparison")
            filetypes = (("Overwatch Workshop Files", "*.ow;*.txt"), ("All Files", "*.*"))
            initial_dir = os.path.dirname(
                self.current_code_file_path) if self.current_code_file_path else self.config.get("last_dir",
                                                                                                 os.getcwd())

        filepath = filedialog.askopenfilename(title=title, filetypes=filetypes, initialdir=initial_dir,
                                              parent=self.root)
        if not filepath:
            return

        try:
            self.progress_bar.pack(side=tk.RIGHT, padx=5, pady=2, before=self.counts_label_widget)
            self.update_statusbar(_("Parsing new file..."), persistent=True)
            self.root.update_idletasks()

            new_strings = []
            new_code_content = None

            if is_po_mode:
                if not filepath.lower().endswith(".pot"):
                    messagebox.showerror(_("Error"), _("Please select a valid .pot template file."), parent=self.root)
                    self.progress_bar.pack_forget()
                    return
                pot_file = polib.pofile(filepath, encoding='utf-8')
                new_strings = [TranslatableString(original_raw=entry.msgid, original_semantic=entry.msgid, line_num=0,
                                                  char_pos_start_in_file=0, char_pos_end_in_file=0, full_code_lines=[])
                               for entry in pot_file if not entry.obsolete]
            else:
                with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                    new_code_content = f.read()
                extraction_patterns = self.config.get("extraction_patterns", DEFAULT_EXTRACTION_PATTERNS)
                new_strings = extract_translatable_strings(new_code_content, extraction_patterns)

            # --- RE-INSTATED UNIVERSAL COMPARISON AND DIALOG LOGIC ---
            self.update_statusbar(_("Comparing versions..."), persistent=True)

            old_strings = self.translatable_objects
            old_map = {s.original_semantic: s for s in old_strings}
            new_map = {s.original_semantic: s for s in new_strings}

            diff_results = {'added': [], 'removed': [], 'modified': [], 'unchanged': []}

            # Find unchanged and modified strings
            for new_obj in new_strings:
                if new_obj.original_semantic in old_map:
                    old_obj = old_map[new_obj.original_semantic]
                    new_obj.translation = old_obj.translation
                    new_obj.comment = old_obj.comment
                    new_obj.is_ignored = old_obj.is_ignored
                    new_obj.is_reviewed = old_obj.is_reviewed
                    diff_results['unchanged'].append({'old_obj': old_obj, 'new_obj': new_obj})
                else:
                    # Fuzzy matching for potentially modified strings
                    best_match_score = 0
                    best_match_old_s = None
                    for old_s in old_strings:
                        if old_s.original_semantic not in new_map:  # Only match with otherwise removed strings
                            score = SequenceMatcher(None, new_obj.original_semantic, old_s.original_semantic).ratio()
                            if score > best_match_score:
                                best_match_score = score
                                best_match_old_s = old_s

                    if best_match_score >= 0.85 and best_match_old_s:
                        new_obj.translation = best_match_old_s.translation
                        new_obj.comment = f"[{_('Inherited from old version')}] {best_match_old_s.comment}".strip()
                        new_obj.is_reviewed = False
                        if new_obj.translation:
                            new_obj.warnings.append("Fuzzy match, please review.")
                        diff_results['modified'].append(
                            {'old_obj': best_match_old_s, 'new_obj': new_obj, 'similarity': best_match_score})
                    else:
                        diff_results['added'].append({'new_obj': new_obj})

            # Find removed strings
            for old_obj in old_strings:
                if old_obj.original_semantic not in new_map:
                    # Check if it was part of a modification
                    was_modified = any(res['old_obj'] is old_obj for res in diff_results['modified'])
                    if not was_modified:
                        diff_results['removed'].append({'old_obj': old_obj})

            self.update_statusbar(_("Comparison complete, generating report..."), persistent=True)

            summary = (
                    _("Comparison complete. Found ")
                    + _("{added} new items, ").format(added=len(diff_results['added']))
                    + _("{removed} removed items, ").format(removed=len(diff_results['removed']))
                    + _("and {modified} modified/inherited items.").format(modified=len(diff_results['modified']))
            )
            diff_results['summary'] = summary

            from dialogs.diff_dialog import DiffDialog
            dialog = DiffDialog(self.root, _("Version Comparison Results"), diff_results)
            # --- END OF RE-INSTATED LOGIC ---

            self.progress_bar.pack_forget()

            if dialog.result:
                self.update_statusbar(_("Applying updates..."), persistent=True)

                self.translatable_objects = new_strings
                if not is_po_mode and new_code_content is not None:
                    self.original_raw_code_content = new_code_content
                    self.current_code_file_path = filepath

                self.apply_tm_to_all_current_strings(silent=True, only_if_empty=True)
                self._run_and_refresh_with_validation()
                self.mark_project_modified()
                self.update_statusbar(_("Project updated to new version."), persistent=True)
            else:
                self.update_statusbar(_("Version update cancelled."))

        except Exception as e:
            self.progress_bar.pack_forget()
            messagebox.showerror(_("Comparison Failed"), _("An error occurred: {error}").format(error=e),
                                 parent=self.root)
            self.update_statusbar(_("Version comparison failed."))