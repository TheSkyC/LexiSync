# Copyright (c) 2025, TheSkyC
# SPDX-License-Identifier: Apache-2.0

import datetime
import html
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import yaml

from lexisync.utils.localization import _


def export_to_html(filepath, translatable_objects, app_instance):
    project_name = "LexiSync Project"
    source_lang = "Unknown"
    target_lang = "Unknown"

    if app_instance.is_project_mode:
        project_name = app_instance.project_config.get("name", project_name)
        source_lang = app_instance.project_config.get("source_language", "en")
        target_lang = app_instance.current_target_language
    else:
        source_lang = app_instance.source_language
        target_lang = app_instance.current_target_language

    # 统计数据
    total = len(translatable_objects)
    translated = len([ts for ts in translatable_objects if ts.translation.strip() and not ts.is_ignored])
    ignored = len([ts for ts in translatable_objects if ts.is_ignored])
    reviewed = len([ts for ts in translatable_objects if ts.is_reviewed])
    progress = int(translated / total * 100) if total > 0 else 0

    # 生成行内容
    rows_html = ""
    for ts in translatable_objects:
        # 状态标签
        status_badge = ""
        row_class = ""
        if ts.is_ignored:
            status_badge = f'<span class="badge badge-gray">{_("Ignored")}</span>'
            row_class = "row-ignored"
        elif ts.is_reviewed:
            status_badge = f'<span class="badge badge-green">{_("Reviewed")}</span>'
        elif ts.is_fuzzy:
            status_badge = f'<span class="badge badge-orange">{_("Fuzzy")}</span>'
        elif not ts.translation.strip():
            status_badge = f'<span class="badge badge-red">{_("Untranslated")}</span>'
        else:
            status_badge = f'<span class="badge badge-blue">{_("Translated")}</span>'

        # 上下文标签
        context_html = f'<div class="context-tag">{html.escape(ts.context)}</div>' if ts.context else ""

        # 处理复数
        source_display = html.escape(ts.original_semantic).replace("\n", "<br>")
        target_display = html.escape(ts.translation).replace("\n", "<br>")

        if ts.is_plural:
            # 原文复数排版
            source_display = (
                f'<div class="plural-row"><span class="plural-tag">Singular</span><span class="plural-text">{source_display}</span></div>'
                f'<div class="plural-row"><span class="plural-tag">Plural</span><span class="plural-text">{html.escape(ts.original_plural).replace(chr(10), "<br>")}</span></div>'
            )
            # 译文复数排版
            plural_parts = []
            for idx, trans in ts.plural_translations.items():
                escaped_trans = html.escape(trans).replace("\n", "<br>")
                plural_parts.append(
                    f'<div class="plural-row"><span class="plural-tag">Form {idx}</span><span class="plural-text">{escaped_trans}</span></div>'
                )
            target_display = "".join(plural_parts)

        # 注释处理
        comment_html = ""
        if ts.comment or ts.po_comment:
            full_comment = (ts.po_comment + "\n" + ts.comment).strip()
            comment_html = f'<div class="comment-box">{html.escape(full_comment).replace(chr(10), "<br>")}</div>'

        # HTML 拼接
        rows_html += f"""
        <tr class="{row_class}">
            <td class="col-source">
                <div class="col-source-inner">
                    {context_html}
                    <div class="text-content">{source_display}</div>
                </div>
                {comment_html}
            </td>
            <td class="col-target">
                <div class="status-wrapper">{status_badge}</div>
                <div class="text-content">{target_display}</div>
            </td>
        </tr>
        """

    # HTML 模板
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    html_template = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>LexiSync Report - {project_name}</title>
    <style>
        :root {{
            /* 基础颜色 */
            --primary-color: #409EFF;
            --success-color: #67C23A;
            --warning-color: #E6A23C;
            --danger-color: #F56C6C;
            --text-main: #303133;
            --text-sec: #909399;
            --bg-body: #F5F7FA;
            --bg-card: #FFFFFF;
            --border-color: #EBEEF5;

            /* 上下文标签与注释颜色 (亮色模式) */
            --ctx-bg: #f4f4f5;
            --ctx-text: #606266;
            --ctx-border: #dcdfe6;
            --comment-bg: #fafafa;

            /* 状态徽章颜色 (亮色模式) */
            --badge-blue-bg: #D9ECFF;   --badge-blue-txt: #409EFF;
            --badge-green-bg: #E1F3D8;  --badge-green-txt: #67C23A;
            --badge-orange-bg: #FDF6EC; --badge-orange-txt: #E6A23C;
            --badge-red-bg: #FEF0F0;    --badge-red-txt: #F56C6C;
            --badge-gray-bg: #F4F4F5;   --badge-gray-txt: #909399;

            /* 复数小徽章颜色 (亮色模式) */
            --plural-bg: #EAE8FF;
            --plural-txt: #5E5CE6;
        }}

        @media (prefers-color-scheme: dark) {{
            :root {{
                /* 基础颜色 (深色模式) */
                --text-main: #E5EAF3;
                --text-sec: #A8ABB2;
                --bg-body: #141414;
                --bg-card: #1D1E1F;
                --border-color: #363637;

                /* 上下文标签与注释颜色 (深色模式) */
                --ctx-bg: #2b2b2c;
                --ctx-text: #a8abb2;
                --ctx-border: #414243;
                --comment-bg: #262727;

                /* 状态徽章颜色 (深色模式) */
                --badge-blue-bg: #1A365D;   --badge-blue-txt: #66B1FF;
                --badge-green-bg: #1B3C2A;  --badge-green-txt: #85CE61;
                --badge-orange-bg: #4A2B12; --badge-orange-txt: #E6A23C;
                --badge-red-bg: #4A1A1A;    --badge-red-txt: #F56C6C;
                --badge-gray-bg: #2B2B2C;   --badge-gray-txt: #A8ABB2;

                /* 复数小徽章颜色 (深色模式) */
                --plural-bg: #2B285E;
                --plural-txt: #8E8CE6;
            }}
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            background-color: var(--bg-body);
            color: var(--text-main);
            font-size: 14px;
            line-height: 1.5;
            margin: 0;
            padding: 15px;
        }}

        .container {{ max-width: 1200px; margin: 0 auto; }}

        header {{
            background: var(--bg-card);
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 1px 4px rgba(0,0,0,0.05);
            margin-bottom: 15px;
        }}

        h1 {{ margin: 0 0 10px 0; color: var(--primary-color); font-size: 22px; }}

        .meta-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 10px;
            font-size: 13px;
        }}

        .meta-item b {{ color: var(--text-sec); margin-right: 6px; }}

        /* Progress Bar */
        .progress-container {{ margin-top: 15px; }}
        .progress-bar-bg {{
            background: var(--border-color);
            height: 6px;
            border-radius: 3px;
            overflow: hidden;
        }}
        .progress-bar-fill {{
            background: var(--success-color);
            height: 100%;
            transition: width 0.5s ease;
        }}
        .progress-text {{
            font-size: 12px;
            color: var(--text-sec);
            margin-top: 4px;
            display: block;
            text-align: right;
        }}

        /* Table Style */
        table {{
            width: 100%;
            border-collapse: collapse;
            background: var(--bg-card);
            border-radius: 8px;
            box-shadow: 0 1px 4px rgba(0,0,0,0.05);
            overflow: hidden;
            table-layout: fixed;
        }}

        th {{
            background: var(--bg-body);
            text-align: left;
            padding: 10px 12px;
            font-size: 13px;
            font-weight: bold;
            color: var(--text-sec);
            border-bottom: 2px solid var(--border-color);
        }}

        td {{
            padding: 10px 12px;
            border-bottom: 1px solid var(--border-color);
            vertical-align: top;
        }}

        tr:hover {{ background-color: rgba(64, 158, 255, 0.03); }}
        .row-ignored {{ opacity: 0.6; background-color: var(--bg-body); }}

        /* 桌面端默认左右等宽 */
        .col-source {{ width: 50%; border-right: 1px solid var(--border-color); }}
        .col-target {{ width: 50%; }}

        /* Flex 布局保证标签和原文在同一行 */
        .col-source-inner {{
            display: flex;
            align-items: flex-start;
            gap: 8px;
        }}

        /* badge 浮动到右上角 (桌面端) */
        .status-wrapper {{
            float: right;
            margin-left: 10px;
            margin-bottom: 2px;
        }}

        .col-target::after {{
            content: "";
            display: table;
            clear: both;
        }}

        .text-content {{ font-size: 14px; word-break: break-word; white-space: pre-wrap; }}

        /* 状态徽章 */
        .badge {{
            display: inline-block;
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 11px;
            font-weight: 600;
        }}
        .badge-blue   {{ background: var(--badge-blue-bg); color: var(--badge-blue-txt); }}
        .badge-green  {{ background: var(--badge-green-bg); color: var(--badge-green-txt); }}
        .badge-orange {{ background: var(--badge-orange-bg); color: var(--badge-orange-txt); }}
        .badge-red    {{ background: var(--badge-red-bg); color: var(--badge-red-txt); }}
        .badge-gray   {{ background: var(--badge-gray-bg); color: var(--badge-gray-txt); }}

        /* 上下文标签 */
        .context-tag {{
            display: inline-block;
            flex-shrink: 0;
            white-space: nowrap;
            font-family: SFMono-Regular, Consolas, "Liberation Mono", Menlo, monospace;
            background-color: var(--ctx-bg);
            color: var(--ctx-text);
            border: 1px solid var(--ctx-border);
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 11px;
            margin-top: 2px;
        }}

        /* 复数表述样式 */
        .plural-row {{
            display: flex;
            align-items: flex-start;
            gap: 8px;
            margin-bottom: 6px;
        }}
        .plural-row:last-child {{ margin-bottom: 0; }}

        .plural-tag {{
            flex-shrink: 0;
            background-color: var(--plural-bg);
            color: var(--plural-txt);
            padding: 1px 6px;
            border-radius: 4px;
            font-size: 11px;
            font-weight: bold;
            margin-top: 2px;
        }}

        .plural-text {{
            word-break: break-word;
        }}

        /* 注释框 */
        .comment-box {{
            margin-top: 8px;
            font-size: 12px;
            color: var(--text-sec);
            background-color: var(--comment-bg);
            border-left: 3px solid var(--border-color);
            padding: 6px 10px;
            border-radius: 0 4px 4px 0;
            white-space: pre-wrap;
            word-break: break-word;
        }}

        footer {{
            text-align: center;
            margin-top: 20px;
            font-size: 12px;
            color: var(--text-sec);
        }}

        /* ===== 移动端适配 ===== */
        @media (max-width: 768px) {{
            table, thead, tbody, th, td, tr {{
                display: block;
            }}
            thead tr {{
                position: absolute;
                top: -9999px;
                left: -9999px;
            }}

            tr {{
                position: relative;
                margin-bottom: 15px;
                border: 1px solid var(--border-color);
                border-radius: 8px;
                background: var(--bg-card);
                box-shadow: 0 1px 4px rgba(0,0,0,0.05);
            }}
            table {{ background: transparent; box-shadow: none; border-radius: 0; }}

            /* 上下堆叠结构 */
            td {{ border: none; padding: 12px; }}

            /* 徽章定位到右上角 */
            .status-wrapper {{
                position: absolute;
                top: 12px;
                right: 12px;
                float: none;
                margin: 0;
            }}

            .col-source {{
                width: 100%;
                border-right: none;
                border-bottom: 1px dashed var(--border-color);
                padding-right: 65px;
                box-sizing: border-box;
            }}

            .col-target {{
                width: 100%;
                background-color: rgba(0, 0, 0, 0.01);
                border-radius: 0 0 8px 8px;
                box-sizing: border-box;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>{project_name}</h1>
            <div class="meta-grid">
                <div class="meta-item"><b>{_("Source Language")}:</b> {source_lang}</div>
                <div class="meta-item"><b>{_("Target Language")}:</b> {target_lang}</div>
                <div class="meta-item"><b>{_("Export Date")}:</b> {now_str}</div>
                <div class="meta-item"><b>{_("Total Items")}:</b> {total}</div>
            </div>
            <div class="progress-container">
                <div class="progress-bar-bg">
                    <div class="progress-bar-fill" style="width: {progress}%"></div>
                </div>
                <span class="progress-text">
                    {translated} / {total} ({progress}%) {_("Translated")} |
                    {reviewed} {_("Reviewed")} |
                    {ignored} {_("Ignored")}
                </span>
            </div>
        </header>

        <table>
            <thead>
                <tr>
                    <th>{_("Source Text")}</th>
                    <th>{_("Translation")}</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>

        <footer>
            Generated by LexiSync v{app_instance.config.get("version", "1.3.0")}
        </footer>
    </div>
</body>
</html>
    """
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html_template)


def export_to_json(filepath, translatable_objects, displayed_ids_order=None, app_instance=None):
    items_to_export_data = []
    export_obj_list = []
    if displayed_ids_order and app_instance:
        obj_map = {obj.id: obj for obj in app_instance.translatable_objects}
        export_obj_list = [obj_map[ts_id] for ts_id in displayed_ids_order if ts_id in obj_map]
    elif app_instance:  # Fallback to all if no specific order/filter is provided
        export_obj_list = app_instance.translatable_objects
    else:
        export_obj_list = translatable_objects

    for ts_obj in export_obj_list:
        items_to_export_data.append(
            {
                "id": ts_obj.id,
                "string_type": ts_obj.string_type,
                "original_semantic": ts_obj.original_semantic,
                "translation": ts_obj.get_translation_for_storage_and_tm(),
                "comment": ts_obj.comment,
                "is_reviewed": ts_obj.is_reviewed,
                "is_ignored": ts_obj.is_ignored,
                "line_num_in_file": ts_obj.line_num_in_file,
                "original_raw": ts_obj.original_raw,
                "is_plural": ts_obj.is_plural,
                "original_plural": ts_obj.original_plural,
                "plural_translations": ts_obj.plural_translations,
            }
        )

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(items_to_export_data, f, indent=4, ensure_ascii=False)


def export_to_yaml(filepath, translatable_objects, displayed_ids_order=None, app_instance=None):
    items_to_export_data = []

    export_obj_list = []
    if displayed_ids_order and app_instance:
        obj_map = {obj.id: obj for obj in app_instance.translatable_objects}
        export_obj_list = [obj_map[ts_id] for ts_id in displayed_ids_order if ts_id in obj_map]
    elif app_instance:  # Fallback to all if no specific order/filter is provided
        export_obj_list = app_instance.translatable_objects
    else:
        export_obj_list = translatable_objects

    for ts_obj in export_obj_list:
        items_to_export_data.append(
            {
                "id": ts_obj.id,
                "string_type": ts_obj.string_type,
                "original_semantic": ts_obj.original_semantic,
                "translation": ts_obj.get_translation_for_storage_and_tm(),
                "comment": ts_obj.comment,
                "is_reviewed": ts_obj.is_reviewed,
                "is_ignored": ts_obj.is_ignored,
                "line_num_in_file": ts_obj.line_num_in_file,
                "original_raw": ts_obj.original_raw,
                "is_plural": ts_obj.is_plural,
                "original_plural": ts_obj.original_plural,
                "plural_translations": ts_obj.plural_translations,
            }
        )

    with open(filepath, "w", encoding="utf-8") as f:
        yaml.dump(items_to_export_data, f, allow_unicode=True, sort_keys=False)


def export_qa_report(filepath, issues, project_name="LexiSync Project"):
    """
    生成带格式的 QA 报告
    issues: List of dicts {severity, file, line, source, translation, type, message, ignored}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "QA Issues"

    # 定义样式
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    error_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # 浅红
    warning_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")  # 浅黄
    info_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # 浅蓝

    headers = [
        _("Severity"),
        _("File"),
        _("Line/ID"),
        _("Source"),
        _("Translation"),
        _("Issue Type"),
        _("Description"),
        _("Status"),
    ]
    ws.append(headers)

    # 应用表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for issue in issues:
        row_data = [
            issue["severity"].upper(),
            issue["file"],
            issue["line"],
            issue["source"],
            issue["translation"],
            issue["type"],
            issue["message"],
            _("Ignored") if issue["ignored"] else _("Active"),
        ]
        ws.append(row_data)

        # 应用行样式
        last_row = ws.max_row
        fill = None
        if issue["severity"] == "error":
            fill = error_fill
        elif issue["severity"] == "warning":
            fill = warning_fill
        elif issue["severity"] == "info":
            fill = info_fill

        if fill:
            for cell in ws[last_row]:
                cell.fill = fill

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)  # 最高50
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)
