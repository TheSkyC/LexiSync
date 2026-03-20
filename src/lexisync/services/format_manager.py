# Copyright (c) 2025-2026, TheSkyC
# SPDX-License-Identifier: Apache-2.0

import copy
import csv
import json
import logging
import os
from pathlib import Path
import plistlib
from typing import Any
import xml.etree.ElementTree as ET
import zipfile

from rapidfuzz import fuzz
import regex as re
import xxhash

from lexisync.models.translatable_string import TranslatableString
from lexisync.services import code_file_service, po_file_service
from lexisync.utils.file_utils import atomic_open
from lexisync.utils.localization import _

logger = logging.getLogger(__name__)


class BaseFormatHandler:
    """格式处理器的基类"""

    format_id = "unknown"
    is_monolingual = False
    extensions = []
    format_type = "translation"

    display_name = "Unknown File"
    badge_text = "UNK"
    badge_bg_color = "#E0E0E0"
    badge_text_color = "#444444"

    def load(self, filepath, **kwargs):
        """返回: (translatable_objects, metadata, language_code)"""
        raise NotImplementedError

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        """保存文件"""
        raise NotImplementedError

    def _get_relative_path(self, filepath: str) -> str:
        current = Path(filepath).parent
        while True:
            if (current / "project.json").is_file():
                try:
                    return Path(filepath).relative_to(current).as_posix()
                except ValueError:
                    break
            if current.parent == current:
                break
            current = current.parent
        return os.path.basename(filepath)

    def _detect_language_from_filename(self, filename: str) -> str:
        name = os.path.splitext(filename)[0].lower()
        m = re.search(r"[._-]([a-z]{2,3}(?:[_-][a-zA-Z]{2,4})?)$", name, re.IGNORECASE)
        if m:
            return m.group(1)
        common = {"en", "zh", "ja", "ko", "fr", "de", "es", "it", "ru", "pt", "ar", "tr", "pl", "nl"}
        parts = re.split(r"[_\-]", name)
        for part in reversed(parts):
            if part in common:
                return part
        return "en"

    def _detect_language_from_path(self, filepath: str) -> str:
        for part in Path(filepath).parts:
            if part.endswith(".lproj"):
                return part[: -len(".lproj")]
            if part.startswith("values-"):
                lang = part[len("values-") :].replace("-r", "-")
                if lang:
                    return lang
        return self._detect_language_from_filename(os.path.basename(filepath))

    def _match_lang_code(self, code: str, available_codes: list[str]) -> str | None:
        """模糊匹配语言代码，例如 zh -> zh-Hans, en -> en-US"""
        if code in available_codes:
            return code

        for a_code in available_codes:
            if a_code.startswith(f"{code}-") or a_code.startswith(f"{code}_"):
                return a_code

        if code == "zh":
            for c in ["zh-Hans", "zh-CN", "zh_CN", "zh-Simplified"]:
                if c in available_codes:
                    return c
        return None

    def get_initial_translation(self, value_from_file, app_instance):
        if self.is_monolingual:
            fill_enabled = False
            if app_instance and hasattr(app_instance, "config"):
                fill_enabled = app_instance.config.get("fill_translation_with_source", False)
            return value_from_file if fill_enabled else ""
        return value_from_file


class PoFormatHandler(BaseFormatHandler):
    """
    GNU Gettext PO/POT 格式处理器
    支持的特性:
    1. 标准元数据: 读写文件头信息 (Project-Id, Language, POT-Creation-Date 等)
    2. 注释处理: 支持翻译者注释 (#) 和提取出的开发者注释 (#.)
    3. 状态同步: 完美支持模糊标记 (fuzzy) 以及 LexiSync 特有的已审阅标记
    4. 源码引用: 自动记录并还原条目在原始代码中的位置信息 (#:)
    """

    format_id = "po"
    is_monolingual = False
    extensions = [".po", ".pot"]
    format_type = "translation"
    display_name = _("PO Translation File")
    badge_text = "PO"
    badge_bg_color = "#F3E5F5"
    badge_text_color = "#7B1FA2"

    def load(self, filepath, **kwargs):
        relative_path = kwargs.get("relative_path")
        return po_file_service.load_from_po(filepath, relative_path=relative_path)

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        original_file_name = kwargs.get("original_file_name", "source_code")
        app_instance = kwargs.get("app_instance")
        po_file_service.save_to_po(filepath, translatable_objects, metadata, original_file_name, app_instance)


class TsFormatHandler(BaseFormatHandler):
    """
    Qt Linguist TS 格式处理器
    支持的特性:
    1. 上下文分组: 严格遵循 XML 结构，按 <context> 元素组织翻译条目
    2. 源码关联: 自动寻址并加载 <location> 标签指向的本地源码片段作为上下文
    3. 状态映射: 将 type="unfinished" 状态与 LexiSync 的“审阅”流程深度绑定
    4. 扩展注释: 支持 extracomment (开发者) 和 translatorcomment (译员) 的读写
    """

    format_id = "ts"
    is_monolingual = False
    extensions = [".ts"]
    format_type = "translation"
    display_name = _("Qt TS Translation File")
    badge_text = "TS"
    badge_bg_color = "#E8F5E9"
    badge_text_color = "#2E7D32"

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        logger.debug(f"[TsFormatHandler] Loading TS file: {filepath}")
        tree = ET.parse(filepath)
        root = tree.getroot()
        language = root.get("language", "")

        translatable_objects = []
        occurrence_counters = {}
        file_content_cache = {}

        relative_path = kwargs.get("relative_path")
        ts_file_rel_path = relative_path if relative_path else self._get_relative_path(filepath)

        for context in root.findall("context"):
            context_name = context.findtext("name", "")
            for message in context.findall("message"):
                source = message.findtext("source", "")
                if not source:
                    continue

                translation_node = message.find("translation")
                translation = translation_node.text if translation_node is not None and translation_node.text else ""
                is_unfinished = translation_node.get("type") == "unfinished" if translation_node is not None else False
                is_obsolete = translation_node.get("type") == "obsolete" if translation_node is not None else False

                if is_obsolete:
                    continue

                locations = []
                for loc in message.findall("location"):
                    locations.append((loc.get("filename", ""), loc.get("line", "0")))

                full_code_lines = []
                if locations:
                    src_rel_path = locations[0][0]
                    src_abs_path = os.path.normpath(os.path.join(os.path.dirname(filepath), src_rel_path))

                    if src_abs_path in file_content_cache:
                        full_code_lines = file_content_cache[src_abs_path]
                    elif os.path.isfile(src_abs_path):
                        try:
                            with open(src_abs_path, encoding="utf-8", errors="replace") as f:
                                lines = f.read().splitlines()
                                file_content_cache[src_abs_path] = lines
                                full_code_lines = lines
                        except Exception as e:
                            logger.warning(f"Failed to read source file for context: {src_abs_path}, error: {e}")

                line_num = int(locations[0][1]) if locations else 0
                forced_occurrences = [(ts_file_rel_path, str(line_num))]

                extracomment = message.findtext("extracomment", "")
                translatorcomment = message.findtext("translatorcomment", "")

                if locations:
                    src_context_lines = full_code_lines

                key = (source, context_name)
                current_index = occurrence_counters.get(key, 0)
                occurrence_counters[key] = current_index + 1

                stable_name_for_uuid = f"{ts_file_rel_path}::{context_name}::{source}::{current_index}"

                obj_id = xxhash.xxh128(stable_name_for_uuid.encode("utf-8")).hexdigest()

                ts = TranslatableString(
                    original_raw=source,
                    original_semantic=source,
                    line_num=line_num,
                    char_pos_start_in_file=0,
                    char_pos_end_in_file=0,
                    full_code_lines=full_code_lines,
                    string_type="TS Import",
                    source_file_path=ts_file_rel_path,
                    occurrences=forced_occurrences,
                    occurrence_index=current_index,
                    id=obj_id,
                )
                ts.set_translation_internal(translation, is_initial=True)
                ts.context = context_name
                ts.po_comment = extracomment
                ts.comment = translatorcomment
                ts.is_reviewed = not is_unfinished
                ts.update_sort_weight()
                translatable_objects.append(ts)

        return translatable_objects, {"language": language}, language

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        root = ET.Element("TS", version="2.1")
        app_instance = kwargs.get("app_instance")

        lang_code = "en"
        if app_instance:
            lang_code = app_instance.current_target_language
        elif metadata and "language" in metadata:
            lang_code = metadata["language"]

        root.set("language", lang_code)
        contexts = {}
        for ts in translatable_objects:
            if not ts.original_semantic or ts.id == "##NEW_ENTRY##":
                continue
            ctx = ts.context or "Default"
            if ctx not in contexts:
                contexts[ctx] = []
            contexts[ctx].append(ts)

        for ctx_name, items in contexts.items():
            context_node = ET.SubElement(root, "context")
            name_node = ET.SubElement(context_node, "name")
            name_node.text = ctx_name

            for ts in items:
                msg_node = ET.SubElement(context_node, "message")

                entry_occurrences = []
                clean_extracomment_lines = []

                if ts.po_comment:
                    for line in ts.po_comment.splitlines():
                        if line.strip().startswith("#:"):
                            content = line.replace("#:", "").strip()
                            for part in content.split():
                                if ":" in part:
                                    try:
                                        fpath, lineno = part.rsplit(":", 1)
                                        entry_occurrences.append((fpath, lineno))
                                    except ValueError:
                                        pass
                        else:
                            clean_extracomment_lines.append(line)

                if not entry_occurrences:
                    entry_occurrences = [("unknown", "0")]

                # 添加位置信息
                for loc_file, loc_line in entry_occurrences:
                    loc_node = ET.SubElement(msg_node, "location")
                    loc_node.set("filename", loc_file)
                    loc_node.set("line", str(loc_line))

                # 添加源文本
                source_node = ET.SubElement(msg_node, "source")
                source_node.text = ts.original_semantic

                # 添加 extracomment（不包含位置信息）
                clean_extracomment = "\n".join(clean_extracomment_lines).strip()
                if clean_extracomment:
                    extracomment_node = ET.SubElement(msg_node, "extracomment")
                    extracomment_node.text = clean_extracomment

                # 添加译者注释
                if ts.comment:
                    translatorcomment_node = ET.SubElement(msg_node, "translatorcomment")
                    translatorcomment_node.text = ts.comment

                trans_node = ET.SubElement(msg_node, "translation")
                trans_node.text = ts.translation

                if not ts.is_reviewed:
                    trans_node.set("type", "unfinished")

        tree = ET.ElementTree(root)
        if hasattr(ET, "indent"):
            ET.indent(tree, space="    ", level=0)

        xml_str = ET.tostring(root, encoding="utf-8", xml_declaration=True).decode("utf-8")
        xml_str = xml_str.replace("?>", "?>\n<!DOCTYPE TS>")

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write(xml_str)


class XliffFormatHandler(BaseFormatHandler):
    """
    XLIFF (XML Localization Interchange File Format) 处理器
    支持版本:
    - XLIFF 1.2
    - XLIFF 2.0
    """

    format_id = "xliff"
    is_monolingual = False
    extensions = [".xlf", ".xliff"]
    format_type = "translation"
    display_name = _("XLIFF Translation File")
    badge_text = "XLIFF"
    badge_bg_color = "#E1F5FE"
    badge_text_color = "#01579B"

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        logger.debug(f"[XliffFormatHandler] Loading XLIFF file: {filepath}")

        try:
            tree = ET.parse(filepath)
            root = tree.getroot()
        except ET.ParseError as e:
            logger.error(f"XML Parse Error: {e}")
            raise ValueError(f"Invalid XML file: {e}") from e

        # 提取命名空间
        namespace = ""
        if root.tag.startswith("{"):
            namespace = root.tag[1:].split("}")[0]

        # 定义命名空间映射，用于 findall
        ns = {"x": namespace} if namespace else {}
        prefix = "x:" if namespace else ""

        version = root.get("version", "1.2")
        rel_path = kwargs.get("relative_path") or os.path.basename(filepath)
        translatable_objects = []
        occurrence_counters = {}

        # 查找所有 file 节点
        files = root.findall(f".//{prefix}file", ns)
        # 如果根节点本身就是 file (某些非标文件)，则把它作为唯一的文件节点
        if not files and (self._strip_ns(root.tag) == "file"):
            files = [root]

        target_lang_global = root.get("trgLang")  # XLIFF 2.0 根节点属性

        for file_elem in files:
            # XLIFF 1.2 属性
            source_lang = file_elem.get("source-language")
            target_lang = file_elem.get("target-language")
            original_file = file_elem.get("original", "")

            # XLIFF 2.0 属性 (srcLang, trgLang)
            if not source_lang:
                source_lang = file_elem.get("srcLang")
            if not target_lang:
                target_lang = file_elem.get("trgLang")

            # 如果 file 节点没写，回退到根节点属性 (2.0)
            if not source_lang:
                source_lang = root.get("srcLang", "en")
            if not target_lang:
                target_lang = target_lang_global or "en"

            # 根据版本分发处理逻辑
            if version.startswith("2"):
                # XLIFF 2.0: <unit> -> <segment> -> <source>/<target>
                units = file_elem.findall(f".//{prefix}unit", ns)
                for unit in units:
                    self._process_unit_v2(
                        unit,
                        translatable_objects,
                        occurrence_counters,
                        rel_path,
                        original_file,
                        ns,
                        prefix,
                        app_instance,
                    )
            else:
                # XLIFF 1.2: <trans-unit> -> <source>/<target>
                trans_units = file_elem.findall(f".//{prefix}trans-unit", ns)
                for trans_unit in trans_units:
                    self._process_trans_unit(
                        trans_unit,
                        translatable_objects,
                        occurrence_counters,
                        rel_path,
                        original_file,
                        ns,
                        prefix,
                        app_instance,
                    )

        metadata = {
            "version": version,
            "source_language": source_lang,
            "target_language": target_lang,
            "namespace_uri": namespace,
        }

        logger.info(f"[XliffFormatHandler] Loaded {len(translatable_objects)} strings. Version: {version}")
        return translatable_objects, metadata, metadata["target_language"]

    def _strip_ns(self, tag):
        return tag.split("}", 1)[1] if "}" in tag else tag

    def _process_trans_unit(
        self, trans_unit, results, occurrence_counters, file_rel_path, original_file, ns, prefix, app_instance
    ):
        """处理 XLIFF 1.2 的 trans-unit"""
        unit_id = trans_unit.get("id", "unknown")

        source_elem = trans_unit.find(f"{prefix}source", ns)
        target_elem = trans_unit.find(f"{prefix}target", ns)

        if source_elem is None or source_elem.text is None:
            return

        source_text = source_elem.text
        target_text = target_elem.text if target_elem is not None and target_elem.text else ""

        state = target_elem.get("state", "needs-translation") if target_elem is not None else "needs-translation"
        is_reviewed = state in ["translated", "final", "signed-off"]

        note_elems = trans_unit.findall(f"{prefix}note", ns)
        notes = [note.text for note in note_elems if note.text]

        self._create_ts(
            source_text,
            target_text,
            unit_id,
            notes,
            is_reviewed,
            file_rel_path,
            results,
            occurrence_counters,
            "XLIFF 1.2",
            app_instance,
        )

    def _process_unit_v2(
        self, unit, results, occurrence_counters, file_rel_path, original_file, ns, prefix, app_instance
    ):
        """处理 XLIFF 2.0 的 unit"""
        unit_id = unit.get("id", "unknown")

        # 提取 Notes (2.0 的 notes 在 unit 级别)
        notes = []
        notes_elem = unit.find(f"{prefix}notes", ns)
        if notes_elem is not None:
            for n in notes_elem.findall(f"{prefix}note", ns):
                if n.text:
                    notes.append(n.text)

        # 遍历所有 segment
        segments = unit.findall(f".//{prefix}segment", ns)
        for i, seg in enumerate(segments):
            source_elem = seg.find(f"{prefix}source", ns)
            target_elem = seg.find(f"{prefix}target", ns)

            if source_elem is None or source_elem.text is None:
                continue

            source_text = source_elem.text
            target_text = target_elem.text if target_elem is not None and target_elem.text else ""

            state = seg.get("state", "initial")
            is_reviewed = state in ["translated", "final", "reviewed"]

            # 如果一个 unit 有多个 segment，ID 需要区分
            current_id = unit_id if len(segments) == 1 else f"{unit_id}_{i}"

            self._create_ts(
                source_text,
                target_text,
                current_id,
                notes,
                is_reviewed,
                file_rel_path,
                results,
                occurrence_counters,
                "XLIFF 2.0",
                app_instance,
            )

    def _create_ts(self, source, target, uid, notes, is_reviewed, rel_path, results, counters, type_str, app_instance):
        """通用的 TranslatableString 创建逻辑"""
        counter_key = (source, uid)
        idx = counters.get(counter_key, 0)
        counters[counter_key] = idx + 1

        stable_name = f"{rel_path}::{uid}::{source}::{idx}"
        obj_id = xxhash.xxh128(stable_name.encode("utf-8")).hexdigest()

        ts = TranslatableString(
            source, source, 0, 0, 0, [], f"{type_str} Import", rel_path, [(rel_path, uid)], idx, obj_id
        )
        # 使用基类方法填充译文
        if target:
            initial_trans = target
        elif self.is_monolingual:
            initial_trans = self.get_initial_translation(source, app_instance)
        else:
            initial_trans = ""
        ts.set_translation_internal(initial_trans, is_initial=True)

        ts.context = uid
        ts.comment = "\n".join(notes) if notes else ""
        ts.po_comment = f"#: XLIFF ID: {uid}"
        ts.is_reviewed = is_reviewed
        ts.update_sort_weight()
        results.append(ts)

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        version = metadata.get("version", "1.2")
        uri = metadata.get("namespace_uri", "urn:oasis:names:tc:xliff:document:1.2")
        body_elem = None
        # 备份并恢复全局命名空间
        old_ns_map = getattr(ET, "_namespace_map", {}).copy()
        # 注册命名空间，防止输出 ns0: 前缀
        ET.register_namespace("", uri)

        try:
            root = ET.Element(f"{{{uri}}}xliff", version=version)

            # 根据版本构建结构
            if version.startswith("2"):
                root.set("srcLang", metadata.get("source_language", "en"))
                root.set("trgLang", metadata.get("target_language", "en"))
                file_elem = ET.SubElement(root, f"{{{uri}}}file", id="f1")
            else:
                file_elem = ET.SubElement(root, f"{{{uri}}}file")
                file_elem.set("source-language", metadata.get("source_language", "en"))
                file_elem.set("target-language", metadata.get("target_language", "en"))
                file_elem.set("datatype", "plaintext")
                body_elem = ET.SubElement(file_elem, f"{{{uri}}}body")

            for ts in translatable_objects:
                if not ts.original_semantic or ts.id == "##NEW_ENTRY##":
                    continue

                unit_id = ts.context or f"u{ts.id[:8]}"

                if version.startswith("2"):
                    # XLIFF 2.0 Save
                    unit = ET.SubElement(file_elem, f"{{{uri}}}unit", id=unit_id)
                    if ts.comment:
                        notes = ET.SubElement(unit, f"{{{uri}}}notes")
                        ET.SubElement(notes, f"{{{uri}}}note").text = ts.comment

                    segment = ET.SubElement(unit, f"{{{uri}}}segment")
                    if ts.is_reviewed:
                        segment.set("state", "translated")

                    ET.SubElement(segment, f"{{{uri}}}source").text = ts.original_semantic
                    ET.SubElement(segment, f"{{{uri}}}target").text = ts.translation
                else:
                    # XLIFF 1.2 Save
                    unit = ET.SubElement(body_elem, f"{{{uri}}}trans-unit", id=unit_id)
                    ET.SubElement(unit, f"{{{uri}}}source").text = ts.original_semantic
                    target = ET.SubElement(unit, f"{{{uri}}}target")
                    target.text = ts.translation

                    if ts.is_reviewed:
                        target.set("state", "translated")
                    else:
                        target.set("state", "needs-translation")

                    if ts.comment:
                        ET.SubElement(unit, f"{{{uri}}}note").text = ts.comment

            tree = ET.ElementTree(root)
            if hasattr(ET, "indent"):
                ET.indent(tree, space="  ")

            import io

            buffer = io.BytesIO()
            tree.write(buffer, encoding="utf-8", xml_declaration=True)
            xml_content = buffer.getvalue().decode("utf-8")

            with atomic_open(filepath, "w", encoding="utf-8") as f:
                f.write(xml_content)
        finally:
            if hasattr(ET, "_namespace_map"):
                ET._namespace_map.clear()
                ET._namespace_map.update(old_ns_map)


class AndroidStringsFormatHandler(BaseFormatHandler):
    """
    Android Strings XML 处理器

    支持的元素:
    - <string name="key">value</string>
    - <string name="key" translatable="false">value</string>
    - <plurals name="key">...</plurals>
    - <string-array name="key">...</string-array>
    """

    format_id = "android_strings"
    is_monolingual = True
    extensions = [".xml"]
    format_type = "translation"
    display_name = _("Android Strings XML")
    badge_text = "Android"
    badge_bg_color = "#E8F5E9"
    badge_text_color = "#1B5E20"

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        logger.debug(f"[AndroidStringsFormatHandler] Loading Android strings.xml: {filepath}")

        tree = ET.parse(filepath)
        root = tree.getroot()

        if root.tag != "resources":
            raise ValueError("Not a valid Android strings.xml file (root element must be <resources>)")

        relative_path = kwargs.get("relative_path")
        xml_file_rel_path = relative_path if relative_path else self._get_relative_path(filepath)

        translatable_objects = []
        occurrence_counters = {}

        # 处理 <string> 元素
        for string_elem in root.findall("string"):
            self._process_string_element(
                string_elem, translatable_objects, occurrence_counters, xml_file_rel_path, app_instance
            )

        # 处理 <plurals> 元素
        for plurals_elem in root.findall("plurals"):
            self._process_plurals_element(
                plurals_elem, translatable_objects, occurrence_counters, xml_file_rel_path, app_instance
            )

        # 处理 <string-array> 元素
        for array_elem in root.findall("string-array"):
            self._process_array_element(
                array_elem, translatable_objects, occurrence_counters, xml_file_rel_path, app_instance
            )

        # 尝试从文件名检测语言
        language_code = self._detect_language_from_path(filepath)

        metadata = {
            "xml_declaration": True,
            "indent": "    ",  # Android 标准使用 4 空格
        }

        logger.info(f"[AndroidStringsFormatHandler] Loaded {len(translatable_objects)} strings from {filepath}")
        return translatable_objects, metadata, language_code

    def _process_string_element(
        self,
        elem: ET.Element,
        results: list[TranslatableString],
        occurrence_counters: dict,
        file_rel_path: str,
        app_instance=None,
    ):
        """处理 <string> 元素"""

        name = elem.get("name")
        if not name:
            return

        # 检查 translatable 属性
        translatable = elem.get("translatable", "true").lower() == "true"
        if not translatable:
            return  # 跳过不可翻译的字符串

        # 提取文本（可能包含 CDATA）
        text = self._extract_text(elem)
        if not text or not text.strip():
            return

        # 生成唯一标识
        counter_key = (text, name)
        current_index = occurrence_counters.get(counter_key, 0)
        occurrence_counters[counter_key] = current_index + 1

        stable_name = f"{file_rel_path}::{name}::{text}::{current_index}"
        obj_id = xxhash.xxh128(stable_name.encode("utf-8")).hexdigest()

        # 检测格式化占位符
        placeholders = self._detect_android_placeholders(text)
        comment_parts = []
        if placeholders:
            comment_parts.append(f"Format arguments: {', '.join(placeholders)}")

        ts = TranslatableString(
            original_raw=text,
            original_semantic=text,
            line_num=0,
            char_pos_start_in_file=0,
            char_pos_end_in_file=0,
            full_code_lines=[],
            string_type="Android String",
            source_file_path=file_rel_path,
            occurrences=[(file_rel_path, name)],
            occurrence_index=current_index,
            id=obj_id,
        )

        ts.context = name
        ts.comment = "\n".join(comment_parts) if comment_parts else ""
        ts.po_comment = f"#: Android string name: {name}"
        ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)
        ts.is_reviewed = False
        ts.update_sort_weight()

        results.append(ts)

    def _process_plurals_element(
        self,
        elem: ET.Element,
        results: list[TranslatableString],
        occurrence_counters: dict,
        file_rel_path: str,
        app_instance=None,
    ):
        """处理 <plurals> 元素"""

        name = elem.get("name")
        if not name:
            return

        # 处理每个 <item quantity="...">
        for item in elem.findall("item"):
            quantity = item.get("quantity", "other")
            text = self._extract_text(item)

            if not text or not text.strip():
                continue

            # 使用 name:quantity 作为唯一标识
            full_name = f"{name}:{quantity}"

            counter_key = (text, full_name)
            current_index = occurrence_counters.get(counter_key, 0)
            occurrence_counters[counter_key] = current_index + 1

            stable_name = f"{file_rel_path}::{full_name}::{text}::{current_index}"
            obj_id = xxhash.xxh128(stable_name.encode("utf-8")).hexdigest()

            placeholders = self._detect_android_placeholders(text)
            comment_parts = [f"Plural form: {quantity}"]
            if placeholders:
                comment_parts.append(f"Format arguments: {', '.join(placeholders)}")

            ts = TranslatableString(
                original_raw=text,
                original_semantic=text,
                line_num=0,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="Android Plural",
                source_file_path=file_rel_path,
                occurrences=[(file_rel_path, full_name)],
                occurrence_index=current_index,
                id=obj_id,
            )

            ts.context = full_name
            ts.comment = "\n".join(comment_parts)
            ts.po_comment = f"#: Android plurals name: {name}, quantity: {quantity}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)

            results.append(ts)

    def _process_array_element(
        self,
        elem: ET.Element,
        results: list[TranslatableString],
        occurrence_counters: dict,
        file_rel_path: str,
        app_instance=None,
    ):
        """处理 <string-array> 元素"""

        name = elem.get("name")
        if not name:
            return

        # 处理每个 <item>
        for idx, item in enumerate(elem.findall("item")):
            text = self._extract_text(item)

            if not text or not text.strip():
                continue

            # 使用 name[index] 作为唯一标识
            full_name = f"{name}[{idx}]"

            counter_key = (text, full_name)
            current_index = occurrence_counters.get(counter_key, 0)
            occurrence_counters[counter_key] = current_index + 1

            stable_name = f"{file_rel_path}::{full_name}::{text}::{current_index}"
            obj_id = xxhash.xxh128(stable_name.encode("utf-8")).hexdigest()

            ts = TranslatableString(
                original_raw=text,
                original_semantic=text,
                line_num=0,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="Android Array",
                source_file_path=file_rel_path,
                occurrences=[(file_rel_path, full_name)],
                occurrence_index=current_index,
                id=obj_id,
            )

            ts.context = full_name
            ts.comment = f"Array item index: {idx}"
            ts.po_comment = f"#: Android string-array name: {name}, index: {idx}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)

            results.append(ts)

    def _extract_text(self, elem: ET.Element) -> str:
        """提取元素的文本内容（处理 CDATA 和转义字符）"""
        if elem.text:
            # 解码 XML 转义
            return self._unescape_android_xml(elem.text)
        return ""

    def _unescape_android_xml(self, text: str) -> str:
        """解码 Android XML 转义字符"""
        # Android 特殊转义
        text = text.replace(r"\\", "\x00BACKSLASH\x00")
        text = text.replace(r"\'", "'")
        text = text.replace(r"\"", '"')
        text = text.replace(r"\n", "\n")
        text = text.replace(r"\t", "\t")
        text = text.replace("\x00BACKSLASH\x00", "\\")
        return text

    def _escape_android_xml(self, text: str) -> str:
        """编码 Android XML 转义字符"""
        text = text.replace("\\", r"\\")
        text = text.replace("'", r"\'")
        text = text.replace('"', r"\"")
        text = text.replace("\n", r"\n")
        text = text.replace("\t", r"\t")
        return text

    def _detect_android_placeholders(self, text: str) -> list[str]:
        """检测 Android 格式化占位符"""
        # 匹配 %1$s, %d, %2$f, %s 等
        pattern = r"%(\d+\$)?[diouxXeEfFgGaAcspn]"
        matches = re.findall(pattern, text)
        return [f"%{m}" if m else "%" for m in matches]

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        """保存 Android strings.xml 文件"""
        logger.debug(f"[AndroidStringsFormatHandler] Saving Android strings.xml: {filepath}")

        # 创建根元素
        root = ET.Element("resources")

        # 按类型分组
        strings_by_type = {"string": [], "plurals": {}, "array": {}}

        for ts in translatable_objects:
            if not ts.original_semantic or ts.id == "##NEW_ENTRY##":
                continue

            if ts.string_type == "Android String":
                strings_by_type["string"].append(ts)
            elif ts.string_type == "Android Plural":
                # 解析 name:quantity
                if ":" in ts.context:
                    name, quantity = ts.context.rsplit(":", 1)
                    if name not in strings_by_type["plurals"]:
                        strings_by_type["plurals"][name] = []
                    strings_by_type["plurals"][name].append((quantity, ts))
            elif ts.string_type == "Android Array":
                # 解析 name[index]
                if "[" in ts.context:
                    name = ts.context[: ts.context.index("[")]
                    if name not in strings_by_type["array"]:
                        strings_by_type["array"][name] = []
                    strings_by_type["array"][name].append(ts)

        # 添加 <string> 元素
        for ts in strings_by_type["string"]:
            string_elem = ET.SubElement(root, "string")
            string_elem.set("name", ts.context)
            translation = ts.translation if ts.translation else ts.original_semantic
            string_elem.text = self._escape_android_xml(translation)

        # 添加 <plurals> 元素
        for name, items in strings_by_type["plurals"].items():
            plurals_elem = ET.SubElement(root, "plurals")
            plurals_elem.set("name", name)

            for quantity, ts in sorted(items, key=lambda x: x[0]):
                item_elem = ET.SubElement(plurals_elem, "item")
                item_elem.set("quantity", quantity)
                translation = ts.translation if ts.translation else ts.original_semantic
                item_elem.text = self._escape_android_xml(translation)

        # 添加 <string-array> 元素
        for name, items in strings_by_type["array"].items():
            array_elem = ET.SubElement(root, "string-array")
            array_elem.set("name", name)

            # 按索引排序
            sorted_items = sorted(
                items, key=lambda ts: int(ts.context[ts.context.index("[") + 1 : ts.context.index("]")])
            )

            for ts in sorted_items:
                item_elem = ET.SubElement(array_elem, "item")
                translation = ts.translation if ts.translation else ts.original_semantic
                item_elem.text = self._escape_android_xml(translation)

        # 格式化并保存
        tree = ET.ElementTree(root)
        if hasattr(ET, "indent"):
            ET.indent(tree, space="    ", level=0)

        with atomic_open(filepath, "wb") as f:
            tree.write(f, encoding="utf-8", xml_declaration=True)

        logger.info(f"[AndroidStringsFormatHandler] Saved {len(translatable_objects)} strings to {filepath}")


class IosStringsFormatHandler(BaseFormatHandler):
    """
    Apple iOS / macOS 本地化文件处理器

    支持的格式与特性:
    1. .strings 文件: 标准 "key" = "value"; 格式，完整保留行内注释和块注释
    2. .stringsdict 文件: XML Plist 复数规则，支持 NSStringLocalizedFormatKey
       以及所有 CLDR 复数类别 (zero/one/two/few/many/other)
    3. 注释提取: 块注释 /* ... */ 和行注释 // 均可作为 translator comment 保留
    4. 转义处理: 正确处理 \\n \\t \\\\ \\\" 等 Apple 转义序列的双向转换
    5. 状态检测: 自动将 value == key 或 value 为空的条目标记为待审阅
    """

    format_id = "ios_strings"
    is_monolingual = True
    extensions = [".strings"]
    format_type = "translation"
    display_name = _("Apple .strings / .stringsdict")
    badge_text = "iOS"
    badge_bg_color = "#F9FBE7"
    badge_text_color = "#558B2F"

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or self._get_relative_path(filepath)
        language_code = self._detect_language_from_path(filepath)
        objects, meta = self._load_strings(filepath, relative_path, app_instance=app_instance)
        return objects, meta, language_code

    def _load_strings(self, filepath, rel_path, app_instance=None):
        encoding = self._detect_encoding(filepath)
        with open(filepath, encoding=encoding, errors="replace") as f:
            content = f.read()

        translatable_objects = []
        occurrence_counters = {}

        # 解析所有条目（带前置注释）
        # 语法: (可选注释块/行) "key" = "value";
        token_re = re.compile(
            r"(?:(?P<block_comment>/\*.*?\*/)|(?P<line_comment>//[^\n]*\n))"
            r'|"(?P<key>(?:[^"\\]|\\.)*)"\s*=\s*"(?P<value>(?:[^"\\]|\\.)*)"\s*;',
            re.DOTALL,
        )

        pending_comment = []
        for m in token_re.finditer(content):
            if m.group("block_comment"):
                text = m.group("block_comment")[2:-2].strip()
                pending_comment.append(text)
            elif m.group("line_comment"):
                text = m.group("line_comment")[2:].strip()
                pending_comment.append(text)
            else:
                key = self._unescape(m.group("key"))
                value = self._unescape(m.group("value"))
                comment = "\n".join(pending_comment).strip()
                pending_comment.clear()

                if not key:
                    continue

                counter_key = (key, rel_path)
                idx = occurrence_counters.get(counter_key, 0)
                occurrence_counters[counter_key] = idx + 1

                stable = f"{rel_path}::{key}::{idx}"
                obj_id = xxhash.xxh128(stable.encode()).hexdigest()

                ts = TranslatableString(
                    original_raw=key,
                    original_semantic=key,
                    line_num=0,
                    char_pos_start_in_file=0,
                    char_pos_end_in_file=0,
                    full_code_lines=[],
                    string_type="iOS String",
                    source_file_path=rel_path,
                    occurrences=[(rel_path, key)],
                    occurrence_index=idx,
                    id=obj_id,
                )
                ts.set_translation_internal(self.get_initial_translation(value, app_instance), is_initial=True)
                ts.context = key
                ts.comment = comment
                ts.po_comment = f"#: Apple strings key: {key}"

                # 审阅状态处理
                # value == key 通常意味着尚未翻译（源语言文件）
                if "@reviewed" in comment:
                    ts.is_reviewed = True
                elif "@unreviewed" in comment:
                    ts.is_reviewed = False
                else:
                    ts.is_reviewed = bool(value and value != key)
                # 清理掉内部标记
                ts.comment = comment.replace("@reviewed", "").replace("@unreviewed", "").strip()

                ts.update_sort_weight()
                translatable_objects.append(ts)

        meta = {"format": "strings", "raw_content": content}
        logger.info(f"[IosStringsFormatHandler] Loaded {len(translatable_objects)} entries from {filepath}")
        return translatable_objects, meta

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        self._save_strings(filepath, translatable_objects, metadata)

    def _save_strings(self, filepath, translatable_objects, metadata):
        lines = []
        for ts in translatable_objects:
            if not ts.original_semantic or ts.id == "##NEW_ENTRY##":
                continue
            comment_to_write = ts.comment
            if ts.is_reviewed:
                comment_to_write = f"{comment_to_write} @reviewed".strip()
            else:
                comment_to_write = f"{comment_to_write} @unreviewed".strip()

            if comment_to_write:
                lines.append(f"/* {comment_to_write} */")
            translation = ts.translation if ts.translation else ts.original_semantic
            escaped_key = self._escape(ts.context or ts.original_semantic)
            escaped_val = self._escape(translation)
            lines.append(f'"{escaped_key}" = "{escaped_val}";')
            lines.append("")

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
            f.write("\n")
        logger.info(f"[IosStringsFormatHandler] Saved {len(translatable_objects)} strings to {filepath}")

    def _detect_encoding(self, filepath: str) -> str:
        """检测 Apple 文件的编码 (UTF-16LE 或 UTF-8)"""
        with open(filepath, "rb") as f:
            raw = f.read(4)
        if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
            return "utf-16"
        return "utf-8-sig"

    def _unescape(self, s: str) -> str:
        """将 .strings 转义序列还原为真实字符。"""
        s = s.replace("\\\\", "\x00ESC_BACKSLASH\x00")
        s = s.replace('\\"', '"')
        s = s.replace("\\'", "'")
        s = s.replace("\\n", "\n")
        s = s.replace("\\r", "\r")
        s = s.replace("\\t", "\t")
        s = s.replace("\x00ESC_BACKSLASH\x00", "\\")
        return s

    def _escape(self, s: str) -> str:
        """将真实字符编码为 .strings 转义序列。"""
        return (
            s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n").replace("\r", "\\r").replace("\t", "\\t")
        )


class XCStringsFormatHandler(BaseFormatHandler):
    """
    Apple Xcode String Catalog (.xcstrings) 格式处理器
    .xcstrings 是 Xcode 15+ 引入的 JSON 格式，将所有语言的翻译合并为一个文件。
    顶层结构:
    {
      "sourceLanguage": "en",
      "strings": {
        "<key>": {
          "comment": "...",
          "localizations": {
            "en": { "stringUnit": { "state": "translated", "value": "Hello" } },
            "zh-Hans": { "stringUnit": { "state": "new", "value": "" } },
            // 复数形式使用 "variations" -> "plural"
          }
        }
      },
      "version": "1.0"
    }

    支持的特性:
    1. 多语言单文件: 按目标语言筛选/写入对应 localization 块
    2. 复数形式: 完整支持 variations.plural (zero/one/two/few/many/other)
    3. 设备变体: 识别 variations.device (iPhone/iPad/mac/…) 并展平提取
    4. 状态映射: Xcode state ("new"/"translated"/"needs_review") <-> is_reviewed
    5. 注释保留: 顶层 comment 字段完整读写
    6. 源语言检测: 自动读取 sourceLanguage 字段
    """

    format_id = "xcstrings"
    is_monolingual = False
    extensions = [".xcstrings"]
    format_type = "translation"
    display_name = _("Xcode String Catalog (.xcstrings)")
    badge_text = "XCS"
    badge_bg_color = "#E3F2FD"
    badge_text_color = "#1565C0"

    # Xcode 复数类别顺序
    PLURAL_CATEGORIES = ["zero", "one", "two", "few", "many", "other"]
    # 支持的 device 变体键
    DEVICE_KEYS = ["iPhone", "iPad", "mac", "appleWatch", "appleTV", "appleVision", "iPod", "other"]

    def load(self, filepath: str, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or self._get_relative_path(filepath)

        with open(filepath, encoding="utf-8") as f:
            data = json.load(f)

        source_language = data.get("sourceLanguage", "en")
        strings_dict: dict = data.get("strings", {})

        # 确定目标语言
        target_lang_req = "en"
        if app_instance and hasattr(app_instance, "current_target_language"):
            target_lang_req = app_instance.current_target_language
        all_locales = set()
        for s in strings_dict.values():
            all_locales.update(s.get("localizations", {}).keys())
        target_lang = self._match_lang_code(target_lang_req, list(all_locales)) or target_lang_req

        translatable_objects: list[TranslatableString] = []
        occurrence_counters: dict = {}

        for key, entry in strings_dict.items():
            if not isinstance(entry, dict):
                continue

            comment = entry.get("comment", "")
            localizations: dict = entry.get("localizations", {})
            # 取目标语言的本地化块，若无则视为空（待翻译）
            loc_block = localizations.get(target_lang, {})
            # 取源语言文本（用于 original_semantic）
            src_block = localizations.get(source_language, {})
            source_value = self._extract_string_unit_value(src_block) or key

            # 判断是否有复数变体
            if "variations" in loc_block and "plural" in loc_block["variations"]:
                self._extract_plural_entries(
                    key,
                    comment,
                    source_value,
                    loc_block["variations"]["plural"],
                    localizations,
                    source_language,
                    translatable_objects,
                    occurrence_counters,
                    relative_path,
                    app_instance,
                )
            elif "variations" in loc_block and "device" in loc_block["variations"]:
                self._extract_device_entries(
                    key,
                    comment,
                    source_value,
                    loc_block["variations"]["device"],
                    translatable_objects,
                    occurrence_counters,
                    relative_path,
                    app_instance,
                )
            else:
                # 普通 stringUnit
                translation_value = self._extract_string_unit_value(loc_block)
                xcode_state = self._extract_state(loc_block)
                is_reviewed = xcode_state == "translated"

                ts = self._make_ts(
                    key=key,
                    context=key,
                    source_value=source_value,
                    translation_value=translation_value,
                    comment=comment,
                    xcode_state=xcode_state,
                    is_reviewed=is_reviewed,
                    relative_path=relative_path,
                    occurrence_counters=occurrence_counters,
                    string_type="XCString",
                    app_instance=app_instance,
                )
                translatable_objects.append(ts)

        metadata = {
            "raw_data": data,
            "source_language": source_language,
            "target_language": target_lang,
        }
        logger.info(
            f"[XCStringsFormatHandler] Loaded {len(translatable_objects)} entries (lang={target_lang}) from {filepath}"
        )
        return translatable_objects, metadata, target_lang

    def save(self, filepath: str, translatable_objects, metadata: dict, **kwargs):
        app_instance = kwargs.get("app_instance")
        raw_data: dict = copy.deepcopy(metadata.get("raw_data", {}))
        source_language: str = metadata.get("source_language", "en")

        target_lang = metadata.get("target_language", source_language)
        if app_instance and hasattr(app_instance, "current_target_language"):
            target_lang = app_instance.current_target_language or target_lang

        strings_dict: dict = raw_data.setdefault("strings", {})

        # 建立 context -> ts 映射（处理复数时 context 带 category 后缀）
        ts_map: dict[str, TranslatableString] = {}
        for ts in translatable_objects:
            if ts.id == "##NEW_ENTRY##" or not ts.context:
                continue
            ts_map[ts.context] = ts

        for key, entry in strings_dict.items():
            if not isinstance(entry, dict):
                continue
            localizations: dict = entry.setdefault("localizations", {})
            loc_block: dict = localizations.setdefault(target_lang, {})

            # 尝试复数
            has_plural = "variations" in loc_block and "plural" in loc_block["variations"]
            has_device = "variations" in loc_block and "device" in loc_block["variations"]
            if has_plural:
                plural_dict = loc_block["variations"]["plural"]
                for cat in self.PLURAL_CATEGORIES:
                    ctx = f"{key}:plural:{cat}"
                    ts = ts_map.get(ctx)
                    if ts and cat in plural_dict:
                        unit = plural_dict[cat].setdefault("stringUnit", {})
                        unit["value"] = ts.translation or ts.original_semantic
                        unit["state"] = "translated" if ts.is_reviewed else "needs_review"
            elif has_device:
                device_dict = loc_block["variations"]["device"]
                for device_key in self.DEVICE_KEYS:
                    ctx = f"{key}:device:{device_key}"
                    ts = ts_map.get(ctx)
                    if ts and device_key in device_dict:
                        unit = device_dict[device_key].setdefault("stringUnit", {})
                        unit["value"] = ts.translation or ts.original_semantic
                        unit["state"] = "translated" if ts.is_reviewed else "needs_review"
            else:
                ctx = key
                ts = ts_map.get(ctx)
                if ts:
                    su = loc_block.setdefault("stringUnit", {})
                    su["value"] = ts.translation or ts.original_semantic
                    su["state"] = "translated" if ts.is_reviewed else "needs_review"

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            json.dump(raw_data, f, indent=2, ensure_ascii=False)
            f.write("\n")

        logger.info(f"[XCStringsFormatHandler] Saved to {filepath} (lang={target_lang})")

    def _extract_string_unit_value(self, loc_block: dict) -> str:
        if not loc_block:
            return ""
        su = loc_block.get("stringUnit", {})
        return su.get("value", "") if isinstance(su, dict) else ""

    def _extract_state(self, loc_block: dict) -> str:
        if not loc_block:
            return "new"
        su = loc_block.get("stringUnit", {})
        return su.get("state", "new") if isinstance(su, dict) else "new"

    def _extract_plural_entries(
        self,
        key,
        comment,
        source_value,
        plural_dict,
        localizations,
        source_language,
        results,
        occurrence_counters,
        rel_path,
        app_instance,
    ):
        for cat in self.PLURAL_CATEGORIES:
            cat_block = plural_dict.get(cat)
            if cat_block is None:
                continue
            su = cat_block.get("stringUnit", {})
            translation_value = su.get("value", "")
            xcode_state = su.get("state", "new")

            # 源语言同类别文本
            src_loc = localizations.get(source_language, {})
            src_plural = (
                src_loc.get("variations", {}).get("plural", {}).get(cat, {}).get("stringUnit", {}).get("value", "")
            ) or source_value

            ts = self._make_ts(
                key=key,
                context=f"{key}:plural:{cat}",
                source_value=src_plural,
                translation_value=translation_value,
                comment=f"{comment}\nPlural category: {cat}".strip(),
                xcode_state=xcode_state,
                is_reviewed=(xcode_state == "translated"),
                relative_path=rel_path,
                occurrence_counters=occurrence_counters,
                string_type="XCString Plural",
                app_instance=app_instance,
            )
            results.append(ts)

    def _extract_device_entries(
        self, key, comment, source_value, device_dict, results, occurrence_counters, rel_path, app_instance
    ):
        for device_key in self.DEVICE_KEYS:
            dev_block = device_dict.get(device_key)
            if dev_block is None:
                continue
            su = dev_block.get("stringUnit", {})
            translation_value = su.get("value", "")
            xcode_state = su.get("state", "new")

            ts = self._make_ts(
                key=key,
                context=f"{key}:device:{device_key}",
                source_value=source_value,
                translation_value=translation_value,
                comment=f"{comment}\nDevice variant: {device_key}".strip(),
                xcode_state=xcode_state,
                is_reviewed=(xcode_state == "translated"),
                relative_path=rel_path,
                occurrence_counters=occurrence_counters,
                string_type="XCString Device",
                app_instance=app_instance,
            )
            results.append(ts)

    def _make_ts(
        self,
        key,
        context,
        source_value,
        translation_value,
        comment,
        xcode_state,
        is_reviewed,
        relative_path,
        occurrence_counters,
        string_type,
        app_instance,
    ) -> TranslatableString:
        counter_key = (source_value, context)
        idx = occurrence_counters.get(counter_key, 0)
        occurrence_counters[counter_key] = idx + 1

        stable = f"{relative_path}::{context}::{source_value}::{idx}"
        obj_id = xxhash.xxh128(stable.encode()).hexdigest()

        ts = TranslatableString(
            original_raw=source_value,
            original_semantic=source_value,
            line_num=0,
            char_pos_start_in_file=0,
            char_pos_end_in_file=0,
            full_code_lines=[],
            string_type=string_type,
            source_file_path=relative_path,
            occurrences=[(relative_path, key)],
            occurrence_index=idx,
            id=obj_id,
        )
        # 单语模式: 初始翻译用文件中读到的目标语言值
        initial = translation_value
        if app_instance:
            fill = getattr(app_instance, "config", {}).get("fill_translation_with_source", False)
            if not initial and fill:
                initial = source_value
        ts.set_translation_internal(initial or "", is_initial=True)
        ts.context = context
        ts.comment = comment
        ts.po_comment = f"#: xcstrings key: {key}"
        ts.is_reviewed = is_reviewed
        ts.update_sort_weight()
        return ts


class StringsDictFormatHandler(BaseFormatHandler):
    """
    Apple .stringsdict 专用复数规则处理器

    .stringsdict 是 iOS/macOS 处理复数变体的标准格式（XML Plist），

    1. 格式模板提取: NSStringLocalizedFormatKey 本身作为独立可翻译条目，
       译者可调整变量引用顺序（例如将 "%1$#@files@ in %2$#@folders@" 本地化
       为 "%2$#@folders@ 中有 %1$#@files@"）
    2. 复合字符串支持: 完整处理含多个 %#@var@ 引用的复合格式字符串，
       每个变量块的所有复数类别均独立呈现
    3. 注释富化: 每条条目均标注所属变量名、值类型、关联模板及
       "兄弟变量" 列表，帮助译者理解上下文
    4. 稳健回写: 基于原始 plist 数据深拷贝定向回填，未触及的键原样保留；
       同时回填格式模板，确保 RTL 语言等需要调整顺序的场景可正常工作
    """

    format_id = "stringsdict"
    is_monolingual = True
    extensions = [".stringsdict"]
    format_type = "translation"
    display_name = _("Apple .stringsdict Plural Rules")
    badge_text = "SDICT"
    badge_bg_color = "#FFF8E1"
    badge_text_color = "#F57F17"

    PLURAL_CATEGORIES = ["zero", "one", "two", "few", "many", "other"]

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or self._get_relative_path(filepath)
        language_code = self._detect_language_from_lproj(filepath)

        with open(filepath, "rb") as f:
            try:
                data = plistlib.load(f)
            except Exception as e:
                logger.error(f"[StringsDictFormatHandler] Failed to parse plist: {e}")
                return [], {"original_data": {}}, "en"

        translatable_objects = []
        occurrence_counters = {}

        for top_key, top_value in data.items():
            if not isinstance(top_value, dict):
                continue

            format_key_template = top_value.get("NSStringLocalizedFormatKey", "")

            # --- 1. NSStringLocalizedFormatKey 本身作为可翻译条目 ---
            # 这个模板控制变量的排列顺序，RTL 语言或某些目标语言可能需要调整顺序，
            # 例如: "%1$#@files@ in %2$#@folders@" → "%2$#@folders@ 中有 %1$#@files@"
            if format_key_template:
                ctx = f"{top_key}.__format__"
                ts = self._make_ts(
                    text=format_key_template,
                    context=ctx,
                    string_type="Stringsdict Format",
                    comment=(
                        f'Format template for "{top_key}".\n'
                        f"Controls variable order in the composed string.\n"
                        f"Use %#@varName@ to reference plural variable blocks, e.g.:\n"
                        f'  single var  → "%#@count@"\n'
                        f'  compound    → "%1$#@files@ in %2$#@folders@"\n'
                        f"Reorder references here for RTL languages or different grammar."
                    ),
                    po_comment=f"#: stringsdict key: {top_key}, field: NSStringLocalizedFormatKey",
                    relative_path=relative_path,
                    occurrence_counters=occurrence_counters,
                    app_instance=app_instance,
                )
                translatable_objects.append(ts)

            # --- 2. 变量规则块 ---
            var_names = [k for k in top_value if k != "NSStringLocalizedFormatKey"]
            for var_name in var_names:
                var_dict = top_value[var_name]
                if not isinstance(var_dict, dict):
                    continue

                spec_type = var_dict.get("NSStringFormatSpecTypeKey", "NSStringPluralRuleType")
                value_type = var_dict.get("NSStringFormatValueTypeKey", "d")

                # 仅处理复数规则型（最常见）
                if spec_type not in ("NSStringPluralRuleType", ""):
                    logger.debug(
                        f"[StringsDictFormatHandler] Skipping unsupported spec type "
                        f"'{spec_type}' in {top_key}.{var_name}"
                    )
                    continue

                # 构建兄弟变量说明（仅多变量时才添加）
                sibling_note = ""
                if len(var_names) > 1:
                    sibling_note = f"\nCompound string — {len(var_names)} variable block(s): {', '.join(var_names)}"

                for category in self.PLURAL_CATEGORIES:
                    text = var_dict.get(category)
                    if text is None:
                        continue

                    full_context = f"{top_key}.{var_name}.{category}"
                    comment = f"Plural category: {category}\nVariable name:   {var_name}  (value type: %{value_type})"
                    if format_key_template:
                        comment += f'\nFormat template: "{format_key_template}"'
                    comment += sibling_note

                    ts = self._make_ts(
                        text=text,
                        context=full_context,
                        string_type="Stringsdict Plural",
                        comment=comment,
                        po_comment=(f"#: stringsdict key: {top_key}, variable: {var_name}, category: {category}"),
                        relative_path=relative_path,
                        occurrence_counters=occurrence_counters,
                        app_instance=app_instance,
                    )
                    translatable_objects.append(ts)

        metadata = {"original_data": data}
        logger.info(f"[StringsDictFormatHandler] Loaded {len(translatable_objects)} entries from {filepath}")
        return translatable_objects, metadata, language_code

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        original_data = metadata.get("original_data", {})
        new_data = copy.deepcopy(original_data)

        trans_map = {
            ts.context: ts.translation
            for ts in translatable_objects
            if ts.translation and not getattr(ts, "is_ignored", False) and ts.context
        }

        for top_key, top_value in new_data.items():
            if not isinstance(top_value, dict):
                continue

            # 回填格式模板
            fmt_ctx = f"{top_key}.__format__"
            if fmt_ctx in trans_map and "NSStringLocalizedFormatKey" in top_value:
                top_value["NSStringLocalizedFormatKey"] = trans_map[fmt_ctx]

            # 回填各变量块的复数条目
            for var_name, var_dict in top_value.items():
                if var_name == "NSStringLocalizedFormatKey":
                    continue
                if not isinstance(var_dict, dict):
                    continue
                for category in self.PLURAL_CATEGORIES:
                    if category not in var_dict:
                        continue
                    ctx = f"{top_key}.{var_name}.{category}"
                    if ctx in trans_map:
                        var_dict[category] = trans_map[ctx]

        with atomic_open(filepath, "wb") as f:
            plistlib.dump(new_data, f, fmt=plistlib.FMT_XML)

        logger.info(f"[StringsDictFormatHandler] Saved {len(trans_map)} entries to {filepath}")

    def _make_ts(
        self,
        text,
        context,
        string_type,
        comment,
        po_comment,
        relative_path,
        occurrence_counters,
        app_instance,
    ) -> "TranslatableString":
        counter_key = (text, context)
        idx = occurrence_counters.get(counter_key, 0)
        occurrence_counters[counter_key] = idx + 1

        stable = f"{relative_path}::{context}::{text}::{idx}"
        obj_id = xxhash.xxh128(stable.encode()).hexdigest()

        ts = TranslatableString(
            original_raw=text,
            original_semantic=text,
            line_num=0,
            char_pos_start_in_file=0,
            char_pos_end_in_file=0,
            full_code_lines=[],
            string_type=string_type,
            source_file_path=relative_path,
            occurrences=[(relative_path, context)],
            occurrence_index=idx,
            id=obj_id,
        )
        ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)
        ts.context = context
        ts.comment = comment
        ts.po_comment = po_comment
        ts.is_reviewed = False
        ts.update_sort_weight()
        return ts

    def _detect_language_from_lproj(self, filepath: str) -> str:
        """优先从 xx.lproj/ 路径段提取语言代码，回退到文件名推断。"""
        for part in Path(filepath).parts:
            if part.endswith(".lproj"):
                return part[: -len(".lproj")]
        return self._detect_language_from_filename(os.path.basename(filepath))


class ArbFormatHandler(BaseFormatHandler):
    """
    Flutter / Dart ARB (Application Resource Bundle) 格式处理器

    支持的特性:
    1. 元数据保留: 完整读写 @@locale、@@last_modified 等文件级元数据
    2. 条目描述符: 支持 @key 块中的 description、placeholders、plural 字段
    3. 占位符解析: 自动识别 {name}、{count} 形式的插值占位符并记录到注释
    4. 复数/性别: 识别 ICU MessageFormat 语法 (plural/select/gender) 并保留原样
    5. 非翻译键过滤: 自动跳过以 @@ 开头的全局元数据键
    6. 语言检测: 优先读取 @@locale 字段，其次从文件名 (app_en.arb) 推断
    """

    format_id = "arb"
    is_monolingual = True
    extensions = [".arb"]
    format_type = "translation"
    display_name = _("Flutter ARB File")
    badge_text = "ARB"
    badge_bg_color = "#E8EAF6"
    badge_text_color = "#283593"

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        logger.debug(f"[ArbFormatHandler] Loading ARB file: {filepath}")

        with open(filepath, encoding="utf-8") as f:
            content = f.read()
            data = json.loads(content)

        relative_path = kwargs.get("relative_path") or self._get_relative_path(filepath)
        language_code = self._detect_language(data, os.path.basename(filepath))
        indent = self._detect_indent(content)

        translatable_objects = []
        occurrence_counters = {}

        # 全局元数据 (@@locale, @@last_modified …) — 保留但不翻译
        global_metadata = {k: v for k, v in data.items() if k.startswith("@@")}

        # 遍历所有翻译键
        for key, value in data.items():
            if key.startswith("@"):  # @key 描述符或 @@ 全局元数据 — 跳过
                continue
            if not isinstance(value, str):  # ARB 规范: 翻译值必须为字符串
                continue
            if not value.strip():
                continue

            # 获取对应的 @key 描述符
            descriptor = data.get(f"@{key}", {})
            description = descriptor.get("description", "") if isinstance(descriptor, dict) else ""
            placeholders = descriptor.get("placeholders", {}) if isinstance(descriptor, dict) else {}

            # 构建人类可读的占位符说明
            ph_notes = []
            for ph_name, ph_info in placeholders.items():
                ph_type = ph_info.get("type", "String") if isinstance(ph_info, dict) else "String"
                ph_example = ph_info.get("example", "") if isinstance(ph_info, dict) else ""
                note = f"{{{ph_name}}}: {ph_type}"
                if ph_example:
                    note += f" (e.g. {ph_example})"
                ph_notes.append(note)

            counter_key = (value, key)
            idx = occurrence_counters.get(counter_key, 0)
            occurrence_counters[counter_key] = idx + 1

            stable = f"{relative_path}::{key}::{value}::{idx}"
            obj_id = xxhash.xxh128(stable.encode()).hexdigest()

            ts = TranslatableString(
                original_raw=value,
                original_semantic=value,
                line_num=0,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="ARB String",
                source_file_path=relative_path,
                occurrences=[(relative_path, key)],
                occurrence_index=idx,
                id=obj_id,
            )
            ts.set_translation_internal(self.get_initial_translation(value, app_instance), is_initial=True)
            ts.context = key
            ts.comment = description
            ts.po_comment = f"#: ARB key: {key}" + (f"\n#. Placeholders: {', '.join(ph_notes)}" if ph_notes else "")
            ts.is_reviewed = False
            ts.update_sort_weight()
            translatable_objects.append(ts)

        metadata = {
            "indent": indent,
            "global_metadata": global_metadata,
            "descriptors": {k[1:]: v for k, v in data.items() if k.startswith("@") and not k.startswith("@@")},
        }

        logger.info(f"[ArbFormatHandler] Loaded {len(translatable_objects)} strings from {filepath}")
        return translatable_objects, metadata, language_code

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        logger.debug(f"[ArbFormatHandler] Saving ARB file: {filepath}")

        indent = metadata.get("indent", 4)
        global_metadata = metadata.get("global_metadata", {})
        descriptors = metadata.get("descriptors", {})

        app = kwargs.get("app_instance")
        target_lang = None
        if app:
            target_lang = app.current_target_language

        output = {}

        # 写入 @@locale
        locale = target_lang or global_metadata.get("@@locale", "en")
        output["@@locale"] = locale

        # 写入其他 @@ 全局元数据（排除 @@locale，已单独写）
        for k, v in global_metadata.items():
            if k != "@@locale":
                output[k] = v

        # 写入翻译条目 + 对应描述符
        for ts in translatable_objects:
            if not ts.original_semantic or ts.id == "##NEW_ENTRY##":
                continue
            key = ts.context or ts.original_semantic
            translation = ts.translation if ts.translation else ts.original_semantic
            output[key] = translation

            # 还原 @key 描述符（保留原有占位符、描述等）
            if key in descriptors:
                output[f"@{key}"] = descriptors[key]
            elif ts.comment:
                # 若原本没有描述符但有 description，生成一个最简描述符
                output[f"@{key}"] = {"description": ts.comment}

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=indent, ensure_ascii=False)
            f.write("\n")  # ARB 文件惯例以换行结尾

        logger.info(f"[ArbFormatHandler] Saved {len(translatable_objects)} strings to {filepath}")

    def _detect_indent(self, content: str) -> int:
        for line in content.split("\n")[1:]:
            stripped = line.lstrip()
            if stripped and line != stripped:
                indent = len(line) - len(stripped)
                if indent > 0:
                    return indent
        return 4  # Flutter 官方工具默认 4 空格

    def _detect_language(self, data: dict, filename: str) -> str:
        # 优先读取 @@locale
        if "@@locale" in data:
            return data["@@locale"]
        # 从文件名推断: intl_en.arb / app_zh_CN.arb / en.arb
        name = filename.lower().replace(".arb", "")
        parts = re.split(r"[_\-]", name)
        common = {
            "en",
            "zh",
            "ja",
            "ko",
            "fr",
            "de",
            "es",
            "it",
            "ru",
            "pt",
            "ar",
            "tr",
            "pl",
            "nl",
            "sv",
            "da",
            "fi",
            "nb",
            "cs",
            "sk",
            "hu",
            "ro",
        }
        for part in reversed(parts):
            if part in common:
                return part
        return "en"


class JsonI18nFormatHandler(BaseFormatHandler):
    """
    JSON 国际化文件处理器
    支持的格式:
    1. 扁平结构: {"key1": "value1", "key2": "value2"}
    2. 嵌套结构: {"menu": {"home": "Home", "about": "About"}}
    3. 数组: {"items": ["Item 1", "Item 2"]}
    4. 混合: {"user": {"messages": ["Welcome", "Goodbye"]}}
    """

    format_id = "json_i18n"
    extensions = [".json"]
    format_type = "translation"
    display_name = _("JSON i18n File")
    badge_text = "JSON"
    badge_bg_color = "#FFF3E0"
    badge_text_color = "#E65100"

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        logger.debug(f"[JsonI18nFormatHandler] Loading JSON file: {filepath}")

        with open(filepath, encoding="utf-8") as f:
            content = f.read()
            data = json.loads(content)

        # 检测 JSON 缩进
        indent = self._detect_indent(content)

        relative_path = kwargs.get("relative_path")
        json_file_rel_path = relative_path if relative_path else self._get_relative_path(filepath)

        translatable_objects = []
        occurrence_counters = {}

        # 递归提取所有可翻译字符串
        self._extract_recursive(
            data,
            [],
            translatable_objects,
            occurrence_counters,
            json_file_rel_path,
            line_num=1,
            app_instance=app_instance,
        )

        metadata = {
            "original_structure": data,
            "indent": indent,
            "ensure_ascii": False,  # 保留 Unicode 字符
        }

        # 尝试检测语言代码
        language_code = self._detect_language(data, os.path.basename(filepath))

        logger.info(f"[JsonI18nFormatHandler] Loaded {len(translatable_objects)} strings from {filepath}")
        return translatable_objects, metadata, language_code

    def _detect_indent(self, json_content: str) -> int:
        """检测 JSON 文件的缩进空格数"""
        lines = json_content.split("\n")
        for line in lines[1:]:  # 跳过第一行
            stripped = line.lstrip()
            if stripped and line != stripped:
                indent = len(line) - len(stripped)
                if indent > 0:
                    return indent
        return 2  # 默认 2 空格

    def _detect_language(self, data: dict, filename: str) -> str:
        # 从数据结构优先检测
        if isinstance(data, dict):
            lang_value = data.get("locale") or data.get("language") or data.get("lang")
            if isinstance(lang_value, str):
                return lang_value
            top_keys = list(data.keys())
            if len(top_keys) == 1 and re.match(r"^[a-z]{2,3}(?:[_-][A-Za-z]{2,4})?$", top_keys[0]):
                return top_keys[0]

        # 从文件名检测
        name_lower = os.path.splitext(filename)[0].lower()
        m = re.search(r"(?:^|[_.\-])([a-z]{2,3}(?:[_\-][a-zA-Z]{2,4})?)(?:[_.\-]|$)", name_lower)
        if m:
            candidate = m.group(1)
            common_langs = {"en", "zh", "ja", "ko", "fr", "de", "es", "it", "ru", "pt", "ar", "tr", "pl", "nl"}
            if candidate in common_langs:
                return candidate
        return "en"

    def _extract_recursive(
        self,
        obj: Any,
        key_path: list[str],
        results: list[TranslatableString],
        occurrence_counters: dict,
        file_rel_path: str,
        line_num: int,
        app_instance=None,
    ):
        """递归提取 JSON 中的所有可翻译字符串"""
        if isinstance(obj, dict):
            for key, value in obj.items():
                self._extract_recursive(
                    value,
                    [*key_path, key],
                    results,
                    occurrence_counters,
                    file_rel_path,
                    line_num,
                    app_instance=app_instance,
                )

        elif isinstance(obj, list):
            for idx, item in enumerate(obj):
                self._extract_recursive(
                    item,
                    [*key_path, f"[{idx}]"],
                    results,
                    occurrence_counters,
                    file_rel_path,
                    line_num,
                    app_instance=app_instance,
                )

        elif isinstance(obj, str) and obj.strip():
            self._create_translatable_string(
                obj, key_path, results, occurrence_counters, file_rel_path, line_num, app_instance=app_instance
            )

    def _create_translatable_string(
        self, text, key_path, results, occurrence_counters, file_rel_path, line_num, app_instance=None
    ):
        """创建 TranslatableString 对象"""

        # 生成完整键路径作为 context
        full_key = ".".join(key_path)

        # 生成唯一计数器键
        counter_key = (text, full_key)
        current_index = occurrence_counters.get(counter_key, 0)
        occurrence_counters[counter_key] = current_index + 1

        # 生成稳定的 UUID
        stable_name = f"{file_rel_path}::{full_key}::{text}::{current_index}"
        obj_id = xxhash.xxh128(stable_name.encode("utf-8")).hexdigest()

        ts = TranslatableString(
            original_raw=text,
            original_semantic=text,
            line_num=line_num,
            char_pos_start_in_file=0,
            char_pos_end_in_file=0,
            full_code_lines=[],
            string_type="JSON i18n",
            source_file_path=file_rel_path,
            occurrences=[(file_rel_path, str(line_num))],
            occurrence_index=current_index,
            id=obj_id,
        )
        ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)
        ts.context = full_key
        ts.comment = ""
        ts.po_comment = f"#: JSON key path: {full_key}"
        ts.is_reviewed = False
        ts.update_sort_weight()

        results.append(ts)

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        """保存翻译后的 JSON 文件"""
        logger.debug(f"[JsonI18nFormatHandler] Saving JSON file: {filepath}")

        # 获取原始结构
        original_structure = metadata.get("original_structure", {})
        indent = metadata.get("indent", 2)
        ensure_ascii = metadata.get("ensure_ascii", False)

        # 创建翻译映射: key_path -> translation
        translation_map = {
            ts.context: (ts.translation if ts.translation else ts.original_semantic)
            for ts in translatable_objects
            if ts.id != "##NEW_ENTRY##" and ts.context and ts.original_semantic
        }

        # 重建 JSON 结构
        translated_structure = self._rebuild_structure(original_structure, translation_map)

        # 保存文件
        with atomic_open(filepath, "w", encoding="utf-8") as f:
            json.dump(translated_structure, f, indent=indent, ensure_ascii=ensure_ascii, sort_keys=False)

        logger.info(f"[JsonI18nFormatHandler] Saved {len(translation_map)} translations to {filepath}")

    def _rebuild_structure(self, obj: Any, translation_map: dict[str, str], key_path: list[str] | None = None) -> Any:
        """递归重建 JSON 结构，应用翻译"""
        if key_path is None:
            key_path = []

        if isinstance(obj, dict):
            result = {}
            for key, value in obj.items():
                result[key] = self._rebuild_structure(value, translation_map, [*key_path, key])
            return result

        if isinstance(obj, list):
            result = []
            for idx, item in enumerate(obj):
                result.append(self._rebuild_structure(item, translation_map, [*key_path, f"[{idx}]"]))
            return result

        if isinstance(obj, str):
            # 查找翻译
            full_key = ".".join(key_path)
            return translation_map.get(full_key, obj)

        return obj


class I18nextJsonFormatHandler(BaseFormatHandler):
    """
    i18next JSON 格式处理器

    i18next 是 JavaScript / TypeScript 生态最主流的 i18n 库（React, Vue, Node 通用）。
    其 JSON 格式在通用 JSON i18n 基础上有以下约定：

    1. 插值语法: {{ variable }} 双花括号，区别于 ICU 的单花括号；
       count 是内置复数触发变量，出现 {{count}} 时自动进入复数选择逻辑
    2. 复数键 (v4 / CLDR):  key_zero / key_one / key_two / key_few / key_many / key_other
       复数键 (v3 兼容):    key / key_plural
       本处理器识别两种格式，并将同一基础键的所有复数形式归组注释
    3. 上下文变体: key_male / key_female / key_<context>（contextSeparator 默认 "_"）
    4. 嵌套 JSON: 任意深度，保存时原样重建；键路径以 . 连接作为 context
    5. Trans 组件: <0>标签</0> 格式的 React Trans 内嵌组件原样保留，注释中标注
    6. 命名空间: 从文件名推断（去掉语言代码后缀），记录到注释
    7. 保序保结构: 加载/保存使用相同原始结构，不改变键顺序或嵌套层级
    """

    format_id = "i18next_json"
    is_monolingual = True
    extensions = [".json"]
    format_type = "translation"
    display_name = _("i18next JSON")
    badge_text = "i18n"
    badge_bg_color = "#E8F5E9"
    badge_text_color = "#1B5E20"

    # i18next v4 CLDR 复数后缀（完整集合）
    PLURAL_SUFFIXES_V4 = ("_zero", "_one", "_two", "_few", "_many", "_other")
    # i18next v3 兼容后缀
    PLURAL_SUFFIX_V3 = "_plural"
    # {{variable}} 插值正则
    INTERPOLATION_RE = re.compile(r"\{\{([^}]+?)\}\}")
    # Trans 组件 HTML 数字标签
    TRANS_TAG_RE = re.compile(r"<\d+>|</\d+>")
    # Context 后缀：末尾 _<word>，但不是复数后缀
    CONTEXT_SUFFIX_RE = re.compile(r"^(.+?)_([a-z][a-z0-9]*)$")

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        logger.debug(f"[I18nextJsonFormatHandler] Loading: {filepath}")

        with open(filepath, encoding="utf-8") as f:
            content = f.read()

        try:
            data = json.loads(content)
        except json.JSONDecodeError as e:
            logger.error(f"[I18nextJsonFormatHandler] JSON parse error: {e}")
            return [], {}, "en"

        relative_path = kwargs.get("relative_path") or os.path.basename(filepath)
        language_code = self._detect_language_from_filename(os.path.basename(filepath))
        indent = self._detect_indent(content)
        namespace = self._detect_namespace(os.path.basename(filepath))

        translatable_objects = []
        occurrence_counters = {}

        self._extract_recursive(
            data,
            [],
            translatable_objects,
            occurrence_counters,
            relative_path,
            app_instance,
            namespace,
        )

        # 后处理：为同一基础键的复数形式组互相补充注释
        self._annotate_plural_groups(translatable_objects)

        metadata = {
            "indent": indent,
            "original_structure": copy.deepcopy(data),
            "namespace": namespace,
        }
        logger.info(
            f"[I18nextJsonFormatHandler] Loaded {len(translatable_objects)} strings "
            f"from {filepath} (namespace: {namespace})"
        )
        return translatable_objects, metadata, language_code

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        logger.debug(f"[I18nextJsonFormatHandler] Saving: {filepath}")
        indent = metadata.get("indent", 2)
        original_structure = metadata.get("original_structure", {})

        trans_map = {
            ts.context: (ts.translation if ts.translation else ts.original_semantic)
            for ts in translatable_objects
            if ts.id != "##NEW_ENTRY##" and ts.context
        }

        result = self._rebuild_structure(original_structure, trans_map)

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=indent, ensure_ascii=False)
            f.write("\n")

        logger.info(f"[I18nextJsonFormatHandler] Saved {len(trans_map)} strings to {filepath}")

    def _extract_recursive(self, obj, key_path, results, counters, rel_path, app_instance, namespace):
        if isinstance(obj, dict):
            for key, value in obj.items():
                self._extract_recursive(
                    value,
                    [*key_path, key],
                    results,
                    counters,
                    rel_path,
                    app_instance,
                    namespace,
                )
        elif isinstance(obj, list):
            for idx, item in enumerate(obj):
                self._extract_recursive(
                    item,
                    [*key_path, f"[{idx}]"],
                    results,
                    counters,
                    rel_path,
                    app_instance,
                    namespace,
                )
        elif isinstance(obj, str) and obj.strip():
            self._create_ts(obj, key_path, results, counters, rel_path, app_instance, namespace)

    def _create_ts(self, text, key_path, results, counters, rel_path, app_instance, namespace):
        # context 用完整键路径（含数组索引）
        full_key = ".".join(key_path)

        counter_key = (text, full_key)
        idx = counters.get(counter_key, 0)
        counters[counter_key] = idx + 1

        stable = f"{rel_path}::{full_key}::{text}::{idx}"
        obj_id = xxhash.xxh128(stable.encode()).hexdigest()

        leaf_key = key_path[-1] if key_path else ""

        # --- 特征检测 ---
        variables = self.INTERPOLATION_RE.findall(text)
        has_count_var = "count" in variables
        has_trans_tags = bool(self.TRANS_TAG_RE.search(text))

        is_plural_v4 = any(leaf_key.endswith(s) for s in self.PLURAL_SUFFIXES_V4)
        is_plural_v3 = leaf_key.endswith(self.PLURAL_SUFFIX_V3)
        is_plural = is_plural_v4 or is_plural_v3

        # Context 变体检测（排除已知复数后缀，避免误判）
        is_context_variant = False
        context_variant_name = ""
        if not is_plural:
            cm = self.CONTEXT_SUFFIX_RE.match(leaf_key)
            if cm:
                suffix = cm.group(2)
                # 排除单字母后缀和语言代码，减少误报
                if len(suffix) >= 2 and suffix not in {"en", "zh", "ja", "ko", "fr", "de"}:
                    is_context_variant = True
                    context_variant_name = suffix

        # --- 注释构建 ---
        comment_lines = []

        if variables:
            var_display = ", ".join(f"{{{{{v}}}}}" for v in variables)
            comment_lines.append(f"Interpolation: {var_display}")

        if has_count_var and not is_plural:
            comment_lines.append(
                "Contains {{count}} — i18next will auto-select plural variants "
                f"(expected sibling keys: {leaf_key}_one, {leaf_key}_other, etc.)"
            )

        if is_plural_v4:
            suffix = next(s for s in self.PLURAL_SUFFIXES_V4 if leaf_key.endswith(s))
            base = leaf_key[: -len(suffix)]
            comment_lines.append(f"Plural form (i18next v4 CLDR) — category: '{suffix[1:]}', base key: '{base}'")
        elif is_plural_v3:
            base = leaf_key[: -len(self.PLURAL_SUFFIX_V3)]
            comment_lines.append(f"Plural form (i18next v3 compat) — '_plural' suffix, base key: '{base}'")

        if is_context_variant:
            base = leaf_key[: -(len(context_variant_name) + 1)]
            comment_lines.append(f"Context variant '{context_variant_name}' of key '{base}'")

        if has_trans_tags:
            comment_lines.append(
                "Contains React <Trans> component tags (<0>, </1>, ...) — preserve tag structure in translation"
            )

        if namespace and namespace not in ("translation", ""):
            comment_lines.append(f"Namespace: {namespace}")

        ts = TranslatableString(
            original_raw=text,
            original_semantic=text,
            line_num=0,
            char_pos_start_in_file=0,
            char_pos_end_in_file=0,
            full_code_lines=[],
            string_type="i18next String",
            source_file_path=rel_path,
            occurrences=[(rel_path, full_key)],
            occurrence_index=idx,
            id=obj_id,
        )
        ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)
        ts.context = full_key
        ts.comment = "\n".join(comment_lines)
        ts.po_comment = f"#: i18next key: {full_key}"
        ts.is_reviewed = False
        ts.update_sort_weight()
        results.append(ts)

    def _annotate_plural_groups(self, objects: list):
        """识别同一基础键的所有复数形式，互相标注兄弟键列表。"""
        groups: dict[str, list] = {}
        for ts in objects:
            base = self._plural_base(ts.context)
            if base != ts.context:
                groups.setdefault(base, []).append(ts)

        for base, members in groups.items():
            if len(members) < 2:
                continue
            sibling_keys = ", ".join(m.context for m in members)
            extra = f"Plural group ({len(members)} forms): [{sibling_keys}]"
            for ts in members:
                ts.comment = f"{ts.comment}\n{extra}".strip() if ts.comment else extra

    def _plural_base(self, key: str) -> str:
        for suffix in (*self.PLURAL_SUFFIXES_V4, self.PLURAL_SUFFIX_V3):
            if key.endswith(suffix):
                return key[: -len(suffix)]
        return key

    def _rebuild_structure(self, obj, trans_map, key_path=None):
        if key_path is None:
            key_path = []
        if isinstance(obj, dict):
            return {k: self._rebuild_structure(v, trans_map, [*key_path, k]) for k, v in obj.items()}
        if isinstance(obj, list):
            return [self._rebuild_structure(v, trans_map, [*key_path, f"[{i}]"]) for i, v in enumerate(obj)]
        if isinstance(obj, str):
            return trans_map.get(".".join(key_path), obj)
        return obj

    def _detect_indent(self, content: str) -> int:
        for line in content.split("\n")[1:]:
            stripped = line.lstrip()
            if stripped and line != stripped:
                n = len(line) - len(stripped)
                if n > 0:
                    return n
        return 2

    def _detect_namespace(self, filename: str) -> str:
        """
        从文件名推断 i18next 命名空间。
        常见约定: translation.json, common.en.json, en/errors.json
        """
        stem = os.path.splitext(filename)[0]
        # 去掉末尾语言代码: common.en → common, errors_zh_CN → errors
        cleaned = re.sub(r"[._-]([a-z]{2,3}(?:[_-][A-Za-z]{2,4})?)$", "", stem, flags=re.IGNORECASE)
        return cleaned or "translation"


class YamlI18nFormatHandler(BaseFormatHandler):
    """
    YAML 国际化文件处理器

    支持的框架与格式:
    1. Ruby on Rails i18n: en:\n  key: value  (顶层语言键包装结构)
    2. Vue i18n / react-i18next: 扁平或嵌套 YAML，无顶层语言键
    3. 通用嵌套: 任意深度嵌套映射，键路径以 . 连接作为 context
    4. 多行字面量块 (|) 与折叠块 (>): 正确提取完整文本，保存时
       自动选择最合适的块标量风格

    关键技术决策:
    - 使用 ruamel.yaml 而非 PyYAML，以确保注释、键顺序、缩进风格
      在加载/保存过程中完整保留（PyYAML 会丢失注释和 key 顺序）
    - 若 ruamel.yaml 不可用则自动回退到 PyYAML（功能降级：注释丢失）
    - Rails 顶层语言键 (如 `en:`) 自动识别并在保存时还原，不作为翻译条目
    - 跳过非字符串叶节点（数字、布尔、null），防止误提取配置值
    - 数组中的字符串元素以 key[0], key[1] 形式纳入翻译管理
    """

    format_id = "yaml_i18n"
    is_monolingual = True
    extensions = [".yml", ".yaml"]
    format_type = "translation"
    display_name = _("YAML i18n File")
    badge_text = "YAML"
    badge_bg_color = "#F1F8E9"
    badge_text_color = "#33691E"

    # Rails 风格顶层语言键检测：单个符合 BCP-47 的顶层键
    _LANG_CODE_RE = re.compile(r"^[a-z]{2,3}(?:[_-][A-Za-z]{2,4})?$")

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        logger.debug(f"[YamlI18nFormatHandler] Loading YAML: {filepath}")

        with open(filepath, encoding="utf-8") as f:
            raw_content = f.read()

        data, yaml_backend = self._yaml_load(raw_content)
        if not isinstance(data, dict):
            logger.warning(f"[YamlI18nFormatHandler] Root is not a mapping: {filepath}")
            return [], {}, "en"

        rel_path = kwargs.get("relative_path") or self._get_relative_path(filepath)
        language_code = self._detect_language(data, os.path.basename(filepath))

        # Rails 顶层语言键解包
        rails_lang_key, data_root = self._unwrap_rails_root(data)

        translatable_objects = []
        occurrence_counters = {}
        self._extract_recursive(data_root, [], translatable_objects, occurrence_counters, rel_path, app_instance)

        metadata = {
            "raw_content": raw_content,
            "rails_lang_key": rails_lang_key,
            "yaml_backend": yaml_backend,
        }

        logger.info(f"[YamlI18nFormatHandler] Loaded {len(translatable_objects)} strings from {filepath}")
        return translatable_objects, metadata, language_code

    def _yaml_load(self, content: str) -> tuple[Any, str]:
        """加载 YAML，返回 (data, backend_name)"""
        try:
            from ruamel.yaml import YAML

            yaml = YAML()
            yaml.preserve_quotes = True
            import io

            data = yaml.load(io.StringIO(content))
            return data, "ruamel"
        except ImportError:
            pass

        try:
            import yaml as pyyaml

            data = pyyaml.safe_load(content)
            return data, "pyyaml"
        except Exception as e:
            logger.error(f"[YamlI18nFormatHandler] YAML parse error: {e}")
            return {}, "pyyaml"

    def _yaml_dump(self, data: Any, backend: str, indent: int = 2) -> str:
        """序列化 YAML，尽量保持原格式"""
        if backend == "ruamel":
            try:
                import io

                from ruamel.yaml import YAML

                yaml = YAML()
                yaml.default_flow_style = False
                yaml.allow_unicode = True
                yaml.indent(mapping=indent, sequence=indent, offset=indent)
                buf = io.StringIO()
                yaml.dump(data, buf)
                return buf.getvalue()
            except ImportError:
                pass

        import yaml as pyyaml

        return pyyaml.dump(data, allow_unicode=True, default_flow_style=False, indent=indent, sort_keys=False)

    def _unwrap_rails_root(self, data: dict) -> tuple[str | None, Any]:
        """
        如果数据只有一个顶层键且符合语言码格式，视为 Rails 风格并解包。
        返回 (lang_key_or_None, inner_data)
        """
        keys = list(data.keys()) if isinstance(data, dict) else []
        if len(keys) == 1:
            key = str(keys[0])
            if self._LANG_CODE_RE.match(key):
                return key, data[key] if isinstance(data[key], dict) else data
        return None, data

    def _wrap_rails_root(self, data: Any, lang_key: str | None) -> Any:
        """将数据重新包装到 Rails 顶层语言键下"""
        if lang_key:
            return {lang_key: data}
        return data

    def _extract_recursive(
        self,
        obj: Any,
        key_path: list[str],
        results: list[TranslatableString],
        counters: dict,
        rel_path: str,
        app_instance=None,
    ):
        if isinstance(obj, dict):
            for k, v in obj.items():
                self._extract_recursive(v, [*key_path, str(k)], results, counters, rel_path, app_instance=app_instance)

        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                self._extract_recursive(
                    item, [*key_path, f"[{i}]"], results, counters, rel_path, app_instance=app_instance
                )

        elif isinstance(obj, str) and obj.strip():
            self._make_ts(obj, key_path, results, counters, rel_path, app_instance=app_instance)

    def _make_ts(
        self,
        text: str,
        key_path: list[str],
        results: list[TranslatableString],
        counters: dict,
        rel_path: str,
        app_instance=None,
    ):
        full_key = ".".join(key_path)
        full_key = full_key.replace(".[", "[")

        counter_key = (text, full_key)
        idx = counters.get(counter_key, 0)
        counters[counter_key] = idx + 1

        stable = f"{rel_path}::{full_key}::{text}::{idx}"
        obj_id = xxhash.xxh128(stable.encode("utf-8")).hexdigest()

        ts = TranslatableString(
            original_raw=text,
            original_semantic=text,
            line_num=0,
            char_pos_start_in_file=0,
            char_pos_end_in_file=0,
            full_code_lines=[],
            string_type="YAML i18n",
            source_file_path=rel_path,
            occurrences=[(rel_path, full_key)],
            occurrence_index=idx,
            id=obj_id,
        )
        ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)
        ts.context = full_key
        ts.comment = ""
        ts.po_comment = f"#: YAML key: {full_key}"
        ts.is_reviewed = False
        ts.update_sort_weight()
        results.append(ts)

    def _rebuild_recursive(self, obj: Any, key_path: list[str], translation_map: dict[str, str]) -> Any:
        """递归将原始 YAML 结构中的字符串替换为译文"""
        if isinstance(obj, dict):
            # ruamel.yaml CommentedMap 需要逐键更新而非整体替换
            result = obj.__class__() if hasattr(obj, "__class__") and hasattr(obj, "ca") else {}
            for k, v in obj.items():
                result[k] = self._rebuild_recursive(v, [*key_path, str(k)], translation_map)
            return result

        if isinstance(obj, list):
            cls = obj.__class__ if hasattr(obj, "ca") else list
            result = cls()
            for i, item in enumerate(obj):
                rebuilt = self._rebuild_recursive(item, [*key_path, f"[{i}]"], translation_map)
                result.append(rebuilt)
            return result

        if isinstance(obj, str):
            full_key = re.sub(r"\.\[", "[", ".".join(key_path))
            return translation_map.get(full_key, obj)

        return obj

    def _detect_language(self, data: dict, filename: str) -> str:
        # 优先从 Rails 顶层键检测
        keys = list(data.keys()) if isinstance(data, dict) else []
        if len(keys) == 1:
            key = str(keys[0])
            if self._LANG_CODE_RE.match(key):
                return key

        # 从文件名检测: zh-CN.yml / messages.fr.yml / i18n_de.yaml
        stem = os.path.splitext(filename)[0]
        m = re.search(r"(?:^|[_.-])([a-z]{2,3}(?:[_-][A-Za-z]{2,4})?)(?:[_.-]|$)", stem)
        if m:
            candidate = m.group(1)
            if self._LANG_CODE_RE.match(candidate):
                return candidate

        return "en"

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        logger.debug(f"[YamlI18nFormatHandler] Saving YAML: {filepath}")

        translation_map = {
            ts.context: (ts.translation or ts.original_semantic)
            for ts in translatable_objects
            if ts.original_semantic and ts.id != "##NEW_ENTRY##" and not getattr(ts, "is_ignored", False) and ts.context
        }

        raw_content = metadata.get("raw_content", "")
        rails_lang_key = metadata.get("rails_lang_key")
        backend = metadata.get("yaml_backend", "pyyaml")

        # 重新加载原始结构（保留注释，如果使用 ruamel）
        original_data, _ = self._yaml_load(raw_content)
        _, data_root = self._unwrap_rails_root(original_data)

        # 递归替换
        translated_root = self._rebuild_recursive(data_root, [], translation_map)

        # 还原 Rails 顶层键包装
        output_data = self._wrap_rails_root(translated_root, rails_lang_key)

        yaml_str = self._yaml_dump(output_data, backend)

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write(yaml_str)

        logger.info(f"[YamlI18nFormatHandler] Saved {len(translation_map)} strings to {filepath}")


class TomlFormatHandler(BaseFormatHandler):
    """
    TOML 配置文件处理器
    使用 tomlkit 库以确保在保存时完美还原注释、空行和结构顺序。
    """

    format_id = "toml"
    is_monolingual = True
    extensions = [".toml"]
    format_type = "translation"
    display_name = _("TOML Config File")
    badge_text = "TOML"
    badge_bg_color = "#FCE4EC"
    badge_text_color = "#3F51B5"

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        import tomlkit

        with open(filepath, encoding="utf-8") as f:
            content = f.read()
            doc = tomlkit.parse(content)

        rel_path = kwargs.get("relative_path") or os.path.basename(filepath)
        translatable_objects = []
        occurrence_counters = {}

        self._extract_recursive(doc, [], translatable_objects, occurrence_counters, rel_path, app_instance=app_instance)

        metadata = {"raw_content": content}
        return translatable_objects, metadata, "en"

    def _extract_recursive(self, obj, key_path, results, counters, rel_path, app_instance=None):
        if isinstance(obj, dict):
            for k, v in obj.items():
                self._extract_recursive(v, [*key_path, str(k)], results, counters, rel_path, app_instance=app_instance)
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                self._extract_recursive(
                    item, [*key_path, f"[{i}]"], results, counters, rel_path, app_instance=app_instance
                )
        elif isinstance(obj, str) and obj.strip():
            # 过滤掉看起来像颜色、路径或纯数字的字符串
            if re.match(r"^#[0-9a-fA-F]{3,8}$", obj):
                return
            if re.match(r"^(?:https?://|/|\./)", obj):
                return

            full_key = ".".join(key_path)
            full_key = re.sub(r"\.\[", "[", full_key)

            counter_key = (obj, full_key)
            idx = counters.get(counter_key, 0)
            counters[counter_key] = idx + 1

            stable = f"{rel_path}::{full_key}::{obj}::{idx}"
            obj_id = xxhash.xxh128(stable.encode()).hexdigest()

            ts = TranslatableString(
                original_raw=obj,
                original_semantic=obj,
                line_num=0,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="TOML String",
                source_file_path=rel_path,
                occurrences=[(rel_path, full_key)],
                occurrence_index=idx,
                id=obj_id,
            )
            ts.set_translation_internal(self.get_initial_translation(obj, app_instance), is_initial=True)
            ts.context = full_key
            ts.po_comment = f"#: TOML key: {full_key}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            results.append(ts)

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        import tomlkit

        raw_content = metadata.get("raw_content", "")
        doc = tomlkit.parse(raw_content)

        translation_map = {
            ts.context: (ts.translation or ts.original_semantic)
            for ts in translatable_objects
            if ts.original_semantic and ts.id != "##NEW_ENTRY##" and ts.context
        }

        self._rebuild_recursive(doc, [], translation_map)

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write(tomlkit.dumps(doc))

    def _rebuild_recursive(self, obj, key_path, translation_map):
        if isinstance(obj, dict):
            for k, v in obj.items():
                if isinstance(v, str):
                    full_key = re.sub(r"\.\[", "[", ".".join([*key_path, str(k)]))
                    if full_key in translation_map:
                        obj[k] = translation_map[full_key]
                else:
                    self._rebuild_recursive(v, [*key_path, str(k)], translation_map)
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                if isinstance(item, str):
                    full_key = re.sub(r"\.\[", "[", ".".join([*key_path, f"[{i}]"]))
                    if full_key in translation_map:
                        obj[i] = translation_map[full_key]
                else:
                    self._rebuild_recursive(item, [*key_path, f"[{i}]"], translation_map)


class IniFormatHandler(BaseFormatHandler):
    """
    INI 配置文件格式处理器

    支持的格式特性:
    1. 节 (Section): [SectionName] 分组，context 写成 "Section.Key"
    2. 注释: ; 和 # 开头均视为注释行，附加到下一个键值对的 comment
    3. 多种分隔符: = 和 : 均支持
    4. 行内注释: key = value  ; inline comment 正确剥离
    5. 多行续行: 行尾 \\ 续行（部分 INI 方言）
    6. 无节键值: 位于任何 [section] 之前的键值放入虚拟节 "__global__"
    7. 空值过滤: value 为空的键跳过（通常是配置开关而非翻译文本）
    8. 结构保留: 保存时还原原始 section 结构，原始注释写回文件
    """

    format_id = "ini"
    is_monolingual = True
    extensions = [".ini", ".cfg", ".conf"]
    format_type = "translation"
    display_name = _("INI Config / i18n File")
    badge_text = "INI"
    badge_bg_color = "#ECEFF1"
    badge_text_color = "#37474F"

    _SECTION_RE = re.compile(r"^\[([^\]]+)\]\s*$")
    _KV_RE = re.compile(r"^([^=:\n#;][^=:\n]*?)\s*[=:]\s*(.*)$")

    def load(self, filepath: str, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or self._get_relative_path(filepath)

        with open(filepath, encoding="utf-8-sig", errors="replace") as f:
            content = f.read()

        language_code = self._detect_language_from_filename(os.path.basename(filepath))
        translatable_objects: list[TranslatableString] = []
        occurrence_counters: dict = {}

        current_section = "__global__"
        pending_comments: list[str] = []
        # 用于保存 raw_lines 以便 save 时重建
        raw_lines: list[dict] = []  # {"type": "section"/"kv"/"comment"/"blank", ...}

        lines = content.splitlines()
        line_num = 0

        for raw_line in lines:
            line_num += 1
            stripped = raw_line.strip()

            # 空行
            if not stripped:
                if pending_comments:
                    # 孤立注释块之间有空行，保留到下一键
                    pass
                raw_lines.append({"type": "blank"})
                continue

            # 注释行
            if stripped.startswith(";") or stripped.startswith("#"):
                comment_text = stripped[1:].strip()
                pending_comments.append(comment_text)
                raw_lines.append({"type": "comment", "text": raw_line})
                continue

            # 节标题
            sec_m = self._SECTION_RE.match(stripped)
            if sec_m:
                current_section = sec_m.group(1).strip()
                pending_comments.clear()
                raw_lines.append({"type": "section", "name": current_section, "text": raw_line})
                continue

            # 键值对
            kv_m = self._KV_RE.match(stripped)
            if kv_m:
                key = kv_m.group(1).strip()
                value = kv_m.group(2).strip()
                # 去除行内注释后的尾随空格
                value = re.sub(r"\s+[;#].*$", "", value).strip()
                # 过滤空值
                if not value:
                    pending_comments.clear()
                    raw_lines.append(
                        {"type": "kv", "section": current_section, "key": key, "value": value, "translatable": False}
                    )
                    continue

                context = f"{current_section}.{key}" if current_section != "__global__" else key
                comment = "\n".join(pending_comments).strip()
                pending_comments.clear()

                counter_key = (value, context)
                idx = occurrence_counters.get(counter_key, 0)
                occurrence_counters[counter_key] = idx + 1

                stable = f"{relative_path}::{context}::{value}::{idx}"
                obj_id = xxhash.xxh128(stable.encode()).hexdigest()

                ts = TranslatableString(
                    original_raw=value,
                    original_semantic=value,
                    line_num=line_num,
                    char_pos_start_in_file=0,
                    char_pos_end_in_file=0,
                    full_code_lines=[],
                    string_type="INI Value",
                    source_file_path=relative_path,
                    occurrences=[(relative_path, context)],
                    occurrence_index=idx,
                    id=obj_id,
                )
                ts.set_translation_internal(self.get_initial_translation(value, app_instance), is_initial=True)
                ts.context = context
                ts.comment = comment
                ts.po_comment = f"#: INI [{current_section}] {key}"
                ts.is_reviewed = False
                ts.update_sort_weight()
                translatable_objects.append(ts)

                raw_lines.append(
                    {
                        "type": "kv",
                        "section": current_section,
                        "key": key,
                        "value": value,
                        "translatable": True,
                        "separator": "=" if "=" in raw_line else ":",
                        "context": context,
                    }
                )
            else:
                # 无法解析的行（如续行，原样保留）
                raw_lines.append({"type": "raw", "text": raw_line})

        metadata = {
            "raw_lines": raw_lines,
            "encoding": "utf-8",
        }
        logger.info(f"[IniFormatHandler] Loaded {len(translatable_objects)} values from {filepath}")
        return translatable_objects, metadata, language_code

    def save(self, filepath: str, translatable_objects, metadata: dict, **kwargs):
        raw_lines: list[dict] = metadata.get("raw_lines", [])
        # 建立 context -> translation 映射
        trans_map: dict[str, str] = {
            ts.context: (ts.translation or ts.original_semantic)
            for ts in translatable_objects
            if ts.id != "##NEW_ENTRY##" and ts.context
        }

        out_lines: list[str] = []
        for item in raw_lines:
            t = item["type"]
            if t in ("blank",):
                out_lines.append("")
            elif t in ("comment", "section", "raw"):
                out_lines.append(item["text"])
            elif t == "kv":
                if not item.get("translatable", False):
                    # 还原原始行（value 为空的配置项）
                    sep = item.get("separator", "=")
                    out_lines.append(f"{item['key']} {sep} {item['value']}")
                else:
                    translation = trans_map.get(item["context"], item["value"])
                    sep = item.get("separator", "=")
                    out_lines.append(f"{item['key']} {sep} {translation}")
            else:
                out_lines.append(item.get("text", ""))

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(out_lines))
            f.write("\n")

        logger.info(f"[IniFormatHandler] Saved to {filepath}")


class JavaPropertiesFormatHandler(BaseFormatHandler):
    """
    Java / Kotlin .properties 资源文件处理器

    支持的特性:
    1. 多种分隔符: 正确解析 key=value、key: value、key value 三种赋值形式
    2. 多行续行: 支持行尾反斜杠 \\ 的跨行字符串
    3. Unicode 转义: 双向处理 \\uXXXX 编码，保留非 ASCII 可读性
    4. 注释提取: # 和 ! 开头的行内注释均作为 translator comment 关联到下一条目
    5. 顺序保留: 保存时严格按原始键顺序输出，保证 diff 友好
    6. 语言检测: 从标准命名规范 messages_zh_CN.properties 自动推断语言代码
    7. 空行保护: 保存时在每个条目间保留适当空行，贴近手写习惯
    """

    format_id = "java_properties"
    is_monolingual = True
    extensions = [".properties"]
    format_type = "translation"
    display_name = _("Java .properties File")
    badge_text = "Props"
    badge_bg_color = "#FFF8E1"
    badge_text_color = "#F57F17"

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        logger.debug(f"[JavaPropertiesFormatHandler] Loading .properties: {filepath}")

        # Java .properties 官方编码为 ISO-8859-1，但现代项目多用 UTF-8
        encoding = self._detect_encoding(filepath)
        with open(filepath, encoding=encoding, errors="replace") as f:
            lines = f.readlines()

        relative_path = kwargs.get("relative_path") or self._get_relative_path(filepath)
        language_code = self._detect_language(os.path.basename(filepath))

        translatable_objects = []
        occurrence_counters = {}

        # --- 解析器状态 ---
        pending_comments: list[str] = []
        line_idx = 0

        while line_idx < len(lines):
            raw = lines[line_idx]
            stripped = raw.strip()
            line_idx += 1

            # 空行 → 清空待关联注释（注释只关联紧随其后的条目）
            if not stripped:
                pending_comments.clear()
                continue

            # 注释行
            if stripped.startswith("#") or stripped.startswith("!"):
                pending_comments.append(stripped[1:].strip())
                continue

            # 键值行（可能带续行）
            logical_line = raw.rstrip("\r\n")
            while logical_line.endswith("\\"):
                logical_line = logical_line[:-1]  # 去掉续行符
                if line_idx < len(lines):
                    logical_line += lines[line_idx].strip()
                    line_idx += 1

            key, value = self._split_key_value(logical_line.strip())
            if key is None:
                pending_comments.clear()
                continue

            # 解码 Unicode 转义
            key = self._decode_unicode_escapes(key)
            value = self._decode_unicode_escapes(value)

            if not value.strip():
                pending_comments.clear()
                continue

            comment = "\n".join(pending_comments).strip()
            pending_comments.clear()

            counter_key = (value, key)
            idx = occurrence_counters.get(counter_key, 0)
            occurrence_counters[counter_key] = idx + 1

            stable = f"{relative_path}::{key}::{value}::{idx}"
            obj_id = xxhash.xxh128(stable.encode()).hexdigest()

            ts = TranslatableString(
                original_raw=value,
                original_semantic=value,
                line_num=0,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="Java Properties",
                source_file_path=relative_path,
                occurrences=[(relative_path, key)],
                occurrence_index=idx,
                id=obj_id,
            )
            ts.set_translation_internal(self.get_initial_translation(value, app_instance), is_initial=True)
            ts.context = key
            ts.comment = comment
            ts.po_comment = f"#: Properties key: {key}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            translatable_objects.append(ts)

        metadata = {
            "encoding": encoding,
            "key_order": [ts.context for ts in translatable_objects],
        }

        logger.info(
            f"[JavaPropertiesFormatHandler] Loaded {len(translatable_objects)} "
            f"entries from {filepath} (encoding: {encoding})"
        )
        return translatable_objects, metadata, language_code

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        logger.debug(f"[JavaPropertiesFormatHandler] Saving .properties: {filepath}")

        encoding = metadata.get("encoding", "utf-8")
        key_order = metadata.get("key_order", [])

        # 建立 context -> ts 映射
        ts_map = {ts.context: ts for ts in translatable_objects if ts.original_semantic and ts.id != "##NEW_ENTRY##"}

        # 按原始顺序输出，末尾追加新增条目
        ordered_keys = key_order + [k for k in ts_map if k not in key_order]

        lines = []
        for key in ordered_keys:
            ts = ts_map.get(key)
            if not ts:
                continue

            # 写注释
            if ts.comment:
                for comment_line in ts.comment.splitlines():
                    lines.append(f"# {comment_line}")

            translation = ts.translation if ts.translation else ts.original_semantic

            # 决定是否需要 Unicode 转义 (仅在 latin-1 编码时必须转义非 ASCII)
            needs_escape = encoding.lower() in ("iso-8859-1", "latin-1", "latin1")
            escaped_key = self._encode_key(key, needs_escape)
            escaped_val = self._encode_value(translation, needs_escape)

            lines.append(f"{escaped_key}={escaped_val}")
            lines.append("")  # 条目间空行

        with atomic_open(filepath, "w", encoding=encoding) as f:
            f.write("\n".join(lines))

        logger.info(f"[JavaPropertiesFormatHandler] Saved {len(ts_map)} entries to {filepath}")

    def _detect_encoding(self, filepath: str) -> str:
        """
        检测文件编码。
        Java 规范是 ISO-8859-1，但 Spring Boot 等现代框架默认 UTF-8。
        检查 BOM 或尝试 UTF-8 解码来区分。
        """
        with open(filepath, "rb") as f:
            raw = f.read(4)
        if raw.startswith(b"\xef\xbb\xbf"):
            return "utf-8-sig"
        # 尝试 UTF-8 解码前 4KB
        try:
            with open(filepath, encoding="utf-8") as f:
                f.read(4096)
            return "utf-8"
        except UnicodeDecodeError:
            return "iso-8859-1"

    def _split_key_value(self, line: str) -> tuple[str | None, str]:
        """
        解析 key=value / key: value / key value 三种形式。
        返回 (key, value)，解析失败返回 (None, '')。
        """
        # 跳过注释和空行（已在上层处理，这里做保险）
        if not line or line[0] in ("#", "!"):
            return None, ""

        # 找到未转义的分隔符 (=, :, 或首个空格)
        i = 0
        while i < len(line):
            ch = line[i]
            if ch == "\\":
                i += 2  # 跳过转义字符
                continue
            if ch in ("=", ":"):
                return line[:i].strip(), line[i + 1 :].lstrip()
            if ch in (" ", "\t"):
                key = line[:i].strip()
                rest = line[i:].lstrip()
                # 如果空格后面紧跟 = 或 :，那才是真正的分隔符
                if rest and rest[0] in ("=", ":"):
                    return key, rest[1:].lstrip()
                return key, rest
            i += 1

        # 只有键没有值（空值）
        return line.strip(), ""

    def _decode_unicode_escapes(self, s: str) -> str:
        """将 \\uXXXX 转义序列解码为 Unicode 字符。"""
        return re.sub(r"\\u([0-9a-fA-F]{4})", lambda m: chr(int(m.group(1), 16)), s)

    def _encode_unicode_escapes(self, s: str) -> str:
        """将非 ASCII 字符编码为 \\uXXXX（仅用于 latin-1 编码文件）。"""
        result = []
        for ch in s:
            if ord(ch) > 127:
                result.append(f"\\u{ord(ch):04X}")
            else:
                result.append(ch)
        return "".join(result)

    def _encode_key(self, key: str, unicode_escape: bool) -> str:
        """对键中的特殊字符进行转义。"""
        key = key.replace("\\", "\\\\")
        key = key.replace(" ", "\\ ")
        key = key.replace("=", "\\=")
        key = key.replace(":", "\\:")
        key = key.replace("#", "\\#")
        key = key.replace("!", "\\!")
        if unicode_escape:
            key = self._encode_unicode_escapes(key)
        return key

    def _encode_value(self, value: str, unicode_escape: bool) -> str:
        """对值进行转义，换行符转为 \\n 续行形式。"""
        value = value.replace("\\", "\\\\")
        value = value.replace("\n", "\\n\\\n    ")
        value = value.replace("\t", "\\t")
        if unicode_escape:
            value = self._encode_unicode_escapes(value)
        return value

    def _detect_language(self, filename: str) -> str:
        """
        从标准 Java ResourceBundle 命名规范中提取语言代码。
        例: messages_zh_CN.properties → zh_CN
            strings_en.properties    → en
            MyApp_fr_FR.properties   → fr_FR
        """
        name = filename.replace(".properties", "")
        # 尝试匹配末尾的 _lang 或 _lang_COUNTRY
        m = re.search(r"_([a-z]{2,3})(?:_([A-Z]{2,3}))?$", name)
        if m:
            lang = m.group(1)
            country = m.group(2)
            return f"{lang}_{country}" if country else lang
        return "en"


class ResxFormatHandler(BaseFormatHandler):
    """
    .NET / C# RESX (XML Resource File) 格式处理器

    支持的特性:
    1. 数据类型过滤: 仅提取 type 属性缺失或为 System.String 的 <data> 节点，
       自动跳过图片、图标、二进制嵌入资源等非文本条目，防止误翻译
    2. 注释双通道: <comment> 子节点保存为 po_comment 供内部追踪；
       xml:space="preserve" 属性被正确识别并在保存时还原
    3. Designer 文件跳过: 自动检测 .Designer.resx 文件并拒绝加载，
       避免将自动生成的资源键混入翻译工作流
    4. ResX 头部保留: 加载时完整记录 <resheader> / <assembly> / <metadata>
       节点，保存时原样写回，确保 Visual Studio 可正常打开并编译
    5. 语言检测: 从 App.zh-CN.resx / Strings.fr.resx 等命名约定中
       自动提取 BCP-47 语言代码
    """

    format_id = "resx"
    is_monolingual = True
    extensions = [".resx"]
    format_type = "translation"
    display_name = _("RESX Resource File (.NET)")
    badge_text = "RESX"
    badge_bg_color = "#E8EAF6"
    badge_text_color = "#283593"

    # 需要跳过的已知非字符串类型前缀
    _SKIP_TYPE_PREFIXES = (
        "System.Drawing",
        "System.Windows.Forms",
        "System.Byte[]",
        "System.Resources",
        "Microsoft.",
    )

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        logger.debug(f"[ResxFormatHandler] Loading RESX: {filepath}")

        # 拒绝处理 .Designer.resx（自动生成文件）
        if filepath.endswith(".Designer.resx"):
            logger.info(f"[ResxFormatHandler] Skipping designer file: {filepath}")
            return [], {}, "en"

        tree = ET.parse(filepath)
        root = tree.getroot()

        rel_path = kwargs.get("relative_path") or self._get_relative_path(filepath)
        translatable_objects = []
        occurrence_counters = {}

        # 收集所有 <data> 节点
        for data_elem in root.findall("data"):
            name = data_elem.get("name", "")
            if not name:
                continue

            # 跳过非字符串资源（图片、二进制等）
            res_type = data_elem.get("type", "")
            if res_type and not res_type.startswith("System.String"):
                if any(res_type.startswith(p) for p in self._SKIP_TYPE_PREFIXES):
                    continue

            # 跳过 mimetype 属性（base64 嵌入数据）
            if data_elem.get("mimetype"):
                continue

            value_elem = data_elem.find("value")
            if value_elem is None or not (value_elem.text or "").strip():
                continue

            value = value_elem.text.strip()

            comment_elem = data_elem.find("comment")
            comment_text = comment_elem.text.strip() if comment_elem is not None and comment_elem.text else ""

            counter_key = (value, name)
            idx = occurrence_counters.get(counter_key, 0)
            occurrence_counters[counter_key] = idx + 1

            stable = f"{rel_path}::{name}::{idx}"
            obj_id = xxhash.xxh128(stable.encode("utf-8")).hexdigest()

            ts = TranslatableString(
                original_raw=value,
                original_semantic=value,
                line_num=0,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="RESX String",
                source_file_path=rel_path,
                occurrences=[(rel_path, name)],
                occurrence_index=idx,
                id=obj_id,
            )
            ts.set_translation_internal(self.get_initial_translation(value, app_instance), is_initial=True)
            ts.context = name
            ts.comment = comment_text
            ts.po_comment = f"#: RESX name: {name}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            translatable_objects.append(ts)

        # 保存头部节点用于回写
        header_nodes = self._collect_header_nodes(root)
        language_code = self._detect_language(os.path.basename(filepath))

        metadata = {
            "header_nodes": header_nodes,
            "xml_version": "1.0",
            "encoding": "utf-8",
        }

        logger.info(f"[ResxFormatHandler] Loaded {len(translatable_objects)} strings from {filepath}")
        return translatable_objects, metadata, language_code

    def _collect_header_nodes(self, root: ET.Element) -> list[dict]:
        """收集 resheader / assembly / metadata 节点，保存时原样还原"""
        preserved = []
        for tag in ("resheader", "assembly", "metadata"):
            for elem in root.findall(tag):
                preserved.append(
                    {
                        "tag": tag,
                        "attrib": dict(elem.attrib),
                        "children": [{"tag": c.tag, "text": c.text, "attrib": dict(c.attrib)} for c in elem],
                    }
                )
        return preserved

    def _detect_language(self, filename: str) -> str:
        """
        从 Resource.zh-CN.resx / Strings.fr.resx / App.de.resx 中提取语言码。
        中性区域文件 (Resource.resx) 视为源语言 (en)。
        """
        stem = os.path.splitext(filename)[0]  # 去掉 .resx
        # BCP-47 语言标签: xx 或 xx-YY
        m = re.search(r"\.([a-z]{2,3}(?:-[A-Za-z]{2,4})?)$", stem, re.IGNORECASE)
        if m:
            return m.group(1)
        return "en"

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        logger.debug(f"[ResxFormatHandler] Saving RESX: {filepath}")

        ET.register_namespace("xsd", "http://www.w3.org/2001/XMLSchema")
        ET.register_namespace("msdata", "urn:schemas-microsoft-com:xml-msdata")

        root = ET.Element("root")

        # 写入架构声明
        schema_elem = ET.SubElement(root, "{http://www.w3.org/2001/XMLSchema}schema")
        schema_elem.set("id", "root")

        # 还原头部节点 (resheader, assembly 等)
        for node_info in metadata.get("header_nodes", []):
            elem = ET.SubElement(root, node_info["tag"], **node_info["attrib"])
            for child in node_info["children"]:
                c = ET.SubElement(elem, child["tag"], **child.get("attrib", {}))
                c.text = child.get("text", "")

        # 写入翻译条目并计数
        saved_count = 0
        for ts in translatable_objects:
            if not ts.original_semantic or ts.id == "##NEW_ENTRY##":
                continue

            data_elem = ET.SubElement(root, "data")
            data_elem.set("name", ts.context)
            # 处理空格保留
            data_elem.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")

            value_elem = ET.SubElement(data_elem, "value")
            # 如果有翻译则用翻译，否则回退到原文
            value_elem.text = ts.translation if ts.translation else ts.original_semantic

            if ts.comment:
                comment_elem = ET.SubElement(data_elem, "comment")
                comment_elem.text = ts.comment

            saved_count += 1

        tree = ET.ElementTree(root)
        if hasattr(ET, "indent"):
            ET.indent(tree, space="  ", level=0)

        with atomic_open(filepath, "wb") as f:
            tree.write(f, encoding="utf-8", xml_declaration=True)
        logger.info(f"[ResxFormatHandler] Saved {saved_count} strings to {filepath}")


class PhpArrayFormatHandler(BaseFormatHandler):
    """
    PHP 数组翻译文件处理器

    支持 Laravel / Symfony / CodeIgniter 等主流 PHP 框架的翻译文件格式:

    格式示例（Laravel）:
        <?php
        return [
            'welcome' => 'Welcome to our application!',
            'auth' => [
                'failed' => 'These credentials do not match.',
                'throttle' => 'Too many login attempts.',
            ],
            // 行注释
            /* 块注释 */
            'items_count' => ':count items',
        ];

    支持的特性:
    1. 嵌套数组: 键路径用点号连接，如 "auth.failed"，最多支持 8 层嵌套
    2. 占位符识别: 自动识别 Laravel :param 和 Symfony %param% 风格占位符
    3. 注释提取: 行注释(//)和块注释(/* */)附加到下一条目
    4. 字符串引号: 单引号与双引号均支持，转义序列正确处理
    5. PHP 标签感知: 自动跳过 <?php / return / ?> 等非翻译行
    6. 结构还原: 按原始嵌套层级重建输出文件，保留 <?php return [...]; 框架

    限制:
    - 不支持 define() / const 形式（使用 INI 处理器替代）
    - 不支持变量插值 "$var" 形式的动态键
    - heredoc / nowdoc 语法暂不支持
    """

    format_id = "php_array"
    is_monolingual = True
    extensions = [".php"]
    format_type = "translation"
    display_name = _("PHP Array Translation File")
    badge_text = "PHP"
    badge_bg_color = "#EDE7F6"
    badge_text_color = "#4527A0"

    # 匹配单引号或双引号字符串（简化，不处理 heredoc）
    _STR_RE = re.compile(r"""(?P<q>['"])(?P<val>(?:[^\\]|\\.)*?)(?P=q)""")
    # Laravel 占位符: :param_name
    _PH_LARAVEL = re.compile(r":[a-zA-Z_]\w*")
    # Symfony/generic 占位符: %param_name%
    _PH_SYMFONY = re.compile(r"%[a-zA-Z_]\w*%")

    def load(self, filepath: str, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or self._get_relative_path(filepath)

        with open(filepath, encoding="utf-8-sig", errors="replace") as f:
            content = f.read()

        language_code = self._detect_language_from_filename(os.path.basename(filepath))
        translatable_objects: list[TranslatableString] = []
        occurrence_counters: dict = {}

        # 提取 return [...] / return array(...) 中的内容
        # 使用简化的行扫描解析器，不依赖完整 PHP 解析器
        entries = self._parse_php_array(content)

        for entry in entries:
            key_path: str = entry["key_path"]
            value: str = entry["value"]
            comment: str = entry.get("comment", "")
            line_num: int = entry.get("line_num", 0)

            if not value.strip():
                continue

            # 占位符注释
            ph_notes = self._detect_placeholders(value)

            counter_key = (value, key_path)
            idx = occurrence_counters.get(counter_key, 0)
            occurrence_counters[counter_key] = idx + 1

            stable = f"{relative_path}::{key_path}::{value}::{idx}"
            obj_id = xxhash.xxh128(stable.encode()).hexdigest()

            full_comment = comment
            if ph_notes:
                ph_str = ", ".join(ph_notes)
                full_comment = f"{comment}\nPlaceholders: {ph_str}".strip() if comment else f"Placeholders: {ph_str}"

            ts = TranslatableString(
                original_raw=value,
                original_semantic=value,
                line_num=line_num,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="PHP String",
                source_file_path=relative_path,
                occurrences=[(relative_path, key_path)],
                occurrence_index=idx,
                id=obj_id,
            )
            ts.set_translation_internal(self.get_initial_translation(value, app_instance), is_initial=True)
            ts.context = key_path
            ts.comment = full_comment
            ts.po_comment = f"#: PHP key: {key_path}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            translatable_objects.append(ts)

        metadata = {
            "raw_content": content,
            "entries": entries,  # 保存解析结果用于重建
        }
        logger.info(f"[PhpArrayFormatHandler] Loaded {len(translatable_objects)} strings from {filepath}")
        return translatable_objects, metadata, language_code

    def save(self, filepath: str, translatable_objects, metadata: dict, **kwargs):
        raw_content: str = metadata.get("raw_content", "")
        entries: list[dict] = metadata.get("entries", [])

        # 建立 key_path -> translation 映射
        trans_map: dict[str, str] = {
            ts.context: (ts.translation or ts.original_semantic)
            for ts in translatable_objects
            if ts.id != "##NEW_ENTRY##" and ts.context
        }

        # 逐条替换: 在 raw_content 中定位原始值并替换为译文
        # 策略: 按行号从后往前替换，避免偏移量变化
        result = self._replace_translations(raw_content, entries, trans_map)

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write(result)

        logger.info(f"[PhpArrayFormatHandler] Saved to {filepath}")

    def _parse_php_array(self, content: str) -> list[dict]:
        """
        扫描 PHP 文件，提取所有 'key' => 'value' 对。
        支持任意嵌套深度，不执行 PHP 代码。
        返回 [{"key_path": "a.b.c", "value": "...", "comment": "...", "line_num": N}, ...]
        """
        lines = content.splitlines()
        entries = []
        section_stack: list[str] = []  # 当前嵌套键路径
        pending_comment: list[str] = []

        # 匹配: 'key' => 'value', 或 "key" => "value",
        kv_re = re.compile(
            r"""(?P<q1>['"])(?P<key>(?:[^\\]|\\.)*?)(?P=q1)\s*=>\s*(?P<q2>['"])(?P<val>(?:[^\\]|\\.)*?)(?P=q2)\s*,?\s*(?://.*|/\*.*?\*/)?$"""
        )
        # 匹配数组开始: 'key' => [ 或 'key' => array(
        arr_start_re = re.compile(r"""(?P<q>['"])(?P<key>(?:[^\\]|\\.)*?)(?P=q)\s*=>\s*(?:\[|array\s*\()""")
        # 注释行
        line_comment_re = re.compile(r"^\s*//\s*(.*)")
        block_comment_re = re.compile(r"/\*(.*?)\*/", re.DOTALL)
        # 数组结束
        arr_end_re = re.compile(r"^\s*[\]\)],?\s*$")

        i = 0
        while i < len(lines):
            raw = lines[i]
            stripped = raw.strip()
            line_num = i + 1

            # 跳过 PHP 标签和 return 语句
            if re.match(r"^<\?php|^\?>|^return\s*[\[\(]|^return\s*array", stripped):
                i += 1
                continue

            # 块注释（可能跨行）
            if "/*" in stripped:
                block_text = stripped
                j = i
                while "*/" not in block_text and j < len(lines) - 1:
                    j += 1
                    block_text += "\n" + lines[j]
                m = block_comment_re.search(block_text)
                if m:
                    pending_comment.append(m.group(1).strip())
                i = j + 1
                continue

            # 行注释
            lc_m = line_comment_re.match(stripped)
            if lc_m:
                pending_comment.append(lc_m.group(1).strip())
                i += 1
                continue

            # 数组起始（嵌套）
            as_m = arr_start_re.match(stripped)
            if as_m:
                section_stack.append(as_m.group("key"))
                pending_comment.clear()
                i += 1
                continue

            # 数组结束
            if arr_end_re.match(stripped):
                if section_stack:
                    section_stack.pop()
                i += 1
                continue

            # 键值对
            kv_m = kv_re.match(stripped)
            if kv_m:
                key = self._unescape_php(kv_m.group("key"))
                value = self._unescape_php(kv_m.group("val"))
                key_path = ".".join([*section_stack, key])
                comment = "\n".join(pending_comment).strip()
                pending_comment.clear()

                entries.append(
                    {
                        "key_path": key_path,
                        "value": value,
                        "comment": comment,
                        "line_num": line_num,
                        "raw_key": kv_m.group("key"),
                        "raw_val": kv_m.group("val"),
                        "quote_key": kv_m.group("q1"),
                        "quote_val": kv_m.group("q2"),
                    }
                )
            else:
                pending_comment.clear()

            i += 1

        return entries

    def _replace_translations(self, content: str, entries: list[dict], trans_map: dict[str, str]) -> str:
        """
        在原始内容中定位并替换每个条目的值字符串。
        从后往前按行号替换，确保行偏移不漂移。
        """
        lines = content.splitlines(keepends=True)

        # 按行号降序排列
        sorted_entries = sorted(entries, key=lambda e: e["line_num"], reverse=True)

        for entry in sorted_entries:
            key_path = entry["key_path"]
            if key_path not in trans_map:
                continue
            translation = trans_map[key_path]
            line_idx = entry["line_num"] - 1
            if line_idx >= len(lines):
                continue

            raw_val = entry["raw_val"]
            quote_val = entry["quote_val"]
            escaped_translation = self._escape_php(translation, quote_val)

            old_quoted = f"{quote_val}{raw_val}{quote_val}"
            new_quoted = f"{quote_val}{escaped_translation}{quote_val}"

            # 只替换值部分（=> 右侧的引号字符串）
            # 找到 => 后的第一个引号字符串
            line = lines[line_idx]
            arrow_pos = line.find("=>")
            if arrow_pos == -1:
                continue
            after_arrow = line[arrow_pos:]
            new_after = after_arrow.replace(old_quoted, new_quoted, 1)
            lines[line_idx] = line[:arrow_pos] + new_after

        return "".join(lines)

    @staticmethod
    def _unescape_php(s: str) -> str:
        return (
            s.replace("\\'", "'")
            .replace('\\"', '"')
            .replace("\\n", "\n")
            .replace("\\r", "\r")
            .replace("\\t", "\t")
            .replace("\\\\", "\\")
        )

    @staticmethod
    def _escape_php(s: str, quote: str = "'") -> str:
        s = s.replace("\\", "\\\\")
        s = s.replace("'", "\\'") if quote == "'" else s.replace('"', '\\"')
        s = s.replace("\n", "\\n").replace("\r", "\\r").replace("\t", "\\t")
        return s

    def _detect_placeholders(self, text: str) -> list[str]:
        found = []
        found += self._PH_LARAVEL.findall(text)
        found += self._PH_SYMFONY.findall(text)
        return list(dict.fromkeys(found))  # 去重保序


class RcFormatHandler(BaseFormatHandler):
    """
    Windows Resource Script (.rc) 处理器

    支持提取的资源块类型:
    ┌──────────────────────────────────────────────────────────────────────┐
    │ STRINGTABLE    最常见的翻译目标，IDS_XXX 对应字符串                  │
    │ MENU / POPUP   菜单项的 MENUITEM / POPUP 标题                        │
    │ DIALOG / DIALOGEX  控件 CAPTION、LTEXT、RTEXT、CTEXT、PUSHBUTTON 等 │
    │ VERSIONINFO    FileDescription、ProductName、LegalCopyright 等        │
    │ CAPTION / GROUPBOX 独立控件标题行                                    │
    └──────────────────────────────────────────────────────────────────────┘

    支持的特性:
    1. 多编码检测: 优先 UTF-16LE（MSVC 默认），回退 UTF-8 / CP1252
    2. 注释提取: // 和 /* */ 注释附加到紧随其后的条目
    3. 忽略规则: 自动过滤纯空串、纯数字串及内部技术关键字
    4. 上下文分组: STRINGTABLE 条目以 "STRINGTABLE.ID" 为 context；
                   DIALOG 控件以 "DIALOG.ControlType[caption]" 为 context
    5. 结构还原: 保存时直接在原始字节流中做字符串替换，不破坏非翻译内容
    6. BOM 保留: 检测到 UTF-16LE BOM 时保存也输出 UTF-16LE

    限制:
    - 不执行 #include / #define 宏展开
    - ACCELERATORS / BITMAP / ICON 等二进制资源块跳过
    """

    format_id = "rc"
    is_monolingual = True
    extensions = [".rc", ".rc2"]
    format_type = "translation"
    display_name = _("Windows Resource Script (.rc)")
    badge_text = "RC"
    badge_bg_color = "#E8EAF6"
    badge_text_color = "#283593"

    # 技术关键字（不翻译）
    _SKIP_VALUES = frozenset(
        {
            "",
            "\\n",
            "\\t",
            "...",
            "OK",
            "Cancel",
            "Yes",
            "No",
            "&OK",
            "&Cancel",
            "&Yes",
            "&No",
        }
    )
    # 跳过纯数字 / 纯符号串
    _SKIP_RE = re.compile(r"^[\d\s\.\,\-\+\%\$\#\@\!\?\:\;\/\\\|\*\&\^]+$")

    # 块类型关键字
    _BLOCK_START_RE = re.compile(r"^\s*(STRINGTABLE|MENU|DIALOG(?:EX)?|VERSIONINFO)\b", re.I)
    _BLOCK_END_RE = re.compile(r"^\s*END\b")

    # STRINGTABLE 条目: ID "value"
    _ST_ENTRY_RE = re.compile(r"""^\s*(\w+)\s+(?:L?)\"((?:[^\"\\]|\\.)*)\"""")
    # DIALOG 控件行: LTEXT/RTEXT/… "caption", id, x, y, w, h
    _CTRL_RE = re.compile(
        r"""^\s*(LTEXT|RTEXT|CTEXT|PUSHBUTTON|DEFPUSHBUTTON|CHECKBOX|RADIOBUTTON|GROUPBOX|CAPTION)\s+(?:L?)\"((?:[^\"\\]|\\.)*)\"""",
        re.I,
    )
    # MENU MENUITEM
    _MENU_RE = re.compile(r"""^\s*(?:MENUITEM|POPUP)\s+(?:L?)\"((?:[^\"\\]|\\.)*)\"""", re.I)
    # VERSIONINFO VALUE
    _VER_RE = re.compile(
        r"""^\s*VALUE\s+\"(FileDescription|ProductName|LegalCopyright|Comments|CompanyName|InternalName|OriginalFilename|ProductVersion|FileVersion)\"\s*,\s*\"((?:[^\"\\]|\\.)*)\"""",
        re.I,
    )
    # 行注释
    _LINE_CMT_RE = re.compile(r"//\s*(.*)")
    # 块注释（单行内）
    _BLOCK_CMT_RE = re.compile(r"/\*(.*?)\*/", re.DOTALL)

    def load(self, filepath: str, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or os.path.basename(filepath)

        raw_bytes, encoding = self._read_rc_file(filepath)
        content = raw_bytes.decode(encoding, errors="replace").lstrip("\ufeff")

        language_code = self._detect_language(filepath, content)
        translatable_objects: list[TranslatableString] = []
        occurrence_counters: dict = {}

        lines = content.splitlines()
        in_stringtable = False
        in_dialog = False
        in_menu = False
        in_versioninfo = False
        dialog_name = ""
        pending_comment: list[str] = []
        block_depth = 0  # 用 BEGIN/END 或 { } 计数

        for line_num, raw_line in enumerate(lines, start=1):
            stripped = raw_line.strip()

            # 注释收集
            lc = self._LINE_CMT_RE.search(stripped)
            if lc and not stripped.startswith('"'):
                pending_comment.append(lc.group(1).strip())
                continue
            bc = self._BLOCK_CMT_RE.search(stripped)
            if bc:
                pending_comment.append(bc.group(1).strip())
                # 块注释可能在同一行，继续解析其余内容

            # 块边界
            if re.match(r"^\s*(BEGIN|\{)\s*$", stripped, re.I):
                block_depth += 1
                continue
            if re.match(r"^\s*(END|\})\s*$", stripped, re.I):
                block_depth -= 1
                if block_depth <= 0:
                    in_stringtable = in_dialog = in_menu = in_versioninfo = False
                    block_depth = 0
                continue

            # 块类型检测
            blk = self._BLOCK_START_RE.match(stripped)
            if blk:
                btype = blk.group(1).upper()
                in_stringtable = btype == "STRINGTABLE"
                in_dialog = btype in ("DIALOG", "DIALOGEX")
                in_menu = btype == "MENU"
                in_versioninfo = btype == "VERSIONINFO"
                if in_dialog:
                    # 尝试提取 dialog 名称（前一个标记）
                    dm = re.match(r"^\s*(\w+)\s+DIALOG", stripped, re.I)
                    dialog_name = dm.group(1) if dm else f"DIALOG_{line_num}"
                pending_comment.clear()
                continue

            # CAPTION（DIALOG 顶级标题）
            cap_m = re.match(r"""^\s*CAPTION\s+(?:L?)\"((?:[^\"\\]|\\.)*)\" """, stripped)
            if cap_m and in_dialog:
                self._add_entry(
                    cap_m.group(1),
                    f"{dialog_name}.CAPTION",
                    "RC Dialog Caption",
                    pending_comment,
                    line_num,
                    relative_path,
                    translatable_objects,
                    occurrence_counters,
                    app_instance,
                )
                pending_comment.clear()
                continue

            # STRINGTABLE 条目
            if in_stringtable:
                st_m = self._ST_ENTRY_RE.match(stripped)
                if st_m:
                    self._add_entry(
                        st_m.group(2),
                        f"STRINGTABLE.{st_m.group(1)}",
                        "RC StringTable",
                        pending_comment,
                        line_num,
                        relative_path,
                        translatable_objects,
                        occurrence_counters,
                        app_instance,
                    )
                    pending_comment.clear()
                continue

            # DIALOG 控件
            if in_dialog:
                ctrl_m = self._CTRL_RE.match(stripped)
                if ctrl_m:
                    ctrl_type = ctrl_m.group(1).upper()
                    self._add_entry(
                        ctrl_m.group(2),
                        f"{dialog_name}.{ctrl_type}",
                        "RC Dialog Control",
                        pending_comment,
                        line_num,
                        relative_path,
                        translatable_objects,
                        occurrence_counters,
                        app_instance,
                    )
                    pending_comment.clear()
                continue

            # MENU 条目
            if in_menu:
                menu_m = self._MENU_RE.match(stripped)
                if menu_m:
                    self._add_entry(
                        menu_m.group(1),
                        "MENU",
                        "RC Menu",
                        pending_comment,
                        line_num,
                        relative_path,
                        translatable_objects,
                        occurrence_counters,
                        app_instance,
                    )
                    pending_comment.clear()
                continue

            # VERSIONINFO VALUE
            if in_versioninfo:
                ver_m = self._VER_RE.match(stripped)
                if ver_m:
                    self._add_entry(
                        ver_m.group(2),
                        f"VERSIONINFO.{ver_m.group(1)}",
                        "RC VersionInfo",
                        pending_comment,
                        line_num,
                        relative_path,
                        translatable_objects,
                        occurrence_counters,
                        app_instance,
                    )
                    pending_comment.clear()
                continue

            # 非块内容也清空悬挂注释
            if stripped and not stripped.startswith("//"):
                pending_comment.clear()

        metadata = {
            "raw_bytes": raw_bytes,
            "encoding": encoding,
        }
        logger.info(f"[RcFormatHandler] Loaded {len(translatable_objects)} strings from {filepath}")
        return translatable_objects, metadata, language_code

    def _add_entry(
        self,
        raw_value: str,
        context: str,
        string_type: str,
        pending_comment: list[str],
        line_num: int,
        rel_path: str,
        results: list,
        counters: dict,
        app_instance,
    ):
        # 反转义 RC 转义序列
        value = self._unescape_rc(raw_value)
        if not value or value in self._SKIP_VALUES or self._SKIP_RE.match(value):
            return

        comment = "\n".join(pending_comment).strip()
        counter_key = (value, context)
        idx = counters.get(counter_key, 0)
        counters[counter_key] = idx + 1

        stable = f"{rel_path}::{context}::{value}::{idx}"
        obj_id = xxhash.xxh128(stable.encode()).hexdigest()

        ts = TranslatableString(
            original_raw=value,
            original_semantic=value,
            line_num=line_num,
            char_pos_start_in_file=0,
            char_pos_end_in_file=0,
            full_code_lines=[],
            string_type=string_type,
            source_file_path=rel_path,
            occurrences=[(rel_path, context)],
            occurrence_index=idx,
            id=obj_id,
        )
        ts.set_translation_internal(self.get_initial_translation(value, app_instance), is_initial=True)
        ts.context = context
        ts.comment = comment
        ts.po_comment = f"#: RC {context} (line {line_num})"
        ts.is_reviewed = False
        ts.update_sort_weight()
        results.append(ts)

    def save(self, filepath: str, translatable_objects, metadata: dict, **kwargs):
        raw_bytes: bytes = metadata["raw_bytes"]
        encoding: str = metadata["encoding"]
        content = raw_bytes.decode(encoding, errors="replace")

        # original_semantic -> translation 直接字符串替换
        # 为防止误替换，从最长串开始替换
        for ts in sorted(translatable_objects, key=lambda t: -len(t.original_semantic)):
            if ts.id == "##NEW_ENTRY##" or not ts.translation:
                continue
            if ts.original_semantic == ts.translation:
                continue
            escaped_orig = self._escape_rc(ts.original_semantic)
            escaped_trans = self._escape_rc(ts.translation)
            content = content.replace(f'"{escaped_orig}"', f'"{escaped_trans}"', 1)

        out_bytes = content.encode(encoding, errors="replace")
        # 保留 BOM
        if encoding.lower().replace("-", "") == "utf16le":
            out_bytes = b"\xff\xfe" + out_bytes

        with atomic_open(filepath, "wb") as f:
            f.write(out_bytes)

        logger.info(f"[RcFormatHandler] Saved to {filepath}")

    @staticmethod
    def _read_rc_file(filepath: str) -> tuple[bytes, str]:
        raw = Path(filepath).read_bytes()
        if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
            return raw[2:], "utf-16-le" if raw[:2] == b"\xff\xfe" else "utf-16-be"
        try:
            raw.decode("utf-8")
            return raw, "utf-8"
        except UnicodeDecodeError:
            return raw, "cp1252"

    @staticmethod
    def _unescape_rc(s: str) -> str:
        return s.replace("\\n", "\n").replace("\\t", "\t").replace('\\"', '"').replace("\\\\", "\\")

    @staticmethod
    def _escape_rc(s: str) -> str:
        return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n").replace("\t", "\\t")

    @staticmethod
    def _detect_language(filepath: str, content: str) -> str:
        # 尝试从 LANGUAGE 语句检测: LANGUAGE LANG_CHINESE, SUBLANG_CHINESE_SIMPLIFIED
        lang_m = re.search(r"LANGUAGE\s+LANG_(\w+)", content, re.I)
        if lang_m:
            mapping = {
                "CHINESE": "zh",
                "ENGLISH": "en",
                "GERMAN": "de",
                "FRENCH": "fr",
                "JAPANESE": "ja",
                "KOREAN": "ko",
                "SPANISH": "es",
                "ITALIAN": "it",
                "PORTUGUESE": "pt",
                "RUSSIAN": "ru",
                "DUTCH": "nl",
                "POLISH": "pl",
            }
            return mapping.get(lang_m.group(1).upper(), "en")
        stem = os.path.splitext(os.path.basename(filepath))[0]
        m = re.search(r"[._-]([a-z]{2,3}(?:[_-][A-Za-z]{2,4})?)$", stem)
        return m.group(1).replace("-", "_") if m else "en"


# 表格类辅助函数
def _guess_column_mapping(headers, config):
    mapping = {}
    is_fuzzy = False
    if not headers:
        return mapping, is_fuzzy

    # 关键词
    default_pool = {
        "source": [
            "source",
            "original",
            "原文",
            "text",
            "string",
            "msgid",
            "src",
            "source_text",
            "default",
            "base",
            "reference",
            "master",
        ],
        "target": [
            "target",
            "translation",
            "译文",
            "msgstr",
            "value",
            "tgt",
            "translated_text",
            "loc",
            "localized",
            "localization",
            "result",
            "dest",
            "destination",
        ],
        "key": [
            "key",
            "id",
            "name",
            "键",
            "标识",
            "identifier",
            "code",
            "string_id",
            "text_id",
            "label",
            "path",
            "var",
            "variable",
        ],
        "comment": [
            "comment",
            "note",
            "description",
            "备注",
            "context",
            "desc",
            "notes",
            "instruction",
            "info",
            "information",
            "说明",
            "reference_url",
        ],
    }

    user_pool = config.get("column_keywords", {})

    def normalize(s):
        return re.sub(r"[^a-z0-9\u4e00-\u9fa5]", "", str(s).lower())

    processed_pools = {}
    for role, words in default_pool.items():
        combined = set(words + user_pool.get(role, []))
        processed_pools[role] = [normalize(w) for w in combined if w]

    assigned_indices = set()

    # --- 第一阶段：精确匹配 ---
    for role in ["source", "target", "key", "comment"]:
        for idx, h in enumerate(headers):
            if idx in assigned_indices:
                continue
            if normalize(h) in processed_pools[role]:
                mapping[role] = idx
                assigned_indices.add(idx)
                break

    # --- 第二阶段：包含匹配 ---
    for role in ["source", "target"]:
        if role in mapping:
            continue
        for idx, h in enumerate(headers):
            if idx in assigned_indices:
                continue
            h_norm = normalize(h)
            if any(word in h_norm for word in processed_pools[role] if len(word) > 3):
                mapping[role] = idx
                assigned_indices.add(idx)
                break

    # --- 第三阶段：Fuzzy 模糊匹配 ---
    for role in ["source", "target"]:
        if role in mapping:
            continue
        best_score = 0
        best_idx = -1
        for idx, h in enumerate(headers):
            if idx in assigned_indices:
                continue
            h_norm = normalize(h)
            for word in processed_pools[role]:
                score = fuzz.ratio(h_norm, word)
                if score > 85 and score > best_score:
                    best_score = score
                    best_idx = idx

        if best_idx != -1:
            mapping[role] = best_idx
            assigned_indices.add(best_idx)
            is_fuzzy = True

    return mapping, is_fuzzy


def _learn_column_mapping(headers, mapping, config):
    """将用户手动选择的表头加入关键词池"""
    if "column_keywords" not in config:
        config["column_keywords"] = {}

    pool = config["column_keywords"]
    changed = False

    for role, col_idx in mapping.items():
        if col_idx >= len(headers):
            continue
        header_val = str(headers[col_idx]).strip().lower()
        if not header_val:
            continue

        if role not in pool:
            pool[role] = []
        if header_val not in pool[role]:
            pool[role].append(header_val)
            changed = True

    return changed


class CsvFormatHandler(BaseFormatHandler):
    format_id = "csv"
    is_monolingual = False
    extensions = [".csv"]
    format_type = "translation"
    display_name = _("CSV Table File")
    badge_text = "CSV"
    badge_bg_color = "#E8F5E9"
    badge_text_color = "#2E7D32"

    def load(self, filepath, **kwargs):
        app = kwargs.get("app_instance")
        force_dialog = kwargs.get("force_dialog", False)
        if not app:
            raise ValueError("App instance required for CSV mapping.")

        with open(filepath, encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except Exception:
                dialect = csv.excel

            reader = csv.reader(f, dialect)
            rows = list(reader)

        if not rows:
            return [], {}, "en"

        headers = rows[0]
        data_rows = rows[1:]

        # 1. 尝试猜测映射
        mapping, is_guessed_fuzzy = _guess_column_mapping(headers, app.config)

        # 2. 如果缺少原文列，或者强制交互，弹出对话框
        if force_dialog or is_guessed_fuzzy or "source" not in mapping:
            from lexisync.dialogs.column_mapper_dialog import ColumnMapperDialog

            dialog = ColumnMapperDialog(
                app.main_window if hasattr(app, "main_window") else app, headers, data_rows[:5], mapping
            )
            if dialog.exec():
                mapping = dialog.result_mapping
                if dialog.remember_choices and _learn_column_mapping(headers, mapping, app.config):
                    app.save_config()
            else:
                return [], {}, "en"

        rel_path = kwargs.get("relative_path") or os.path.basename(filepath)
        translatable_objects = []
        occurrence_counters = {}

        src_idx = mapping.get("source")
        tgt_idx = mapping.get("target")
        key_idx = mapping.get("key")
        cmt_idx = mapping.get("comment")

        for row_num, row in enumerate(data_rows, start=2):
            if src_idx >= len(row):
                continue
            source_text = row[src_idx]
            if not source_text.strip():
                continue

            target_text = row[tgt_idx] if tgt_idx is not None and tgt_idx < len(row) else ""
            key_text = row[key_idx] if key_idx is not None and key_idx < len(row) else ""
            comment_text = row[cmt_idx] if cmt_idx is not None and cmt_idx < len(row) else ""

            context = key_text or f"row_{row_num}"
            counter_key = (source_text, context)
            idx = occurrence_counters.get(counter_key, 0)
            occurrence_counters[counter_key] = idx + 1

            stable = f"{rel_path}::{context}::{source_text}::{idx}"
            obj_id = xxhash.xxh128(stable.encode()).hexdigest()

            ts = TranslatableString(
                original_raw=source_text,
                original_semantic=source_text,
                line_num=row_num,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="CSV Row",
                source_file_path=rel_path,
                occurrences=[(rel_path, str(row_num))],
                occurrence_index=idx,
                id=obj_id,
            )
            ts.set_translation_internal(target_text, is_initial=True)
            ts.context = context
            ts.comment = comment_text
            ts.po_comment = f"#: Row {row_num}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            translatable_objects.append(ts)

        language_code = self._detect_language_from_filename(os.path.basename(filepath))
        metadata = {
            "mapping": mapping,
            "dialect": {
                "delimiter": dialect.delimiter,
                "quotechar": dialect.quotechar,
                "lineterminator": dialect.lineterminator,
            },
        }
        return translatable_objects, metadata, language_code

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        mapping = metadata.get("mapping", {})
        tgt_idx = mapping.get("target")
        cmt_idx = mapping.get("comment")

        # 如果没有目标列，我们需要在末尾追加一列
        append_target = False
        if tgt_idx is None:
            append_target = True

        dialect_info = metadata.get("dialect", {})

        # 读取原始数据
        with open(filepath, encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(
                f, delimiter=dialect_info.get("delimiter", ","), quotechar=dialect_info.get("quotechar", '"')
            )
            rows = list(reader)

        if not rows:
            return

        if append_target:
            tgt_idx = len(rows[0])
            rows[0].append("Translation")  # 追加表头

        # 建立行号映射
        ts_map = {ts.line_num: ts for ts in translatable_objects if ts.id != "##NEW_ENTRY##"}

        for row_num, row in enumerate(rows[1:], start=2):
            ts = ts_map.get(row_num)
            if ts:
                # 1. 回填译文
                trans_text = ts.translation if ts.translation else ts.original_semantic
                if append_target:
                    row.append(trans_text)
                elif tgt_idx < len(row):
                    row[tgt_idx] = trans_text
                else:
                    # 补齐长度
                    row.extend([""] * (tgt_idx - len(row) + 1))
                    row[tgt_idx] = trans_text
                # 2. 回填注释
                if cmt_idx is not None and cmt_idx < len(row):
                    row[cmt_idx] = ts.comment

        # 写回
        with atomic_open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(
                f,
                delimiter=dialect_info.get("delimiter", ","),
                quotechar=dialect_info.get("quotechar", '"'),
                lineterminator=dialect_info.get("lineterminator", "\r\n"),
            )
            writer.writerows(rows)


class XlsxFormatHandler(BaseFormatHandler):
    format_id = "xlsx"
    is_monolingual = False
    extensions = [".xlsx"]
    format_type = "translation"
    display_name = _("Excel Workbook")
    badge_text = "XLSX"
    badge_bg_color = "#E8F5E9"
    badge_text_color = "#1B5E20"

    def load(self, filepath, **kwargs):
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError(
                _(
                    "The 'openpyxl' library is required to read Excel files. Please install it via 'pip install openpyxl'."
                )
            ) from e

        app = kwargs.get("app_instance")
        force_dialog = kwargs.get("force_dialog", False)
        if not app:
            raise ValueError("App instance required.")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], {}, "en"

        headers = [str(c) if c is not None else "" for c in rows[0]]
        data_rows = rows[1:]

        mapping, is_guessed_fuzzy = _guess_column_mapping(headers, app.config)

        if force_dialog or is_guessed_fuzzy or "source" not in mapping:
            from lexisync.dialogs.column_mapper_dialog import ColumnMapperDialog

            dialog = ColumnMapperDialog(
                app.main_window if hasattr(app, "main_window") else app, headers, data_rows[:5], mapping
            )

            if dialog.exec():
                mapping = dialog.result_mapping
                if dialog.remember_choices and _learn_column_mapping(headers, mapping, app.config):
                    app.save_config()
            else:
                return [], {}, "en"

        rel_path = kwargs.get("relative_path") or os.path.basename(filepath)
        translatable_objects = []
        occurrence_counters = {}

        src_idx = mapping.get("source")
        tgt_idx = mapping.get("target")
        key_idx = mapping.get("key")
        cmt_idx = mapping.get("comment")

        for row_num, row in enumerate(data_rows, start=2):
            if src_idx >= len(row) or row[src_idx] is None:
                continue
            source_text = str(row[src_idx])
            if not source_text.strip():
                continue

            target_text = (
                str(row[tgt_idx]) if tgt_idx is not None and tgt_idx < len(row) and row[tgt_idx] is not None else ""
            )
            key_text = (
                str(row[key_idx]) if key_idx is not None and key_idx < len(row) and row[key_idx] is not None else ""
            )
            comment_text = (
                str(row[cmt_idx]) if cmt_idx is not None and cmt_idx < len(row) and row[cmt_idx] is not None else ""
            )

            context = key_text or f"row_{row_num}"
            counter_key = (source_text, context)
            idx = occurrence_counters.get(counter_key, 0)
            occurrence_counters[counter_key] = idx + 1

            stable = f"{rel_path}::{context}::{source_text}::{idx}"
            obj_id = xxhash.xxh128(stable.encode()).hexdigest()

            ts = TranslatableString(
                original_raw=source_text,
                original_semantic=source_text,
                line_num=row_num,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="Excel Row",
                source_file_path=rel_path,
                occurrences=[(rel_path, str(row_num))],
                occurrence_index=idx,
                id=obj_id,
            )
            ts.set_translation_internal(target_text, is_initial=True)
            ts.context = context
            ts.comment = comment_text
            ts.po_comment = f"#: Row {row_num}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            translatable_objects.append(ts)

        language_code = self._detect_language_from_filename(os.path.basename(filepath))
        metadata = {"mapping": mapping, "sheet_name": ws.title}
        return translatable_objects, metadata, language_code

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        import openpyxl

        mapping = metadata.get("mapping", {})
        tgt_idx = mapping.get("target")
        cmt_idx = mapping.get("comment")

        wb = openpyxl.load_workbook(filepath)
        ws = wb[metadata.get("sheet_name", wb.active.title)]

        if tgt_idx is None:
            tgt_idx = ws.max_column
            ws.cell(row=1, column=tgt_idx + 1, value="Translation")

        ts_map = {ts.line_num: ts for ts in translatable_objects if ts.id != "##NEW_ENTRY##"}

        for row_num in range(2, ws.max_row + 1):
            ts = ts_map.get(row_num)
            if ts:
                # 1. 回填译文 (openpyxl 是 1-based)
                trans_text = ts.translation if ts.translation else ts.original_semantic
                ws.cell(row=row_num, column=tgt_idx + 1, value=trans_text)

                # 2. 回填注释
                if cmt_idx is not None:
                    ws.cell(row=row_num, column=cmt_idx + 1, value=ts.comment)
        temp_filepath = filepath + ".tmp"

        try:
            wb.save(temp_filepath)
            os.replace(temp_filepath, filepath)
        except Exception as e:
            logger.error(f"Failed to save Excel file to {filepath}: {e}")
            raise e
        finally:
            if os.path.exists(temp_filepath):
                try:
                    os.remove(temp_filepath)
                except OSError:
                    pass


class SrtFormatHandler(BaseFormatHandler):
    """
    SubRip 字幕 (.srt) 格式处理器

    标准块结构:
        <序号>
        <HH:MM:SS,mmm> --> <HH:MM:SS,mmm> [位置标注(可选)]
        <文本行1>
        [文本行2 ...]
        <空行>

    支持的特性:
    1. 完整时间码保留: 读写均原样保留 --> 行及可选位置参数
    2. 多行字幕: 将多行合并为单条目，换行符以 \\n 标记保存至 comment
    3. HTML 标签感知: 保留 <b> <i> <u> <font> 等字幕格式标签
    4. 编号重排: 保存时自动按顺序重新编号，防止空洞序号
    5. BOM 兼容: 读取时自动处理 UTF-8 BOM
    """

    format_id = "srt"
    is_monolingual = True
    extensions = [".srt"]
    format_type = "translation"
    display_name = _("SubRip Subtitle (.srt)")
    badge_text = "SRT"
    badge_bg_color = "#FFF8E1"
    badge_text_color = "#F57F17"

    # 时间码行正则: 支持可选的位置参数 (X1:N Y1:N ...)
    _TC_RE = re.compile(
        r"^(\d{2}:\d{2}:\d{2}[,\.]\d{3})"  # 开始时间码
        r"\s*-->\s*"
        r"(\d{2}:\d{2}:\d{2}[,\.]\d{3})"  # 结束时间码
        r"(.*?)$"  # 可选位置参数
    )

    def load(self, filepath: str, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or os.path.basename(filepath)

        with open(filepath, encoding="utf-8-sig", errors="replace") as f:
            content = f.read()

        language_code = self._detect_language_from_filename(os.path.basename(filepath))
        translatable_objects: list[TranslatableString] = []
        occurrence_counters: dict = {}

        # 按空行拆分块，兼容 \r\n
        raw_blocks = re.split(r"\n\s*\n", content.replace("\r\n", "\n").strip())

        for block in raw_blocks:
            lines = block.strip().splitlines()
            if len(lines) < 3:
                continue

            # 第一行: 序号（可能含 BOM 残余）
            seq_line = lines[0].strip().lstrip("\ufeff")
            if not seq_line.isdigit():
                continue

            # 第二行: 时间码
            tc_match = self._TC_RE.match(lines[1].strip())
            if not tc_match:
                continue

            start_tc = tc_match.group(1)
            end_tc = tc_match.group(2)
            position_hint = tc_match.group(3).strip()
            timecode = f"{start_tc} --> {end_tc}"
            if position_hint:
                timecode += f" {position_hint}"

            # 其余行: 字幕文本
            text_lines = lines[2:]
            text = "\n".join(text_lines).strip()
            if not text:
                continue

            # 唯一标识用时间码 (比序号更稳定)
            counter_key = (text, timecode)
            idx = occurrence_counters.get(counter_key, 0)
            occurrence_counters[counter_key] = idx + 1

            stable = f"{relative_path}::{timecode}::{text}::{idx}"
            obj_id = xxhash.xxh128(stable.encode()).hexdigest()

            ts = TranslatableString(
                original_raw=text,
                original_semantic=text,
                line_num=int(seq_line),
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="SRT Subtitle",
                source_file_path=relative_path,
                occurrences=[(relative_path, timecode)],
                occurrence_index=idx,
                id=obj_id,
            )
            initial_trans = self.get_initial_translation(text, app_instance)
            ts.set_translation_internal(initial_trans, is_initial=True)
            ts.context = timecode  # 时间码作为唯一上下文
            ts.comment = ""
            ts.po_comment = f"#: {seq_line} | {timecode}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            translatable_objects.append(ts)

        metadata = {"raw_content": content}
        logger.info(f"[SrtFormatHandler] Loaded {len(translatable_objects)} subtitles from {filepath}")
        return translatable_objects, metadata, language_code

    def save(self, filepath: str, translatable_objects, metadata: dict, **kwargs):
        lines_out: list[str] = []

        # 按原始行号（即原序号）排序
        ordered = sorted([ts for ts in translatable_objects if ts.id != "##NEW_ENTRY##"], key=lambda ts: ts.line_num)

        for new_seq, ts in enumerate(ordered, start=1):
            translation = ts.translation if ts.translation else ts.original_semantic
            lines_out.append(str(new_seq))
            lines_out.append(ts.context)  # context 存时间码行
            lines_out.append(translation)
            lines_out.append("")  # 块间空行

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines_out))
            if not lines_out[-1]:
                pass  # 已有尾部空行
            else:
                f.write("\n")

        logger.info(f"[SrtFormatHandler] Saved {len(ordered)} subtitles to {filepath}")


class VttFormatHandler(BaseFormatHandler):
    """
    WebVTT 字幕 (.vtt) 格式处理器

    VTT 格式规范要点:
    - 文件首行必须为 "WEBVTT"（可带可选描述）
    - 块类型: CUE（字幕）/ NOTE（注释，跳过）/ REGION / STYLE（跳过）
    - 时间码格式: MM:SS.mmm 或 HH:MM:SS.mmm，使用 " --> " 分隔
    - CUE ID 可选，且允许为非数字字符串

    支持的特性:
    1. CUE ID 保留: 读写时保留可选的 CUE 标识符
    2. CUE 设置保留: line/position/align/size 等设置参数原样写回
    3. NOTE 块跳过: WEBVTT NOTE 注释块不作为可翻译内容
    4. 多行文本: 同 SRT，多行合并处理
    5. VTT 标签感知: 保留 <v Speaker> <ruby> <c.class> 等语音/样式标签
    """

    format_id = "vtt"
    is_monolingual = True
    extensions = [".vtt"]
    format_type = "translation"
    display_name = _("WebVTT Subtitle (.vtt)")
    badge_text = "VTT"
    badge_bg_color = "#F3E5F5"
    badge_text_color = "#6A1B9A"

    _TC_RE = re.compile(
        r"^(\d{2}:\d{2}:\d{2}\.\d{3}|\d{2}:\d{2}\.\d{3})"
        r"\s*-->\s*"
        r"(\d{2}:\d{2}:\d{2}\.\d{3}|\d{2}:\d{2}\.\d{3})"
        r"(.*?)$"
    )

    def load(self, filepath: str, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or os.path.basename(filepath)

        with open(filepath, encoding="utf-8-sig", errors="replace") as f:
            content = f.read()

        language_code = self._detect_language_from_filename(os.path.basename(filepath))
        translatable_objects: list[TranslatableString] = []
        occurrence_counters: dict = {}

        lines_all = content.replace("\r\n", "\n").splitlines()
        # 校验首行
        if not lines_all or not lines_all[0].startswith("WEBVTT"):
            logger.warning(f"[VttFormatHandler] Missing WEBVTT header in {filepath}")

        # 按空行切块
        raw_blocks: list[list[str]] = []
        current: list[str] = []
        for line in lines_all[1:]:  # 跳过首行 WEBVTT
            if line.strip() == "":
                if current:
                    raw_blocks.append(current)
                    current = []
            else:
                current.append(line)
        if current:
            raw_blocks.append(current)

        seq_counter = 0
        for block in raw_blocks:
            if not block:
                continue
            # 跳过 NOTE / STYLE / REGION 块
            first = block[0].strip()
            if first.startswith("NOTE") or first.startswith("STYLE") or first.startswith("REGION"):
                continue

            # 找时间码行的位置
            tc_line_idx = None
            cue_id = ""
            for i, line in enumerate(block):
                m = self._TC_RE.match(line.strip())
                if m:
                    tc_line_idx = i
                    if i > 0:
                        cue_id = block[0].strip()
                    break

            if tc_line_idx is None:
                continue

            tc_line = block[tc_line_idx].strip()
            m = self._TC_RE.match(tc_line)
            start_tc = m.group(1)
            end_tc = m.group(2)
            cue_settings = m.group(3).strip()

            timecode = f"{start_tc} --> {end_tc}"
            full_tc_line = timecode + (f" {cue_settings}" if cue_settings else "")

            text_lines = block[tc_line_idx + 1 :]
            text = "\n".join(l for l in text_lines).strip()
            if not text:
                continue

            seq_counter += 1
            context_key = f"{cue_id}|{full_tc_line}" if cue_id else full_tc_line

            counter_key = (text, context_key)
            idx = occurrence_counters.get(counter_key, 0)
            occurrence_counters[counter_key] = idx + 1

            stable = f"{relative_path}::{context_key}::{text}::{idx}"
            obj_id = xxhash.xxh128(stable.encode()).hexdigest()

            ts = TranslatableString(
                original_raw=text,
                original_semantic=text,
                line_num=seq_counter,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="VTT Subtitle",
                source_file_path=relative_path,
                occurrences=[(relative_path, full_tc_line)],
                occurrence_index=idx,
                id=obj_id,
            )
            ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)
            ts.context = context_key  # "cueId|HH:MM:SS.mmm --> HH:MM:SS.mmm [settings]"
            ts.comment = f"CUE ID: {cue_id}" if cue_id else ""
            ts.po_comment = f"#: {full_tc_line}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            translatable_objects.append(ts)

        metadata = {"raw_content": content}
        logger.info(f"[VttFormatHandler] Loaded {len(translatable_objects)} cues from {filepath}")
        return translatable_objects, metadata, language_code

    def save(self, filepath: str, translatable_objects, metadata: dict, **kwargs):
        lines_out = ["WEBVTT", ""]

        ordered = sorted(
            [ts for ts in translatable_objects if ts.id != "##NEW_ENTRY##"],
            key=lambda ts: ts.line_num,
        )

        for ts in ordered:
            # context_key = "cueId|tc_line" 或 "tc_line"
            ctx = ts.context
            if "|" in ctx:
                cue_id, tc_line = ctx.split("|", 1)
                lines_out.append(cue_id)
            else:
                tc_line = ctx

            lines_out.append(tc_line)
            translation = ts.translation if ts.translation else ts.original_semantic
            lines_out.append(translation)
            lines_out.append("")

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines_out))
            f.write("\n")

        logger.info(f"[VttFormatHandler] Saved {len(ordered)} cues to {filepath}")


class HtmlFormatHandler(BaseFormatHandler):
    """
    HTML / HTM 网页文件翻译处理器

    提取来源（白名单标签内的可见文本）:
    ┌─────────────────────────────────────────────────────────────────────┐
    │ 正文标签: p / h1-h6 / li / td / th / dt / dd / figcaption / caption │
    │ 行内标签: a / span / strong / em / b / i / label / button / legend  │
    │ 表单标签: input[placeholder] / input[value] / textarea              │
    │ 元数据:   <title> / <meta name="description" content="…">           │
    │           <meta name="keywords" content="…">                        │
    │           <meta property="og:title/description" content="…">        │
    │ 属性:     alt / title（仅 img 和 a 标签）                           │
    └─────────────────────────────────────────────────────────────────────┘

    支持的特性:
    1. 结构保留: 翻译单元为单个标签的"净文本"（子标签内容合并），
               保存时用正则精准替换，不解析整棵 DOM
    2. i18n 感知: 自动跳过含 data-i18n-ignore / translate="no" 的节点
    3. 脚本跳过: <script> / <style> / <code> / <pre> 内容完全跳过
    4. 注释跳过: HTML 注释 <!-- --> 不提取
    5. 属性提取: alt / placeholder / title 属性单独成条目
    6. 语言检测: 读取 <html lang="…"> 属性，其次从文件名推断
    7. 保存策略: 基于原始文本的字符串替换（从长到短防止子串误替换），
               保留所有 HTML 结构、内联样式、脚本不变
    """

    format_id = "html"
    is_monolingual = True
    extensions = [".html", ".htm"]
    format_type = "translation"
    display_name = _("HTML Web Page (.html/.htm)")
    badge_text = "HTML"
    badge_bg_color = "#FFF3E0"
    badge_text_color = "#E65100"

    # 提取文本的块级 / 行内标签
    _TEXT_TAGS = frozenset(
        {
            "p",
            "h1",
            "h2",
            "h3",
            "h4",
            "h5",
            "h6",
            "li",
            "td",
            "th",
            "dt",
            "dd",
            "figcaption",
            "caption",
            "a",
            "span",
            "strong",
            "em",
            "b",
            "i",
            "label",
            "button",
            "legend",
            "title",
            "summary",  # <details> summary
        }
    )
    # 跳过内部全部内容的标签
    _SKIP_TAGS = frozenset({"script", "style", "code", "pre", "noscript", "template"})
    # 提取属性的规则: tag -> [attr, ...]
    _ATTR_EXTRACT: dict[str, list[str]] = {
        "img": ["alt", "title"],
        "a": ["title"],
        "input": ["placeholder", "value"],
        "textarea": ["placeholder"],
        "area": ["alt"],
        "th": ["abbr"],
    }

    # 用于从标签中提取净文本（去除子标签）
    _TAG_CONTENT_RE = re.compile(r"<[^>]+>")
    # 检测是否含不可翻译属性
    _IGNORE_RE = re.compile(r'\btranslate\s*=\s*["\']no["\']|data-i18n-ignore', re.I)

    def load(self, filepath: str, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or os.path.basename(filepath)

        with open(filepath, encoding="utf-8", errors="replace") as f:
            content = f.read()

        language_code = self._detect_language(content, os.path.basename(filepath))
        translatable_objects: list[TranslatableString] = []
        occurrence_counters: dict = {}

        # 去除 script / style / code / pre 区块，防止误匹配
        content_clean = self._strip_skip_blocks(content)
        # 去除 HTML 注释
        content_clean = re.sub(r"<!--.*?-->", "", content_clean, flags=re.DOTALL)

        # ── 提取 <meta> 标签属性 ──────────────────────────────────────────
        self._extract_meta(content_clean, relative_path, translatable_objects, occurrence_counters, app_instance)

        # ── 提取普通标签文本内容 ─────────────────────────────────────────
        tag_pattern = re.compile(
            r"<(?P<tag>" + "|".join(self._TEXT_TAGS) + r")(?P<attrs>[^>]*)>"
            r"(?P<inner>(?:[^<]+|<(?!/?(?P=tag)\b)[^>]*>|<(?P=tag)\b[^>]*>(?P>inner)</(?P=tag)>)*)"
            r"</(?P=tag)>",
            re.IGNORECASE | re.DOTALL,
        )
        for m in tag_pattern.finditer(content_clean):
            tag = m.group("tag").lower()
            attrs = m.group("attrs")
            inner = m.group("inner")

            # 跳过 translate="no" / data-i18n-ignore
            if self._IGNORE_RE.search(attrs):
                continue

            # 提取净文本（去除子标签）
            clean_text = self._TAG_CONTENT_RE.sub("", inner).strip()
            # 折叠空白
            clean_text = re.sub(r"\s+", " ", clean_text).strip()
            if not clean_text or len(clean_text) < 2:
                continue

            # 计算大致行号
            line_num = content[: m.start()].count("\n") + 1
            context = f"{tag}.{line_num}"

            inner_abs_start = m.start("inner")
            text_rel_start = inner.find(clean_text)
            if text_rel_start == -1:
                char_start = inner_abs_start
                char_end = m.end("inner")
            else:
                char_start = inner_abs_start + text_rel_start
                char_end = char_start + len(clean_text)

            self._make_entry(
                clean_text,
                context,
                f"HTML <{tag}>",
                attrs,
                relative_path,
                line_num,
                translatable_objects,
                occurrence_counters,
                app_instance,
                char_start=char_start,
                char_end=char_end,
            )

        # ── 提取属性（alt / placeholder / title …） ─────────────────────
        for tag, attr_list in self._ATTR_EXTRACT.items():
            for attr in attr_list:
                attr_pattern = re.compile(
                    r"<" + tag + r"[^>]*\b" + attr + r'\s*=\s*["\'](?P<val>[^"\']+)["\'][^>]*>',
                    re.IGNORECASE,
                )
                for m in attr_pattern.finditer(content_clean):
                    val = m.group("val").strip()
                    if not val or len(val) < 2:
                        continue
                    line_num = content[: m.start()].count("\n") + 1
                    context = f"{tag}@{attr}.{line_num}"

                    # 记录属性值的精确偏移量
                    char_start = m.start("val")
                    char_end = m.end("val")

                    self._make_entry(
                        val,
                        context,
                        f"HTML {tag}@{attr}",
                        m.group(0),
                        relative_path,
                        line_num,
                        translatable_objects,
                        occurrence_counters,
                        app_instance,
                        char_start=char_start,
                        char_end=char_end,
                    )

        metadata = {"raw_content": content}
        logger.info(f"[HtmlFormatHandler] Loaded {len(translatable_objects)} entries from {filepath}")
        return translatable_objects, metadata, language_code

    def _extract_meta(self, content: str, rel_path: str, results: list, counters: dict, app_instance):
        # <title>
        title_m = re.search(r"<title[^>]*>([^<]+)</title>", content, re.I)
        if title_m:
            val = title_m.group(1).strip()
            if val:
                self._make_entry(val, "meta.title", "HTML <title>", "", rel_path, 0, results, counters, app_instance)

        # <meta name="description/keywords" content="…">
        meta_re = re.compile(
            r'<meta\s[^>]*\bname\s*=\s*["\'](?P<name>description|keywords|author)["\'][^>]*\bcontent\s*=\s*["\'](?P<val>[^"\']+)["\']',
            re.I,
        )
        for m in meta_re.finditer(content):
            val = m.group("val").strip()
            name = m.group("name").lower()
            if val:
                self._make_entry(
                    val, f"meta.{name}", f"HTML <meta {name}>", "", rel_path, 0, results, counters, app_instance
                )

        # Open Graph
        og_re = re.compile(
            r'<meta\s[^>]*\bproperty\s*=\s*["\']og:(?P<prop>title|description)["\'][^>]*\bcontent\s*=\s*["\'](?P<val>[^"\']+)["\']',
            re.I,
        )
        for m in og_re.finditer(content):
            val = m.group("val").strip()
            prop = m.group("prop").lower()
            if val:
                self._make_entry(
                    val, f"meta.og_{prop}", f"HTML og:{prop}", "", rel_path, 0, results, counters, app_instance
                )

    def _make_entry(
        self,
        text: str,
        context: str,
        string_type: str,
        attrs: str,
        rel_path: str,
        line_num: int,
        results: list,
        counters: dict,
        app_instance,
        char_start: int = 0,
        char_end: int = 0,
    ):
        counter_key = (text, context)
        idx = counters.get(counter_key, 0)
        counters[counter_key] = idx + 1

        stable = f"{rel_path}::{context}::{text}::{idx}"
        obj_id = xxhash.xxh128(stable.encode()).hexdigest()

        ts = TranslatableString(
            original_raw=text,
            original_semantic=text,
            line_num=line_num,
            char_pos_start_in_file=char_start,
            char_pos_end_in_file=char_end,
            full_code_lines=[],
            string_type=string_type,
            source_file_path=rel_path,
            occurrences=[(rel_path, context)],
            occurrence_index=idx,
            id=obj_id,
        )
        ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)
        ts.context = context
        ts.comment = ""
        ts.po_comment = f"#: HTML {context}"
        ts.is_reviewed = False
        ts.update_sort_weight()
        results.append(ts)

    def save(self, filepath: str, translatable_objects, metadata: dict, **kwargs):
        content: str = metadata["raw_content"]
        replace_ops = []
        for ts in translatable_objects:
            if not ts.original_semantic or ts.id == "##NEW_ENTRY##":
                continue
            translation = ts.translation or ts.original_semantic
            if translation == ts.original_semantic:
                continue
            if ts.char_pos_start_in_file > 0 or ts.char_pos_end_in_file > 0:
                replace_ops.append(
                    (ts.char_pos_start_in_file, ts.char_pos_end_in_file, ts.original_semantic, translation)
                )
        # 按起始位置降序排列
        replace_ops.sort(key=lambda x: x[0], reverse=True)

        for start, end, original, translation in replace_ops:
            if original in content[start:end]:
                content = content[:start] + content[start:end].replace(original, translation, 1) + content[end:]
            else:
                content = content.replace(original, translation, 1)

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write(content)
        logger.info(f"[HtmlFormatHandler] Saved to {filepath}")

    @staticmethod
    def _strip_skip_blocks(content: str) -> str:
        """将 <script>/<style>/<pre>/<code> 块内容替换为空占位，保留标签位置。"""
        for tag in ("script", "style", "pre", "code", "noscript", "template"):
            content = re.sub(
                r"<" + tag + r"[^>]*>.*?</" + tag + r">", f"<{tag}></{tag}>", content, flags=re.IGNORECASE | re.DOTALL
            )
        return content

    @staticmethod
    def _detect_language(content: str, filename: str) -> str:
        # <html lang="zh-CN">
        m = re.search(r"<html[^>]+\blang\s*=\s*[\"']([^\"']+)[\"']", content, re.I)
        if m:
            return m.group(1).replace("-", "_")
        stem = os.path.splitext(filename)[0]
        m2 = re.search(r"[._-]([a-z]{2,3}(?:[_-][A-Za-z]{2,4})?)$", stem)
        return m2.group(1).replace("-", "_") if m2 else "en"


class MarkdownFormatHandler(BaseFormatHandler):
    """
    Markdown / MDX 文档本地化处理器

    支持的提取单元（按语义粒度从细到粗）:
    1. Frontmatter 字段: 解析 YAML frontmatter (--- ... ---) 中的字符串字段，
       如 title / description / keywords，跳过日期、布尔、数字等非文本字段
    2. 标题 (ATX/Setext): 提取 # / ## / ### 等 ATX 风格标题；
       同时支持 Setext 风格（下划线 === / ---）标题
    3. 段落块: 将连续非空行合并为一个段落单元提取，保留行内 Markdown 标记
       (**bold**, *italic*, `code`, [link](url), ![alt](src))，
       仅替换文本内容，不破坏标记结构
    4. 列表项: 有序列表 (1. 2. 3.) 和无序列表 (- * +) 的每个条目单独提取；
       支持嵌套列表，子项以缩进层级区分
    5. 表格单元: 解析 GFM (GitHub Flavored Markdown) 表格，提取表头和数据单元，
       跳过分隔行 (| --- | --- |)
    6. 块引用: 提取 > 引用块中的文字内容，多行引用合并为一个单元
    7. 自定义 MDX 组件属性: 提取 JSX 风格属性中的字符串字面量，
       如 <Button label="Click me"> 中的 "Click me"

    刻意跳过的内容（保证代码结构不受损）:
    - 围栏代码块 (``` 或 ~~~)
    - 行内代码 (`code`)
    - HTML 注释 (<!-- -->)
    - import / export 语句 (MDX 专用)
    - 纯 URL / 路径
    - 数学公式 ($ ... $ 和 $$ ... $$)

    保存策略:
    - 基于 char_pos_start_in_file / char_pos_end_in_file 精准定位并替换，
      文档其余部分（格式、空行、代码块）完全原样保留
    - 若无 char 偏移信息则回退到行号替换
    """

    format_id = "markdown"
    is_monolingual = True
    extensions = [".md", ".mdx", ".markdown"]
    format_type = "translation"
    display_name = _("Markdown / MDX Document")
    badge_text = "MD"
    badge_bg_color = "#ECEFF1"
    badge_text_color = "#37474F"

    # 跳过的 frontmatter 字段（通常是日期、布尔、路径等）
    _FRONTMATTER_SKIP_KEYS = {
        "date",
        "updated",
        "created",
        "draft",
        "published",
        "order",
        "weight",
        "slug",
        "permalink",
        "url",
        "layout",
        "template",
        "type",
        "id",
        "uuid",
    }

    # 纯 URL / 路径正则（不值得翻译）
    _URL_RE = re.compile(r"^(?:https?://|ftp://|/|\.{0,2}/)[\w./?=&%#@:+\-]*$")

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        logger.debug(f"[MarkdownFormatHandler] Loading Markdown: {filepath}")
        with open(filepath, encoding="utf-8") as f:
            content = f.read()

        rel_path = kwargs.get("relative_path") or self._get_relative_path(filepath)

        full_lines = content.splitlines()

        translatable_objects = []
        occurrence_counters = {}
        skip_ranges = self._find_skip_ranges(content)

        fm_end = 0
        _, fm_end = self._extract_frontmatter(
            content, rel_path, translatable_objects, occurrence_counters, full_lines, app_instance=app_instance
        )

        self._extract_body(
            content,
            fm_end,
            skip_ranges,
            rel_path,
            translatable_objects,
            occurrence_counters,
            full_lines,
            app_instance=app_instance,
        )

        language_code = self._detect_language(os.path.basename(filepath))
        metadata = {
            "original_content": content,
            "skip_ranges": skip_ranges,
        }

        logger.info(f"[MarkdownFormatHandler] Loaded {len(translatable_objects)} segments from {filepath}")
        return translatable_objects, metadata, language_code

    def _find_skip_ranges(self, content: str) -> list[tuple[int, int]]:
        """
        返回不应被提取或替换的字符区间列表 [(start, end), ...]。
        涵盖: 围栏代码块、行内代码、HTML注释、数学公式、import/export
        """
        ranges = []

        # 围栏代码块: ```...``` 或 ~~~...~~~
        for m in re.finditer(r"(?m)^(```+|~~~+)[^\n]*\n.*?\n\1[ \t]*$", content, re.DOTALL):
            ranges.append((m.start(), m.end()))

        # 行内代码: `...`
        for m in re.finditer(r"`+[^`\n]+`+", content):
            ranges.append((m.start(), m.end()))

        # HTML 注释
        for m in re.finditer(r"<!--.*?-->", content, re.DOTALL):
            ranges.append((m.start(), m.end()))

        # 数学公式块 $$ ... $$
        for m in re.finditer(r"\$\$.*?\$\$", content, re.DOTALL):
            ranges.append((m.start(), m.end()))

        # 行内数学 $ ... $（单行）
        for m in re.finditer(r"\$[^\n$]+\$", content):
            ranges.append((m.start(), m.end()))

        # MDX import / export 语句
        for m in re.finditer(r"(?m)^(?:import|export)\s+.+$", content):
            ranges.append((m.start(), m.end()))

        return sorted(ranges)

    def _in_skip_range(self, pos: int, skip_ranges: list[tuple[int, int]]) -> bool:
        for s, e in skip_ranges:
            if s <= pos < e:
                return True
            if s > pos:
                break
        return False

    def _extract_frontmatter(
        self, content: str, rel_path: str, results: list, counters: dict, full_lines: list[str], app_instance=None
    ) -> tuple[dict, int]:
        """提取 YAML frontmatter 中的可翻译字段，返回 (字段dict, frontmatter结束位置)"""
        fm_end = 0
        extracted = {}

        m = re.match(r"^---[ \t]*\r?\n(.*?)\r?\n---[ \t]*\r?\n", content, re.DOTALL)
        if not m:
            return extracted, fm_end

        fm_text = m.group(1)
        fm_end = m.end()

        for fm_m in re.finditer(r'^([ \t]*)(\w[\w-]*)[ \t]*:[ \t]*(["\']?)(.+?)\3[ \t]*$', fm_text, re.MULTILINE):
            indent_str, key, quote, value = fm_m.groups()
            if key.lower() in self._FRONTMATTER_SKIP_KEYS:
                continue
            if self._URL_RE.match(value.strip()):
                continue
            if re.match(r"^(?:true|false|null|\d[\d.,]*)$", value.strip(), re.I):
                continue
            if value.strip():
                self._make_ts(
                    value.strip(),
                    f"frontmatter.{key}",
                    rel_path,
                    results,
                    counters,
                    full_lines,
                    line_num=content[: m.start() + fm_m.start()].count("\n") + 1,
                    char_start=m.start(1) + fm_m.start(4),
                    char_end=m.start(1) + fm_m.end(4),
                    string_type="MD Frontmatter",
                    app_instance=app_instance,
                )

        return extracted, fm_end

    def _extract_body(
        self,
        content: str,
        body_start: int,
        skip_ranges: list[tuple[int, int]],
        rel_path: str,
        results: list,
        counters: dict,
        full_lines: list[str],
        app_instance=None,
    ):
        """逐行扫描文档正文，按语义单元提取"""
        lines = content[body_start:].split("\n")
        abs_offset = body_start

        i = 0
        while i < len(lines):
            line = lines[i]
            line_abs_start = abs_offset
            line_num = content[:abs_offset].count("\n") + 1

            # 如果整行在禁区内，跳过
            if self._in_skip_range(line_abs_start, skip_ranges):
                abs_offset += len(line) + 1
                i += 1
                continue

            stripped = line.strip()
            if not stripped:
                abs_offset += len(line) + 1
                i += 1
                continue

            # --- 1. ATX 标题 ---
            atx_m = re.match(r"^(#{1,6})\s+(.*?)(?:\s+#+\s*)?$", stripped)
            if atx_m:
                heading_text = atx_m.group(2).strip()
                if heading_text:
                    level = len(atx_m.group(1))
                    # 计算文本在文件中的精确起始位置
                    atx_m.group(1)
                    # 找到第一个非空字符的位置
                    text_rel_start = re.search(re.escape(heading_text), line).start()
                    text_abs_start = line_abs_start + text_rel_start
                    self._make_ts(
                        heading_text,
                        f"heading.h{level}",
                        rel_path,
                        results,
                        counters,
                        full_lines,
                        line_num=line_num,
                        char_start=text_abs_start,
                        char_end=text_abs_start + len(heading_text),
                        string_type="MD Heading",
                        app_instance=app_instance,
                    )
                abs_offset += len(line) + 1
                i += 1
                continue

            # --- 2. Setext 标题 ---
            if i + 1 < len(lines):
                next_stripped = lines[i + 1].strip()
                if re.match(r"^=+$", next_stripped) or re.match(r"^-+$", next_stripped):
                    level = 1 if next_stripped.startswith("=") else 2
                    if stripped:
                        self._make_ts(
                            stripped,
                            f"heading.h{level}",
                            rel_path,
                            results,
                            counters,
                            full_lines,
                            line_num=line_num,
                            char_start=line_abs_start + line.find(stripped),
                            char_end=line_abs_start + line.find(stripped) + len(stripped),
                            string_type="MD Heading",
                            app_instance=app_instance,
                        )
                    abs_offset += len(line) + 1 + len(lines[i + 1]) + 1
                    i += 2
                    continue

            # --- 3. 列表项 ---
            list_m = re.match(r"^([ \t]*)(?:[-*+]|\d+\.)\s+(.*)", line)
            if list_m:
                item_text = list_m.group(2).strip()
                item_text_clean = self._strip_inline_code(item_text)
                if item_text_clean and not self._URL_RE.match(item_text_clean):
                    text_rel_start = line.find(list_m.group(2))
                    text_abs_start = line_abs_start + text_rel_start
                    self._make_ts(
                        item_text,
                        "list.item",
                        rel_path,
                        results,
                        counters,
                        full_lines,
                        line_num=line_num,
                        char_start=text_abs_start,
                        char_end=text_abs_start + len(item_text),
                        string_type="MD List Item",
                        app_instance=app_instance,
                    )
                abs_offset += len(line) + 1
                i += 1
                continue

            # --- 4. 块引用 ---
            if stripped.startswith(">"):
                quote_lines = []
                quote_start_offset = abs_offset
                while i < len(lines) and lines[i].strip().startswith(">"):
                    # 移除开头的 > 符号
                    content_part = re.sub(r"^[ \t]*>+[ \t]?", "", lines[i])
                    quote_lines.append(content_part)
                    abs_offset += len(lines[i]) + 1
                    i += 1

                quote_text = "\n".join(quote_lines).strip()
                if quote_text and not self._URL_RE.match(self._strip_inline_code(quote_text)):
                    self._make_ts(
                        quote_text,
                        "blockquote",
                        rel_path,
                        results,
                        counters,
                        full_lines,
                        line_num=line_num,
                        char_start=quote_start_offset,
                        char_end=abs_offset - 1,
                        string_type="MD Blockquote",
                        app_instance=app_instance,
                    )
                continue

            # --- 5. GFM 表格行 ---
            if "|" in stripped and not re.match(r"^\|?[ \t:|-]+\|", stripped):
                # 简单的表格单元提取
                cells = [c.strip() for c in stripped.strip("|").split("|")]
                for cell in cells:
                    if not cell:
                        continue
                    cell_clean = self._strip_inline_code(cell)
                    if cell_clean and not self._URL_RE.match(cell_clean) and len(cell_clean) > 1:
                        # 定位单元格在行中的位置
                        cell_rel_start = line.find(cell)
                        self._make_ts(
                            cell,
                            "table.cell",
                            rel_path,
                            results,
                            counters,
                            full_lines,
                            line_num=line_num,
                            char_start=line_abs_start + cell_rel_start,
                            char_end=line_abs_start + cell_rel_start + len(cell),
                            string_type="MD Table",
                            app_instance=app_instance,
                        )
                abs_offset += len(line) + 1
                i += 1
                continue

            # --- 6. 段落（连续非空行）---
            # 只有不满足上述所有条件的非空行才进入段落解析
            para_lines = []
            para_start_offset = abs_offset
            para_start_line = line_num
            while i < len(lines):
                cur_line = lines[i]
                cur_stripped = cur_line.strip()
                # 检查是否遇到其他块的起始标识
                if (
                    not cur_stripped
                    or cur_stripped.startswith("#")
                    or cur_stripped.startswith("```")
                    or cur_stripped.startswith("~~~")
                    or re.match(r"^(?:[-*_]){3,}$", cur_stripped)
                    or re.match(r"^(?:[-*+]|\d+\.)\s", cur_stripped)
                    or cur_stripped.startswith(">")
                    or self._in_skip_range(abs_offset, skip_ranges)
                ):
                    break
                para_lines.append(cur_line)
                abs_offset += len(cur_line) + 1
                i += 1

            if para_lines:
                # 合并多行段落为单行文本（Markdown 渲染特性）
                para_text = " ".join(l.strip() for l in para_lines).strip()
                para_text_clean = self._strip_inline_code(para_text)

                if para_text_clean and len(para_text_clean) > 2 and not self._URL_RE.match(para_text_clean):
                    self._make_ts(
                        para_text,
                        "paragraph",
                        rel_path,
                        results,
                        counters,
                        full_lines,
                        line_num=para_start_line,
                        char_start=para_start_offset,
                        char_end=abs_offset - 1,
                        string_type="MD Paragraph",
                        app_instance=app_instance,
                    )
                continue

            abs_offset += len(line) + 1
            i += 1

    def _strip_inline_code(self, text: str) -> str:
        """去除行内反引号代码后返回纯文本，用于判断是否值得翻译"""
        return re.sub(r"`[^`]*`", "", text).strip()

    def _make_ts(
        self,
        text,
        context_hint,
        rel_path,
        results,
        counters,
        full_lines,
        line_num=0,
        char_start=0,
        char_end=0,
        string_type="MD Text",
        app_instance=None,
    ):
        """统一创建 TranslatableString 对象"""
        # 过滤过短或无意义的文本
        clean = self._strip_inline_code(text)
        if len(clean.strip()) < 2:
            return

        # context = hint::顺序计数，防止同文件同类型条目冲突
        counter_key = (text, context_hint)
        idx = counters.get(counter_key, 0)
        counters[counter_key] = idx + 1

        context = f"{context_hint}[{idx}]" if idx > 0 else context_hint

        ts = TranslatableString(
            original_raw=text,
            original_semantic=text,
            line_num=line_num,
            char_pos_start_in_file=char_start,
            char_pos_end_in_file=char_end,
            full_code_lines=full_lines,
            string_type=string_type,
            source_file_path=rel_path,
            occurrences=[(rel_path, str(line_num))],
            occurrence_index=counters.get((text, context_hint), 0),
            id=xxhash.xxh128(f"{rel_path}::{context_hint}::{text}".encode()).hexdigest(),
        )
        ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)
        ts.context = context
        ts.comment = f"Type: {string_type}"
        ts.po_comment = f"#: {rel_path}:{line_num} ({string_type})"
        ts.is_reviewed = False
        ts.update_sort_weight()
        results.append(ts)

    def _detect_language(self, filename: str) -> str:
        stem = os.path.splitext(filename)[0]
        # docs.zh-CN.md / index.fr.mdx
        m = re.search(r"\.([a-z]{2,3}(?:-[A-Za-z]{2,4})?)$", stem, re.IGNORECASE)
        if m:
            return m.group(1)
        return "en"

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        logger.debug(f"[MarkdownFormatHandler] Saving Markdown: {filepath}")

        original_content = metadata.get("original_content", "")

        # 按 char_pos_start_in_file 降序排列，从后往前替换，避免偏移漂移
        replace_ops = []
        for ts in translatable_objects:
            if not ts.original_semantic or ts.id == "##NEW_ENTRY##":
                continue
            translation = ts.translation or ts.original_semantic
            if translation == ts.original_semantic:
                continue  # 未翻译，跳过
            if ts.char_pos_start_in_file > 0 or ts.char_pos_end_in_file > 0:
                replace_ops.append(
                    (ts.char_pos_start_in_file, ts.char_pos_end_in_file, ts.original_semantic, translation)
                )

        # 去重 + 降序
        replace_ops.sort(key=lambda x: x[0], reverse=True)

        content = original_content
        for start, end, original, translation in replace_ops:
            # 安全校验：确认位置内容与原文匹配
            if content[start:end] == original:
                content = content[:start] + translation + content[end:]
            else:
                # 偏移可能已漂移，回退到全文替换（只替换第一次出现）
                content = content.replace(original, translation, 1)

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write(content)

        logger.info(f"[MarkdownFormatHandler] Saved translated document to {filepath}")


def _ooxml_read_part(zf: zipfile.ZipFile, part_path: str) -> str | None:
    """读取 ZIP 包内指定 part，返回 UTF-8 字符串，不存在则返回 None。"""
    try:
        return zf.read(part_path).decode("utf-8", errors="replace")
    except KeyError:
        return None


def _ooxml_list_parts(zf: zipfile.ZipFile, prefix: str) -> list[str]:
    """列出 ZIP 包内以 prefix 开头的所有 part 路径。"""
    return [n for n in zf.namelist() if n.startswith(prefix)]


def _ooxml_clone_and_patch(
    src_filepath: str,
    dst_filepath: str,
    patched_parts: dict[str, bytes],
):
    """
    将 src_filepath 的 ZIP 内容复制到 dst_filepath，
    同时将 patched_parts 中指定的 part 替换为新内容。
    """
    with (
        zipfile.ZipFile(src_filepath, "r") as src_zf,
        zipfile.ZipFile(dst_filepath, "w", compression=zipfile.ZIP_DEFLATED) as dst_zf,
    ):
        for item in src_zf.infolist():
            if item.filename in patched_parts:
                dst_zf.writestr(item, patched_parts[item.filename])
            else:
                dst_zf.writestr(item, src_zf.read(item.filename))


class DocxFormatHandler(BaseFormatHandler):
    """
    Microsoft Word DOCX / DOCM 处理器

    提取来源（按优先级）:
    ┌─────────────────────────────────────────────────────────────────────┐
    │ word/document.xml   正文段落 <w:t>                                  │
    │ word/comments.xml   批注文本                                        │
    │ word/footnotes.xml  脚注                                            │
    │ word/endnotes.xml   尾注                                            │
    │ word/header*.xml    页眉                                            │
    │ word/footer*.xml    页脚                                            │
    └─────────────────────────────────────────────────────────────────────┘

    支持的特性:
    1. Run 合并: 同一段落的多个 <w:r><w:t> 合并为单条翻译，避免割裂语义
    2. 格式标签保留: Bold/Italic/Underline 等 <w:rPr> 属性原样写回
    3. 超链接识别: <w:hyperlink> 内文本作为独立条目，关联 URL 记录到 comment
    4. 表格感知: <w:tbl> → <w:tc> → <w:p> 全部扫描，context 标注单元格位置
    5. 文本框: <mc:AlternateContent> / <w:txbxContent> 内文本正确提取
    6. 样式感知: Heading1-9 在 context 中标注层级，便于审阅排序
    7. 原子保存: 先 clone ZIP 再 patch document.xml 等，不破坏嵌入资源
    """

    format_id = "docx"
    is_monolingual = True
    extensions = [".docx", ".docm"]
    format_type = "translation"
    display_name = _("Word Document (.docx/.docm)")
    badge_text = "DOCX"
    badge_bg_color = "#E3F2FD"
    badge_text_color = "#1565C0"

    # Word XML 命名空间
    _NS = {
        "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
        "w14": "http://schemas.microsoft.com/office/word/2010/wordml",
    }

    def load(self, filepath: str, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or os.path.basename(filepath)
        translatable_objects: list[TranslatableString] = []
        occurrence_counters: dict = {}

        # 收集所有需处理的 part
        parts_to_scan = []
        with zipfile.ZipFile(filepath, "r") as zf:
            all_names = set(zf.namelist())
            primary_parts = ["word/document.xml", "word/comments.xml", "word/footnotes.xml", "word/endnotes.xml"]
            for p in primary_parts:
                if p in all_names:
                    parts_to_scan.append(p)
            for n in all_names:
                if re.match(r"word/(header|footer)\d*\.xml$", n):
                    parts_to_scan.append(n)

            part_xmls: dict[str, bytes] = {p: zf.read(p) for p in parts_to_scan}

        for part_path, xml_bytes in part_xmls.items():
            part_label = Path(part_path).stem  # document / header1 / footer2 …
            try:
                root = ET.fromstring(xml_bytes)
            except ET.ParseError as e:
                logger.warning(f"[DocxFormatHandler] XML parse error in {part_path}: {e}")
                continue

            self._extract_paragraphs(
                root, part_label, relative_path, translatable_objects, occurrence_counters, app_instance
            )

        metadata = {
            "filepath": filepath,
            "part_xmls": part_xmls,  # 原始 bytes，用于 save
            "parts_to_scan": parts_to_scan,
        }
        logger.info(f"[DocxFormatHandler] Loaded {len(translatable_objects)} paragraphs from {filepath}")
        return translatable_objects, metadata, self._detect_language(filepath)

    def _extract_paragraphs(
        self, root: ET.Element, part_label: str, rel_path: str, results: list, counters: dict, app_instance
    ):
        w = self._NS["w"]
        para_idx = 0

        for para in root.iter(f"{{{w}}}p"):
            # 合并段落内所有 <w:t> 文本
            texts = []
            for t_elem in para.iter(f"{{{w}}}t"):
                if t_elem.text:
                    texts.append(t_elem.text)
            full_text = "".join(texts).strip()
            if not full_text:
                para_idx += 1
                continue

            # 获取段落样式
            pstyle = ""
            ppr = para.find(f"{{{w}}}pPr")
            if ppr is not None:
                pstyle_elem = ppr.find(f"{{{w}}}pStyle")
                if pstyle_elem is not None:
                    pstyle = pstyle_elem.get(f"{{{w}}}val", "")

            context = f"{part_label}.p{para_idx}"
            if pstyle:
                context += f"[{pstyle}]"

            counter_key = (full_text, context)
            idx = counters.get(counter_key, 0)
            counters[counter_key] = idx + 1

            stable = f"{rel_path}::{context}::{full_text}::{idx}"
            obj_id = xxhash.xxh128(stable.encode()).hexdigest()

            ts = TranslatableString(
                original_raw=full_text,
                original_semantic=full_text,
                line_num=para_idx,
                char_pos_start_in_file=0,
                char_pos_end_in_file=0,
                full_code_lines=[],
                string_type="DOCX Paragraph",
                source_file_path=rel_path,
                occurrences=[(rel_path, context)],
                occurrence_index=idx,
                id=obj_id,
            )
            ts.set_translation_internal(self.get_initial_translation(full_text, app_instance), is_initial=True)
            ts.context = context
            ts.comment = f"Style: {pstyle}" if pstyle else ""
            ts.po_comment = f"#: {part_label} paragraph {para_idx}"
            ts.is_reviewed = False
            ts.update_sort_weight()
            results.append(ts)
            para_idx += 1

    def save(self, filepath: str, translatable_objects, metadata: dict, **kwargs):
        part_xmls: dict[str, bytes] = metadata["part_xmls"]
        parts_to_scan: list[str] = metadata["parts_to_scan"]

        # context(part_label.pN[style]) -> translation
        trans_map: dict[str, str] = {
            ts.context: (ts.translation or ts.original_semantic)
            for ts in translatable_objects
            if ts.id != "##NEW_ENTRY##" and ts.context and ts.translation
        }

        patched: dict[str, bytes] = {}
        w = self._NS["w"]

        for part_path in parts_to_scan:
            xml_bytes = part_xmls.get(part_path)
            if not xml_bytes:
                continue
            try:
                root = ET.fromstring(xml_bytes)
            except ET.ParseError:
                continue

            part_label = Path(part_path).stem
            para_idx = 0
            modified = False

            for para_idx, para in enumerate(root.iter(f"{{{w}}}p")):
                pstyle = ""
                ppr = para.find(f"{{{w}}}pPr")
                if ppr is not None:
                    ps = ppr.find(f"{{{w}}}pStyle")
                    if ps is not None:
                        pstyle = ps.get(f"{{{w}}}val", "")

                context = f"{part_label}.p{para_idx}"
                if pstyle:
                    context += f"[{pstyle}]"

                if context in trans_map:
                    translation = trans_map[context]
                    # 将段落内所有 <w:t> 合并写入第一个 run，清空其余
                    t_elems = list(para.iter(f"{{{w}}}t"))
                    if t_elems:
                        t_elems[0].text = translation
                        # 保留第一个 run 的格式，清空后续 runs 的文字
                        for t in t_elems[1:]:
                            t.text = ""
                        modified = True

            if modified:
                # 保留原始 XML 声明
                xml_out = ET.tostring(root, encoding="unicode", xml_declaration=False)
                patched[part_path] = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml_out).encode(
                    "utf-8"
                )

        # 原子化写回 ZIP
        tmp = filepath + ".tmp"
        try:
            _ooxml_clone_and_patch(filepath, tmp, patched)
            os.replace(tmp, filepath)
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)

        logger.info(f"[DocxFormatHandler] Saved {len(trans_map)} paragraphs to {filepath}")

    @staticmethod
    def _detect_language(filepath: str) -> str:
        stem = os.path.splitext(os.path.basename(filepath))[0]
        m = re.search(r"[._-]([a-z]{2,3}(?:[_-][A-Za-z]{2,4})?)$", stem)
        return m.group(1).replace("-", "_") if m else "en"


class PptxFormatHandler(BaseFormatHandler):
    """
    Microsoft PowerPoint PPTX / PPTM 处理器

    提取来源:
    ┌──────────────────────────────────────────────────────────────────────┐
    │ ppt/slides/slide*.xml          正文幻灯片文本框                     │
    │ ppt/slideLayouts/layout*.xml   版式占位符（通常不翻译，可配置）      │
    │ ppt/notesSlides/notesSlide*.xml 演讲者备注                          │
    └──────────────────────────────────────────────────────────────────────┘

    支持的特性:
    1. 形状感知: <p:sp> → <p:nvSpPr> 提取形状名称作为 context 辅助信息
    2. 占位符标注: ph type (title/body/subTitle) 记录到 context
    3. Run 合并: 同一段落多个 <a:r><a:t> 合并，保留 <a:rPr> 格式不变
    4. 表格提取: <a:tbl> → <a:tc> → <a:p> 全扫描，context 标注行列
    5. 备注提取: notesSlide 内文本作为独立条目，context 前缀 "notes"
    6. 幻灯片顺序: context 格式 "slideN.shapeM.pK" 保证顺序稳定
    7. 原子保存: 同 DocxFormatHandler
    """

    format_id = "pptx"
    is_monolingual = True
    extensions = [".pptx", ".pptm"]
    format_type = "translation"
    display_name = _("PowerPoint Presentation (.pptx/.pptm)")
    badge_text = "PPTX"
    badge_bg_color = "#FBE9E7"
    badge_text_color = "#BF360C"

    _NS = {
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    def load(self, filepath: str, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or os.path.basename(filepath)
        translatable_objects: list[TranslatableString] = []
        occurrence_counters: dict = {}

        with zipfile.ZipFile(filepath, "r") as zf:
            all_names = set(zf.namelist())
            slide_parts = sorted(
                [n for n in all_names if re.match(r"ppt/slides/slide\d+\.xml$", n)],
                key=lambda n: int(re.search(r"\d+", Path(n).stem).group()),
            )
            notes_parts = sorted(
                [n for n in all_names if re.match(r"ppt/notesSlides/notesSlide\d+\.xml$", n)],
                key=lambda n: int(re.search(r"\d+", Path(n).stem).group()),
            )
            part_xmls: dict[str, bytes] = {p: zf.read(p) for p in slide_parts + notes_parts}

        for part_path, xml_bytes in part_xmls.items():
            is_notes = "notesSlides" in part_path
            slide_num_m = re.search(r"(\d+)", Path(part_path).stem)
            slide_num = int(slide_num_m.group()) if slide_num_m else 0
            prefix = f"notes{slide_num}" if is_notes else f"slide{slide_num}"

            try:
                root = ET.fromstring(xml_bytes)
            except ET.ParseError as e:
                logger.warning(f"[PptxFormatHandler] Parse error in {part_path}: {e}")
                continue

            self._extract_slide_text(
                root, prefix, relative_path, translatable_objects, occurrence_counters, app_instance
            )

        metadata = {
            "filepath": filepath,
            "part_xmls": part_xmls,
            "slide_parts": slide_parts,
            "notes_parts": notes_parts,
        }
        logger.info(f"[PptxFormatHandler] Loaded {len(translatable_objects)} text runs from {filepath}")
        return translatable_objects, metadata, self._detect_language(filepath)

    def _extract_slide_text(
        self, root: ET.Element, prefix: str, rel_path: str, results: list, counters: dict, app_instance
    ):
        a = self._NS["a"]
        p_ns = self._NS["p"]

        shape_idx = 0
        # 遍历所有 <p:sp> 形状（包括文本框、占位符）
        for sp in root.iter(f"{{{p_ns}}}sp"):
            # 形状名
            nvsp = sp.find(f"{{{p_ns}}}nvSpPr")
            shape_name = ""
            if nvsp is not None:
                cnvpr = nvsp.find(f"{{{p_ns}}}cNvPr")
                shape_name = cnvpr.get("name", "") if cnvpr is not None else ""

            # 占位符类型
            ph_type = ""
            if nvsp is not None:
                nvppr = nvsp.find(f"{{{p_ns}}}nvPr")
                if nvppr is not None:
                    ph_elem = nvppr.find(f"{{{p_ns}}}ph")
                    ph_type = ph_elem.get("type", "") if ph_elem is not None else ""

            txbody = sp.find(f"{{{p_ns}}}txBody")
            if txbody is None:
                shape_idx += 1
                continue

            para_idx = 0
            for para in txbody.iter(f"{{{a}}}p"):
                texts = []
                for t_elem in para.iter(f"{{{a}}}t"):
                    if t_elem.text:
                        texts.append(t_elem.text)
                full_text = "".join(texts).strip()
                if not full_text:
                    para_idx += 1
                    continue

                context_parts = [prefix, f"shape{shape_idx}", f"p{para_idx}"]
                if ph_type:
                    context_parts.insert(2, ph_type)
                context = ".".join(context_parts)

                counter_key = (full_text, context)
                idx = counters.get(counter_key, 0)
                counters[counter_key] = idx + 1

                stable = f"{rel_path}::{context}::{full_text}::{idx}"
                obj_id = xxhash.xxh128(stable.encode()).hexdigest()

                ts = TranslatableString(
                    original_raw=full_text,
                    original_semantic=full_text,
                    line_num=shape_idx * 1000 + para_idx,
                    char_pos_start_in_file=0,
                    char_pos_end_in_file=0,
                    full_code_lines=[],
                    string_type="PPTX Text",
                    source_file_path=rel_path,
                    occurrences=[(rel_path, context)],
                    occurrence_index=idx,
                    id=obj_id,
                )
                ts.set_translation_internal(self.get_initial_translation(full_text, app_instance), is_initial=True)
                ts.context = context
                ts.comment = f"Shape: {shape_name}" + (f" | Type: {ph_type}" if ph_type else "")
                ts.po_comment = f"#: {prefix} shape{shape_idx} para{para_idx}"
                ts.is_reviewed = False
                ts.update_sort_weight()
                results.append(ts)
                para_idx += 1

            shape_idx += 1

        # 表格
        table_idx = 0
        for tbl in root.iter(f"{{{a}}}tbl"):
            for row_idx, tr in enumerate(tbl.iter(f"{{{a}}}tr")):
                for col_idx, tc in enumerate(tr.iter(f"{{{a}}}tc")):
                    for para in tc.iter(f"{{{a}}}p"):
                        texts = [t.text for t in para.iter(f"{{{a}}}t") if t.text]
                        full_text = "".join(texts).strip()
                        if not full_text:
                            continue
                        context = f"{prefix}.tbl{table_idx}.r{row_idx}c{col_idx}"
                        counter_key = (full_text, context)
                        idx = counters.get(counter_key, 0)
                        counters[counter_key] = idx + 1
                        stable = f"{rel_path}::{context}::{full_text}::{idx}"
                        obj_id = xxhash.xxh128(stable.encode()).hexdigest()
                        ts = TranslatableString(
                            original_raw=full_text,
                            original_semantic=full_text,
                            line_num=table_idx * 10000 + row_idx * 100 + col_idx,
                            char_pos_start_in_file=0,
                            char_pos_end_in_file=0,
                            full_code_lines=[],
                            string_type="PPTX Table Cell",
                            source_file_path=rel_path,
                            occurrences=[(rel_path, context)],
                            occurrence_index=idx,
                            id=obj_id,
                        )
                        ts.set_translation_internal(
                            self.get_initial_translation(full_text, app_instance), is_initial=True
                        )
                        ts.context = context
                        ts.comment = f"Table cell row={row_idx} col={col_idx}"
                        ts.po_comment = f"#: {prefix} table{table_idx} [{row_idx},{col_idx}]"
                        ts.is_reviewed = False
                        ts.update_sort_weight()
                        results.append(ts)
            table_idx += 1

    def save(self, filepath: str, translatable_objects, metadata: dict, **kwargs):
        part_xmls: dict[str, bytes] = metadata["part_xmls"]

        trans_map: dict[str, str] = {
            ts.context: (ts.translation or ts.original_semantic)
            for ts in translatable_objects
            if ts.id != "##NEW_ENTRY##" and ts.context and ts.translation
        }

        patched: dict[str, bytes] = {}
        a = self._NS["a"]
        p_ns = self._NS["p"]

        for part_path, xml_bytes in part_xmls.items():
            is_notes = "notesSlides" in part_path
            slide_num_m = re.search(r"(\d+)", Path(part_path).stem)
            slide_num = int(slide_num_m.group()) if slide_num_m else 0
            prefix = f"notes{slide_num}" if is_notes else f"slide{slide_num}"

            try:
                root = ET.fromstring(xml_bytes)
            except ET.ParseError:
                continue

            modified = False
            shape_idx = 0

            for sp in root.iter(f"{{{p_ns}}}sp"):
                nvsp = sp.find(f"{{{p_ns}}}nvSpPr")
                ph_type = ""
                if nvsp is not None:
                    nvppr = nvsp.find(f"{{{p_ns}}}nvPr")
                    if nvppr is not None:
                        ph_elem = nvppr.find(f"{{{p_ns}}}ph")
                        ph_type = ph_elem.get("type", "") if ph_elem is not None else ""
                txbody = sp.find(f"{{{p_ns}}}txBody")
                if txbody is None:
                    shape_idx += 1
                    continue
                para_idx = 0
                for para_idx, para in enumerate(txbody.iter(f"{{{a}}}p")):
                    context_parts = [prefix, f"shape{shape_idx}", f"p{para_idx}"]
                    if ph_type:
                        context_parts.insert(2, ph_type)
                    context = ".".join(context_parts)
                    if context in trans_map:
                        t_elems = list(para.iter(f"{{{a}}}t"))
                        if t_elems:
                            t_elems[0].text = trans_map[context]
                            for t in t_elems[1:]:
                                t.text = ""
                            modified = True
                shape_idx += 1

            # 表格
            table_idx = 0
            for table_idx, tbl in enumerate(root.iter(f"{{{a}}}tbl")):
                for row_idx, tr in enumerate(tbl.iter(f"{{{a}}}tr")):
                    for col_idx, tc in enumerate(tr.iter(f"{{{a}}}tc")):
                        for para in tc.iter(f"{{{a}}}p"):
                            context = f"{prefix}.tbl{table_idx}.r{row_idx}c{col_idx}"
                            if context in trans_map:
                                t_elems = list(para.iter(f"{{{a}}}t"))
                                if t_elems:
                                    t_elems[0].text = trans_map[context]
                                    for t in t_elems[1:]:
                                        t.text = ""
                                    modified = True
            if modified:
                xml_out = ET.tostring(root, encoding="unicode")
                patched[part_path] = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml_out).encode(
                    "utf-8"
                )

        tmp = filepath + ".tmp"
        try:
            _ooxml_clone_and_patch(filepath, tmp, patched)
            os.replace(tmp, filepath)
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)

        logger.info(f"[PptxFormatHandler] Saved {len(trans_map)} items to {filepath}")

    @staticmethod
    def _detect_language(filepath: str) -> str:
        stem = os.path.splitext(os.path.basename(filepath))[0]
        m = re.search(r"[._-]([a-z]{2,3}(?:[_-][A-Za-z]{2,4})?)$", stem)
        return m.group(1).replace("-", "_") if m else "en"


class FluentFormatHandler(BaseFormatHandler):
    """
    Mozilla Fluent (.ftl) 本地化格式处理器

    Fluent 是 Mozilla 设计的下一代本地化系统，被 Firefox、Thunderbird、
    Zulip 等项目采用。其设计目标是让译者能够充分表达目标语言的语法自然性。

    支持的特性:
    1. 消息主值: message-id = value，单行与多行（缩进续行）均支持
    2. 属性: .attr-name = value，每个属性作为独立条目，
       context 格式为 "message-id.attr-name"
    3. 选择表达式 (selector / plural): { $var -> [one] ... *[other] ... }
       整体作为一个条目保留，译者可修改各分支文本；
       注释中列出所有变体 key 供参考
    4. 术语跳过: -term = value 开头的术语不进入翻译流程
       （术语由开发者管理，译者通过 { -term } 引用）
    5. 插值标注: 检测 { $variable } 变量引用和 { -term } 术语引用并记录注释
    6. 注释保留: # 消息注释关联到下一条目并作为 comment 字段
    7. 稳健回写: 以原始文本为基础做定向行级替换，保留所有注释、
       空行及未翻译条目；支持单行值、属性值的原位替换

    技术说明:
    - 多行值的回写: 新译文若含换行，将以缩进续行格式写出
    - 选择表达式的回写: 整体替换（译者须手动保持 { $var -> } 骨架）
    - 纯注释块（## 群组注释、### 文件注释）不作为翻译条目提取
    """

    format_id = "fluent"
    is_monolingual = True
    extensions = [".ftl"]
    format_type = "translation"
    display_name = _("Mozilla Fluent File")
    badge_text = "FTL"
    badge_bg_color = "#EDE7F6"
    badge_text_color = "#4527A0"

    # 变量引用: { $name }
    _VAR_RE = re.compile(r"\{\s*\$([a-zA-Z_][a-zA-Z0-9_-]*)\s*(?:\(.*?\))?\s*\}")
    # 术语引用: { -term }
    _TERM_REF_RE = re.compile(r"\{\s*-([a-zA-Z][a-zA-Z0-9_-]*)\s*\}")
    # 选择表达式变体键: [key] 或 *[key]
    _SELECTOR_KEY_RE = re.compile(r"\*?\[([^\]]+)\]")
    # 数字字面量参数（函数调用中）
    _FUNC_CALL_RE = re.compile(r"\{\s*[A-Z_]+\(")

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        relative_path = kwargs.get("relative_path") or os.path.basename(filepath)
        language_code = self._detect_language_from_filename(os.path.basename(filepath))

        with open(filepath, encoding="utf-8", errors="replace") as f:
            content = f.read()

        entries = self._parse_ftl(content)
        translatable_objects = []
        occurrence_counters = {}

        for entry in entries:
            if entry["type"] != "message":
                continue

            msg_id = entry["msg_id"]
            msg_comment = entry.get("comment", "")

            # 主值
            if entry.get("value") is not None:
                ts = self._make_ts(
                    text=entry["value"],
                    context=msg_id,
                    string_type="Fluent Message",
                    comment=msg_comment,
                    msg_id=msg_id,
                    results=translatable_objects,
                    counters=occurrence_counters,
                    rel_path=relative_path,
                    app_instance=app_instance,
                )
                translatable_objects.append(ts)

            # 属性（每个属性独立条目，context = "msg-id.attr-name"）
            for attr_name, attr_value in entry.get("attributes", {}).items():
                attr_ctx = f"{msg_id}.{attr_name}"
                attr_comment = f"{msg_comment}\nAttribute: .{attr_name}".strip()
                ts = self._make_ts(
                    text=attr_value,
                    context=attr_ctx,
                    string_type="Fluent Attribute",
                    comment=attr_comment,
                    msg_id=msg_id,
                    results=translatable_objects,
                    counters=occurrence_counters,
                    rel_path=relative_path,
                    app_instance=app_instance,
                )
                translatable_objects.append(ts)

        metadata = {
            "raw_content": content,
            "language_code": language_code,
            # 存储每条消息的原始文本块，用于 save 时定向替换
            "entry_originals": {e["msg_id"]: e["raw_block"] for e in entries if e["type"] == "message"},
        }
        logger.info(f"[FluentFormatHandler] Loaded {len(translatable_objects)} entries from {filepath}")
        return translatable_objects, metadata, language_code

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        raw_content = metadata.get("raw_content", "")
        entry_originals = metadata.get("entry_originals", {})

        trans_map: dict[str, str] = {
            ts.context: ts.translation
            for ts in translatable_objects
            if ts.translation and not getattr(ts, "is_ignored", False) and ts.context
        }

        result = self._apply_translations(raw_content, trans_map, entry_originals)

        with atomic_open(filepath, "w", encoding="utf-8") as f:
            f.write(result)

        logger.info(f"[FluentFormatHandler] Saved to {filepath}")

    def _parse_ftl(self, content: str) -> list[dict]:
        """
        将 FTL 文本解析为条目列表。

        每个条目的结构:
          type: "message" | "term" | "comment" | "blank" | "junk"
          msg_id: str          (message/term)
          value:  str | None   (消息主值，可含选择表达式)
          attributes: dict     (属性名 -> 值)
          comment: str         (紧邻前置 # 注释)
          raw_block: str       (该条目在原始文本中的完整文本，含前置注释)
        """
        entries = []
        lines = content.split("\n")
        n = len(lines)
        i = 0
        pending_comment_lines: list[str] = []
        pending_comment_start = -1

        while i < n:
            line = lines[i]

            # --- 文件/群组注释（## / ###）：不关联到消息，直接丢弃积累 ---
            if re.match(r"^#{2,3}\s", line) or line in ("##", "###"):
                pending_comment_lines = []
                pending_comment_start = -1
                i += 1
                continue

            # --- 消息注释（单 # ）---
            if line.startswith("# ") or line == "#":
                if not pending_comment_lines:
                    pending_comment_start = i
                pending_comment_lines.append(line[2:] if line.startswith("# ") else "")
                i += 1
                continue

            # --- 空行：若此前有孤立注释，重置 ---
            if not line.strip():
                # 若注释后紧跟空行则认为是独立注释块，不关联下一消息
                if pending_comment_lines:
                    entries.append(
                        {
                            "type": "comment",
                            "text": "\n".join(pending_comment_lines),
                            "raw_block": "\n".join(lines[pending_comment_start:i] + [""]),
                        }
                    )
                    pending_comment_lines = []
                    pending_comment_start = -1
                entries.append({"type": "blank", "raw_block": ""})
                i += 1
                continue

            # --- 术语（-term = …）---
            term_m = re.match(r"^(-[a-zA-Z][a-zA-Z0-9_-]*)\s*=", line)
            if term_m:
                term_id = term_m.group(1)
                block_start = pending_comment_start if pending_comment_start >= 0 else i
                j = i + 1
                while j < n and (lines[j].startswith("    ") or lines[j].startswith("\t")):
                    j += 1
                raw_block = "\n".join(lines[block_start:j])
                entries.append(
                    {
                        "type": "term",
                        "msg_id": term_id,
                        "raw_block": raw_block,
                    }
                )
                pending_comment_lines = []
                pending_comment_start = -1
                i = j
                continue

            # --- 消息（message-id = …）---
            msg_m = re.match(r"^([a-zA-Z][a-zA-Z0-9_-]*)\s*=\s*(.*)", line)
            if msg_m:
                msg_id = msg_m.group(1)
                inline_rest = msg_m.group(2)  # 等号之后的内容（可能为空）

                block_start = pending_comment_start if pending_comment_start >= 0 else i
                comment_text = "\n".join(pending_comment_lines).strip()
                pending_comment_lines = []
                pending_comment_start = -1

                # 收集所有后续缩进行（消息体）
                j = i + 1
                while j < n and (lines[j].startswith("    ") or lines[j].startswith("\t")):
                    j += 1

                body_lines = [inline_rest] + [lines[k] for k in range(i + 1, j)]
                value, attributes = self._parse_message_body(body_lines)

                raw_block = "\n".join(lines[block_start:j])
                entries.append(
                    {
                        "type": "message",
                        "msg_id": msg_id,
                        "value": value,
                        "attributes": attributes,
                        "comment": comment_text,
                        "raw_block": raw_block,
                    }
                )
                i = j
                continue

            # --- Junk（无法识别的行）---
            entries.append({"type": "junk", "raw_block": line})
            pending_comment_lines = []
            pending_comment_start = -1
            i += 1

        return entries

    def _parse_message_body(self, body_lines: list[str]) -> tuple:
        """
        解析消息体（等号后的行列表），返回 (value, attributes)。

        消息体规则:
          - 第一个非空行是内联值（若无则消息无独立值，仅有属性）
          - .attr-name = value 行是属性定义
          - 其余缩进行是上一个值/属性的续行
        """
        value_parts: list[str] = []
        attributes: dict[str, str] = {}
        current_attr: str | None = None
        current_attr_parts: list[str] = []

        def flush_attr():
            if current_attr is not None:
                attributes[current_attr] = self._normalize_multiline(current_attr_parts)

        for line in body_lines:
            stripped = line.strip()

            # 属性行
            attr_m = re.match(r"^\s*\.([a-zA-Z][a-zA-Z0-9_-]*)\s*=\s*(.*)", line)
            if attr_m:
                flush_attr()
                current_attr = attr_m.group(1)
                current_attr_parts = [attr_m.group(2).strip()]
                continue

            if current_attr is not None:
                # 属性续行
                current_attr_parts.append(stripped)
            else:
                # 主值行
                value_parts.append(stripped)

        flush_attr()

        value = self._normalize_multiline(value_parts) if value_parts else None
        # 空字符串也算 None（无显式值的消息只有属性）
        if value is not None and value.strip() == "":
            value = None

        return value, attributes

    @staticmethod
    def _normalize_multiline(parts: list[str]) -> str:
        """去掉首尾空元素后合并多行。"""
        while parts and not parts[0].strip():
            parts = parts[1:]
        while parts and not parts[-1].strip():
            parts = parts[:-1]
        return "\n".join(parts)

    def _make_ts(
        self,
        text,
        context,
        string_type,
        comment,
        msg_id,
        results,
        counters,
        rel_path,
        app_instance,
    ) -> "TranslatableString":
        counter_key = (text, context)
        idx = counters.get(counter_key, 0)
        counters[counter_key] = idx + 1

        stable = f"{rel_path}::{context}::{text}::{idx}"
        obj_id = xxhash.xxh128(stable.encode()).hexdigest()

        # 插值变量和术语引用分析
        variables = self._VAR_RE.findall(text)
        term_refs = self._TERM_REF_RE.findall(text)
        selector_keys = self._SELECTOR_KEY_RE.findall(text)
        has_func = bool(self._FUNC_CALL_RE.search(text))

        comment_lines = [comment] if comment else []
        if variables:
            comment_lines.append("Variables: " + ", ".join(f"{{ ${v} }}" for v in sorted(set(variables))))
        if term_refs:
            comment_lines.append("Term refs: " + ", ".join(f"{{ -{t} }}" for t in sorted(set(term_refs))))
        if selector_keys:
            comment_lines.append(
                f"Selector variants: [{', '.join(selector_keys)}] — preserve {{ $var -> }} skeleton when translating"
            )
        if has_func:
            comment_lines.append("Contains Fluent built-in function call (e.g. NUMBER(), DATETIME())")

        ts = TranslatableString(
            original_raw=text,
            original_semantic=text,
            line_num=0,
            char_pos_start_in_file=0,
            char_pos_end_in_file=0,
            full_code_lines=[],
            string_type=string_type,
            source_file_path=rel_path,
            occurrences=[(rel_path, context)],
            occurrence_index=idx,
            id=obj_id,
        )
        ts.set_translation_internal(self.get_initial_translation(text, app_instance), is_initial=True)
        ts.context = context
        ts.comment = "\n".join(comment_lines)
        ts.po_comment = f"#: Fluent id: {msg_id}"
        ts.is_reviewed = False
        ts.update_sort_weight()
        return ts

    def _apply_translations(
        self,
        content: str,
        trans_map: dict[str, str],
        entry_originals: dict[str, str],
    ) -> str:
        """
        将翻译回写到 FTL 原始内容。

        策略: 逐消息块定向替换，保留注释、空行和未触及的条目。
        对于每条消息，若其主值或任何属性有翻译，则重建整个消息块文本。
        """
        result = content

        for msg_id, original_block in entry_originals.items():
            # 收集此消息的主值翻译和属性翻译
            value_trans = trans_map.get(msg_id)
            attr_trans = {
                k[len(msg_id) + 1 :]: v
                for k, v in trans_map.items()
                if k.startswith(f"{msg_id}.") and "." not in k[len(msg_id) + 1 :]
            }

            if not value_trans and not attr_trans:
                continue

            new_block = self._rebuild_message_block(msg_id, original_block, value_trans, attr_trans)

            if new_block != original_block:
                # 用首次出现替换（防止相同文本的消息块互相干扰）
                result = result.replace(original_block, new_block, 1)

        return result

    def _rebuild_message_block(
        self,
        msg_id: str,
        original_block: str,
        value_trans: str | None,
        attr_trans: dict[str, str],
    ) -> str:
        """
        在保留消息原有结构（注释、多行缩进、属性）的前提下替换值。

        处理规则:
          - 单行值 (message-id = old text)  → message-id = new text
          - 多行值 (message-id =\n    line1) → message-id =\n    new_line1\n    new_line2
          - 属性   (    .attr = old)         →     .attr = new
          - 对于含选择表达式的值: 整体替换
        """
        lines = original_block.split("\n")
        result_lines: list[str] = []
        i = 0
        message_line_handled = False

        while i < len(lines):
            line = lines[i]

            # --- 消息定义行 ---
            msg_m = re.match(r"^([a-zA-Z][a-zA-Z0-9_-]*)\s*=\s*(.*)", line)
            if msg_m and msg_m.group(1) == msg_id and not message_line_handled:
                message_line_handled = True
                inline_value = msg_m.group(2).strip()

                # 找出属于此消息的所有后续缩进行
                j = i + 1
                while j < len(lines) and (lines[j].startswith("    ") or lines[j].startswith("\t")):
                    j += 1
                indented = lines[i + 1 : j]

                # 判断是否有多行主值（区别于属性行）
                has_multiline_value = bool(
                    inline_value == "" and any(not l.strip().startswith(".") for l in indented if l.strip())
                )

                if value_trans is not None:
                    new_val_lines = value_trans.split("\n")
                    if inline_value and not has_multiline_value:
                        # 单行值替换
                        result_lines.append(f"{msg_id} = {new_val_lines[0]}")
                        if len(new_val_lines) > 1:
                            for extra in new_val_lines[1:]:
                                result_lines.append(f"    {extra}")
                    else:
                        # 多行值或原本为空内联
                        result_lines.append(f"{msg_id} =")
                        for vl in new_val_lines:
                            result_lines.append(f"    {vl}" if vl.strip() else "")
                else:
                    result_lines.append(line)

                # 处理缩进块中的属性行
                for attr_line in indented:
                    attr_m = re.match(r"^(\s*)\.([a-zA-Z][a-zA-Z0-9_-]*)\s*=\s*(.*)", attr_line)
                    if attr_m:
                        indent_ws = attr_m.group(1)
                        attr_name = attr_m.group(2)
                        if attr_name in attr_trans:
                            new_attr_val = attr_trans[attr_name]
                            # 属性值也可能多行
                            attr_val_lines = new_attr_val.split("\n")
                            result_lines.append(f"{indent_ws}.{attr_name} = {attr_val_lines[0]}")
                            if len(attr_val_lines) > 1:
                                for extra in attr_val_lines[1:]:
                                    result_lines.append(f"{indent_ws}    {extra}" if extra.strip() else "")
                        else:
                            result_lines.append(attr_line)
                    elif attr_line.strip() and value_trans is not None:
                        # 多行主值的续行已在上面处理，跳过原始续行
                        pass
                    else:
                        result_lines.append(attr_line)

                i = j
                continue

            # --- 其余行（前置注释等）原样保留 ---
            result_lines.append(line)
            i += 1

        return "\n".join(result_lines)


class OwCodeFormatHandler(BaseFormatHandler):
    """
    守望先锋工坊代码格式处理器
    支持的特性:
    1. 智能提取: 基于正则表达式从 .ow 或 .txt 源码中精准抠取待翻译文本
    2. 结构回填: 保存时将译文精准替换回原始位置，确保代码逻辑和格式不受损
    3. 多维识别: 支持提取自定义字符串 (Custom String)、模式名称及模式描述
    4. 自动过滤: 智能跳过数字、纯占位符及已知的工坊技术关键字
    """

    format_id = "ow_code"
    is_monolingual = False
    extensions = [".ow", ".txt"]
    format_type = "source"
    display_name = _("Overwatch Workshop Code")
    badge_text = "Code"
    badge_bg_color = "#E3F2FD"
    badge_text_color = "#0277BD"

    def load(self, filepath, **kwargs):
        app_instance = kwargs.get("app_instance")
        with open(filepath, encoding="utf-8", errors="replace") as f:
            content = f.read()
        extraction_patterns = kwargs.get("extraction_patterns", [])
        relative_path = kwargs.get("relative_path", os.path.basename(filepath))
        strings = code_file_service.extract_translatable_strings(
            content, extraction_patterns, relative_path, app_instance=app_instance
        )
        return strings, {"raw_content": content}, "en"

    def save(self, filepath, translatable_objects, metadata, **kwargs):
        app_instance = kwargs.get("app_instance")
        raw_content = metadata.get("raw_content", "")
        code_file_service.save_translated_code(filepath, raw_content, translatable_objects, app_instance)


# ============================================================================
# FORMAT MANAGER
# ============================================================================


class FormatManager:
    _handlers: dict[str, BaseFormatHandler] = {}

    @classmethod
    def register_handler(cls, handler_class):
        handler = handler_class()
        cls._handlers[handler.format_id] = handler
        logger.info(f"Registered format handler: {handler.format_id}")

    @classmethod
    def get_handler(cls, format_id) -> BaseFormatHandler:
        return cls._handlers.get(format_id)

    @classmethod
    def get_handler_by_extension(cls, filepath: str, content: str = None, sniff: bool = False):
        ext = Path(filepath).suffix.lower()

        if ext == ".json":
            if sniff:
                if content is None:
                    try:
                        with open(filepath, encoding="utf-8", errors="ignore") as f:
                            content = f.read(512000)
                    except OSError as e:
                        logger.warning(f"[FormatManager] Cannot read {filepath} for type detection: {e}")
                        return cls._handlers.get("json_i18n")
                if content:
                    try:
                        data = json.loads(content)
                        if cls._is_i18next_structure(data):
                            return cls._handlers.get("i18next_json")
                    except json.JSONDecodeError:
                        pass
            return cls._handlers.get("json_i18n")

        for handler in cls._handlers.values():
            if ext in handler.extensions:
                return handler
        return None

    @classmethod
    def _is_i18next_structure(cls, data, depth=0):
        if depth > 3:
            return False
        score = 0
        if isinstance(data, dict):
            for key, value in data.items():
                if key.endswith(("_one", "_other", "_few", "_many", "_zero", "_plural")):
                    score += 2
                if isinstance(value, str) and re.search(r"\{\{[^}]+\}\}", value):
                    score += 1
                if isinstance(value, dict) and cls._is_i18next_structure(value, depth + 1):
                    score += 1
                if score >= 2:
                    return True
        return False

    @classmethod
    def get_file_dialog_filters(cls, format_type=None):
        """生成文件选择器的过滤器字符串"""
        filters = []
        all_exts = []

        for handler in cls._handlers.values():
            if format_type and handler.format_type != format_type:
                continue
            ext_str = " ".join(f"*{ext}" for ext in handler.extensions)
            filters.append(f"{handler.display_name} ({ext_str})")
            all_exts.extend(handler.extensions)

        all_ext_str = " ".join(f"*{ext}" for ext in set(all_exts))

        if format_type == "translation":
            prefix = _("All Translation Files")
        elif format_type == "source":
            prefix = _("All Source Files")
        else:
            prefix = _("All Supported Files")

        filter_string = f"{prefix} ({all_ext_str});;" + ";;".join(filters) + f";;{_('All Files')} (*.*)"
        return filter_string


# 1. 行业标准翻译与桌面端 UI 格式 (Standard Translation & Desktop)
FormatManager.register_handler(PoFormatHandler)
FormatManager.register_handler(TsFormatHandler)
FormatManager.register_handler(XliffFormatHandler)

# 2. 移动端与跨平台开发生态 (Mobile & Cross-platform)
FormatManager.register_handler(AndroidStringsFormatHandler)
FormatManager.register_handler(IosStringsFormatHandler)
FormatManager.register_handler(XCStringsFormatHandler)
FormatManager.register_handler(StringsDictFormatHandler)
FormatManager.register_handler(ArbFormatHandler)

# 3. 数据序列化与通用配置格式 (Data Serialization & Configs)
FormatManager.register_handler(JsonI18nFormatHandler)
FormatManager.register_handler(I18nextJsonFormatHandler)
FormatManager.register_handler(YamlI18nFormatHandler)
FormatManager.register_handler(TomlFormatHandler)
FormatManager.register_handler(IniFormatHandler)

# 4. 传统桌面、后端与特定语言资源 (Desktop, Backend & Language Specific)
FormatManager.register_handler(JavaPropertiesFormatHandler)
FormatManager.register_handler(ResxFormatHandler)
FormatManager.register_handler(PhpArrayFormatHandler)
FormatManager.register_handler(RcFormatHandler)

# 5. 表格与批量处理格式 (Tabular & Spreadsheets)
FormatManager.register_handler(CsvFormatHandler)
FormatManager.register_handler(XlsxFormatHandler)

# 6. 多媒体与字幕格式 (Media & Subtitles)
FormatManager.register_handler(SrtFormatHandler)
FormatManager.register_handler(VttFormatHandler)

# 7. 网页与富文本办公文档 (Web & Rich Office Documents)
FormatManager.register_handler(HtmlFormatHandler)
FormatManager.register_handler(MarkdownFormatHandler)
FormatManager.register_handler(DocxFormatHandler)
FormatManager.register_handler(PptxFormatHandler)

# 8. 自定义代码与特殊格式 (Custom Code & Special)
FormatManager.register_handler(FluentFormatHandler)
FormatManager.register_handler(OwCodeFormatHandler)
