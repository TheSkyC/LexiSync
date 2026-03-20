# Copyright (c) 2025-2026, TheSkyC
# SPDX-License-Identifier: Apache-2.0

from contextlib import contextmanager
from datetime import datetime
from functools import lru_cache
import hashlib
import json
import logging
import os
import re
import sqlite3
import threading

from rapidfuzz import fuzz

from lexisync.utils.localization import _

logger = logging.getLogger(__name__)

MANIFEST_FILE = "manifest.json"
DB_FILE = "tm.db"


class TMService:
    def __init__(self):
        self.project_db_path = None
        self.global_db_path = None
        self._lock = threading.RLock()
        self._conns = {}

        self.get_fuzzy_matches = lru_cache(maxsize=128)(self._do_actual_search)

    def _get_conn(self, db_path):
        """获取或创建持久化连接"""
        if db_path not in self._conns:
            conn = sqlite3.connect(db_path, timeout=10.0, check_same_thread=False)
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA synchronous=NORMAL")
            conn.execute("PRAGMA cache_size=-2000")  # 2MB 缓存
            conn.row_factory = sqlite3.Row
            self._conns[db_path] = conn
        return self._conns[db_path]

    def connect_databases(self, global_tm_path: str | None, project_tm_path: str | None = None):
        with self._lock:
            self.disconnect_databases()

            if global_tm_path:
                os.makedirs(global_tm_path, exist_ok=True)
                self.global_db_path = os.path.join(global_tm_path, DB_FILE)
                with self._get_db_connection(self.global_db_path) as conn:
                    self._create_schema(conn)

            if project_tm_path:
                os.makedirs(project_tm_path, exist_ok=True)
                self.project_db_path = os.path.join(project_tm_path, DB_FILE)
                with self._get_db_connection(self.project_db_path) as conn:
                    self._create_schema(conn)

    def disconnect_databases(self):
        with self._lock:
            self.project_db_path = None
            self.global_db_path = None

    @contextmanager
    def _get_db_connection(self, db_path: str):
        conn = None
        try:
            conn = sqlite3.connect(db_path, timeout=30.0, check_same_thread=False, isolation_level=None)
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA synchronous=NORMAL")
            conn.row_factory = sqlite3.Row
            yield conn
        except Exception as e:
            if conn:
                try:
                    conn.rollback()
                except Exception:
                    pass
            logger.error(f"TM Database connection error for {db_path}: {e}")
            raise
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

    def _create_schema(self, conn: sqlite3.Connection):
        try:
            cursor = conn.cursor()
            cursor.execute("BEGIN IMMEDIATE TRANSACTION")

            # 原始数据表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS translation_units (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_lang TEXT NOT NULL,
                    target_lang TEXT NOT NULL,
                    source_text TEXT NOT NULL,
                    target_text TEXT NOT NULL,
                    source_manifest_key TEXT NOT NULL,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(source_lang, target_lang, source_text)
                );
            """)

            # FTS5 全文搜索虚拟表
            # 使用 unicode61 分词器，支持多语言
            cursor.execute("""
                CREATE VIRTUAL TABLE IF NOT EXISTS tm_search_index USING fts5(
                    source_text,
                    content='translation_units',
                    content_rowid='id',
                    tokenize='unicode61'
                );
            """)

            # 创建触发器
            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS trg_tm_insert AFTER INSERT ON translation_units BEGIN
                    INSERT INTO tm_search_index(rowid, source_text) VALUES (new.id, new.source_text);
                END;
            """)

            # UPDATE 触发器
            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS trg_tm_update AFTER UPDATE ON translation_units BEGIN
                    INSERT INTO tm_search_index(tm_search_index, rowid, source_text) VALUES('delete', old.id, old.source_text);
                    INSERT INTO tm_search_index(rowid, source_text) VALUES (new.id, new.source_text);
                END;
            """)

            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS trg_tm_delete AFTER DELETE ON translation_units BEGIN
                    INSERT INTO tm_search_index(tm_search_index, rowid, source_text)
                    VALUES('delete', old.id, old.source_text);
                END;
            """)

            cursor.execute("COMMIT")
        except Exception as e:
            cursor.execute("ROLLBACK")
            logger.error(f"TM Schema creation failed: {e}")

    def find_conflicts(
        self, db_path: str, source_texts: list[str], source_lang: str, target_lang: str
    ) -> dict[str, dict]:
        """
        Batch check for existing TM entries.
        Returns: { 'source_text': {'id': id, 'original_text': source_text, 'existing_targets': ['target_text']} }
        Note: TM usually enforces unique (source, source_lang, target_lang), so existing_targets will have at most 1 item.
        """
        if not db_path or not os.path.exists(db_path):
            return {}

        conflicts = {}
        unique_sources = list(set(source_texts))

        try:
            with self._get_db_connection(db_path) as conn:
                self._create_schema(conn)

                cursor = conn.cursor()
                chunk_size = 900
                for i in range(0, len(unique_sources), chunk_size):
                    chunk = unique_sources[i : i + chunk_size]
                    placeholders = ",".join("?" for _ in chunk)

                    query = f"""
                        SELECT id, source_text, target_text
                        FROM translation_units
                        WHERE source_text IN ({placeholders})
                          AND source_lang = ?
                          AND target_lang = ?
                    """
                    cursor.execute(query, [*chunk, source_lang, target_lang])

                    rows = cursor.fetchall()
                    for row in rows:
                        term_id = row["id"]
                        src_text = row["source_text"]
                        tgt_text = row["target_text"]

                        # Use exact source text as key
                        if src_text not in conflicts:
                            conflicts[src_text] = {"id": term_id, "original_text": src_text, "existing_targets": []}

                        conflicts[src_text]["existing_targets"].append(tgt_text)

        except Exception as e:
            logger.error(f"Failed to find TM conflicts: {e}")

        return conflicts

    def delete_tm_entry(self, db_path: str, source_text: str, source_lang: str, target_lang: str) -> bool:
        if not source_text.strip() or not db_path:
            return False

        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")
                    cursor.execute(
                        """
                        DELETE FROM translation_units
                        WHERE source_text = ? AND source_lang = ? AND target_lang = ?
                    """,
                        (source_text, source_lang, target_lang),
                    )
                    rowcount = cursor.rowcount
                    cursor.execute("COMMIT")

                    if rowcount > 0:
                        logger.info(f"Deleted TM entry for '{source_text}'")
                        return True
                    return False
            except Exception as e:
                logger.error(f"Failed to delete TM entry: {e}", exc_info=True)
                return False

    def get_translation(
        self, source_text: str, source_lang: str, target_lang: str, db_to_check: str = "all"
    ) -> str | None:
        with self._lock:
            if db_to_check in ("all", "project") and self.project_db_path:
                try:
                    with self._get_db_connection(self.project_db_path) as conn:
                        result = self._query_translation_in_db(conn, source_text, source_lang, target_lang)
                        if result is not None:
                            return result
                except Exception:
                    pass

            if db_to_check in ("all", "global") and self.global_db_path:
                try:
                    with self._get_db_connection(self.global_db_path) as conn:
                        return self._query_translation_in_db(conn, source_text, source_lang, target_lang)
                except Exception:
                    pass
            return None

    def get_entry_count_by_source(self, dir_path: str, source_key: str) -> int:
        if not dir_path or not os.path.exists(dir_path):
            return 0
        db_path = os.path.join(dir_path, DB_FILE)
        if not os.path.exists(db_path):
            return 0

        try:
            with self._get_db_connection(db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM translation_units WHERE source_manifest_key = ?", (source_key,))
                result = cursor.fetchone()
                return result[0] if result else 0
        except Exception as e:
            logger.error(f"Failed to get entry count: {e}")
            return 0

    def get_translations_batch(self, words: list[str], source_lang: str, target_lang: str) -> dict[str, str]:
        """
        Batch query for translations. Returns a dictionary mapping source_text to target_text.
        """
        if not words:
            return {}

        with self._lock:
            all_matches = {}

            # Deduplicate words to query
            unique_words = list(set(words))

            # Query project DB first
            if self.project_db_path:
                try:
                    with self._get_db_connection(self.project_db_path) as conn:
                        matches = self._query_translations_batch_in_db(conn, unique_words, source_lang, target_lang)
                        all_matches.update(matches)
                except Exception as e:
                    logger.warning(f"Project TM batch query failed: {e}")

            # Query global DB for remaining words
            if self.global_db_path:
                remaining_words = [w for w in unique_words if w not in all_matches]
                if remaining_words:
                    try:
                        with self._get_db_connection(self.global_db_path) as conn:
                            matches = self._query_translations_batch_in_db(
                                conn, remaining_words, source_lang, target_lang
                            )
                            all_matches.update(matches)
                    except Exception as e:
                        logger.warning(f"Global TM batch query failed: {e}")

            return all_matches

    def _query_translation_in_db(
        self, conn: sqlite3.Connection, source_text: str, source_lang: str, target_lang: str
    ) -> str | None:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT target_text FROM translation_units WHERE source_text = ? AND source_lang = ? AND target_lang = ?",
            (source_text, source_lang, target_lang),
        )
        row = cursor.fetchone()
        return row["target_text"] if row else None

    def _query_translations_batch_in_db(
        self, conn: sqlite3.Connection, words: list[str], source_lang: str, target_lang: str
    ) -> dict[str, str]:
        """Helper for batch querying a single DB file."""
        if not words:
            return {}

        matches = {}
        chunk_size = 900

        for i in range(0, len(words), chunk_size):
            chunk = words[i : i + chunk_size]
            placeholders = ",".join("?" for _ in chunk)

            query = f"""
                SELECT source_text, target_text
                FROM translation_units
                WHERE source_text IN ({placeholders})
                  AND source_lang = ?
                  AND target_lang = ?
            """

            params = [*chunk, source_lang, target_lang]

            cursor = conn.cursor()
            cursor.execute(query, params)

            for row in cursor.fetchall():
                matches[row["source_text"]] = row["target_text"]

        return matches

    def update_tm_entry(
        self,
        db_path: str,
        source_text: str,
        target_text: str,
        source_lang: str,
        target_lang: str,
        source_key: str = "manual",
    ):
        if not source_text.strip() or not db_path:
            return

        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")
                    cursor.execute(
                        """
                        INSERT OR REPLACE INTO translation_units
                        (source_lang, target_lang, source_text, target_text, source_manifest_key)
                        VALUES (?, ?, ?, ?, ?)
                    """,
                        (source_lang, target_lang, source_text, target_text, source_key),
                    )
                    cursor.execute("COMMIT")
            except Exception as e:
                logger.error(f"Failed to update TM entry: {e}")

        self.get_fuzzy_matches.cache_clear()

    def import_from_file(
        self, source_filepath: str, tm_dir_path: str, source_lang: str, target_lang: str, progress_callback=None
    ) -> tuple[bool, str]:
        with self._lock:
            db_path = os.path.join(tm_dir_path, DB_FILE)
            manifest_path = os.path.join(tm_dir_path, MANIFEST_FILE)

            try:
                os.makedirs(tm_dir_path, exist_ok=True)
                manifest = self._read_manifest(manifest_path)
                file_stats = self._get_file_stats(source_filepath)
                filename = os.path.basename(source_filepath)

                if filename in manifest.get("imported_sources", {}):
                    existing_stats = manifest["imported_sources"][filename]
                    if all(
                        existing_stats.get(k) == file_stats.get(k) for k in ["filesize", "last_modified", "checksum"]
                    ):
                        return True, _("This TM file has already been imported and has not changed.")

                if progress_callback:
                    progress_callback(_("Parsing TM file..."))
                tus_to_import = self._parse_tm_file(source_filepath, source_lang, target_lang)

                if progress_callback:
                    progress_callback(_("Connecting to database..."))
                with self._get_db_connection(db_path) as conn:
                    self._create_schema(conn)
                    self._merge_tus_into_db(conn, tus_to_import, filename, progress_callback)

                file_stats["import_date"] = datetime.now().isoformat() + "Z"
                file_stats["tu_count"] = len(tus_to_import)
                file_stats["source_lang"] = source_lang
                file_stats["target_lang"] = target_lang
                manifest.setdefault("imported_sources", {})[filename] = file_stats
                self._write_manifest(manifest_path, manifest)
                self.get_fuzzy_matches.cache_clear()
                return True, _("Successfully imported {count} TM entries.").format(count=len(tus_to_import))
            except Exception as e:
                logger.error(f"Failed to import TM file '{source_filepath}': {e}", exc_info=True)
                return False, str(e)

    def _query_fuzzy_in_db(self, conn, source_text, source_lang, target_lang, limit):
        tokens = [t for t in re.findall(r"\w+", source_text) if len(t) > 1]
        if not tokens:
            return []
        fts_query = " OR ".join(tokens[:10])

        query = """
            SELECT u.source_text, u.target_text
            FROM tm_search_index i
            JOIN translation_units u ON i.rowid = u.id
            WHERE i.source_text MATCH ?
              AND u.source_lang = ?
              AND u.target_lang = ?
            ORDER BY bm25(tm_search_index)
            LIMIT 100;
        """
        cursor = conn.cursor()
        cursor.execute(query, (fts_query, source_lang, target_lang))
        return cursor.fetchall()

    def _do_actual_search(
        self, source_text: str, source_lang: str, target_lang: str, limit: int = 5, threshold: float = 0.7
    ) -> list[dict]:
        all_candidates = []
        with self._lock:
            for db_path in filter(None, [self.project_db_path, self.global_db_path]):
                try:
                    with self._get_db_connection(db_path) as conn:
                        all_candidates.extend(
                            self._query_fuzzy_in_db(conn, source_text, source_lang, target_lang, limit)
                        )
                except sqlite3.Error as e:
                    logger.warning(f"Database error at {db_path}: {e}")
                    continue

        if not all_candidates:
            return []

        # 2. 精排阶段
        scored_matches = []
        query_len = len(source_text)

        # 打分逻辑
        for cand in all_candidates:
            cand_src = cand["source_text"]
            cand_len = len(cand_src)

            # A. 计算基础相似度
            base_score = fuzz.token_set_ratio(source_text, cand_src) / 100.0

            # B. 快速长度惩罚公式
            # 惩罚因子 alpha = 0.5 (值越大对长度差异越敏感)
            ratio = min(query_len, cand_len) / max(query_len, cand_len)
            penalty = ratio**0.5

            final_score = base_score * penalty

            if final_score >= threshold:
                scored_matches.append(
                    {"score": final_score, "source_text": cand_src, "target_text": cand["target_text"]}
                )

        # 排序并
        scored_matches.sort(key=lambda x: x["score"], reverse=True)

        # 去重
        seen = set()
        unique_results = []
        for m in scored_matches:
            if m["target_text"] not in seen:
                unique_results.append(m)
                seen.add(m["target_text"])
            if len(unique_results) >= limit:
                break

        return unique_results

    def _parse_tm_file(self, filepath: str, source_lang: str, target_lang: str) -> list[dict]:
        __, ext = os.path.splitext(filepath)
        ext = ext.lower()

        if ext == ".xlsx":
            return self._parse_xlsx(filepath, source_lang, target_lang)
        raise ValueError(_("Unsupported file extension: {ext}").format(ext=ext))

    def _parse_xlsx(self, filepath: str, source_lang: str, target_lang: str) -> list[dict]:
        from openpyxl import load_workbook

        tus = []
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2 and row[0] is not None and row[1] is not None:
                source_text = str(row[0])
                target_text = str(row[1])
                tus.append(
                    {
                        "source_lang": source_lang,
                        "target_lang": target_lang,
                        "source_text": source_text,
                        "target_text": target_text,
                    }
                )
        return tus

    def _merge_tus_into_db(self, conn: sqlite3.Connection, tus: list[dict], source_key: str, progress_callback=None):
        cursor = conn.cursor()
        data_to_insert = [
            (tu["source_lang"], tu["target_lang"], tu["source_text"], tu["target_text"], source_key)
            for tu in tus
            if tu.get("source_text") and tu.get("target_text")
        ]

        if not data_to_insert:
            return

        try:
            cursor.execute("BEGIN IMMEDIATE TRANSACTION;")
            if progress_callback:
                progress_callback(_("Inserting {count} TM entries...").format(count=len(data_to_insert)))

            cursor.executemany(
                """
                INSERT OR REPLACE INTO translation_units
                (source_lang, target_lang, source_text, target_text, source_manifest_key)
                VALUES (?, ?, ?, ?, ?)
            """,
                data_to_insert,
            )

            cursor.execute("COMMIT")
        except Exception as e:
            cursor.execute("ROLLBACK")
            raise OSError(f"Database operation failed: {e}") from e

    def remove_source(self, source_key: str, tm_dir_path: str) -> tuple[bool, str]:
        with self._lock:
            db_path = os.path.join(tm_dir_path, DB_FILE)
            manifest_path = os.path.join(tm_dir_path, MANIFEST_FILE)

            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION;")
                    cursor.execute("DELETE FROM translation_units WHERE source_manifest_key = ?", (source_key,))
                    rows_deleted = cursor.rowcount
                    cursor.execute("COMMIT")

                manifest = self._read_manifest(manifest_path)
                if "imported_sources" in manifest and source_key in manifest["imported_sources"]:
                    del manifest["imported_sources"][source_key]
                    self._write_manifest(manifest_path, manifest)

                return True, _("Successfully removed {count} TM entries from source '{source}'.").format(
                    count=rows_deleted, source=source_key
                )
            except Exception as e:
                logger.error(f"Failed to remove TM source '{source_key}': {e}", exc_info=True)
                return False, str(e)

    def query_entries(
        self,
        db_path: str,
        page: int = 1,
        page_size: int = 50,
        source_key: str | None = None,
        src_lang: str | None = None,
        tgt_lang: str | None = None,
        search_term: str | None = None,
    ) -> list[dict]:
        if not db_path or not os.path.exists(db_path):
            return []

        offset = (page - 1) * page_size
        query = "SELECT id, source_text, target_text, source_lang, target_lang, source_manifest_key, created_at FROM translation_units WHERE 1=1"
        params = []

        if source_key and source_key != "All":
            query += " AND source_manifest_key = ?"
            params.append(source_key)
        if src_lang and src_lang != "All":
            query += " AND source_lang = ?"
            params.append(src_lang)
        if tgt_lang and tgt_lang != "All":
            query += " AND target_lang = ?"
            params.append(tgt_lang)
        if search_term:
            query += " AND (source_text LIKE ? OR target_text LIKE ?)"
            wildcard = f"%{search_term}%"
            params.extend([wildcard, wildcard])

        query += " ORDER BY id DESC LIMIT ? OFFSET ?"
        params.extend([page_size, offset])

        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute(query, params)
                    return [dict(row) for row in cursor.fetchall()]
            except Exception as e:
                logger.error(f"TM query failed: {e}")
                return []

    def count_entries(
        self,
        db_path: str,
        source_key: str | None = None,
        src_lang: str | None = None,
        tgt_lang: str | None = None,
        search_term: str | None = None,
    ) -> int:
        if not db_path or not os.path.exists(db_path):
            return 0

        query = "SELECT COUNT(*) FROM translation_units WHERE 1=1"
        params = []
        if source_key and source_key != "All":
            query += " AND source_manifest_key = ?"
            params.append(source_key)
        if src_lang and src_lang != "All":
            query += " AND source_lang = ?"
            params.append(src_lang)
        if tgt_lang and tgt_lang != "All":
            query += " AND target_lang = ?"
            params.append(tgt_lang)
        if search_term:
            query += " AND (source_text LIKE ? OR target_text LIKE ?)"
            wildcard = f"%{search_term}%"
            params.extend([wildcard, wildcard])

        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute(query, params)
                    return cursor.fetchone()[0]
            except Exception as e:
                logger.error(f"TM count failed: {e}")
                return 0

    def update_entry_target(self, db_path: str, entry_id: int, new_target: str) -> bool:
        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")
                    cursor.execute("UPDATE translation_units SET target_text = ? WHERE id = ?", (new_target, entry_id))
                    cursor.execute("COMMIT")
                    return True
            except Exception as e:
                logger.error(f"TM update failed: {e}")
                return False

    def update_entry_source(self, db_path: str, entry_id: int, new_source: str) -> bool:
        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")
                    cursor.execute("UPDATE translation_units SET source_text = ? WHERE id = ?", (new_source, entry_id))
                    cursor.execute("COMMIT")
                    return True
            except sqlite3.IntegrityError:
                logger.warning(f"TM update failed: Duplicate entry would result from changing source to '{new_source}'")
                return False
            except Exception as e:
                logger.error(f"TM source update failed: {e}")
                return False

    def delete_entry_by_id(self, db_path: str, entry_id: int) -> bool:
        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")
                    cursor.execute("DELETE FROM translation_units WHERE id = ?", (entry_id,))
                    cursor.execute("COMMIT")
                    return True
            except Exception as e:
                logger.error(f"TM delete failed: {e}")
                return False

    def get_distinct_languages(self, db_path: str) -> tuple[list[str], list[str]]:
        if not db_path or not os.path.exists(db_path):
            return [], []
        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT DISTINCT source_lang FROM translation_units")
                    srcs = [r[0] for r in cursor.fetchall()]
                    cursor.execute("SELECT DISTINCT target_lang FROM translation_units")
                    tgts = [r[0] for r in cursor.fetchall()]
                    return sorted(srcs), sorted(tgts)
            except Exception:
                return [], []

    def batch_update_tm(
        self,
        db_path: str,
        entries: list[dict],
        source_lang: str,
        target_lang: str,
        source_key: str,
        display_name: str,
        strategy: str = "overwrite",
    ) -> tuple[bool, str]:
        """
        Batch update TM.
        entries: List of {'source': str, 'target': str, 'action': 'new'|'overwrite'|'skip'}
        strategy: Global strategy fallback ('overwrite' or 'skip')
        """
        if not db_path or not os.path.exists(db_path):
            return False, _("Target database not found.")

        if not entries:
            return False, _("No entries to save.")

        # Prepare data
        data_to_insert = []

        for e in entries:
            action = e.get("action", strategy)
            if action == "skip":
                continue

            data_to_insert.append((source_lang, target_lang, e["source"], e["target"], source_key))

        if not data_to_insert:
            return True, _("No entries to update (all skipped).")

        with self._lock:
            try:
                with self._get_db_connection(db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("BEGIN IMMEDIATE TRANSACTION")

                    sql = """
                        INSERT OR REPLACE INTO translation_units
                        (source_lang, target_lang, source_text, target_text, source_manifest_key)
                        VALUES (?, ?, ?, ?, ?)
                    """

                    cursor.executemany(sql, data_to_insert)

                    cursor.execute(
                        "SELECT COUNT(*) FROM translation_units WHERE source_manifest_key = ?", (source_key,)
                    )
                    actual_count = cursor.fetchone()[0]

                    # Update manifest
                    manifest_path = os.path.join(os.path.dirname(db_path), MANIFEST_FILE)
                    manifest = self._read_manifest(manifest_path)

                    manifest.setdefault("imported_sources", {})[source_key] = {
                        "filepath": display_name,
                        "filesize": 0,
                        "last_modified": datetime.utcnow().isoformat() + "Z",
                        "checksum": "batch_save",
                        "import_date": datetime.utcnow().isoformat() + "Z",
                        "tu_count": actual_count,
                        "source_lang": source_lang,
                        "target_lang": target_lang,
                    }

                    self._write_manifest(manifest_path, manifest)

                    cursor.execute("COMMIT")

                return True, _("Successfully processed {count} entries.").format(count=len(data_to_insert))
            except Exception as e:
                logger.error(f"Batch TM update failed: {e}", exc_info=True)
                return False, str(e)

    def _read_manifest(self, manifest_path: str) -> dict:
        try:
            with open(manifest_path, encoding="utf-8") as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            return {"version": 1, "imported_sources": {}}

    def _write_manifest(self, manifest_path: str, manifest_data: dict):
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest_data, f, indent=4, ensure_ascii=False)

    def _get_file_stats(self, filepath: str) -> dict:
        stat = os.stat(filepath)
        normalized_filepath = filepath.replace("\\", "/")
        return {
            "filepath": normalized_filepath,
            "filesize": stat.st_size,
            "last_modified": datetime.fromtimestamp(stat.st_mtime).isoformat() + "Z",
            "checksum": self._calculate_checksum(filepath),
        }

    def _calculate_checksum(self, filepath: str, hash_algo="sha256") -> str:
        h = hashlib.new(hash_algo)
        with open(filepath, "rb") as f:
            while chunk := f.read(8192):
                h.update(chunk)
        return h.hexdigest()
